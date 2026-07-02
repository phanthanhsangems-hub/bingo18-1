import os
import json
import math
import time as _time

# Auto-load .env khi chạy local (không có DATABASE_URL trong env)
if not os.environ.get('DATABASE_URL'):
    try:
        from dotenv import load_dotenv
        load_dotenv()
    except ImportError:
        pass
import requests
from collections import Counter
from datetime import datetime, timezone
from functools import wraps
from flask import Flask, jsonify, request, send_from_directory, make_response, render_template, Response
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

import pandas as pd
import config
from database import DatabaseManager, USE_POSTGRES

# ── Khởi tạo app ─────────────────────────────────────────────
app = Flask(__name__)
# Restrict CORS: chỉ cho phép domain của bạn. Thay "*" bằng domain thật khi deploy.
_CORS_ORIGINS = os.environ.get("CORS_ORIGINS", "*")
CORS(app, origins=_CORS_ORIGINS)
db  = DatabaseManager()

# ── In-memory response cache (GET only, TTL-based) ───────────
_resp_cache: dict = {}  # full_path -> (payload_bytes, expiry_ts)

def cache_resp(ttl: int = 120):
    """Cache a GET endpoint's JSON response for `ttl` seconds per unique URL."""
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if request.method != 'GET':
                return fn(*args, **kwargs)
            key = request.full_path  # includes query string
            now = _time.time()
            hit = _resp_cache.get(key)
            if hit and now < hit[1]:
                resp = make_response(hit[0])
                resp.headers['Content-Type'] = 'application/json'
                resp.headers['X-Cache'] = 'HIT'
                return resp
            result = fn(*args, **kwargs)
            # Only cache 200 responses; result may be (Response, status) tuple
            resp_obj = result[0] if isinstance(result, tuple) else result
            if not isinstance(result, tuple) or result[1] == 200:
                try:
                    _resp_cache[key] = (resp_obj.get_data(), now + ttl)
                except Exception:
                    pass
            return result
        return wrapper
    return decorator

# ── Telegram dedup ────────────────────────────────────────────
_PROCESSED_UPDATES: set = set()
_PROCESSED_MAX = 1000

# ── Telegram per-user command throttle ────────────────────────
import _thread as _threading_mod
_TG_CMD_COOLDOWN: dict = {}   # chat_id → last_command_ts (monotonic)
_TG_CMD_COOLDOWN_SEC = 3      # min seconds between commands per user
_TG_SLOW_CMDS = {"/compare", "/voters", "/calibration", "/votertrend", "/top", "/wincal"}
_TG_SLOW_COOLDOWN_SEC = 8     # heavier commands get longer cooldown

# ── P0 Alert throttle ─────────────────────────────────────────
_last_alert_ts: float      = 0.0
_ALERT_COOLDOWN_SEC        = 600   # 10 minutes between Telegram alerts
_consecutive_db_errors: int = 0   # debounce: chỉ alert khi >= 2 lần fail liên tiếp

# ── Sync lag alert ────────────────────────────────────────────
_last_sync_alert_ts: float = 0.0
_SYNC_ALERT_COOLDOWN_SEC   = 1800  # max 1 alert/30 phút
_SYNC_LAG_THRESHOLD_MIN    = 15    # alert nếu lag > 15 phút trong giờ game

# ── Size bias alert ───────────────────────────────────────────
_last_bias_alert_ts: float = 0.0
_BIAS_ALERT_COOLDOWN_SEC   = 7200  # max 1 alert/2 hours
_BIAS_HOA_THRESHOLD        = 30.0  # alert nếu predicted HOA% > 30 (actual ~20%)

# ── P70: LON excess alert ─────────────────────────────────────
_last_lon_excess_alert_ts: float = 0.0
_LON_EXCESS_ALERT_COOLDOWN_SEC   = 1800   # max 1 alert/30 phút
_LON_EXCESS_CONSEC_THRESHOLD     = 3      # alert khi consecutive_excess >= 3

# ── Triple drought alert ───────────────────────────────────────
_last_triple_alert_ts: float = 0.0
_TRIPLE_ALERT_COOLDOWN_SEC   = 3600   # max 1 alert/1h
_TRIPLE_WARN_START           = 20     # alert đầu tiên — trước đa số triple (median=25)
_TRIPLE_WARN_STEP            = 25     # re-alert mỗi 25 kỳ (~2.5h)
_TRIPLE_DROUGHT_P90          = 95     # p90 gap lịch sử (hạn hán hiếm)
_TRIPLE_DROUGHT_P95          = 139    # p95 gap lịch sử (cực hiếm)
_triple_drought_notified_gap: int = 0  # gap lúc gửi alert gần nhất (reset khi triple xảy ra)


_CHECKPOINT_TS = '2026-05-15 16:15:00'
_CHECKPOINT_N  = 200
_CHECKPOINT_ALERT_KEY = 'checkpoint_200_reached'


def _get_checkpoint_config(cur):
    """Đọc checkpoint_ts và checkpoint_n từ system_config, fallback về hardcoded defaults."""
    try:
        cur.execute(
            "SELECT config_key, config_value FROM system_config "
            "WHERE config_key IN ('checkpoint_ts','checkpoint_n')"
        )
        cfg = {r[0]: r[1] for r in cur.fetchall()}
        ts = cfg.get('checkpoint_ts', _CHECKPOINT_TS)
        n  = int(cfg.get('checkpoint_n', _CHECKPOINT_N))
        return ts, n
    except Exception:
        return _CHECKPOINT_TS, _CHECKPOINT_N


def _save_checkpoint_config(cur, ts: str, n: int):
    """Lưu checkpoint_ts và checkpoint_n vào system_config (upsert)."""
    for key, val, desc in [
        ('checkpoint_ts', ts,    'Checkpoint start timestamp (UTC)'),
        ('checkpoint_n',  str(n), 'Checkpoint target n'),
    ]:
        cur.execute("""
            INSERT INTO system_config (config_key, config_value, description)
            VALUES (%s, %s, %s)
            ON CONFLICT (config_key) DO UPDATE
              SET config_value = EXCLUDED.config_value,
                  updated_at   = NOW()
        """, (key, val, desc))


def _check_checkpoint_alert():
    """P152: Gửi Telegram alert khi đạt 200 fresh predictions. Dùng alert_log để dedup qua restarts."""
    if not USE_POSTGRES:
        return
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM predictions WHERE created_at > %s", (_CHECKPOINT_TS,))
        n_fresh = int(cur.fetchone()[0])
        if n_fresh < _CHECKPOINT_N:
            conn.close()
            return
        cur.execute("SELECT 1 FROM alert_log WHERE alert_key = %s LIMIT 1", (_CHECKPOINT_ALERT_KEY,))
        already_sent = cur.fetchone() is not None
        if already_sent:
            conn.close()
            return
        # Fetch ml stats for the alert body
        import math as _math
        _BASELINE = 0.375
        ctrl_str = over_str = z_ctrl = None
        z_lon_str = "n/a"
        ship_p171 = False
        try:
            cur.execute("""
                SELECT
                  SUM(CASE WHEN (p.vote_breakdown->'all_votes'->>'ml') =
                               (p.vote_breakdown->>'final_size') THEN 1 ELSE 0 END),
                  SUM(CASE WHEN (p.vote_breakdown->'all_votes'->>'ml') =
                               (p.vote_breakdown->>'final_size')
                              AND pr.is_win_size THEN 1 ELSE 0 END),
                  SUM(CASE WHEN (p.vote_breakdown->'all_votes'->>'ml') <>
                               (p.vote_breakdown->>'final_size') THEN 1 ELSE 0 END),
                  SUM(CASE WHEN (p.vote_breakdown->'all_votes'->>'ml') <>
                               (p.vote_breakdown->>'final_size')
                              AND pr.is_win_size THEN 1 ELSE 0 END),
                  SUM(CASE WHEN p.vote_breakdown->'all_votes'->>'ml' = 'LON'
                              AND p.vote_breakdown->>'final_size' = 'LON' THEN 1 ELSE 0 END),
                  SUM(CASE WHEN p.vote_breakdown->'all_votes'->>'ml' = 'LON'
                              AND p.vote_breakdown->>'final_size' = 'LON'
                              AND pr.is_win_size THEN 1 ELSE 0 END)
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE p.created_at > %s AND pr.is_win_size IS NOT NULL
                  AND p.vote_breakdown IS NOT NULL
                  AND p.vote_breakdown->'all_votes'->>'ml' IS NOT NULL
            """, (_CHECKPOINT_TS,))
            r2 = cur.fetchone()
            ctrl_t, ctrl_w, over_t, over_w, lon_t, lon_w = (int(x or 0) for x in r2)
            wr_ctrl = ctrl_w / ctrl_t if ctrl_t > 0 else None
            z_ctrl_val = ((wr_ctrl - _BASELINE) / _math.sqrt(_BASELINE * (1 - _BASELINE) / ctrl_t)
                          if wr_ctrl is not None and ctrl_t >= 10 else None)
            ctrl_str = (f"{wr_ctrl*100:.1f}% (n={ctrl_t}, z={z_ctrl_val:+.2f})"
                        if z_ctrl_val is not None else f"{ctrl_w}/{ctrl_t}")
            over_str = f"{over_w/over_t*100:.1f}% (n={over_t})" if over_t > 0 else "n/a"
            if z_ctrl_val is not None and z_ctrl_val <= -2.0:
                ship_p171 = True
            if lon_t >= 10:
                wr_lon = lon_w / lon_t
                z_lon  = (wr_lon - _BASELINE) / _math.sqrt(_BASELINE * (1 - _BASELINE) / lon_t)
                z_lon_str = f"{wr_lon*100:.1f}% (n={lon_t}, z={z_lon:+.2f})"
                if z_lon <= -2.0:
                    ship_p171 = True
        except Exception:
            ctrl_str = ctrl_str or "n/a"
            over_str = over_str or "n/a"
        action_line = ("🚨 <b>ĐIỀU KIỆN ĐẠT — SHIP P171: xóa ml voter!</b>"
                       if ship_p171 else
                       "⏳ Chưa đủ z ≤ −2.0 — chạy /checkpoint kiểm tra chi tiết")
        cur.execute(
            "INSERT INTO alert_log (alert_key, message) VALUES (%s, %s)",
            (_CHECKPOINT_ALERT_KEY, f"Checkpoint reached: {n_fresh}/{_CHECKPOINT_N} fresh predictions")
        )
        conn.commit()
        conn.close()
        from telegram_bot import TelegramBot
        from zoneinfo import ZoneInfo
        now_vn = datetime.now(ZoneInfo("Asia/Ho_Chi_Minh")).strftime("%H:%M %d/%m/%Y")
        TelegramBot().send_message(
            f"✅ <b>CHECKPOINT ĐẠT {_CHECKPOINT_N} DỰ ĐOÁN!</b>\n\n"
            f"Đã có <b>{n_fresh}</b> predictions kể từ p128 deploy.\n\n"
            f"1️⃣ <b>ML-LON (pre-reg #1)</b>\n"
            f"   {z_lon_str}\n\n"
            f"2️⃣ <b>ML controls (pre-reg #2)</b>\n"
            f"   ctrl: {ctrl_str}\n"
            f"   override: {over_str}\n\n"
            f"3️⃣ <b>WR by SIZE</b> → /checkpoint\n\n"
            f"{action_line}\n"
            f"Time: {now_vn} VN"
        )
    except Exception:
        pass


def _check_sync_lag():
    """Gửi Telegram alert nếu sync lag > threshold trong giờ game. Gọi từ cron endpoint."""
    global _last_sync_alert_ts
    try:
        from zoneinfo import ZoneInfo
        vn_hour = datetime.now(ZoneInfo("Asia/Ho_Chi_Minh")).hour
        if vn_hour < 6 or vn_hour >= 22:
            return  # ngoài giờ game, không cần alert
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("SELECT draw_time FROM draw_history ORDER BY draw_number DESC LIMIT 1")
        row = cur.fetchone()
        conn.close()
        if not row:
            return
        draw_dt = row[0]
        if isinstance(draw_dt, str):
            draw_dt = datetime.strptime(draw_dt, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
        elif not getattr(draw_dt, 'tzinfo', None):
            draw_dt = draw_dt.replace(tzinfo=timezone.utc)
        lag_min = (datetime.now(timezone.utc) - draw_dt).total_seconds() / 60
        if lag_min < _SYNC_LAG_THRESHOLD_MIN:
            return
        now_t = _time.monotonic()
        if now_t - _last_sync_alert_ts < _SYNC_ALERT_COOLDOWN_SEC:
            return
        _last_sync_alert_ts = now_t
        from telegram_bot import TelegramBot
        severity = "🔴 <b>CRITICAL</b>" if lag_min >= 60 else "⚠️ <b>WARNING</b>"
        TelegramBot().send_message(
            f"{severity} — Sync Lag\n"
            f"Lag: <b>{round(lag_min)} phút</b> (ngưỡng {_SYNC_LAG_THRESHOLD_MIN} phút)\n"
            f"Hãy kiểm tra <code>sync_to_supabase.py --mode watch</code> trên máy local.\n"
            f"Time: {datetime.now(ZoneInfo('Asia/Ho_Chi_Minh')).strftime('%H:%M %d/%m/%Y')} VN"
        )
    except Exception:
        pass

def _check_lon_excess_alert(prediction_result: dict):
    """P70: Gửi Telegram alert khi consecutive_excess >= threshold (LON bias escalating)."""
    global _last_lon_excess_alert_ts
    try:
        at = prediction_result.get('adaptive_thresholds') or {}
        consec  = at.get('consecutive_excess', 0)
        tune_k  = at.get('tune_k', 0.0)
        excess  = at.get('pred_lon_excess', 0.0)
        nho_min = at.get('nho_share_min', 0.0)
        if consec < _LON_EXCESS_CONSEC_THRESHOLD:
            return
        now_t = _time.monotonic()
        if now_t - _last_lon_excess_alert_ts < _LON_EXCESS_ALERT_COOLDOWN_SEC:
            return
        _last_lon_excess_alert_ts = now_t
        draw_num = prediction_result.get('draw_number', '?')
        from telegram_bot import TelegramBot
        from zoneinfo import ZoneInfo
        TelegramBot().send_message(
            f"⚠️ <b>LON Excess Alert · #{draw_num}</b>\n"
            f"Model đang lệch LON liên tục <b>{consec}</b> chu kỳ\n"
            f"pred_lon_excess: <b>{excess*100:+.1f}%</b>\n"
            f"TUNE_K đang ở: <b>{tune_k}</b> (max 1.0)\n"
            f"nho_share_min tự động: <b>{nho_min:.0%}</b>\n"
            f"Time: {datetime.now(ZoneInfo('Asia/Ho_Chi_Minh')).strftime('%H:%M %d/%m/%Y')} VN"
        )
    except Exception:
        pass


def _check_triple_drought_alert():
    """Alert khi đã lâu chưa có triple (tất cả 3 số giống nhau).
    p90 gap = 95 kỳ, p95 gap = 139 kỳ (từ 64k draws lịch sử).
    """
    global _last_triple_alert_ts, _triple_drought_notified_gap
    try:
        now_t = _time.monotonic()
        if now_t - _last_triple_alert_ts < _TRIPLE_ALERT_COOLDOWN_SEC:
            return
        conn = db.get_connection()
        cur  = conn.cursor()
        if USE_POSTGRES:
            cur.execute("""
                SELECT draw_number,
                       (numbers::json->>0)::int AS n1,
                       (numbers::json->>1)::int AS n2,
                       (numbers::json->>2)::int AS n3
                FROM draw_history
                ORDER BY draw_number DESC
                LIMIT 200
            """)
        else:
            cur.execute("""
                SELECT draw_number,
                       json_extract(numbers, '$[0]') AS n1,
                       json_extract(numbers, '$[1]') AS n2,
                       json_extract(numbers, '$[2]') AS n3
                FROM draw_history
                ORDER BY draw_number DESC
                LIMIT 200
            """)
        rows = cur.fetchall()
        conn.close()
        if not rows:
            return
        latest_draw = rows[0][0]
        last_triple_draw = None
        for draw_num, n1, n2, n3 in rows:
            if n1 == n2 == n3:
                last_triple_draw = draw_num
                break
        if last_triple_draw is None:
            gap = 200  # ít nhất 200 kỳ
        else:
            gap = latest_draw - last_triple_draw
        # Triple vừa ra → reset
        if gap < _TRIPLE_WARN_START:
            _triple_drought_notified_gap = 0
            return
        # Re-alert chỉ khi gap tăng thêm ≥ STEP kỳ so với lần alert trước
        if gap < _triple_drought_notified_gap + _TRIPLE_WARN_STEP:
            return
        _last_triple_alert_ts = now_t
        _triple_drought_notified_gap = gap
        q_val   = 0.9722
        cum_prob = round((1 - q_val ** gap) * 100, 1)
        next10   = round((1 - q_val ** 10) * 100, 1)
        if gap >= _TRIPLE_DROUGHT_P95:
            level, note = "🚨", "cực hiếm — top 5% hạn hán lịch sử"
        elif gap >= _TRIPLE_DROUGHT_P90:
            level, note = "🔥", "hiếm — top 10% hạn hán lịch sử"
        elif gap >= 47:
            level, note = "⚠️", "trên trung bình (avg=47kỳ)"
        else:
            level, note = "🎲", "đang tích lũy (median=25kỳ)"
        # Milestone projections from NOW (memoryless)
        log_q = math.log(q_val)
        milestones = []
        for pct, label in [(50, "Median"), (75, "75%"), (90, "P90"), (95, "P95")]:
            n = math.ceil(math.log(1 - pct / 100) / log_q)
            draw_at = latest_draw + n
            hrs = n * 6 / 60
            hrs_str = f"{hrs:.1f}h" if hrs < 24 else f"{hrs/24:.1f}ngày"
            milestones.append(f"  {label}: ~{n} kỳ → <b>#{draw_at}</b> (~{hrs_str})")
        milestone_text = "\n".join(milestones)
        from telegram_bot import TelegramBot
        from zoneinfo import ZoneInfo
        TelegramBot().send_message(
            f"{level} <b>Triple Alert · {gap} kỳ chưa có triple</b>\n"
            f"━━━━━━━━━━━━━━━━\n"
            f"Kỳ triple cuối: <b>#{last_triple_draw}</b> · {note}\n"
            f"📊 Xác suất đã qua: <b>{cum_prob}%</b>  |  P(10 kỳ tới): <b>{next10}%</b>\n"
            f"\n📅 Dự kiến triple tiếp theo (từ kỳ #{latest_draw}):\n"
            f"{milestone_text}\n"
            f"\n💡 Mỗi kỳ luôn 2.78% — cân nhắc <b>Bộ 3 bất kỳ ×20</b>\n"
            f"⏱ {datetime.now(ZoneInfo('Asia/Ho_Chi_Minh')).strftime('%H:%M %d/%m/%Y')} VN"
        )
    except Exception:
        pass


# ── Rate Limiting ─────────────────────────────────────────────
# Dùng memory storage (Cloud Run stateless – mỗi instance độc lập)
limiter = Limiter(
    key_func=get_remote_address,
    app=app,
    default_limits=["200 per minute"],
    storage_uri="memory://",
)


# ── API: Scheduler / Cron ─────────────────────────────────────
@app.route('/api/fetch-latest', methods=['GET'])
@limiter.limit("20 per minute")
def fetch_latest_result():
    """Fetch kết quả mới nhất từ Vietlott và lưu vào DB. Cloud Scheduler gọi mỗi 6 phút."""
    try:
        from vietlott_fetcher import get_latest_result
        result = get_latest_result()
        if not result:
            return jsonify({"status": "error", "message": "Không lấy được dữ liệu từ Vietlott."}), 400

        draw_number = result.get("draw_number")
        numbers     = result.get("numbers")

        if not draw_number or not numbers:
            return jsonify({"status": "error", "message": "Dữ liệu không hợp lệ.", "raw": result}), 400

        # Kiểm tra đã có trong DB chưa
        df = db.get_recent_draws(1)
        if not df.empty and int(df.iloc[0]['draw_number']) == int(draw_number):
            return jsonify({"status": "skipped", "message": f"Kỳ #{draw_number} đã có trong DB."})

        # Lưu vào DB
        db.insert_draw(draw_number, numbers)
        return jsonify({
            "status": "success",
            "message": f"Đã lưu kỳ #{draw_number}",
            "draw_number": draw_number,
            "numbers": numbers
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/sync-github', methods=['GET'])
@limiter.limit("10 per minute")
def sync_from_github():
    """Sync dữ liệu từ GitHub (vietvudanh/vietlott-data). Cloud Scheduler gọi mỗi sáng."""
    try:
        import requests as req

        GITHUB_URL = "https://raw.githubusercontent.com/vietvudanh/vietlott-data/main/data/bingo18.jsonl"
        resp = req.get(GITHUB_URL, timeout=30)
        if resp.status_code != 200:
            return jsonify({"status": "error", "message": f"GitHub HTTP {resp.status_code}"}), 500

        # Parse JSONL
        records = []
        for line in resp.text.strip().splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                records.append(json.loads(line))
            except Exception:
                continue

        if not records:
            return jsonify({"status": "error", "message": "Không parse được dữ liệu từ GitHub"}), 500

        # Lấy kỳ mới nhất trong DB
        df = db.get_recent_draws(1)
        latest_in_db = int(df.iloc[0]['draw_number']) if not df.empty else 0

        # Chỉ insert kỳ mới hơn
        new_count = 0
        conn = db.get_connection()
        cur = conn.cursor()
        for rec in records:
            draw_number = rec.get("id") or rec.get("draw_number")
            numbers = rec.get("result") or rec.get("numbers")
            if not draw_number or not numbers:
                continue
            draw_number = int(draw_number)
            if draw_number <= latest_in_db:
                continue
            if isinstance(numbers, list) and len(numbers) == 3:
                total = sum(numbers)
                size = "NHO" if total <= 9 else ("HOA" if total <= 11 else "LON")
                draw_time = rec.get("date") or rec.get("draw_time") or datetime.now().isoformat()
                if USE_POSTGRES:
                    cur.execute(
                        "INSERT INTO draw_history (draw_number, numbers, draw_time, sum_value, size_category) "
                        "VALUES (%s, %s, %s, %s, %s) ON CONFLICT (draw_number) DO NOTHING",
                        (draw_number, json.dumps(numbers), draw_time, total, size)
                    )
                new_count += 1
        conn.commit()
        conn.close()

        return jsonify({
            "status": "success",
            "message": f"Đã sync {new_count} kỳ mới từ GitHub",
            "total_in_file": len(records),
            "new_inserted": new_count,
            "latest_was": latest_in_db,
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/predict', methods=['GET'])
@limiter.limit("20 per minute")
def auto_predict_cron():
    """Dành cho Cloud Scheduler gọi mỗi 6 phút — 1 cycle duy nhất mỗi lần."""
    try:
        from prediction_service import run_prediction_cycle
        result = run_prediction_cycle()
        _check_sync_lag()
        _check_lon_excess_alert(result)
        _check_checkpoint_alert()
        _check_triple_drought_alert()
        return jsonify({"status": "success", "cycles": 1, "data": [result]})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/trigger-prediction', methods=['POST'])
@limiter.limit("10 per minute")
def trigger_prediction():
    """Dự phòng – bảo mật bằng Secret Key."""
    secret = request.headers.get("X-Trigger-Secret")
    if secret != config.TRIGGER_SECRET:
        return jsonify({"error": "Unauthorized"}), 401
    try:
        from prediction_service import run_prediction_cycle
        result = run_prediction_cycle()
        return jsonify({"success": True, "prediction": result})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/last-draw-id', methods=['GET'])
@limiter.limit("60 per minute")
def last_draw_id():
    """Trả về draw_number lớn nhất trong DB — dùng bởi GitHub Actions."""
    secret = request.headers.get("X-Trigger-Secret")
    if secret != config.TRIGGER_SECRET:
        return jsonify({"error": "Unauthorized"}), 401
    try:
        draws = db.get_recent_draws(limit=1)
        last_id = draws[0]['draw_number'] if draws else 0
        return jsonify({"draw_number": last_id})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/ingest-draws', methods=['POST'])
@limiter.limit("30 per minute")
def ingest_draws():
    """
    Nhận draw data từ GitHub Actions (scrape từ Vietlott) và ghi vào DB.
    Body JSON: {"draws": [{"draw_id": X, "numbers": [a,b,c], "draw_date": "..."}]}
    """
    secret = request.headers.get("X-Trigger-Secret")
    if secret != config.TRIGGER_SECRET:
        return jsonify({"error": "Unauthorized"}), 401

    draws = request.json.get("draws", []) if request.json else []
    inserted = 0
    for draw in draws:
        try:
            draw_id = int(draw.get("draw_id") or draw.get("draw_number", 0))
            numbers = [int(n) for n in draw.get("numbers", [])]
            draw_date_str = draw.get("draw_date") or draw.get("draw_time") or ""
            if not draw_id or len(numbers) != 3:
                continue
            draw_time = None
            if draw_date_str:
                for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
                    try:
                        draw_time = datetime.strptime(draw_date_str[:19], fmt)
                        break
                    except Exception:
                        pass
            row_id = db.insert_draw(draw_id, numbers, draw_time)
            if row_id and row_id > 0:
                inserted += 1
                logger.info("ingest-draws: inserted #%d %s", draw_id, numbers)
        except Exception as e:
            logger.warning("ingest-draws: skip draw %s — %s", draw, e)

    if inserted > 0:
        try:
            from prediction_service import run_prediction_cycle
            run_prediction_cycle()
        except Exception as e:
            logger.warning("ingest-draws: prediction trigger failed: %s", e)

    return jsonify({"inserted": inserted, "received": len(draws)})


# ── PWA helpers (#51) ────────────────────────────────────────
import struct as _struct, zlib as _zlib

def _make_png(size: int, r: int, g: int, b: int) -> bytes:
    """Generate minimal solid-color PNG (no PIL needed)."""
    def _chunk(tag: bytes, data: bytes) -> bytes:
        return (_struct.pack('>I', len(data)) + tag + data +
                _struct.pack('>I', _zlib.crc32(tag + data) & 0xFFFFFFFF))
    scanline = b'\x00' + bytes([r, g, b] * size)
    raw = scanline * size
    return (b'\x89PNG\r\n\x1a\n'
            + _chunk(b'IHDR', _struct.pack('>IIBBBBB', size, size, 8, 2, 0, 0, 0))
            + _chunk(b'IDAT', _zlib.compress(raw))
            + _chunk(b'IEND', b''))

@app.route('/icon-<int:sz>.png')
def pwa_icon(sz):
    png = _make_png(min(sz, 512), 0, 229, 255)  # cyan #00e5ff
    resp = make_response(png)
    resp.headers['Content-Type'] = 'image/png'
    resp.headers['Cache-Control'] = 'public, max-age=86400'
    return resp

@app.route('/manifest.json')
def pwa_manifest():
    return jsonify({
        "name": "Bingo18 Predictor",
        "short_name": "Bingo18",
        "description": "Dự đoán kết quả xổ số Bingo18 real-time",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#0d0f14",
        "theme_color": "#00e5ff",
        "icons": [
            {"src": "/icon-192.png", "sizes": "192x192", "type": "image/png", "purpose": "any maskable"},
            {"src": "/icon-512.png", "sizes": "512x512", "type": "image/png", "purpose": "any maskable"},
        ],
    })

@app.route('/sw.js')
def service_worker():
    # Kill SW: xóa cache + tự unregister → dashboard chạy như plain web app
    sw = r"""
self.addEventListener('install', e => { self.skipWaiting(); });
self.addEventListener('activate', e => {
  e.waitUntil(
    caches.keys()
      .then(keys => Promise.all(keys.map(k => caches.delete(k))))
      .then(() => self.registration.unregister())
  );
  self.clients.claim();
});
"""
    resp = make_response(sw)
    resp.headers['Content-Type'] = 'application/javascript'
    resp.headers['Cache-Control'] = 'no-store'
    return resp


# ── Draw history search (#34) ────────────────────────────────
@app.route('/api/draw-search')
@limiter.limit("30 per minute")
def draw_search():
    """#34 Tìm kiếm lịch sử kỳ theo bộ số, size, hoặc khoảng kỳ."""
    q      = request.args.get('q', '').strip()       # VD: "1,2,3" or "3"
    size   = request.args.get('size', '').upper()    # NHO / HOA / LON
    draw_from = request.args.get('from', type=int)
    draw_to   = request.args.get('to',   type=int)
    limit_n   = min(int(request.args.get('limit', 50)), 200)
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        conditions = []
        params     = []
        if size in ('NHO', 'HOA', 'LON'):
            conditions.append("size_category = %s")
            params.append(size)
        if draw_from:
            conditions.append("draw_number >= %s")
            params.append(draw_from)
        if draw_to:
            conditions.append("draw_number <= %s")
            params.append(draw_to)
        # Number filter: any of the queried numbers must appear
        nums_filter = []
        if q:
            for tok in q.replace(' ', '').split(','):
                if tok.isdigit() and 1 <= int(tok) <= 6:
                    nums_filter.append(int(tok))
        if nums_filter:
            # numbers column is JSON array — cast and check containment
            placeholders = ','.join(['%s'] * len(nums_filter))
            conditions.append(
                f"EXISTS (SELECT 1 FROM json_array_elements_text(numbers::json) x "
                f"WHERE x::int IN ({placeholders}))"
            )
            params.extend(nums_filter)
        where = ('WHERE ' + ' AND '.join(conditions)) if conditions else ''
        cur.execute(f"""
            SELECT draw_number, numbers, size_category, sum_value, draw_time
            FROM draw_history
            {where}
            ORDER BY draw_number DESC
            LIMIT %s
        """, params + [limit_n])
        rows = cur.fetchall()
        conn.close()
        results = []
        for dn, nums_raw, sz, sv, dt in rows:
            try:
                nums = json.loads(nums_raw) if isinstance(nums_raw, str) else nums_raw
            except Exception:
                nums = []
            results.append({
                'draw_number': dn, 'numbers': nums,
                'size': sz, 'sum': sv,
                'draw_time': str(dt) if dt else None,
            })
        return jsonify({'results': results, 'count': len(results)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Number frequency by VN hour (#38) ────────────────────────
@app.route('/api/number-by-hour')
@limiter.limit("20 per minute")
def number_by_hour():
    """#38 Heatmap: VN hour × number (1-6) → frequency."""
    limit_draws = min(int(request.args.get('n', 10000)), 50000)
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute(f"""
            SELECT
                EXTRACT(HOUR FROM draw_time AT TIME ZONE 'UTC'
                                             AT TIME ZONE 'Asia/Ho_Chi_Minh')::int AS vn_hour,
                x.num::int AS number,
                COUNT(*) AS cnt
            FROM (
                SELECT draw_time, numbers FROM draw_history
                WHERE draw_time IS NOT NULL
                ORDER BY draw_number DESC LIMIT %s
            ) sub,
            json_array_elements_text(sub.numbers::json) AS x(num)
            GROUP BY vn_hour, number
            ORDER BY vn_hour, number
        """, [limit_draws])
        rows = cur.fetchall()
        conn.close()
        # Build {hour: {num: cnt}}
        matrix: dict = {}
        for hour, num, cnt in rows:
            h = int(hour)
            matrix.setdefault(h, {})[int(num)] = int(cnt)
        # Compute hour totals for normalization
        hour_totals = {h: sum(v.values()) for h, v in matrix.items()}
        return jsonify({'matrix': matrix, 'hour_totals': hour_totals, 'n': limit_draws})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Sum distribution (#43) ───────────────────────────────────
@app.route('/api/sum-distribution')
@limiter.limit("20 per minute")
def sum_distribution():
    """#43 Histogram of actual sum (3-18) vs theoretical distribution."""
    limit_n = min(int(request.args.get('n', 10000)), 100000)
    # Theoretical: number of ways to get sum k with 3 dice (1-6, with repetition)
    theory: dict = {}
    for a in range(1, 7):
        for b in range(1, 7):
            for c in range(1, 7):
                s = a + b + c
                theory[s] = theory.get(s, 0) + 1
    total_theory = sum(theory.values())  # 216
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT sum_value, COUNT(*) AS cnt
            FROM draw_history
            WHERE sum_value IS NOT NULL
            ORDER BY draw_number DESC LIMIT %s
        """, [limit_n])
        rows = cur.fetchall()
        # Use draw_history's sum_value; fallback: compute from numbers
        if not rows:
            cur.execute("""
                SELECT (SELECT SUM(x::int) FROM json_array_elements_text(numbers::json) x) AS s,
                       COUNT(*) FROM draw_history GROUP BY s
            """)
            rows = cur.fetchall()
        conn.close()
        actual: dict = {int(r[0]): int(r[1]) for r in rows if r[0] is not None}
        total_actual = sum(actual.values()) or 1
        result = []
        for s in range(3, 19):
            result.append({
                'sum': s,
                'actual_cnt': actual.get(s, 0),
                'actual_pct': round(actual.get(s, 0) / total_actual * 100, 2),
                'theory_cnt': theory.get(s, 0),
                'theory_pct': round(theory.get(s, 0) / total_theory * 100, 2),
            })
        return jsonify({'data': result, 'total_draws': total_actual})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Drawdown monitor (#45) ───────────────────────────────────
@app.route('/api/drawdown')
@limiter.limit("20 per minute")
@cache_resp(ttl=120)
def drawdown():
    """#45 Current drawdown, max drawdown, longest losing streak."""
    n = min(int(request.args.get('n', 500)), 5000)
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT p.draw_number, pr.is_win_size
            FROM predictions p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            WHERE pr.is_win_size IS NOT NULL
            ORDER BY p.draw_number DESC
            LIMIT %s
        """, [n])
        rows = cur.fetchall()
        conn.close()
        if not rows:
            return jsonify({'error': 'no_data'})
        results = [(r[0], bool(r[1])) for r in rows]  # newest first
        # Current streak from newest
        cur_win   = results[0][1]
        cur_streak = 1
        for _, w in results[1:]:
            if w == cur_win:
                cur_streak += 1
            else:
                break
        # Current drawdown (consecutive losses from newest draw)
        cur_dd = 0
        for _, w in results:
            if not w:
                cur_dd += 1
            else:
                break
        # Max losing streak in window
        max_loss = 0
        run = 0
        for _, w in results:
            if not w:
                run += 1
                max_loss = max(max_loss, run)
            else:
                run = 0
        # Max winning streak
        max_win = 0
        run = 0
        for _, w in results:
            if w:
                run += 1
                max_win = max(max_win, run)
            else:
                run = 0
        total = len(results)
        wins  = sum(1 for _, w in results if w)
        return jsonify({
            'current_streak': cur_streak,
            'current_streak_type': 'WIN' if cur_win else 'LOSS',
            'current_drawdown': cur_dd,
            'max_loss_streak': max_loss,
            'max_win_streak':  max_win,
            'window': total,
            'wins': wins,
            'losses': total - wins,
            'wr_pct': round(wins / total * 100, 1) if total else 0,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Calibration chart (#35) ──────────────────────────────────
@app.route('/api/calibration-chart')
@limiter.limit("20 per minute")
def calibration_chart():
    """#35 Bin predictions by confidence, compute actual WR per bin."""
    n = min(int(request.args.get('n', 2000)), 10000)
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT p.confidence, pr.is_win_size
            FROM predictions p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            WHERE pr.is_win_size IS NOT NULL AND p.confidence IS NOT NULL
            ORDER BY p.draw_number DESC
            LIMIT %s
        """, [n])
        rows = cur.fetchall()
        conn.close()
        if not rows:
            return jsonify({'error': 'no_data'})
        # Bin by 0.025 width (fine-grained since our range is narrow 0.30-0.55)
        from collections import defaultdict
        bins: dict = defaultdict(lambda: {'wins': 0, 'total': 0, 'conf_sum': 0.0})
        for conf, is_win in rows:
            c = float(conf)
            b = round(int(c / 0.025) * 0.025, 3)
            bins[b]['total']    += 1
            bins[b]['conf_sum'] += c
            if is_win:
                bins[b]['wins'] += 1
        result = []
        for b in sorted(bins.keys()):
            d = bins[b]
            if d['total'] >= 5:
                result.append({
                    'bin':      b,
                    'avg_conf': round(d['conf_sum'] / d['total'], 4),
                    'actual_wr':round(d['wins'] / d['total'], 4),
                    'n':        d['total'],
                })
        baseline = sum(1 for _, w in rows if w) / len(rows)
        return jsonify({'bins': result, 'baseline': round(baseline, 4), 'total': len(rows)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Model vs random chart (#39) ───────────────────────────────
@app.route('/api/model-vs-random')
@limiter.limit("20 per minute")
def model_vs_random():
    """#39 Rolling WR (20 and 50 draws) vs baseline 37.5%."""
    n = min(int(request.args.get('n', 500)), 2000)
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT p.draw_number,
                   AVG(CASE WHEN pr.is_win_size THEN 1.0 ELSE 0.0 END)
                     OVER (ORDER BY p.draw_number ROWS BETWEEN 19 PRECEDING AND CURRENT ROW) AS r20,
                   AVG(CASE WHEN pr.is_win_size THEN 1.0 ELSE 0.0 END)
                     OVER (ORDER BY p.draw_number ROWS BETWEEN 49 PRECEDING AND CURRENT ROW) AS r50
            FROM predictions p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            WHERE pr.is_win_size IS NOT NULL
            ORDER BY p.draw_number DESC
            LIMIT %s
        """, [n])
        rows = cur.fetchall()
        conn.close()
        data = [{'draw': int(r[0]),
                 'r20':  round(float(r[1]), 4) if r[1] is not None else None,
                 'r50':  round(float(r[2]), 4) if r[2] is not None else None}
                for r in reversed(rows)]
        return jsonify({'data': data, 'baseline': 0.375})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Monthly WR trend (#48) ────────────────────────────────────
@app.route('/api/monthly-wr')
@limiter.limit("20 per minute")
@cache_resp(ttl=300)
def monthly_wr():
    """#48 WR by calendar month (VN timezone)."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT
                TO_CHAR(p.created_at AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh',
                        'YYYY-MM') AS month,
                COUNT(*) FILTER (WHERE pr.is_win_size) AS wins,
                COUNT(*) AS total
            FROM predictions p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            WHERE pr.is_win_size IS NOT NULL AND p.created_at IS NOT NULL
            GROUP BY month
            ORDER BY month
        """)
        rows = cur.fetchall()
        conn.close()
        data = [{'month': r[0],
                 'wins':  int(r[1] or 0),
                 'total': int(r[2] or 0),
                 'wr':    round(int(r[1] or 0) / int(r[2] or 1) * 100, 1)}
                for r in rows if r[2]]
        return jsonify({'data': data, 'baseline': 37.5})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── HOA status (#49) ─────────────────────────────────────────
@app.route('/api/hoa-status')
@limiter.limit("20 per minute")
def hoa_status():
    """#49 HOA actual rate across time windows + P142 block recommendation."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        windows = [('7d', "NOW() - INTERVAL '7 days'"),
                   ('30d', "NOW() - INTERVAL '30 days'"),
                   ('90d', "NOW() - INTERVAL '90 days'"),
                   ('all', None)]
        result = {}
        for label, since in windows:
            where = f"WHERE draw_time >= {since}" if since else ''
            cur.execute(f"""
                SELECT
                    COUNT(*) FILTER (WHERE size_category='HOA') AS hoa_cnt,
                    COUNT(*) FILTER (WHERE size_category='NHO') AS nho_cnt,
                    COUNT(*) FILTER (WHERE size_category='LON') AS lon_cnt,
                    COUNT(*) AS total
                FROM draw_history {where}
            """)
            row = cur.fetchone()
            hoa_c, nho_c, lon_c, tot = (int(x or 0) for x in row)
            result[label] = {
                'hoa_pct': round(hoa_c / tot * 100, 2) if tot else 0,
                'nho_pct': round(nho_c / tot * 100, 2) if tot else 0,
                'lon_pct': round(lon_c / tot * 100, 2) if tot else 0,
                'n': tot,
            }
        # HOA prediction WR (if any HOA predictions slipped through)
        cur.execute("""
            SELECT COUNT(*) FILTER (WHERE pr.is_win_size) AS wins, COUNT(*) AS total
            FROM predictions p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            WHERE (p.vote_breakdown->>'final_size') = 'HOA'
              AND pr.is_win_size IS NOT NULL
        """)
        hoa_pred = cur.fetchone()
        conn.close()
        hoa_w, hoa_t = int(hoa_pred[0] or 0), int(hoa_pred[1] or 0)
        delta = result['30d']['hoa_pct'] - result['all']['hoa_pct']
        recommend = 'RECHECK' if delta >= 3.0 and result['30d']['n'] >= 100 else 'BLOCK_OK'
        return jsonify({
            'windows': result,
            'hoa_prediction_wr': round(hoa_w / hoa_t * 100, 1) if hoa_t else None,
            'hoa_predictions_total': hoa_t,
            'p142_block': True,
            'delta_30d_vs_all': round(delta, 2),
            'recommendation': recommend,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Voter overrides (#50) ─────────────────────────────────────
@app.route('/api/voter-overrides', methods=['GET'])
@limiter.limit("30 per minute")
def voter_overrides_get():
    """#50 Get current voter overrides + computed multipliers."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        # Read manual overrides from system_config
        cur.execute(
            "SELECT config_key, config_value FROM system_config "
            "WHERE config_key LIKE 'voter_override_%'"
        )
        manual = {}
        for ck, cv in cur.fetchall():
            vname = ck[len('voter_override_'):]
            try: manual[vname] = float(cv)
            except ValueError: pass
        # Read computed multipliers from voter weight cache via recent predictions
        cur.execute("""
            SELECT
                vname,
                SUM(CASE WHEN vsize = actual_size THEN 1 ELSE 0 END)::float / NULLIF(COUNT(*),0) AS wr,
                COUNT(*) AS n
            FROM (
                SELECT
                    json_object_keys(p.vote_breakdown->'all_votes') AS vname,
                    p.vote_breakdown->'all_votes'->>json_object_keys(p.vote_breakdown->'all_votes') AS vsize,
                    pr.actual_size
                FROM predictions p
                JOIN (
                    SELECT pr2.prediction_id,
                           CASE WHEN (SELECT SUM(x::int) FROM json_array_elements_text(pr2.actual_numbers::json) x) <= 9 THEN 'NHO'
                                WHEN (SELECT SUM(x::int) FROM json_array_elements_text(pr2.actual_numbers::json) x) <= 11 THEN 'HOA'
                                ELSE 'LON' END AS actual_size
                    FROM prediction_results pr2
                    WHERE pr2.actual_numbers IS NOT NULL
                ) pr ON pr.prediction_id = p.id
                WHERE p.vote_breakdown IS NOT NULL
                  AND p.draw_number > (SELECT MAX(draw_number) - 500 FROM predictions)
            ) sub
            GROUP BY vname
            HAVING COUNT(*) >= 10
            ORDER BY wr DESC
        """)
        computed = {}
        baseline = 0.375
        for vname, wr, n in cur.fetchall():
            if wr is not None:
                computed[vname] = {'wr': round(float(wr) * 100, 1), 'n': int(n),
                                   'auto_mult': round(max(0.4, min(float(wr) / baseline, 2.5)), 3)}
        conn.close()
        return jsonify({'manual': manual, 'computed': computed})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/voter-overrides', methods=['POST'])
@limiter.limit("10 per minute")
def voter_overrides_post():
    """#50 Save manual voter multiplier override to system_config."""
    data = request.get_json(force=True, silent=True) or {}
    voter = str(data.get('voter', '')).strip()
    try:
        mult = float(data.get('mult', 1.0))
    except (TypeError, ValueError):
        return jsonify({'error': 'invalid mult'}), 400
    if not voter or not (0.1 <= mult <= 5.0):
        return jsonify({'error': 'voter required, mult must be 0.1–5.0'}), 400
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        key = f'voter_override_{voter}'
        cur.execute("""
            INSERT INTO system_config (config_key, config_value, description)
            VALUES (%s, %s, %s)
            ON CONFLICT (config_key) DO UPDATE SET config_value = EXCLUDED.config_value
        """, (key, str(mult), f'#50 manual voter mult override for {voter}'))
        conn.commit()
        conn.close()
        # Invalidate voter weight cache so next prediction picks up change
        from prediction_service import _voter_weight_cache
        _voter_weight_cache.clear()
        return jsonify({'ok': True, 'voter': voter, 'mult': mult})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── DB monitor (#53) ─────────────────────────────────────────
@app.route('/api/db-monitor')
@limiter.limit("20 per minute")
def db_monitor():
    import time as _t
    try:
        conn = db.get_connection()
        cur  = conn.cursor()

        t0 = _t.time()
        cur.execute("SELECT 1")
        ping_ms = round((_t.time() - t0) * 1000, 1)

        t0 = _t.time()
        cur.execute("SELECT COUNT(*) FROM draw_history")
        total_draws = cur.fetchone()[0]
        draw_query_ms = round((_t.time() - t0) * 1000, 1)

        cur.execute("SELECT COUNT(*) FROM predictions")
        total_preds = cur.fetchone()[0]

        cur.execute("SELECT COUNT(*) FROM prediction_results")
        total_results = cur.fetchone()[0]

        cur.execute("SELECT MAX(draw_number), MAX(draw_time) FROM draw_history")
        last_draw_n, last_draw_t = cur.fetchone()

        if USE_POSTGRES:
            cur.execute("""
                SELECT pg_size_pretty(pg_total_relation_size('draw_history')),
                       pg_size_pretty(pg_total_relation_size('predictions'))
            """)
            sz_row = cur.fetchone()
            draws_size  = sz_row[0] if sz_row else None
            preds_size  = sz_row[1] if sz_row else None
        else:
            draws_size = preds_size = None

        conn.close()
        return jsonify({
            'ping_ms':       ping_ms,
            'draw_query_ms': draw_query_ms,
            'total_draws':   int(total_draws),
            'total_predictions': int(total_preds),
            'total_results': int(total_results),
            'last_draw_number': last_draw_n,
            'last_draw_time':   str(last_draw_t) if last_draw_t else None,
            'draws_table_size': draws_size,
            'preds_table_size': preds_size,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── #36 Multi-window WR ───────────────────────────────────────
@app.route('/api/multi-window-wr')
@limiter.limit("30 per minute")
def multi_window_wr():
    windows = [7, 14, 30, 50, 100]
    conn = db.get_connection()
    cur  = conn.cursor()
    if USE_POSTGRES:
        cur.execute("SELECT is_win_size FROM prediction_results ORDER BY id DESC LIMIT 100")
    else:
        cur.execute("SELECT is_win_size FROM prediction_results ORDER BY id DESC LIMIT 100")
    rows = [bool(r[0]) for r in cur.fetchall() if r[0] is not None]
    conn.close()
    result = []
    for w in windows:
        batch = rows[:w]
        wins  = sum(batch)
        total = len(batch)
        result.append({'window': w, 'wins': wins, 'total': total,
                       'win_rate': round(wins / total, 4) if total else 0})
    return jsonify({'windows': result, 'baseline': 0.375})


# ── #40 Number frequency last N draws ─────────────────────────
@app.route('/api/number-freq-n')
@limiter.limit("30 per minute")
def number_freq_n():
    n = max(7, min(int(request.args.get('n', 30)), 500))
    conn = db.get_connection()
    cur  = conn.cursor()
    if USE_POSTGRES:
        cur.execute("SELECT numbers FROM draw_history ORDER BY draw_number DESC LIMIT %s", (n,))
    else:
        cur.execute("SELECT numbers FROM draw_history ORDER BY draw_number DESC LIMIT ?", (n,))
    rows = cur.fetchall()
    conn.close()
    counts = Counter({i: 0 for i in range(1, 7)})
    total_draws = len(rows)
    for row in rows:
        nums = row[0]
        if isinstance(nums, str):
            nums = json.loads(nums)
        for num in (nums or []):
            counts[int(num)] += 1
    total_slots = total_draws * 3 or 1
    freq = [{'number': i, 'count': counts[i],
              'pct': round(counts[i] / total_slots * 100, 1)} for i in range(1, 7)]
    return jsonify({'n': total_draws, 'total_slots': total_slots, 'freq': freq})


# ── API: Dashboard ────────────────────────────────────────────
@app.route('/')
def home():
    resp = make_response(_load_dashboard())
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate'
    return resp


@app.route('/art')
def art_page():
    resp = make_response(render_template('art.html'))
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate'
    return resp


@app.route('/api/health')
def health_check():
    global _last_alert_ts
    result = {
        "status":     "ok",
        "timestamp":  datetime.now(timezone.utc).isoformat(),
        "components": {},
    }

    # ── Database health (retry 1× trước khi báo lỗi) ─────────
    db_start = _time.monotonic()
    conn = None
    _db_exc = None
    for _attempt in range(2):
        try:
            conn     = db.get_connection()
            cur      = conn.cursor()
            cur.execute("SELECT 1")
            cur.fetchone()
            _db_exc  = None
            break
        except Exception as e:
            _db_exc = e
            if conn:
                try: conn.close()
                except Exception: pass
                conn = None
            if _attempt == 0:
                _time.sleep(1)   # chờ 1s rồi thử lần 2
    if _db_exc:
        result["components"]["database"] = {
            "connected":        False,
            "query_latency_ms": None,
            "status":           "error",
        }
        result["status"] = "error"
        return jsonify(result), 503
    latency_ms = round((_time.monotonic() - db_start) * 1000)
    db_status  = "error" if latency_ms > 1000 else "ok"
    result["components"]["database"] = {
        "connected":        True,
        "query_latency_ms": latency_ms,
        "status":           db_status,
    }

    # ── Sync health ───────────────────────────────────────────
    try:
        cur.execute(
            "SELECT draw_number, draw_time FROM draw_history ORDER BY draw_number DESC LIMIT 1"
        )
        row = cur.fetchone()
        if row:
            last_draw_num, last_draw_dt = row
            # Deprecated parsing – replaced below
            # New robust parsing for draw_time
            if isinstance(last_draw_dt, str):
                try:
                    parsed_dt = datetime.strptime(last_draw_dt, "%Y-%m-%d %H:%M:%S")
                    parsed_dt = parsed_dt.replace(tzinfo=timezone.utc)
                except Exception:
                    parsed_dt = datetime.now(timezone.utc)
            else:
                parsed_dt = last_draw_dt if getattr(last_draw_dt, 'tzinfo', None) else last_draw_dt.replace(tzinfo=timezone.utc)
            lag_sec = (datetime.now(timezone.utc) - parsed_dt).total_seconds()
            lag_min = round(lag_sec / 60, 1)
            # Bingo18 chỉ quay 06:00–22:00 VN — ngoài giờ này lag cao là bình thường
            from zoneinfo import ZoneInfo
            vn_hour = datetime.now(ZoneInfo("Asia/Ho_Chi_Minh")).hour
            _is_game_off = vn_hour < 6 or vn_hour >= 22
            if _is_game_off:
                sync_status = "ok"
            else:
                sync_status = "error" if lag_min > 60 else ("warning" if lag_min > 15 else "ok")
            draw_time_str = parsed_dt.isoformat()
        else:
            last_draw_num = 0
            draw_time_str = None
            lag_min       = None
            sync_status   = "error"
        result["components"]["sync"] = {
            "last_draw_number": last_draw_num,
            "last_draw_time":   draw_time_str,
            "lag_minutes":      lag_min,
            "status":           sync_status,
        }
    except Exception:
        result["components"]["sync"] = {
            "last_draw_number": None,
            "last_draw_time":   None,
            "lag_minutes":      None,
            "status":           "error",
        }

    # ── Sequence gap check (recent 500 draws only — old gaps from watcher downtime are not actionable)
    try:
        if USE_POSTGRES:
            cur.execute("""
                SELECT COUNT(*) FROM (
                    SELECT LEAD(draw_number) OVER (ORDER BY draw_number) - draw_number - 1 AS gap
                    FROM (SELECT draw_number FROM draw_history ORDER BY draw_number DESC LIMIT 500) s
                ) g WHERE gap > 0
            """)
        else:
            cur.execute("""
                SELECT COUNT(*) FROM (
                    SELECT LEAD(draw_number) OVER (ORDER BY draw_number) - draw_number - 1 AS gap
                    FROM (SELECT draw_number FROM draw_history ORDER BY draw_number DESC LIMIT 500)
                ) WHERE gap > 0
            """)
        gap_count = (cur.fetchone() or [0])[0] or 0
        result["components"]["sequence"] = {
            "gaps_in_last_500": gap_count,
            "status": "warning" if gap_count > 0 else "ok",
        }
    except Exception:
        result["components"]["sequence"] = {"gaps_in_last_500": None, "status": "ok"}

    # ── Model health ──────────────────────────────────────────
    _MS_THRESHOLD   = 3.0   # % multiset exact match
    _SIZE_THRESHOLD = 25.0  # % size category match (baseline ~33%, alert only if significantly below)
    try:
        cur.execute(
            "SELECT draw_number, full_time_vietnam FROM predictions_vn ORDER BY draw_number DESC LIMIT 1"
        )
        row = cur.fetchone()
        last_pred_num = row[0] if row else 0
        pred_time_str = str(row[1]) if row and row[1] else None

        # Both win rates over last 100 evaluated predictions
        cur.execute("""
            SELECT COUNT(*),
                   COALESCE(SUM(CASE WHEN is_win      THEN 1 ELSE 0 END), 0),
                   COALESCE(SUM(CASE WHEN is_win_size THEN 1 ELSE 0 END), 0)
            FROM (
                SELECT is_win, is_win_size FROM prediction_results
                WHERE actual_numbers IS NOT NULL
                ORDER BY created_at DESC LIMIT 100
            ) sub
        """)
        row100 = cur.fetchone()
        if row100 and row100[0] > 0:
            n = row100[0]
            ms_rate   = round(row100[1] / n * 100, 1)
            size_rate = round(row100[2] / n * 100, 1)
        else:
            ms_rate = size_rate = None

        if ms_rate is None:
            model_status = "warning"
        elif ms_rate < _MS_THRESHOLD or size_rate < _SIZE_THRESHOLD:
            model_status = "error"
        else:
            model_status = "ok"

        result["components"]["model"] = {
            "last_prediction_number":      last_pred_num,
            "last_prediction_time":        pred_time_str,
            "multiset_win_rate_last100":   ms_rate,
            "size_win_rate_last100":       size_rate,
            "multiset_threshold":          _MS_THRESHOLD,
            "size_threshold":              _SIZE_THRESHOLD,
            "status":                      model_status,
        }
    except Exception:
        result["components"]["model"] = {
            "last_prediction_number":    None,
            "last_prediction_time":      None,
            "multiset_win_rate_last100": None,
            "size_win_rate_last100":     None,
            "multiset_threshold":        _MS_THRESHOLD,
            "size_threshold":            _SIZE_THRESHOLD,
            "status":                    "warning",
        }

    # ── Prediction staleness check ────────────────────────────
    try:
        last_draw_num = result["components"].get("sync", {}).get("last_draw_number") or 0
        last_pred_num = result["components"].get("model", {}).get("last_prediction_number") or 0
        pred_gap      = max(0, last_draw_num - last_pred_num)

        from zoneinfo import ZoneInfo as _ZI
        _vn_hour = datetime.now(_ZI("Asia/Ho_Chi_Minh")).hour
        _game_on = 6 <= _vn_hour < 22

        if not _game_on:
            gap_status = "ok"
        elif pred_gap >= 5:
            gap_status = "error"
        elif pred_gap >= 2:
            gap_status = "warning"
        else:
            gap_status = "ok"

        result["components"]["prediction_gap"] = {
            "last_draw_number":      last_draw_num,
            "last_prediction_number": last_pred_num,
            "gap_draws":             pred_gap,
            "game_hours_active":     _game_on,
            "status":                gap_status,
        }
    except Exception:
        pass

    conn.close()

    # ── Aggregate overall status ──────────────────────────────
    statuses = [c["status"] for c in result["components"].values()]
    if "error" in statuses:
        result["status"] = "error"
    elif "warning" in statuses:
        result["status"] = "warning"

    # ── Size bias alert (background, non-blocking) ───────────
    try:
        _check_size_bias()
    except Exception:
        pass

    # ── Throttled Telegram alert on ERROR (debounce: 2 lần liên tiếp) ──
    global _consecutive_db_errors
    if result["status"] == "error":
        _consecutive_db_errors += 1
    else:
        _consecutive_db_errors = 0
    if result["status"] == "error" and _consecutive_db_errors >= 2:
        now_t = _time.monotonic()
        if now_t - _last_alert_ts > _ALERT_COOLDOWN_SEC:
            _last_alert_ts = now_t
            try:
                from telegram_bot import TelegramBot
                bad  = [k for k, v in result["components"].items() if v["status"] == "error"]
                gap  = result["components"].get("prediction_gap", {})
                extra = (f"\n⚠️ Prediction gap: {gap.get('gap_draws')} kỳ "
                         f"(draw #{gap.get('last_draw_number')} vs pred #{gap.get('last_prediction_number')})"
                         if gap.get("status") == "error" else "")
                TelegramBot().send_message(
                    f"🚨 <b>Bingo18 P0 Alert</b>\n"
                    f"Status: ERROR\n"
                    f"Components: {', '.join(bad)}"
                    f"{extra}\n"
                    f"Time: {result['timestamp']}"
                )
            except Exception:
                pass

    return jsonify(result), (503 if result["status"] == "error" else 200)


@app.route('/api/recent_draws')
@limiter.limit("60 per minute")
def get_recent_draws_api():
    try:
        limit = int(request.args.get('limit', 15))
        df    = db.get_recent_draws(limit)
        return df.to_json(orient='records')
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/draw-gaps')
@limiter.limit("20 per minute")
def draw_gaps():
    """P134: Detect holes in draw_number sequence (draws missed by sync_to_supabase.py)."""
    try:
        look_back = min(int(request.args.get('n', 5000)), 20000)
        conn = db.get_connection()
        cur  = conn.cursor()
        if USE_POSTGRES:
            cur.execute("""
                SELECT gap_start, gap_end, gap_size
                FROM (
                    SELECT draw_number + 1                              AS gap_start,
                           LEAD(draw_number) OVER (ORDER BY draw_number) - 1 AS gap_end,
                           LEAD(draw_number) OVER (ORDER BY draw_number) - draw_number - 1 AS gap_size
                    FROM (
                        SELECT draw_number FROM draw_history
                        ORDER BY draw_number DESC LIMIT %s
                    ) sub
                ) g
                WHERE gap_size > 0
                ORDER BY gap_start DESC
                LIMIT 50
            """, (look_back,))
        else:
            cur.execute("""
                SELECT draw_number + 1 AS gap_start,
                       next_dn - 1    AS gap_end,
                       next_dn - draw_number - 1 AS gap_size
                FROM (
                    SELECT draw_number,
                           LEAD(draw_number) OVER (ORDER BY draw_number) AS next_dn
                    FROM (SELECT draw_number FROM draw_history
                          ORDER BY draw_number DESC LIMIT ?)
                ) g
                WHERE gap_size > 0
                ORDER BY gap_start DESC
                LIMIT 50
            """, (look_back,))
        gaps = [{"gap_start": r[0], "gap_end": r[1], "gap_size": r[2]}
                for r in cur.fetchall()]
        total_missing = sum(g["gap_size"] for g in gaps)

        cur.execute("SELECT MIN(draw_number), MAX(draw_number), COUNT(*) FROM draw_history")
        row = cur.fetchone()
        conn.close()

        expected = (row[1] - row[0] + 1) if row and row[0] and row[1] else None
        actual   = row[2] if row else None
        coverage = round(actual / expected * 100, 2) if expected else None

        return jsonify({
            "gaps":          gaps,
            "total_missing": total_missing,
            "gap_count":     len(gaps),
            "draw_min":      row[0] if row else None,
            "draw_max":      row[1] if row else None,
            "draw_count":    actual,
            "expected":      expected,
            "coverage_pct":  coverage,
            "look_back":     look_back,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/prediction-coverage')
@limiter.limit("20 per minute")
def prediction_coverage():
    """P135: % draws có prediction + evaluation coverage."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        if USE_POSTGRES:
            cur.execute("""
                SELECT
                    (SELECT COUNT(*) FROM draw_history)                          AS total_draws,
                    (SELECT COUNT(DISTINCT draw_number) FROM predictions)        AS predicted_draws,
                    (SELECT COUNT(*) FROM prediction_results
                     WHERE actual_numbers IS NOT NULL)                           AS evaluated_draws,
                    (SELECT MIN(draw_number) FROM draw_history)                  AS draw_min,
                    (SELECT MAX(draw_number) FROM draw_history)                  AS draw_max,
                    (SELECT MAX(draw_number) FROM predictions)                   AS last_pred_draw,
                    (SELECT COUNT(*) FROM predictions WHERE draw_number >
                        (SELECT COALESCE(MAX(p2.draw_number),0) FROM predictions p2
                         JOIN prediction_results pr ON pr.prediction_id = p2.id
                         WHERE pr.actual_numbers IS NOT NULL))                   AS pending_eval
            """)
        else:
            cur.execute("""
                SELECT
                    (SELECT COUNT(*) FROM draw_history)                        AS total_draws,
                    (SELECT COUNT(DISTINCT draw_number) FROM predictions)      AS predicted_draws,
                    (SELECT COUNT(*) FROM prediction_results
                     WHERE actual_numbers IS NOT NULL)                         AS evaluated_draws,
                    (SELECT MIN(draw_number) FROM draw_history)                AS draw_min,
                    (SELECT MAX(draw_number) FROM draw_history)                AS draw_max,
                    (SELECT MAX(draw_number) FROM predictions)                 AS last_pred_draw,
                    0                                                          AS pending_eval
            """)
        row = cur.fetchone()
        conn.close()
        total, predicted, evaluated, draw_min, draw_max, last_pred, pending = row

        pred_cov  = round(predicted / total * 100, 1)  if total  else 0
        eval_cov  = round(evaluated / total * 100, 1)  if total  else 0
        backlog   = max(0, (draw_max or 0) - (last_pred or 0))

        return jsonify({
            "total_draws":      total,
            "predicted_draws":  predicted,
            "evaluated_draws":  evaluated,
            "draw_min":         draw_min,
            "draw_max":         draw_max,
            "last_pred_draw":   last_pred,
            "pending_eval":     pending,
            "backlog_draws":    backlog,
            "pred_coverage_pct": pred_cov,
            "eval_coverage_pct": eval_cov,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/statistics')
@limiter.limit("60 per minute")
def get_statistics():
    try:
        stats = db.get_statistics()
        # Bổ sung today_draws và today_win_rate mà dashboard cần
        conn = db.get_connection()
        try:
            cur = conn.cursor()
            today = datetime.now().strftime('%Y-%m-%d')
            if db.__class__.__name__ == 'DatabaseManager':
                from database import USE_POSTGRES
            else:
                USE_POSTGRES = False
            from database import USE_POSTGRES as _USE_PG
            if _USE_PG:
                cur.execute(
                    "SELECT COUNT(*) FROM draw_history WHERE draw_time::date = CURRENT_DATE")
            else:
                cur.execute(
                    "SELECT COUNT(*) FROM draw_history WHERE date(draw_time) = date('now')")
            today_draws = cur.fetchone()[0] or 0

            # Win rate hôm nay (so sánh prediction vs actual)
            if _USE_PG:
                cur.execute("""
                    SELECT COUNT(*), SUM(CASE WHEN pr.is_win THEN 1 ELSE 0 END)
                    FROM prediction_results pr
                    WHERE pr.created_at::date = CURRENT_DATE
                """)
            else:
                cur.execute("""
                    SELECT COUNT(*), SUM(CASE WHEN is_win THEN 1 ELSE 0 END)
                    FROM prediction_results
                    WHERE date(created_at) = date('now')
                """)
            row = cur.fetchone()
            today_total = row[0] or 0
            today_wins  = row[1] or 0
            today_win_rate = (today_wins / today_total) if today_total > 0 else None

            # Win rate multiset vs size (overall)
            cur.execute("""
                SELECT COUNT(*),
                       COALESCE(SUM(CASE WHEN is_win      THEN 1 ELSE 0 END), 0),
                       COALESCE(SUM(CASE WHEN is_win_size THEN 1 ELSE 0 END), 0)
                FROM prediction_results
                WHERE actual_numbers IS NOT NULL
            """)
            row2 = cur.fetchone()
            total_r   = row2[0] or 0
            wins_ms   = row2[1] or 0
            wins_size = row2[2] or 0
        finally:
            conn.close()

        stats['today_draws']       = today_draws
        stats['today_win_rate']    = today_win_rate
        stats['win_rate_multiset'] = round(wins_ms   / total_r, 4) if total_r else None
        stats['win_rate_size']     = round(wins_size / total_r, 4) if total_r else None
        return jsonify(stats)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/hot_cold_numbers')
def get_hot_cold_numbers():
    try:
        window = int(request.args.get('window', 50))
        return jsonify(db.get_hot_cold_numbers(window))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/number_frequency')
def get_number_frequency():
    try:
        window = int(request.args.get('window', 100))
        return jsonify(db.get_number_frequency(window))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/cold-streaks')
def cold_streaks():
    """Kỳ chưa ra cho số 1-6 và tất cả bộ 3 số (sorted combo)."""
    try:
        conn = db.get_connection()
        cur = conn.cursor()

        if USE_POSTGRES:
            cur.execute("""
                WITH draws AS (
                  SELECT draw_number,
                    (SELECT array_agg(v::int) FROM jsonb_array_elements_text(numbers::jsonb) v) AS nums
                  FROM draw_history
                ),
                exploded AS (SELECT draw_number, unnest(nums) AS num FROM draws),
                last_seen AS (
                  SELECT num, MAX(draw_number) AS last_draw
                  FROM exploded WHERE num BETWEEN 1 AND 6 GROUP BY num
                ),
                mx AS (SELECT MAX(draw_number) AS v FROM draw_history)
                SELECT ls.num, mx.v - ls.last_draw AS streak
                FROM last_seen ls, mx ORDER BY num
            """)
            num_rows = cur.fetchall()

            # sorted canonical — 1-1-4 == 4-1-1 == 1-4-1 (nhóm hoán vị cùng cụm)
            cur.execute("""
                WITH draws AS (
                  SELECT draw_number,
                    array_to_string(
                      ARRAY(SELECT v::int FROM jsonb_array_elements_text(numbers::jsonb) v ORDER BY v::int),
                      ''
                    ) AS combo
                  FROM draw_history
                ),
                combo_stats AS (
                  SELECT combo, COUNT(*) AS freq, MAX(draw_number) AS last_draw
                  FROM draws GROUP BY combo
                ),
                mx AS (SELECT MAX(draw_number) AS v FROM draw_history)
                SELECT cs.combo, mx.v - cs.last_draw AS streak, cs.freq
                FROM combo_stats cs, mx
                ORDER BY streak DESC
            """)
            combo_rows = cur.fetchall()
        else:
            cur.execute("SELECT draw_number, numbers FROM draw_history ORDER BY draw_number DESC")
            all_draws = cur.fetchall()
            max_draw = all_draws[0][0] if all_draws else 0
            num_last, combo_last, combo_freq = {}, {}, {}
            for draw_num, nums_raw in all_draws:
                try:
                    nums = json.loads(nums_raw)  # preserve draw order
                except Exception:
                    continue
                for n in nums:
                    if n not in num_last:
                        num_last[n] = draw_num
                key = ''.join(str(n) for n in sorted(nums))
                combo_freq[key] = combo_freq.get(key, 0) + 1
                if key not in combo_last:
                    combo_last[key] = draw_num
            num_rows = [(n, max_draw - num_last.get(n, max_draw)) for n in range(1, 7)]
            combo_rows = sorted(
                [(c, max_draw - v, combo_freq.get(c, 0)) for c, v in combo_last.items()],
                key=lambda x: -x[1]
            )

        conn.close()
        return jsonify({
            "numbers": {str(r[0]): r[1] for r in num_rows},
            "combos":  [{"combo": r[0], "streak": r[1], "freq": r[2]} for r in combo_rows],
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/transition-stats')
def get_transition_stats():
    """P(next_size | prev_sum) and top-3 next sums, computed from full draw history."""
    try:
        from prediction_service import _transition_cache, _query_transition_probs
        # Use live cache if available, else query fresh
        cache = _transition_cache
        if cache and cache.get('probs'):
            probs    = cache['probs']
            top_sums = cache.get('top_sums', {})
            loaded_at = cache.get('loaded_at', 0)
        else:
            probs, top_sums = _query_transition_probs(db)
            loaded_at = 0
        return jsonify({
            "probs":     {str(k): v for k, v in probs.items()},
            "top_sums":  {str(k): v for k, v in top_sums.items()},
            "loaded_at": loaded_at,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/multi-preview')
@limiter.limit("30 per minute")
def multi_preview():
    """P119: Current prediction (N+1) + Markov-projected SIZE dist for N+2 and N+3."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()

        # ── 1. Current prediction (N+1) from DB ──
        cur.execute(
            "SELECT predicted_numbers, confidence, model_name, draw_number, vote_breakdown "
            "FROM predictions ORDER BY draw_number DESC LIMIT 1"
        )
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "no predictions"}), 404

        pred_nums_raw, conf, model_name, draw_n1, vb_raw = row
        try:
            pred_nums = json.loads(pred_nums_raw) if isinstance(pred_nums_raw, str) else pred_nums_raw
        except Exception:
            pred_nums = []
        pred_sum  = sum(pred_nums) if pred_nums else 0
        pred_size = 'NHO' if pred_sum <= 9 else ('HOA' if pred_sum <= 11 else 'LON')

        # Vote breakdown SIZE weights for N+1 probabilities
        n1_probs = {'NHO': None, 'HOA': None, 'LON': None}
        ema_flipped = False
        ema_fracs = {}
        try:
            vb = json.loads(vb_raw) if isinstance(vb_raw, str) else (vb_raw or {})
            sw = vb.get('size_weights', {})
            total_w = sum(sw.values()) if sw else 0
            if total_w > 0:
                n1_probs = {sz: round(sw.get(sz, 0) / total_w, 4) for sz in ('NHO', 'HOA', 'LON')}
            ema_flipped = bool(vb.get('ema_flipped', False))
            ema_fracs   = vb.get('size_weights_ema', {})
        except Exception:
            pass

        # ── 2. SIZE→SIZE transition matrix from last 2000 draws ──
        if USE_POSTGRES:
            cur.execute("""
                WITH sized AS (
                    SELECT draw_number, size_category AS sz
                    FROM draw_history
                    ORDER BY draw_number DESC LIMIT 2000
                )
                SELECT a.sz AS prev_sz, b.sz AS next_sz, COUNT(*)::int AS cnt
                FROM sized a
                JOIN sized b ON b.draw_number = a.draw_number + 1
                GROUP BY a.sz, b.sz
            """)
        else:
            cur.execute("""
                WITH sized AS (
                    SELECT draw_number, size_category AS sz
                    FROM draw_history
                    ORDER BY draw_number DESC LIMIT 2000
                )
                SELECT a.sz, b.sz, COUNT(*) AS cnt
                FROM sized a
                JOIN sized b ON b.draw_number = a.draw_number + 1
                GROUP BY a.sz, b.sz
            """)

        # Build SIZE→SIZE transition: trans[from_sz][to_sz] = probability
        from collections import defaultdict
        raw: dict = defaultdict(lambda: {'NHO': 0, 'HOA': 0, 'LON': 0})
        for prev_sz, next_sz, cnt in cur.fetchall():
            if prev_sz in raw and next_sz in ('NHO', 'HOA', 'LON'):
                raw[prev_sz][next_sz] += cnt
        trans: dict = {}
        for sz in ('NHO', 'HOA', 'LON'):
            total = sum(raw[sz].values())
            trans[sz] = {k: round(v / total, 4) if total else 1/3 for k, v in raw[sz].items()}

        # ── 3. Project N+2 and N+3 ──
        def apply_transition(probs: dict) -> dict:
            out = {'NHO': 0.0, 'HOA': 0.0, 'LON': 0.0}
            for from_sz, p_from in probs.items():
                if p_from is None: continue
                for to_sz, p_trans in trans.get(from_sz, {}).items():
                    out[to_sz] += p_from * p_trans
            total = sum(out.values())
            return {k: round(v / total, 4) if total else 1/3 for k, v in out.items()}

        # If vote breakdown unavailable, bootstrap from predicted SIZE
        if any(v is None for v in n1_probs.values()):
            n1_probs = {sz: (0.7 if sz == pred_size else 0.15) for sz in ('NHO', 'HOA', 'LON')}

        n2_probs = apply_transition(n1_probs)
        n3_probs = apply_transition(n2_probs)
        conn.close()

        def top_size(probs): return max(probs, key=probs.get)

        return jsonify({
            'n1': {
                'draw_number': draw_n1,
                'predicted_numbers': pred_nums,
                'predicted_size': pred_size,
                'confidence': round(float(conf), 4) if conf else None,
                'model_name': model_name,
                'size_probs': n1_probs,
                'confidence_label': 'HIGH',
                'ema_flipped': ema_flipped,
                'ema_fracs': ema_fracs,
            },
            'n2': {
                'draw_number': draw_n1 + 1,
                'predicted_size': top_size(n2_probs),
                'size_probs': n2_probs,
                'confidence_label': 'LOW',
                'method': 'Markov¹',
            },
            'n3': {
                'draw_number': draw_n1 + 2,
                'predicted_size': top_size(n3_probs),
                'size_probs': n3_probs,
                'confidence_label': 'VERY LOW',
                'method': 'Markov²',
            },
            'transition': trans,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/next-prediction')
def next_prediction_hyphen():
    return multi_preview()


@app.route('/api/next_prediction')
def get_next_prediction():
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute(
            "SELECT predicted_numbers, confidence, model_name, draw_number, "
            "full_time_vietnam, display_time_vietnam, vote_breakdown "
            "FROM predictions_vn ORDER BY draw_number DESC LIMIT 1")
        row = conn.cursor().fetchone() if False else cur.fetchone()
        conn.close()
        if not row:
            return jsonify({})

        raw_conf   = float(row[1]) if row[1] else 0.0
        model_name = row[2]
        vb_raw     = row[6]

        try:
            vb = json.loads(vb_raw) if isinstance(vb_raw, str) else vb_raw
        except Exception:
            vb = None

        # Apply calibration: replace raw model score with historical win rate,
        # bucketed by vote_share (consensus strength) when available.
        try:
            from calibration import get_calibrator
            calibrator = get_calibrator(db)
            vote_share = (vb or {}).get('vote_share', 0.5)
            win_prob, cal_meta = calibrator.calibrate_by_vote_share(vote_share, model_name, raw_conf)
        except Exception:
            win_prob, cal_meta = raw_conf, {}

        return jsonify({
            "predicted_numbers":    json.loads(row[0]) if isinstance(row[0], str) else row[0],
            "confidence":           win_prob,
            "raw_confidence":       raw_conf,
            "calibration":          cal_meta,
            "is_confident":         cal_meta.get("is_confident", False),
            "model_name":           model_name,
            "draw_number":          row[3],
            "prediction_time":      str(row[4]) if row[4] else None,
            "display_time_vietnam": str(row[5]) if row[5] else None,
            "vote_breakdown":       vb,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/calibration')
def get_calibration_stats():
    """Trả về win rate thực tế của từng model — nguồn gốc của calibrated confidence."""
    try:
        from calibration import get_calibrator, SIZE_WIN_BASELINE
        calibrator = get_calibrator(db)
        result = {}
        for model_name in calibrator._rates:
            rates  = calibrator._rates[model_name]
            counts = calibrator._counts[model_name]
            cal, _ = calibrator.calibrate(model_name, 0.5)
            result[model_name] = {
                "calibrated_win_prob": round(cal, 4),
                "win_rate_last_50":    round(rates.get("last_50")  or 0, 4),
                "win_rate_last_100":   round(rates.get("last_100") or 0, 4),
                "win_rate_all_time":   round(rates.get("all_time") or 0, 4),
                "n_predictions":       counts.get("all_time", 0),
            }
        vote_share_buckets = {}
        for bucket in calibrator._vs_rates:
            rates  = calibrator._vs_rates[bucket]
            counts = calibrator._vs_counts[bucket]
            cal, _ = calibrator.calibrate_by_vote_share(
                {"weak": 0.20, "low": 0.45, "moderate": 0.55, "strong": 0.65, "dominant": 0.85}.get(bucket, 0.5),
                "majority_vote", 0.5)
            vote_share_buckets[bucket] = {
                "calibrated_win_prob": round(cal, 4),
                "win_rate_last_50":    round(rates.get("last_50")  or 0, 4),
                "win_rate_last_100":   round(rates.get("last_100") or 0, 4),
                "win_rate_all_time":   round(rates.get("all_time") or 0, 4),
                "n_predictions":       counts.get("all_time", 0),
            }
        return jsonify({
            "models":             result,
            "vote_share_buckets": vote_share_buckets,
            "random_baseline":    SIZE_WIN_BASELINE,
            "note":               "confidence trong /api/next_prediction = calibrated_win_prob (bucketed by vote_share)",
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/predictions')
@limiter.limit("60 per minute")
def get_predictions_history():
    """
    Trả về lịch sử dự đoán kèm kết quả thực tế.
    Dashboard gọi endpoint này để hiển thị bảng WIN/LOSS.
    FIX: LIMIT ? → LIMIT %s (PostgreSQL dùng %s, không phải ?)
    """
    try:
        limit = max(1, min(int(request.args.get('limit', 20)), 200))
        conn  = db.get_connection()
        cur   = conn.cursor()
        ph    = db._ph()   # '%s' cho Postgres, '?' cho SQLite
        cur.execute(f"""
            SELECT p.draw_number,
                   p.predicted_numbers,
                   p.model_name,
                   p.confidence,
                   p.full_time_vietnam,
                   pr.actual_numbers,
                   pr.match_count,
                   pr.is_win,
                   pr.is_win_size
            FROM predictions_vn p
            LEFT JOIN prediction_results pr ON pr.prediction_id = p.id
            ORDER BY p.draw_number DESC
            LIMIT {ph}
        """, (limit,))
        rows = cur.fetchall()
        conn.close()
        result = []
        for row in rows:
            result.append({
                "draw_number":       row[0],
                "predicted_numbers": json.loads(row[1]) if isinstance(row[1], str) else row[1],
                "model_name":        row[2],
                "confidence":        row[3],
                "prediction_time":   str(row[4]) if row[4] else None,
                "actual_numbers":    json.loads(row[5]) if isinstance(row[5], str) and row[5] else (row[5] or []),
                "match_count":       row[6],
                "is_win":            bool(row[7]) if row[7] is not None else None,
                "is_win_size":       bool(row[8]) if row[8] is not None else None,
            })
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/win-rate-history')
@limiter.limit("30 per minute")
def win_rate_history():
    """Win rate theo ngày trong N ngày gần nhất."""
    try:
        days = max(1, min(int(request.args.get('days', 7)), 90))
        conn = db.get_connection()
        cur  = conn.cursor()

        if USE_POSTGRES:
            cur.execute("""
                SELECT DATE(p.prediction_time)                                  AS day,
                       COUNT(*)                                                 AS total,
                       COALESCE(SUM(CASE WHEN pr.is_win THEN 1 ELSE 0 END), 0) AS wins
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE p.prediction_time >= NOW() - INTERVAL '1 day' * %s
                GROUP BY DATE(p.prediction_time)
                ORDER BY day ASC
            """, (days,))
        else:
            cur.execute("""
                SELECT date(p.prediction_time)                                  AS day,
                       COUNT(*)                                                 AS total,
                       COALESCE(SUM(CASE WHEN pr.is_win THEN 1 ELSE 0 END), 0) AS wins
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE p.prediction_time >= datetime('now', ? || ' days')
                GROUP BY date(p.prediction_time)
                ORDER BY day ASC
            """, (f'-{days}',))
        day_rows = cur.fetchall()

        cur.execute("""
            SELECT COUNT(*),
                   COALESCE(SUM(CASE WHEN is_win THEN 1 ELSE 0 END), 0)
            FROM prediction_results
        """)
        overall = cur.fetchone()
        conn.close()

        history = []
        for day, total, wins in day_rows:
            wins = int(wins or 0)
            history.append({
                "date":     str(day),
                "total":    total,
                "wins":     wins,
                "win_rate": round(wins / total * 100, 1) if total else 0.0,
            })

        total_all = overall[0] or 0
        wins_all  = int(overall[1] or 0)
        return jsonify({
            "history":          history,
            "days":             days,
            "overall_total":    total_all,
            "overall_wins":     wins_all,
            "overall_win_rate": round(wins_all / total_all * 100, 1) if total_all else 0.0,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/size-distribution')
@limiter.limit("30 per minute")
def size_distribution():
    """P60: Predicted vs actual SIZE distribution from evaluated draws (aligned window).
    Uses JOIN so both come from the same set of draws."""
    try:
        n    = min(int(request.args.get('n', 100)), 1000)
        conn = db.get_connection()
        cur  = conn.cursor()

        if USE_POSTGRES:
            cur.execute("""
                SELECT
                    pred_cat, actual_cat, COUNT(*) AS n
                FROM (
                    SELECT
                        CASE
                            WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 9  THEN 'NHO'
                            WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 11 THEN 'HOA'
                            ELSE 'LON'
                        END AS pred_cat,
                        CASE
                            WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 9  THEN 'NHO'
                            WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 11 THEN 'HOA'
                            ELSE 'LON'
                        END AS actual_cat
                    FROM predictions p
                    JOIN prediction_results pr ON pr.prediction_id = p.id
                    WHERE pr.actual_numbers IS NOT NULL
                    ORDER BY p.draw_number DESC LIMIT %s
                ) sub
                GROUP BY pred_cat, actual_cat
            """, (n,))
        else:
            cur.execute("""
                SELECT 'NHO' AS pred_cat, 'NHO' AS actual_cat, 0 AS n
            """)  # SQLite stub
        rows = cur.fetchall()
        conn.close()

        pred_cnt   = {'NHO': 0, 'HOA': 0, 'LON': 0}
        actual_cnt = {'NHO': 0, 'HOA': 0, 'LON': 0}
        total = 0
        for pred_cat, actual_cat, cnt in rows:
            pred_cnt[pred_cat]   = pred_cnt.get(pred_cat, 0) + cnt
            actual_cnt[actual_cat] = actual_cnt.get(actual_cat, 0) + cnt
            total += cnt

        tp = sum(pred_cnt.values()) or 1
        ta = sum(actual_cnt.values()) or 1
        baselines = {'NHO': 37.5, 'HOA': 25.0, 'LON': 37.5}

        dist = {}
        for cat in ['NHO', 'HOA', 'LON']:
            pp = round(pred_cnt.get(cat, 0) / tp * 100, 1)
            ap = round(actual_cnt.get(cat, 0) / ta * 100, 1)
            dist[cat] = {
                'actual_n':   actual_cnt.get(cat, 0),
                'actual_pct': ap,
                'pred_n':     pred_cnt.get(cat, 0),
                'pred_pct':   pp,
                'baseline':   baselines[cat],
                'pred_excess': round(pp - baselines[cat], 1),
            }
        return jsonify({'n': n, 'evaluated': total, 'distribution': dist,
                        'total_actual': ta, 'total_predicted': tp})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/calibration-report')
@limiter.limit("20 per minute")
def calibration_report():
    """P61: Brier score + ECE + reliability diagram data from evaluated draws.

    Pulls confidence + is_win_size from predictions JOIN prediction_results,
    bins confidence into 10 buckets, computes per-bin mean_conf / actual_wr / count.
    Brier score = mean((conf - is_win)^2).  ECE = weighted mean |mean_conf - actual_wr|.
    """
    try:
        n    = min(int(request.args.get('n', 500)), 2000)
        conn = db.get_connection()
        cur  = conn.cursor()

        if USE_POSTGRES:
            cur.execute("""
                SELECT p.confidence, pr.is_win_size
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE p.confidence IS NOT NULL
                  AND pr.is_win_size IS NOT NULL
                ORDER BY p.draw_number DESC
                LIMIT %s
            """, (n,))
        else:
            cur.execute("SELECT 0.5 AS confidence, 1 AS is_win_size WHERE 0=1")
        rows = cur.fetchall()
        conn.close()

        if not rows:
            return jsonify({'error': 'no data'}), 404

        confs  = [float(r[0]) for r in rows]
        wins   = [int(bool(r[1]))  for r in rows]
        total  = len(confs)

        # Brier score
        brier = round(sum((c - w) ** 2 for c, w in zip(confs, wins)) / total, 5)

        # 10 equal-width bins [0.0, 0.1) … [0.9, 1.0]
        bins = [[] for _ in range(10)]
        for c, w in zip(confs, wins):
            idx = min(int(c * 10), 9)
            bins[idx].append((c, w))

        reliability = []
        ece_num = 0.0
        for i, bucket in enumerate(bins):
            if not bucket:
                reliability.append({
                    'bin_low': round(i * 0.1, 1),
                    'bin_high': round((i + 1) * 0.1, 1),
                    'mean_conf': None,
                    'actual_wr': None,
                    'count': 0,
                    'gap': None,
                })
                continue
            bc   = [x[0] for x in bucket]
            bw   = [x[1] for x in bucket]
            mc   = sum(bc) / len(bc)
            wr   = sum(bw) / len(bw)
            gap  = abs(mc - wr)
            ece_num += gap * len(bucket)
            reliability.append({
                'bin_low':   round(i * 0.1, 1),
                'bin_high':  round((i + 1) * 0.1, 1),
                'mean_conf': round(mc, 4),
                'actual_wr': round(wr, 4),
                'count':     len(bucket),
                'gap':       round(gap, 4),
            })

        ece = round(ece_num / total, 5)
        overall_wr = round(sum(wins) / total, 4)

        # Skill score: fraction of Brier improvement vs reference (always predict baseline)
        baseline_brier = round(overall_wr * (1 - overall_wr), 5)
        brier_skill = round(1 - brier / baseline_brier, 4) if baseline_brier else 0

        return jsonify({
            'n':            total,
            'brier_score':  brier,
            'ece':          ece,
            'brier_skill':  brier_skill,
            'overall_wr':   overall_wr,
            'baseline_brier': baseline_brier,
            'reliability':  reliability,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/calibration-by-size')
@limiter.limit("20 per minute")
def calibration_by_size():
    """P80: Per-SIZE Brier score, avg confidence vs actual WR, calibration gap."""
    try:
        n    = min(int(request.args.get('n', 500)), 2000)
        conn = db.get_connection()
        cur  = conn.cursor()

        if USE_POSTGRES:
            cur.execute("""
                SELECT
                    CASE WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 9  THEN 'NHO'
                         WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 11 THEN 'HOA'
                         ELSE 'LON' END AS pred_size,
                    p.confidence,
                    COALESCE(pr.is_win_size, FALSE) AS is_win
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE p.confidence IS NOT NULL AND pr.actual_numbers IS NOT NULL
                ORDER BY p.draw_number DESC LIMIT %s
            """, (n,))
        else:
            cur.execute("""
                SELECT
                    CASE WHEN (SELECT SUM(CAST(value AS INTEGER)) FROM json_each(p.predicted_numbers)) <= 9  THEN 'NHO'
                         WHEN (SELECT SUM(CAST(value AS INTEGER)) FROM json_each(p.predicted_numbers)) <= 11 THEN 'HOA'
                         ELSE 'LON' END AS pred_size,
                    p.confidence,
                    COALESCE(pr.is_win_size, 0) AS is_win
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE p.confidence IS NOT NULL AND pr.actual_numbers IS NOT NULL
                ORDER BY p.draw_number DESC LIMIT ?
            """, (n,))

        rows = cur.fetchall()
        conn.close()

        from collections import defaultdict
        buckets = defaultdict(list)  # {pred_size: [(conf, is_win), ...]}
        for pred_sz, conf, is_win in rows:
            if pred_sz in ('NHO', 'HOA', 'LON'):
                buckets[pred_sz].append((float(conf), bool(is_win)))

        result = {}
        for sz, pairs in buckets.items():
            if len(pairs) < 5:
                result[sz] = {'n': len(pairs), 'error': 'insufficient data'}
                continue
            confs   = [p[0] for p in pairs]
            wins    = [p[1] for p in pairs]
            t       = len(pairs)
            avg_conf = sum(confs) / t
            actual_wr = sum(wins) / t
            brier    = sum((c - int(w)) ** 2 for c, w in pairs) / t
            baseline_brier = actual_wr * (1 - actual_wr)
            brier_skill = round(1 - brier / baseline_brier, 3) if baseline_brier > 0 else None
            gap = round(avg_conf - actual_wr, 4)

            # 5-bin reliability (20% width each)
            bins = [{'conf_lo': i/5, 'conf_hi': (i+1)/5, 'count': 0, 'wins': 0,
                     'mean_conf': 0.0, 'actual_wr': None} for i in range(5)]
            for c, w in pairs:
                bi = min(int(c * 5), 4)
                bins[bi]['count'] += 1
                bins[bi]['wins']  += int(w)
                bins[bi]['mean_conf'] += c
            for b in bins:
                if b['count'] > 0:
                    b['mean_conf'] = round(b['mean_conf'] / b['count'], 3)
                    b['actual_wr'] = round(b['wins'] / b['count'], 3)

            result[sz] = {
                'n':             t,
                'avg_conf':      round(avg_conf, 4),
                'actual_wr':     round(actual_wr, 4),
                'gap':           gap,
                'brier':         round(brier, 4),
                'brier_skill':   brier_skill,
                'baseline_brier': round(baseline_brier, 4),
                'bins':          bins,
            }

        return jsonify({'sizes': result, 'n': len(rows)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/conf-hist-by-size')
@limiter.limit("20 per minute")
def conf_hist_by_size():
    """P107: Confidence distribution split by predicted SIZE (NHO/HOA/LON), 10 bins each."""
    try:
        n    = min(int(request.args.get('n', 500)), 2000)
        conn = db.get_connection()
        cur  = conn.cursor()
        if USE_POSTGRES:
            cur.execute("""
                SELECT p.confidence,
                       p.predicted_size,
                       COALESCE(pr.is_win_size, FALSE) AS won
                FROM predictions p
                LEFT JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE p.confidence IS NOT NULL AND p.predicted_size IS NOT NULL
                ORDER BY p.draw_number DESC LIMIT %s
            """, (n,))
        else:
            cur.execute("SELECT 0.5,'NHO',0 WHERE 0=1")
        rows = cur.fetchall()
        conn.close()

        N_BINS = 10
        from collections import defaultdict
        size_bins = {sz: [{'count': 0, 'wins': 0, 'evaluated': 0} for _ in range(N_BINS)]
                     for sz in ('NHO', 'HOA', 'LON')}
        size_totals = defaultdict(int)

        for conf, pred_sz, won in rows:
            if pred_sz not in size_bins:
                continue
            idx = min(int(float(conf) * N_BINS), N_BINS - 1)
            b   = size_bins[pred_sz][idx]
            b['count'] += 1
            if won is not None:
                b['evaluated'] += 1
                if won:
                    b['wins'] += 1
            size_totals[pred_sz] += 1

        result = {}
        for sz, bins in size_bins.items():
            hist = []
            for i, b in enumerate(bins):
                ev = b['evaluated']
                hist.append({
                    'bin_low':  round(i / N_BINS, 1),
                    'bin_high': round((i + 1) / N_BINS, 1),
                    'count':    b['count'],
                    'wins':     b['wins'],
                    'win_rate': round(b['wins'] / ev, 3) if ev else None,
                })
            result[sz] = {'bins': hist, 'total': size_totals[sz]}

        return jsonify({'sizes': result, 'n': len(rows)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/confidence-histogram')
@limiter.limit("20 per minute")
def confidence_histogram():
    """P62: Distribution of confidence scores for last N predictions.

    Returns 20 equal-width bins [0, 0.05, 0.10, …, 1.0] with count, win_count, win_rate.
    Also returns mean, median, std, and mode_bin for anomaly detection.
    """
    try:
        n    = min(int(request.args.get('n', 200)), 2000)
        conn = db.get_connection()
        cur  = conn.cursor()

        if USE_POSTGRES:
            cur.execute("""
                SELECT p.confidence, pr.is_win_size
                FROM predictions p
                LEFT JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE p.confidence IS NOT NULL
                ORDER BY p.draw_number DESC
                LIMIT %s
            """, (n,))
        else:
            cur.execute("SELECT 0.5, 1 WHERE 0=1")
        rows = cur.fetchall()
        conn.close()

        if not rows:
            return jsonify({'error': 'no data'}), 404

        confs  = [float(r[0]) for r in rows]
        wins   = [int(bool(r[1])) if r[1] is not None else None for r in rows]
        total  = len(confs)

        # 20 bins width 0.05
        N_BINS = 20
        bins   = [{'conf_wins': [], 'conf_vals': []} for _ in range(N_BINS)]
        for c, w in zip(confs, wins):
            idx = min(int(c * N_BINS), N_BINS - 1)
            bins[idx]['conf_vals'].append(c)
            if w is not None:
                bins[idx]['conf_wins'].append(w)

        histogram = []
        for i, b in enumerate(bins):
            cnt  = len(b['conf_vals'])
            ev   = len(b['conf_wins'])
            wcnt = sum(b['conf_wins'])
            histogram.append({
                'bin_low':   round(i * (1 / N_BINS), 2),
                'bin_high':  round((i + 1) * (1 / N_BINS), 2),
                'count':     cnt,
                'evaluated': ev,
                'wins':      wcnt,
                'win_rate':  round(wcnt / ev, 4) if ev else None,
            })

        # Summary stats
        sorted_c  = sorted(confs)
        mean_c    = sum(confs) / total
        mid       = total // 2
        median_c  = (sorted_c[mid - 1] + sorted_c[mid]) / 2 if total % 2 == 0 else sorted_c[mid]
        variance  = sum((c - mean_c) ** 2 for c in confs) / total
        std_c     = variance ** 0.5
        mode_bin  = max(range(N_BINS), key=lambda i: len(bins[i]['conf_vals']))

        return jsonify({
            'n':          total,
            'mean':       round(mean_c, 4),
            'median':     round(median_c, 4),
            'std':        round(std_c, 4),
            'mode_bin':   histogram[mode_bin]['bin_low'],
            'histogram':  histogram,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/size-accuracy-trend')
@limiter.limit("20 per minute")
def size_accuracy_trend():
    """P79: NHO/HOA/LON win rate per batch over time."""
    try:
        n     = min(int(request.args.get('n', 500)), 2000)
        batch = max(10, min(int(request.args.get('batch', 50)), 200))
        conn  = db.get_connection()
        cur   = conn.cursor()

        if USE_POSTGRES:
            cur.execute("""
                SELECT p.draw_number,
                    CASE WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 9  THEN 'NHO'
                         WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 11 THEN 'HOA'
                         ELSE 'LON' END AS pred_size,
                    COALESCE(pr.is_win_size, FALSE) AS is_win
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE p.predicted_numbers IS NOT NULL AND pr.actual_numbers IS NOT NULL
                ORDER BY p.draw_number DESC LIMIT %s
            """, (n,))
        else:
            cur.execute("""
                SELECT p.draw_number,
                    CASE WHEN (SELECT SUM(CAST(value AS INTEGER)) FROM json_each(p.predicted_numbers)) <= 9  THEN 'NHO'
                         WHEN (SELECT SUM(CAST(value AS INTEGER)) FROM json_each(p.predicted_numbers)) <= 11 THEN 'HOA'
                         ELSE 'LON' END AS pred_size,
                    COALESCE(pr.is_win_size, 0) AS is_win
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE p.predicted_numbers IS NOT NULL AND pr.actual_numbers IS NOT NULL
                ORDER BY p.draw_number DESC LIMIT ?
            """, (n,))

        rows = list(reversed(cur.fetchall()))
        conn.close()

        if not rows:
            return jsonify({'batches': [], 'n': 0})

        batches = []
        for i in range(0, len(rows), batch):
            chunk = rows[i:i + batch]
            acc = {'NHO': [0, 0], 'HOA': [0, 0], 'LON': [0, 0]}  # [total, wins]
            for _, pred_sz, is_win in chunk:
                if pred_sz in acc:
                    acc[pred_sz][0] += 1
                    if is_win: acc[pred_sz][1] += 1
            b = {
                'draw_from': chunk[0][0],
                'draw_to':   chunk[-1][0],
                'count':     len(chunk),
            }
            for sz in ('NHO', 'HOA', 'LON'):
                t, w = acc[sz]
                b[sz] = {
                    'total':    t,
                    'wins':     w,
                    'win_rate': round(w / t, 4) if t >= 3 else None,
                }
            batches.append(b)

        return jsonify({'batches': batches, 'n': len(rows), 'batch_size': batch})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/hoa-trend')
@limiter.limit("20 per minute")
def hoa_trend():
    """P68: Predicted vs actual HOA/NHO/LON% per batch, to monitor HOA correction trend."""
    try:
        n     = min(int(request.args.get('n', 500)), 2000)
        batch = max(10, min(int(request.args.get('batch', 50)), 200))
        conn  = db.get_connection()
        cur   = conn.cursor()
        ph    = db._ph()

        if USE_POSTGRES:
            cur.execute("""
                SELECT p.draw_number,
                    CASE WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 9  THEN 'NHO'
                         WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 11 THEN 'HOA'
                         ELSE 'LON' END AS pred_size,
                    CASE WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 9  THEN 'NHO'
                         WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 11 THEN 'HOA'
                         ELSE 'LON' END AS actual_size
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE p.predicted_numbers IS NOT NULL AND pr.actual_numbers IS NOT NULL
                ORDER BY p.draw_number DESC LIMIT %s
            """, (n,))
        else:
            cur.execute("""
                SELECT p.draw_number,
                    CASE WHEN (SELECT SUM(CAST(value AS INTEGER)) FROM json_each(p.predicted_numbers)) <= 9  THEN 'NHO'
                         WHEN (SELECT SUM(CAST(value AS INTEGER)) FROM json_each(p.predicted_numbers)) <= 11 THEN 'HOA'
                         ELSE 'LON' END AS pred_size,
                    CASE WHEN (SELECT SUM(CAST(value AS INTEGER)) FROM json_each(pr.actual_numbers)) <= 9  THEN 'NHO'
                         WHEN (SELECT SUM(CAST(value AS INTEGER)) FROM json_each(pr.actual_numbers)) <= 11 THEN 'HOA'
                         ELSE 'LON' END AS actual_size
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE p.predicted_numbers IS NOT NULL AND pr.actual_numbers IS NOT NULL
                ORDER BY p.draw_number DESC LIMIT ?
            """, (n,))

        rows = cur.fetchall()
        conn.close()

        if not rows:
            return jsonify({'batches': [], 'n': 0})

        # Reverse to chronological order (oldest first)
        rows = list(reversed(rows))
        total = len(rows)

        batches = []
        for i in range(0, total, batch):
            chunk = rows[i:i + batch]
            cnt   = len(chunk)
            draw_from = chunk[0][0]
            draw_to   = chunk[-1][0]
            pred_counts  = {'NHO': 0, 'HOA': 0, 'LON': 0}
            actual_counts = {'NHO': 0, 'HOA': 0, 'LON': 0}
            for draw_num, pred_sz, actual_sz in chunk:
                if pred_sz   in pred_counts:   pred_counts[pred_sz]   += 1
                if actual_sz in actual_counts: actual_counts[actual_sz] += 1
            batches.append({
                'draw_from':       draw_from,
                'draw_to':         draw_to,
                'count':           cnt,
                'pred_nho_pct':    round(pred_counts['NHO']   / cnt * 100, 1),
                'pred_hoa_pct':    round(pred_counts['HOA']   / cnt * 100, 1),
                'pred_lon_pct':    round(pred_counts['LON']   / cnt * 100, 1),
                'actual_nho_pct':  round(actual_counts['NHO'] / cnt * 100, 1),
                'actual_hoa_pct':  round(actual_counts['HOA'] / cnt * 100, 1),
                'actual_lon_pct':  round(actual_counts['LON'] / cnt * 100, 1),
            })

        return jsonify({'batches': batches, 'n': total, 'batch_size': batch})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/number-frequency')
@limiter.limit("20 per minute")
def number_frequency():
    """P82: Actual vs predicted frequency per number (1-6) for last N draws."""
    try:
        n    = min(int(request.args.get('n', 500)), 5000)
        conn = db.get_connection()
        cur  = conn.cursor()

        if USE_POSTGRES:
            # Actual draws
            cur.execute("""
                WITH recent AS (
                    SELECT numbers FROM draw_history ORDER BY draw_number DESC LIMIT %s
                )
                SELECT CAST(v AS INTEGER) AS num, COUNT(*) AS cnt
                FROM recent, json_array_elements_text(numbers::json) AS v
                GROUP BY num ORDER BY num
            """, (n,))
            actual_rows = cur.fetchall()

            # Predicted numbers (last N predictions with results)
            cur.execute("""
                WITH recent AS (
                    SELECT p.predicted_numbers
                    FROM predictions p
                    JOIN prediction_results pr ON pr.prediction_id = p.id
                    WHERE p.predicted_numbers IS NOT NULL
                    ORDER BY p.draw_number DESC LIMIT %s
                )
                SELECT CAST(v AS INTEGER) AS num, COUNT(*) AS cnt
                FROM recent, json_array_elements_text(predicted_numbers::json) AS v
                GROUP BY num ORDER BY num
            """, (n,))
            pred_rows = cur.fetchall()
        else:
            cur.execute("""
                SELECT CAST(value AS INTEGER) AS num, COUNT(*) AS cnt
                FROM (SELECT numbers FROM draw_history ORDER BY draw_number DESC LIMIT ?)
                     AS recent, json_each(numbers)
                GROUP BY num ORDER BY num
            """, (n,))
            actual_rows = cur.fetchall()
            cur.execute("""
                SELECT CAST(value AS INTEGER) AS num, COUNT(*) AS cnt
                FROM (SELECT p.predicted_numbers FROM predictions p
                      JOIN prediction_results pr ON pr.prediction_id = p.id
                      WHERE p.predicted_numbers IS NOT NULL
                      ORDER BY p.draw_number DESC LIMIT ?) AS recent,
                     json_each(predicted_numbers)
                GROUP BY num ORDER BY num
            """, (n,))
            pred_rows = cur.fetchall()

        conn.close()

        actual = {int(r[0]): int(r[1]) for r in actual_rows}
        pred   = {int(r[0]): int(r[1]) for r in pred_rows}

        act_total  = sum(actual.values()) or 1
        pred_total = sum(pred.values())   or 1
        baseline   = round(1 / 6, 4)  # uniform expected frequency per number

        numbers_out = []
        for num in range(1, 7):
            a_cnt = actual.get(num, 0)
            p_cnt = pred.get(num, 0)
            a_pct = round(a_cnt / act_total, 4)
            p_pct = round(p_cnt / pred_total, 4)
            numbers_out.append({
                'num':          num,
                'actual_count': a_cnt,
                'pred_count':   p_cnt,
                'actual_pct':   a_pct,
                'pred_pct':     p_pct,
                'diff':         round(p_pct - a_pct, 4),  # + = over-predict, - = under-predict
            })

        return jsonify({
            'numbers':    numbers_out,
            'n_draws':    n,
            'act_total':  act_total,
            'pred_total': pred_total,
            'baseline':   baseline,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/prediction-heatmap')
@limiter.limit("20 per minute")
def prediction_heatmap():
    """P81: Prediction count heatmap — (day-of-week) × (hour-of-day) grid."""
    try:
        n    = min(int(request.args.get('n', 2000)), 10000)
        conn = db.get_connection()
        cur  = conn.cursor()

        if USE_POSTGRES:
            cur.execute("""
                SELECT
                    EXTRACT(DOW FROM dh.draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::int AS dow,
                    EXTRACT(HOUR FROM dh.draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::int AS vn_hour,
                    COUNT(*) AS cnt
                FROM predictions p
                JOIN draw_history dh ON dh.draw_number = p.draw_number
                WHERE p.draw_number IS NOT NULL
                GROUP BY dow, vn_hour
                ORDER BY dow, vn_hour
            """)
        else:
            cur.execute("""
                SELECT
                    CAST(strftime('%w', dh.draw_time) AS INTEGER) AS dow,
                    CAST(strftime('%H', dh.draw_time) AS INTEGER) AS vn_hour,
                    COUNT(*) AS cnt
                FROM predictions p
                JOIN draw_history dh ON dh.draw_number = p.draw_number
                GROUP BY dow, vn_hour
                ORDER BY dow, vn_hour
            """)

        rows = cur.fetchall()
        conn.close()

        # Build grid: dow 0–6, hour 6–22
        from collections import defaultdict
        grid = defaultdict(int)  # (dow, hour) -> count
        total = 0
        for dow, hour, cnt in rows:
            grid[(int(dow), int(hour))] += int(cnt)
            total += int(cnt)

        max_val = max(grid.values()) if grid else 1

        # Serialize as list of {dow, hour, count, intensity}
        cells = []
        for dow in range(7):
            for hour in range(6, 23):
                cnt = grid.get((dow, hour), 0)
                cells.append({
                    'dow':       dow,
                    'hour':      hour,
                    'count':     cnt,
                    'intensity': round(cnt / max_val, 3) if max_val else 0,
                })

        DOW_LABELS = {0: 'CN', 1: 'T2', 2: 'T3', 3: 'T4', 4: 'T5', 5: 'T6', 6: 'T7'}
        return jsonify({
            'cells':      cells,
            'total':      total,
            'max_count':  max_val,
            'dow_labels': DOW_LABELS,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/wr-by-dow')
@limiter.limit("20 per minute")
def wr_by_dow():
    """P73: Win rate (SIZE) per day-of-week in VN time."""
    try:
        n    = min(int(request.args.get('n', 1000)), 5000)
        conn = db.get_connection()
        cur  = conn.cursor()

        if USE_POSTGRES:
            cur.execute("""
                SELECT
                    EXTRACT(DOW FROM dh.draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::int AS dow,
                    COALESCE(pr.is_win_size, pr.is_win, FALSE) AS is_win,
                    CASE WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 9  THEN 'NHO'
                         WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 11 THEN 'HOA'
                         ELSE 'LON' END AS pred_size
                FROM prediction_results pr
                JOIN draw_history dh ON dh.draw_number = pr.draw_number
                JOIN predictions p ON p.id = pr.prediction_id
                WHERE pr.actual_numbers IS NOT NULL AND p.predicted_numbers IS NOT NULL
                ORDER BY pr.draw_number DESC LIMIT %s
            """, (n,))
        else:
            cur.execute("""
                SELECT
                    CAST(strftime('%w', dh.draw_time) AS INTEGER) AS dow,
                    COALESCE(pr.is_win_size, pr.is_win, 0) AS is_win,
                    CASE WHEN (SELECT SUM(CAST(value AS INTEGER)) FROM json_each(p.predicted_numbers)) <= 9  THEN 'NHO'
                         WHEN (SELECT SUM(CAST(value AS INTEGER)) FROM json_each(p.predicted_numbers)) <= 11 THEN 'HOA'
                         ELSE 'LON' END AS pred_size
                FROM prediction_results pr
                JOIN draw_history dh ON dh.draw_number = pr.draw_number
                JOIN predictions p ON p.id = pr.prediction_id
                WHERE pr.actual_numbers IS NOT NULL AND p.predicted_numbers IS NOT NULL
                ORDER BY pr.draw_number DESC LIMIT ?
            """, (n,))

        rows = cur.fetchall()
        conn.close()

        from collections import defaultdict
        # dow: 0=Sun,1=Mon,...,6=Sat
        dow_data: dict = defaultdict(lambda: {'total': 0, 'wins': 0, 'nho': 0, 'hoa': 0, 'lon': 0})
        for dow, is_win, pred_sz in rows:
            d = dow_data[int(dow)]
            d['total'] += 1
            if is_win: d['wins'] += 1
            if pred_sz == 'NHO': d['nho'] += 1
            elif pred_sz == 'HOA': d['hoa'] += 1
            elif pred_sz == 'LON': d['lon'] += 1

        # Vietnamese day order: Mon(1)..Sat(6), Sun(0) → display T2..T7,CN
        DOW_LABELS = {0: 'CN', 1: 'T2', 2: 'T3', 3: 'T4', 4: 'T5', 5: 'T6', 6: 'T7'}
        DOW_ORDER  = [1, 2, 3, 4, 5, 6, 0]  # Mon→Sun

        days_out = []
        for d in DOW_ORDER:
            dd = dow_data.get(d, {'total': 0, 'wins': 0, 'nho': 0, 'hoa': 0, 'lon': 0})
            t  = dd['total']
            w  = dd['wins']
            days_out.append({
                'dow':       d,
                'label':     DOW_LABELS[d],
                'total':     t,
                'wins':      w,
                'win_rate':  round(w / t, 4) if t >= 5 else None,
                'nho_pct':   round(dd['nho'] / t * 100, 1) if t else 0,
                'hoa_pct':   round(dd['hoa'] / t * 100, 1) if t else 0,
                'lon_pct':   round(dd['lon'] / t * 100, 1) if t else 0,
            })

        return jsonify({'days': days_out, 'n': len(rows), 'baseline': 0.375})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/daily-wincal')
@limiter.limit("20 per minute")
def daily_wincal():
    """P128: Win-rate per calendar day in VN time — last N days for heatmap calendar."""
    try:
        days = min(int(request.args.get('days', 90)), 365)
        conn = db.get_connection()
        cur  = conn.cursor()

        if USE_POSTGRES:
            cur.execute("""
                SELECT
                    (dh.draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date AS vn_date,
                    COUNT(*)                                              AS total,
                    SUM(CASE WHEN COALESCE(pr.is_win_size, pr.is_win, FALSE) THEN 1 ELSE 0 END) AS wins
                FROM prediction_results pr
                JOIN draw_history dh ON dh.draw_number = pr.draw_number
                WHERE pr.actual_numbers IS NOT NULL
                  AND dh.draw_time >= NOW() - (%s || ' days')::INTERVAL
                GROUP BY 1
                ORDER BY 1 DESC
            """, (days,))
        else:
            cur.execute("""
                SELECT
                    DATE(datetime(dh.draw_time, '+7 hours')) AS vn_date,
                    COUNT(*)                                AS total,
                    SUM(COALESCE(pr.is_win_size, pr.is_win, 0)) AS wins
                FROM prediction_results pr
                JOIN draw_history dh ON dh.draw_number = pr.draw_number
                WHERE pr.actual_numbers IS NOT NULL
                  AND dh.draw_time >= DATE(datetime('now', ?))
                GROUP BY 1
                ORDER BY 1 DESC
            """, (f'-{days} days',))

        rows = cur.fetchall()
        conn.close()

        import datetime as _dt
        today_vn = _dt.datetime.utcnow() + _dt.timedelta(hours=7)
        today_str = today_vn.strftime('%Y-%m-%d')

        result = []
        for vn_date, total, wins in rows:
            date_str = str(vn_date)[:10]
            d = _dt.date.fromisoformat(date_str)
            result.append({
                'date':     date_str,
                'total':    int(total),
                'wins':     int(wins),
                'win_rate': round(int(wins) / int(total), 4) if total >= 5 else None,
                'dow':      d.weekday(),   # 0=Mon..6=Sun
                'is_today': date_str == today_str,
            })

        return jsonify({'days': result, 'baseline': 0.375, 'today': today_str})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/hoa-by-hour')
@limiter.limit("20 per minute")
def hoa_by_hour():
    """P72: HOA predicted/win rate per VN hour-of-day (6–22)."""
    try:
        n    = min(int(request.args.get('n', 1000)), 5000)
        conn = db.get_connection()
        cur  = conn.cursor()

        if USE_POSTGRES:
            cur.execute("""
                SELECT
                    EXTRACT(HOUR FROM dh.draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::int AS vn_hour,
                    CASE WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 9  THEN 'NHO'
                         WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 11 THEN 'HOA'
                         ELSE 'LON' END AS pred_size,
                    CASE WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 9  THEN 'NHO'
                         WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 11 THEN 'HOA'
                         ELSE 'LON' END AS actual_size
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                JOIN draw_history dh ON dh.draw_number = p.draw_number
                WHERE p.predicted_numbers IS NOT NULL AND pr.actual_numbers IS NOT NULL
                ORDER BY p.draw_number DESC LIMIT %s
            """, (n,))
        else:
            cur.execute("""
                SELECT
                    CAST(strftime('%H', dh.draw_time) AS INTEGER) AS vn_hour,
                    CASE WHEN (SELECT SUM(CAST(value AS INTEGER)) FROM json_each(p.predicted_numbers)) <= 9  THEN 'NHO'
                         WHEN (SELECT SUM(CAST(value AS INTEGER)) FROM json_each(p.predicted_numbers)) <= 11 THEN 'HOA'
                         ELSE 'LON' END AS pred_size,
                    CASE WHEN (SELECT SUM(CAST(value AS INTEGER)) FROM json_each(pr.actual_numbers)) <= 9  THEN 'NHO'
                         WHEN (SELECT SUM(CAST(value AS INTEGER)) FROM json_each(pr.actual_numbers)) <= 11 THEN 'HOA'
                         ELSE 'LON' END AS actual_size
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                JOIN draw_history dh ON dh.draw_number = p.draw_number
                WHERE p.predicted_numbers IS NOT NULL AND pr.actual_numbers IS NOT NULL
                ORDER BY p.draw_number DESC LIMIT ?
            """, (n,))

        rows = cur.fetchall()
        conn.close()

        # Aggregate per hour
        from collections import defaultdict
        hour_data: dict = defaultdict(lambda: {'total': 0, 'pred_hoa': 0, 'hoa_wins': 0, 'actual_hoa': 0})
        for vn_hour, pred_sz, actual_sz in rows:
            h = int(vn_hour)
            hour_data[h]['total']      += 1
            if pred_sz   == 'HOA': hour_data[h]['pred_hoa']   += 1
            if actual_sz == 'HOA': hour_data[h]['actual_hoa'] += 1
            if pred_sz == 'HOA' and actual_sz == 'HOA':
                hour_data[h]['hoa_wins'] += 1

        hours_out = []
        for h in range(6, 23):
            d = hour_data.get(h, {'total': 0, 'pred_hoa': 0, 'hoa_wins': 0, 'actual_hoa': 0})
            t = d['total']
            p = d['pred_hoa']
            hours_out.append({
                'hour':           h,
                'total':          t,
                'pred_hoa':       p,
                'hoa_wins':       d['hoa_wins'],
                'actual_hoa':     d['actual_hoa'],
                'pred_hoa_rate':  round(p / t, 4) if t else None,
                'actual_hoa_rate': round(d['actual_hoa'] / t, 4) if t else None,
                'hoa_win_rate':   round(d['hoa_wins'] / p, 4) if p >= 5 else None,
            })

        return jsonify({'hours': hours_out, 'n': len(rows)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/wr-by-hour')
@limiter.limit("20 per minute")
def wr_by_hour():
    """P115: Overall win rate (SIZE) by VN hour of day (6-22)."""
    try:
        n    = min(int(request.args.get('n', 1000)), 5000)
        conn = db.get_connection()
        cur  = conn.cursor()

        if USE_POSTGRES:
            cur.execute("""
                SELECT
                    EXTRACT(HOUR FROM dh.draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::int AS vn_hour,
                    COALESCE(pr.is_win_size, pr.is_win, FALSE) AS is_win,
                    CASE WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 9  THEN 'NHO'
                         WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 11 THEN 'HOA'
                         ELSE 'LON' END AS pred_size
                FROM prediction_results pr
                JOIN draw_history dh ON dh.draw_number = pr.draw_number
                JOIN predictions p ON p.id = pr.prediction_id
                WHERE pr.actual_numbers IS NOT NULL AND p.predicted_numbers IS NOT NULL
                ORDER BY pr.draw_number DESC LIMIT %s
            """, (n,))
        else:
            cur.execute("""
                SELECT
                    CAST(strftime('%H', dh.draw_time) AS INTEGER) AS vn_hour,
                    COALESCE(pr.is_win_size, pr.is_win, 0) AS is_win,
                    CASE WHEN (SELECT SUM(CAST(value AS INTEGER)) FROM json_each(p.predicted_numbers)) <= 9  THEN 'NHO'
                         WHEN (SELECT SUM(CAST(value AS INTEGER)) FROM json_each(p.predicted_numbers)) <= 11 THEN 'HOA'
                         ELSE 'LON' END AS pred_size
                FROM prediction_results pr
                JOIN draw_history dh ON dh.draw_number = pr.draw_number
                JOIN predictions p ON p.id = pr.prediction_id
                WHERE pr.actual_numbers IS NOT NULL AND p.predicted_numbers IS NOT NULL
                ORDER BY pr.draw_number DESC LIMIT ?
            """, (n,))

        rows = cur.fetchall()
        conn.close()

        from collections import defaultdict
        hour_data: dict = defaultdict(lambda: {
            'total': 0, 'wins': 0,
            'NHO': {'total': 0, 'wins': 0},
            'HOA': {'total': 0, 'wins': 0},
            'LON': {'total': 0, 'wins': 0},
        })

        for vn_hour, is_win, pred_size in rows:
            h = int(vn_hour)
            w = bool(is_win)
            hour_data[h]['total'] += 1
            if w: hour_data[h]['wins'] += 1
            if pred_size in ('NHO', 'HOA', 'LON'):
                hour_data[h][pred_size]['total'] += 1
                if w: hour_data[h][pred_size]['wins'] += 1

        hours_out = []
        for h in range(6, 23):
            d = hour_data.get(h, {'total': 0, 'wins': 0,
                                   'NHO': {'total': 0, 'wins': 0},
                                   'HOA': {'total': 0, 'wins': 0},
                                   'LON': {'total': 0, 'wins': 0}})
            t, w = d['total'], d['wins']
            entry = {
                'hour':  h,
                'total': t,
                'wins':  w,
                'wr':    round(w / t, 4) if t >= 5 else None,
                'by_size': {},
            }
            for sz in ('NHO', 'HOA', 'LON'):
                st = d[sz]['total']
                sw = d[sz]['wins']
                entry['by_size'][sz] = {
                    'total': st,
                    'wins':  sw,
                    'wr':    round(sw / st, 4) if st >= 3 else None,
                }
            hours_out.append(entry)

        return jsonify({'hours': hours_out, 'n': len(rows)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/autotune-history')
@limiter.limit("20 per minute")
def autotune_history():
    """P100: Adaptive threshold history — extracted from vote_breakdown.adaptive per batch."""
    import json as _json
    try:
        n          = min(int(request.args.get('n', 500)), 2000)
        batch_size = min(int(request.args.get('batch', 25)), 100)
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute(
            "SELECT draw_number, vote_breakdown FROM predictions "
            "WHERE vote_breakdown IS NOT NULL ORDER BY draw_number DESC LIMIT %s" if USE_POSTGRES else
            "SELECT draw_number, vote_breakdown FROM predictions "
            "WHERE vote_breakdown IS NOT NULL ORDER BY draw_number DESC LIMIT ?",
            (n,)
        )
        rows = cur.fetchall()
        conn.close()

        FIELDS = ['tune_k', 'hoa_suppress', 'nho_share_min',
                  'pred_lon_excess', 'pred_nho_excess', 'consecutive_excess']

        rows = list(reversed(rows))  # oldest first
        batches = []
        for i in range(0, len(rows), batch_size):
            chunk = rows[i:i + batch_size]
            if len(chunk) < max(3, batch_size // 3):
                continue
            sums  = {f: 0.0 for f in FIELDS}
            count = 0
            for dn, vb_raw in chunk:
                try:
                    vb  = _json.loads(vb_raw) if isinstance(vb_raw, str) else (vb_raw or {})
                    at  = vb.get('adaptive') or {}
                    for f in FIELDS:
                        v = at.get(f)
                        if v is not None:
                            sums[f] += float(v)
                    count += 1
                except Exception:
                    pass
            if count == 0:
                continue
            entry = {
                'batch':      len(batches) + 1,
                'draw_start': chunk[0][0],
                'draw_end':   chunk[-1][0],
                'n':          count,
            }
            for f in FIELDS:
                entry[f] = round(sums[f] / count, 4)
            batches.append(entry)

        return jsonify({'batches': batches, 'fields': FIELDS})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/voter-wr-trend')
@limiter.limit("20 per minute")
def voter_wr_trend():
    """P129: Per-voter win rate per batch — actual WR trajectory for each voter over time."""
    import json as _json
    try:
        n          = min(int(request.args.get('n', 500)), 2000)
        batch_size = min(int(request.args.get('batch', 25)), 100)
        conn = db.get_connection()
        cur  = conn.cursor()
        if USE_POSTGRES:
            cur.execute("""
                SELECT p.draw_number, p.vote_breakdown, pr.actual_numbers
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE p.vote_breakdown IS NOT NULL AND pr.actual_numbers IS NOT NULL
                ORDER BY p.draw_number DESC LIMIT %s
            """, (n,))
        else:
            cur.execute("""
                SELECT p.draw_number, p.vote_breakdown, pr.actual_numbers
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE p.vote_breakdown IS NOT NULL AND pr.actual_numbers IS NOT NULL
                ORDER BY p.draw_number DESC LIMIT ?
            """, (n,))
        rows = cur.fetchall()
        conn.close()

        def _actual_size(nums_raw):
            try:
                ns = nums_raw if not isinstance(nums_raw, str) else __import__('ast').literal_eval(nums_raw)
                s = sum(int(x) for x in ns)
                if s <= 9:  return 'NHO'
                if s <= 11: return 'HOA'
                return 'LON'
            except Exception:
                return None

        rows = list(reversed(rows))  # oldest first
        batches = []
        for i in range(0, len(rows), batch_size):
            chunk = rows[i:i + batch_size]
            if len(chunk) < max(3, batch_size // 3):
                continue
            voter_stats: dict = {}
            for dn, vb_raw, actual_raw in chunk:
                actual_size = _actual_size(actual_raw)
                if not actual_size:
                    continue
                try:
                    vb = _json.loads(vb_raw) if isinstance(vb_raw, str) else (vb_raw or {})
                    all_votes = vb.get('all_votes') or {}
                    for vname, vsize in all_votes.items():
                        if vname not in voter_stats:
                            voter_stats[vname] = {'w': 0, 't': 0}
                        voter_stats[vname]['t'] += 1
                        if vsize == actual_size:
                            voter_stats[vname]['w'] += 1
                except Exception:
                    pass
            if not voter_stats:
                continue
            entry = {
                'batch':      len(batches) + 1,
                'draw_start': chunk[0][0],
                'draw_end':   chunk[-1][0],
                'n':          len(chunk),
                'voters':     {
                    v: round(d['w'] / d['t'], 4) if d['t'] >= 3 else None
                    for v, d in voter_stats.items()
                },
            }
            batches.append(entry)

        # Collect voter names that appear in enough batches
        all_voters: dict = {}
        for b in batches:
            for v, wr in b['voters'].items():
                if v not in all_voters:
                    all_voters[v] = 0
                if wr is not None:
                    all_voters[v] += 1
        min_batches = max(2, len(batches) // 4)
        voter_names = sorted(
            [v for v, cnt in all_voters.items() if cnt >= min_batches],
            key=lambda v: -all_voters[v]
        )

        return jsonify({
            'batches':     batches,
            'voter_names': voter_names,
            'baseline':    0.375,
            'n':           len(rows),
            'batch_size':  batch_size,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/voter-conf-trend')
@limiter.limit("20 per minute")
def voter_conf_trend():
    """P96: Average confidence per voter per batch (from all_votes_detail in vote_breakdown)."""
    try:
        n          = min(int(request.args.get('n', 500)), 2000)
        batch_size = min(int(request.args.get('batch', 25)), 100)
        conn = db.get_connection()
        cur  = conn.cursor()
        if USE_POSTGRES:
            cur.execute("""
                SELECT p.draw_number, p.vote_breakdown
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE p.vote_breakdown IS NOT NULL AND pr.actual_numbers IS NOT NULL
                ORDER BY p.draw_number DESC LIMIT %s
            """, (n,))
        else:
            cur.execute("SELECT NULL, NULL WHERE 0=1")
        rows = cur.fetchall()
        conn.close()

        if not rows:
            return jsonify({'error': 'no data'}), 404

        # Oldest first for chronological batches
        rows = list(reversed(rows))
        VOTER_ORDER = ['ml', 'markov', 'prior_nho', 'prior_lon', 'prior_hoa']

        batches = []
        for i in range(0, len(rows), batch_size):
            chunk = rows[i:i + batch_size]
            voter_sums: dict = {v: {'conf': 0.0, 'n': 0} for v in VOTER_ORDER}
            for draw_num, vb_raw in chunk:
                try:
                    vb     = json.loads(vb_raw) if isinstance(vb_raw, str) else vb_raw
                    detail = (vb or {}).get('all_votes_detail') or {}
                    for vname, d in detail.items():
                        if vname in voter_sums and d.get('conf') is not None:
                            voter_sums[vname]['conf'] += float(d['conf'])
                            voter_sums[vname]['n']    += 1
                except Exception:
                    continue
            voter_avgs = {}
            for vname, s in voter_sums.items():
                if s['n'] >= max(1, batch_size // 4):
                    voter_avgs[vname] = round(s['conf'] / s['n'], 4)
            if voter_avgs:
                batches.append({
                    'batch':      i // batch_size + 1,
                    'draw_start': chunk[0][0],
                    'draw_end':   chunk[-1][0],
                    'n':          len(chunk),
                    'voters':     voter_avgs,
                })

        return jsonify({
            'batches':    batches,
            'batch_size': batch_size,
            'n_draws':    len(rows),
            'voters':     VOTER_ORDER,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/daily-card')
@limiter.limit("30 per minute")
def daily_card():
    """P90: Today's detailed stats — WR per SIZE, confidence avg, streak today, vs yesterday."""
    from zoneinfo import ZoneInfo
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        if not USE_POSTGRES:
            conn.close()
            return jsonify({'error': 'PostgreSQL only'}), 400

        cur.execute("""
            WITH today AS (
                SELECT
                    CASE
                        WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 9  THEN 'NHO'
                        WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 11 THEN 'HOA'
                        ELSE 'LON'
                    END AS pred_size,
                    COALESCE(pr.is_win_size, pr.is_win, FALSE) AS is_win,
                    p.confidence
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE (pr.created_at AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                      = (NOW() AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                  AND pr.actual_numbers IS NOT NULL
            ),
            yesterday AS (
                SELECT COALESCE(pr.is_win_size, pr.is_win, FALSE) AS is_win
                FROM prediction_results pr
                JOIN predictions p ON p.id = pr.prediction_id
                WHERE (pr.created_at AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                      = (NOW() AT TIME ZONE 'Asia/Ho_Chi_Minh')::date - INTERVAL '1 day'
                  AND pr.actual_numbers IS NOT NULL
            )
            SELECT
                COUNT(*)                                                        AS total,
                SUM(CASE WHEN is_win THEN 1 ELSE 0 END)                       AS wins,
                AVG(confidence)                                                 AS avg_conf,
                SUM(CASE WHEN pred_size = 'NHO' AND is_win THEN 1 ELSE 0 END) AS nho_w,
                SUM(CASE WHEN pred_size = 'NHO' THEN 1 ELSE 0 END)            AS nho_t,
                SUM(CASE WHEN pred_size = 'HOA' AND is_win THEN 1 ELSE 0 END) AS hoa_w,
                SUM(CASE WHEN pred_size = 'HOA' THEN 1 ELSE 0 END)            AS hoa_t,
                SUM(CASE WHEN pred_size = 'LON' AND is_win THEN 1 ELSE 0 END) AS lon_w,
                SUM(CASE WHEN pred_size = 'LON' THEN 1 ELSE 0 END)            AS lon_t,
                (SELECT COUNT(*) FROM yesterday)                               AS yest_total,
                (SELECT SUM(CASE WHEN is_win THEN 1 ELSE 0 END) FROM yesterday) AS yest_wins
            FROM today
        """)
        row = cur.fetchone()
        conn.close()

        total, wins, avg_conf, nho_w, nho_t, hoa_w, hoa_t, lon_w, lon_t, yt, yw = [
            (int(x) if isinstance(x, float) and x == int(x) else x) if x is not None else 0
            for x in row
        ]
        wr      = wins / total if total else 0
        wr_yest = yw / yt if yt else None

        sizes = []
        for sz, w, t in [('NHO', nho_w, nho_t), ('HOA', hoa_w, hoa_t), ('LON', lon_w, lon_t)]:
            if t:
                sizes.append({'size': sz, 'wins': w, 'total': t, 'wr': round(w / t, 4)})

        best_size = max(sizes, key=lambda s: s['wr']) if sizes else None

        return jsonify({
            'total':     total,
            'wins':      wins,
            'wr':        round(wr, 4),
            'avg_conf':  round(float(avg_conf), 4) if avg_conf else None,
            'sizes':     sizes,
            'best_size': best_size,
            'wr_yest':   round(wr_yest, 4) if wr_yest is not None else None,
            'yt':        yt,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/prediction-timeline')
@limiter.limit("30 per minute")
def prediction_timeline():
    """P89: Last N predictions with predicted/actual SIZE and win/loss for timeline strip."""
    try:
        n    = min(int(request.args.get('n', 40)), 100)
        conn = db.get_connection()
        cur  = conn.cursor()
        if USE_POSTGRES:
            cur.execute("""
                SELECT p.draw_number, p.confidence,
                    CASE
                        WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 9  THEN 'NHO'
                        WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 11 THEN 'HOA'
                        ELSE 'LON'
                    END AS pred_size,
                    CASE
                        WHEN pr.actual_numbers IS NULL THEN NULL
                        WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 9  THEN 'NHO'
                        WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 11 THEN 'HOA'
                        ELSE 'LON'
                    END AS actual_size,
                    COALESCE(pr.is_win_size, pr.is_win) AS is_win
                FROM predictions p
                LEFT JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE p.predicted_numbers IS NOT NULL
                ORDER BY p.draw_number DESC LIMIT %s
            """, (n,))
        else:
            cur.execute("SELECT NULL, NULL, NULL, NULL, NULL WHERE 0=1")
        rows = cur.fetchall()
        conn.close()

        draws = []
        for draw_num, conf, pred_sz, act_sz, is_win in rows:
            draws.append({
                'draw_number': draw_num,
                'confidence':  round(float(conf), 3) if conf else None,
                'pred_size':   pred_sz,
                'actual_size': act_sz,
                'is_win':      bool(is_win) if is_win is not None else None,
            })

        # Reverse so oldest is first (left → right = older → newer)
        draws.reverse()
        return jsonify({'draws': draws, 'n': len(draws)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/adaptive-state')
@limiter.limit("30 per minute")
def adaptive_state():
    """P88: Current adaptive threshold cache + recent predicted vs actual SIZE distribution."""
    from prediction_service import _adaptive_thresh_cache as _atc
    try:
        at = dict(_atc) if _atc else {}

        # Also fetch recent predicted vs actual SIZE distribution (last 100 majority_vote)
        conn = db.get_connection()
        cur  = conn.cursor()
        if USE_POSTGRES:
            cur.execute("""
                SELECT
                    SUM(CASE WHEN pred_sum <= 9  THEN 1 ELSE 0 END)::float / COUNT(*) AS pred_nho,
                    SUM(CASE WHEN pred_sum BETWEEN 10 AND 11 THEN 1 ELSE 0 END)::float / COUNT(*) AS pred_hoa,
                    SUM(CASE WHEN pred_sum >= 12 THEN 1 ELSE 0 END)::float / COUNT(*) AS pred_lon,
                    SUM(CASE WHEN act_sum  <= 9  THEN 1 ELSE 0 END)::float / COUNT(*) AS act_nho,
                    SUM(CASE WHEN act_sum  BETWEEN 10 AND 11 THEN 1 ELSE 0 END)::float / COUNT(*) AS act_hoa,
                    SUM(CASE WHEN act_sum  >= 12 THEN 1 ELSE 0 END)::float / COUNT(*) AS act_lon,
                    COUNT(*) AS n
                FROM (
                    SELECT
                        (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) AS pred_sum,
                        (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v)  AS act_sum
                    FROM predictions p
                    JOIN prediction_results pr ON pr.prediction_id = p.id
                    WHERE p.model_name = 'majority_vote'
                      AND pr.actual_numbers IS NOT NULL
                    ORDER BY p.draw_number DESC
                    LIMIT 100
                ) sub
            """)
            row = cur.fetchone()
            if row and row[6] and int(row[6]) >= 10:
                at['dist_pred_nho'] = round(float(row[0] or 0), 3)
                at['dist_pred_hoa'] = round(float(row[1] or 0), 3)
                at['dist_pred_lon'] = round(float(row[2] or 0), 3)
                at['dist_act_nho']  = round(float(row[3] or 0), 3)
                at['dist_act_hoa']  = round(float(row[4] or 0), 3)
                at['dist_act_lon']  = round(float(row[5] or 0), 3)
                at['dist_n']        = int(row[6])
        conn.close()

        # Add ml_mult from latest prediction's vote_breakdown
        try:
            conn2 = db.get_connection()
            cur2  = conn2.cursor()
            ph = '%s' if USE_POSTGRES else '?'
            cur2.execute(
                f"SELECT vote_breakdown FROM predictions "
                f"WHERE vote_breakdown IS NOT NULL ORDER BY draw_number DESC LIMIT 1"
            )
            vb_row = cur2.fetchone()
            conn2.close()
            if vb_row:
                import json as _j
                _vb = _j.loads(vb_row[0]) if isinstance(vb_row[0], str) else (vb_row[0] or {})
                _ml = (_vb.get('all_votes_detail') or {}).get('ml') or {}
                at['ml_mult']  = round(float(_ml.get('mult', 1.0)), 3)
                at['ml_decay'] = round(float(_ml.get('decay', 1.0)), 2)
                at['ml_streak'] = int(_ml.get('streak', 0))
        except Exception:
            pass

        # Defaults for display when cache is empty
        defaults = {
            'hoa_suppress':   0.70, 'nho_share_min': 0.45,
            'prior_nho_conf': 0.44, 'prior_lon_conf': 0.40,
            'tune_k': 0.40, 'consecutive_excess': 0, 'pred_lon_excess': 0.0, 'pred_nho_excess': 0.0,
            'ml_mult': 1.0, 'ml_decay': 1.0, 'ml_streak': 0,
        }
        for k, v in defaults.items():
            at.setdefault(k, v)

        return jsonify(at)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/checkpoint')
@limiter.limit("30 per minute")
def checkpoint_api():
    """P150: Validation checkpoint progress — fresh predictions since p128 deploy."""
    CHECKPOINT_TS = _CHECKPOINT_TS
    CHECKPOINT_N  = _CHECKPOINT_N
    try:
        if not USE_POSTGRES:
            return jsonify({'n_fresh': 0, 'n_target': CHECKPOINT_N, 'ready': False, 'n_evaluated': 0, 'progress': 0.0})
        conn = db.get_connection()
        cur  = conn.cursor()
        CHECKPOINT_TS, CHECKPOINT_N = _get_checkpoint_config(cur)
        cur.execute("SELECT COUNT(*) FROM predictions WHERE created_at > %s", (CHECKPOINT_TS,))
        n_fresh = int(cur.fetchone()[0])
        cur.execute("""
            SELECT COUNT(*), SUM(CASE WHEN pr.is_win_size THEN 1 ELSE 0 END)
            FROM predictions p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            WHERE p.created_at > %s AND pr.is_win_size IS NOT NULL
        """, (CHECKPOINT_TS,))
        row   = cur.fetchone()
        n_eval = int(row[0])
        n_wins = int(row[1] or 0)
        # Pre-reg hypothesis #1 + #2: ml_controls vs overridden + ML-LON specific
        cur.execute("""
            SELECT
              SUM(CASE WHEN (p.vote_breakdown->'all_votes'->>'ml') =
                             (p.vote_breakdown->>'final_size') THEN 1 ELSE 0 END) AS ctrl_t,
              SUM(CASE WHEN (p.vote_breakdown->'all_votes'->>'ml') =
                             (p.vote_breakdown->>'final_size')
                            AND pr.is_win_size THEN 1 ELSE 0 END)               AS ctrl_w,
              SUM(CASE WHEN (p.vote_breakdown->'all_votes'->>'ml') <>
                             (p.vote_breakdown->>'final_size') THEN 1 ELSE 0 END) AS over_t,
              SUM(CASE WHEN (p.vote_breakdown->'all_votes'->>'ml') <>
                             (p.vote_breakdown->>'final_size')
                            AND pr.is_win_size THEN 1 ELSE 0 END)               AS over_w,
              SUM(CASE WHEN p.vote_breakdown->'all_votes'->>'ml' = 'LON'
                            AND p.vote_breakdown->>'final_size' = 'LON' THEN 1 ELSE 0 END) AS lon_t,
              SUM(CASE WHEN p.vote_breakdown->'all_votes'->>'ml' = 'LON'
                            AND p.vote_breakdown->>'final_size' = 'LON'
                            AND pr.is_win_size THEN 1 ELSE 0 END)               AS lon_w
            FROM predictions p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            WHERE p.created_at > %s AND pr.is_win_size IS NOT NULL
              AND p.vote_breakdown IS NOT NULL
              AND p.vote_breakdown->'all_votes'->>'ml' IS NOT NULL
        """, (CHECKPOINT_TS,))
        r2 = cur.fetchone()
        ctrl_t, ctrl_w, over_t, over_w, lon_t, lon_w = (int(x or 0) for x in r2)
        # WR per batch of 25 (trend)
        cur.execute("""
            WITH fresh AS (
                SELECT pr.is_win_size,
                    ROW_NUMBER() OVER (ORDER BY p.created_at) AS rn
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE p.created_at > %s AND pr.is_win_size IS NOT NULL
            )
            SELECT ((rn-1)/25)+1 AS batch,
                COUNT(*)::int AS n,
                SUM(CASE WHEN is_win_size THEN 1 ELSE 0 END)::int AS wins
            FROM fresh GROUP BY batch ORDER BY batch
        """, (CHECKPOINT_TS,))
        batch_rows = cur.fetchall()
        conn.close()
        batches = [{'batch': r[0], 'n': r[1], 'wins': r[2],
                    'wr': round(r[2]/r[1], 4) if r[1] else 0} for r in batch_rows]
        import math as _math
        _B = 0.375
        wr_fresh  = round(n_wins / n_eval, 4) if n_eval > 0 else None
        wr_ctrl   = round(ctrl_w / ctrl_t, 4) if ctrl_t > 0 else None
        wr_over   = round(over_w / over_t, 4) if over_t > 0 else None
        wr_lon    = round(lon_w  / lon_t,  4) if lon_t  > 0 else None
        z_ctrl    = round((wr_ctrl - _B) / _math.sqrt(_B * (1 - _B) / ctrl_t), 2) if wr_ctrl and ctrl_t >= 10 else None
        z_lon     = round((wr_lon  - _B) / _math.sqrt(_B * (1 - _B) / lon_t),  2) if wr_lon  and lon_t  >= 10 else None
        remain    = max(0, CHECKPOINT_N - n_fresh)
        eta_h     = round(remain / 10, 1)   # ~10 draws/h during game hours
        return jsonify({
            'n_fresh':     n_fresh,
            'n_target':    CHECKPOINT_N,
            'n_evaluated': n_eval,
            'n_wins':      n_wins,
            'wr_fresh':    wr_fresh,
            'eta_hours':   eta_h,
            'ready':       n_fresh >= CHECKPOINT_N,
            'progress':    round(min(n_fresh / CHECKPOINT_N, 1.0), 3),
            'ml_controls':  {'n': ctrl_t, 'wins': ctrl_w, 'wr': wr_ctrl, 'z': z_ctrl},
            'ml_overridden': {'n': over_t, 'wins': over_w, 'wr': wr_over},
            'ml_lon':       {'n': lon_t,  'wins': lon_w,  'wr': wr_lon,  'z': z_lon},
            'batches':      batches[-10:],
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/alert-log')
@limiter.limit("20 per minute")
def alert_log_api():
    """P104: Recent alert history from alert_log table."""
    import json as _json
    try:
        n    = min(int(request.args.get('n', 50)), 200)
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute(
            "SELECT id, alert_key, fired_at, message, metadata "
            "FROM alert_log ORDER BY fired_at DESC LIMIT %s" if USE_POSTGRES else
            "SELECT id, alert_key, fired_at, message, metadata "
            "FROM alert_log ORDER BY fired_at DESC LIMIT ?",
            (n,)
        )
        rows = cur.fetchall()
        conn.close()
        entries = []
        for row in rows:
            _id, key, fired_at, msg, meta = row
            entries.append({
                'id':       _id,
                'key':      key,
                'fired_at': fired_at.isoformat() if hasattr(fired_at, 'isoformat') else str(fired_at),
                'message':  msg or '',
                'metadata': meta if isinstance(meta, dict) else (_json.loads(meta) if meta else None),
            })
        return jsonify({'alerts': entries, 'total': len(entries)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/system-health-badge')
@limiter.limit("30 per minute")
def system_health_badge():
    """P103: WR50 + worst voter decay + draw gap → system status good/warn/bad."""
    from zoneinfo import ZoneInfo
    try:
        conn = db.get_connection()
        cur  = conn.cursor()

        # WR50
        cur.execute(
            "SELECT COALESCE(is_win_size, is_win, FALSE) "
            "FROM prediction_results ORDER BY draw_number DESC LIMIT 50"
        )
        wr_rows = cur.fetchall()
        wr50 = sum(1 for (w,) in wr_rows if w) / len(wr_rows) if wr_rows else None

        # Last draw gap
        gap_min = None
        active  = False
        if USE_POSTGRES:
            cur.execute("SELECT draw_time FROM draw_history ORDER BY draw_number DESC LIMIT 1")
            row = cur.fetchone()
            if row and row[0]:
                dt = row[0]
                if hasattr(dt, 'tzinfo') and dt.tzinfo is None:
                    dt = dt.replace(tzinfo=__import__('datetime').timezone.utc)
                from datetime import timezone as _tz
                gap_min = (datetime.now(_tz.utc) - dt.astimezone(_tz.utc)).total_seconds() / 60
                active  = 6 <= datetime.now(ZoneInfo('Asia/Ho_Chi_Minh')).hour < 22

        conn.close()

        # Worst voter decay
        worst_decay  = 1.0
        worst_voter  = None
        try:
            from prediction_service import _voter_decay_cache as _vdc
            for vname, info in (_vdc or {}).items():
                d = info.get('decay', 1.0) if isinstance(info, dict) else 1.0
                if d < worst_decay:
                    worst_decay  = d
                    worst_voter  = vname
        except Exception:
            pass

        # Status
        bad  = ((wr50  is not None and wr50 < 0.30)
                or worst_decay < 0.55
                or (active and gap_min is not None and gap_min > 20))
        warn = ((wr50  is not None and wr50 < 0.375)
                or worst_decay < 0.70
                or (active and gap_min is not None and gap_min > 12))

        status = 'bad' if bad else ('warn' if warn else 'good')

        return jsonify({
            'status':      status,
            'wr50':        round(wr50, 4) if wr50 is not None else None,
            'worst_decay': round(worst_decay, 3),
            'worst_voter': worst_voter,
            'gap_min':     round(gap_min, 1) if gap_min is not None else None,
            'active':      active,
        })
    except Exception as e:
        return jsonify({'error': str(e), 'status': 'unknown'}), 500


@app.route('/api/algo-impact')
@limiter.limit("20 per minute")
def algo_impact():
    """Compare win rate before vs after algorithm deployment.

    Auto-detects cutoff as the first draw with bocpd_dist in vote_breakdown
    (i.e., first draw after BOCPD/Hedge/Conformal went live).
    Optional: ?since_draw=N to override cutoff manually.
    """
    try:
        conn = db.get_connection()
        cur = conn.cursor()
        since_draw = request.args.get('since_draw', type=int)

        if USE_POSTGRES:
            if since_draw is None:
                cur.execute("""
                    SELECT MIN(draw_number) FROM predictions
                    WHERE vote_breakdown IS NOT NULL
                      AND vote_breakdown::jsonb->>'bocpd_dist' IS NOT NULL
                """)
                row = cur.fetchone()
                since_draw = row[0] if row and row[0] else None

            if since_draw is None:
                conn.close()
                return jsonify({'error': 'no_algo_data', 'message': 'Chưa có draw nào sau khi deploy thuật toán'}), 200

            cur.execute("""
                SELECT
                    p.draw_number < %s AS is_before,
                    COALESCE(pr.is_win_size, pr.is_win, FALSE) AS win
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE pr.actual_numbers IS NOT NULL
                ORDER BY p.draw_number
            """, (since_draw,))
        else:
            conn.close()
            return jsonify({'error': 'postgres_only'}), 200

        rows = cur.fetchall()
        conn.close()

        before = [w for is_b, w in rows if is_b]
        after  = [w for is_b, w in rows if not is_b]

        import math as _math

        def stats(wins_list):
            n = len(wins_list)
            if n == 0:
                return {'n': 0, 'wins': 0, 'wr': None, 'ci95': None}
            w = sum(wins_list)
            wr = w / n
            se = _math.sqrt(wr * (1 - wr) / n) if n > 0 else 0
            return {'n': n, 'wins': w, 'wr': round(wr, 4), 'ci95': round(1.96 * se, 4)}

        sb = stats(before)
        sa = stats(after)
        delta = None
        z = None
        if sb['wr'] is not None and sa['wr'] is not None:
            delta = round(sa['wr'] - sb['wr'], 4)
            pooled_se = _math.sqrt(
                (sb['wr'] * (1 - sb['wr']) / sb['n']) +
                (sa['wr'] * (1 - sa['wr']) / sa['n'])
            ) if sb['n'] > 0 and sa['n'] > 0 else None
            z = round(delta / pooled_se, 2) if pooled_se else None

        return jsonify({
            'since_draw': since_draw,
            'baseline': 0.375,
            'before': sb,
            'after': sa,
            'delta': delta,
            'z_score': z,
            'significant': abs(z) >= 1.96 if z is not None else False,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/hedge-weights')
@limiter.limit("30 per minute")
def hedge_weights_api():
    """Return current Hedge voter log-weights and multipliers from system_config."""
    try:
        from hedge_voter import load_hedge_weights, _HEDGE_WARMUP
        hw = load_hedge_weights(db)
        if hw is None:
            return jsonify({'available': False, 'n_updates': 0})
        mults = hw.get_multipliers()
        voters = sorted(
            [{'name': v, 'log_w': round(lw, 4), 'mult': mults.get(v, 1.0)}
             for v, lw in hw.log_weights.items()],
            key=lambda x: x['mult'], reverse=True
        )
        return jsonify({
            'available': True,
            'n_updates': hw.n_updates,
            'eta': hw.eta,
            'warmup_active': hw.n_updates >= _HEDGE_WARMUP,
            'voters': voters,
        })
    except Exception as e:
        return jsonify({'error': str(e), 'available': False}), 500


@app.route('/api/voter-health')
@limiter.limit("30 per minute")
def voter_health():
    """P86: Per-voter live health — WR (200 & 50 draws), streak, decay, effective multiplier."""
    from prediction_service import _get_voter_decay
    from collections import defaultdict
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        if USE_POSTGRES:
            cur.execute("""
                SELECT p.vote_breakdown,
                    CASE
                        WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 9  THEN 'NHO'
                        WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 11 THEN 'HOA'
                        ELSE 'LON'
                    END AS actual_size
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE p.vote_breakdown IS NOT NULL AND pr.actual_numbers IS NOT NULL
                ORDER BY p.draw_number DESC LIMIT 200
            """)
        else:
            cur.execute("""
                SELECT p.vote_breakdown, 'NHO' AS actual_size
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE p.vote_breakdown IS NOT NULL AND pr.actual_numbers IS NOT NULL
                ORDER BY p.draw_number DESC LIMIT 200
            """)
        rows = cur.fetchall()
        conn.close()

        acc200 = defaultdict(lambda: {'correct': 0, 'total': 0})
        acc50  = defaultdict(lambda: {'correct': 0, 'total': 0})
        for idx, (vb_raw, actual_size) in enumerate(rows):
            try:
                vb = json.loads(vb_raw) if isinstance(vb_raw, str) else vb_raw
                all_votes = (vb or {}).get('all_votes') or {}
                for vname, vsize in all_votes.items():
                    acc200[vname]['total']   += 1
                    if vsize == actual_size:
                        acc200[vname]['correct'] += 1
                    if idx < 50:
                        acc50[vname]['total']   += 1
                        if vsize == actual_size:
                            acc50[vname]['correct'] += 1
            except Exception:
                continue

        decay_cache = _get_voter_decay()
        BASELINE    = 0.375
        VOTER_ORDER = ['ml', 'markov', 'prior_nho', 'prior_lon', 'prior_hoa']

        voters_out = []
        for vname in VOTER_ORDER:
            a200 = acc200.get(vname, {'correct': 0, 'total': 0})
            a50  = acc50.get(vname,  {'correct': 0, 'total': 0})
            t200, w200 = a200['total'], a200['correct']
            t50,  w50  = a50['total'],  a50['correct']
            wr200 = w200 / t200 if t200 else None
            wr50  = w50  / t50  if t50  else None

            # WR-based multiplier (same formula as prediction_service)
            wr_mult = 1.0
            if t200 >= 20:
                wr_mult = round(max(0.4, min((wr200 or BASELINE) / BASELINE, 2.5)), 3)
            elif t200 >= 10 and wr200 is not None and wr200 < BASELINE - 0.08:
                wr_mult = round(max(0.4, wr200 / BASELINE), 3)

            dk      = decay_cache.get(vname, {'streak': 0, 'decay': 1.0})
            streak  = dk['streak']
            decay   = dk['decay']
            eff_mult = round(wr_mult * decay, 3)

            # Health status
            if eff_mult >= 1.2:
                status = 'good'
            elif eff_mult >= 0.8:
                status = 'ok'
            elif eff_mult >= 0.6:
                status = 'warn'
            else:
                status = 'bad'

            voters_out.append({
                'name':     vname,
                'wr200':    round(wr200, 4) if wr200 is not None else None,
                'wr50':     round(wr50,  4) if wr50  is not None else None,
                'n200':     t200,
                'n50':      t50,
                'wr_mult':  wr_mult,
                'streak':   streak,
                'decay':    decay,
                'eff_mult': eff_mult,
                'status':   status,
            })

        return jsonify({'voters': voters_out, 'baseline': BASELINE})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/voter-stats')
@limiter.limit("30 per minute")
def voter_stats():
    """Per-voter SIZE accuracy from stored vote_breakdown (last N predictions).
    Optional: ?since=2026-05-15T16:15:00 to filter to predictions after that timestamp.
    """
    try:
        n     = min(int(request.args.get('n', 200)), 1000)
        since = request.args.get('since', '')   # ISO timestamp string or empty
        conn  = db.get_connection()
        cur   = conn.cursor()

        # Fetch vote_breakdown + actual SIZE for each evaluated prediction
        if USE_POSTGRES:
            if since:
                cur.execute("""
                    SELECT p.vote_breakdown,
                        CASE
                            WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 9  THEN 'NHO'
                            WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 11 THEN 'HOA'
                            ELSE 'LON'
                        END AS actual_size
                    FROM predictions p
                    JOIN prediction_results pr ON pr.prediction_id = p.id
                    WHERE p.vote_breakdown IS NOT NULL
                      AND pr.actual_numbers IS NOT NULL
                      AND p.created_at > %s
                    ORDER BY p.draw_number DESC LIMIT %s
                """, (since, n))
            else:
                cur.execute("""
                    SELECT p.vote_breakdown,
                        CASE
                            WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 9  THEN 'NHO'
                            WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 11 THEN 'HOA'
                            ELSE 'LON'
                        END AS actual_size
                    FROM predictions p
                    JOIN prediction_results pr ON pr.prediction_id = p.id
                    WHERE p.vote_breakdown IS NOT NULL
                      AND pr.actual_numbers IS NOT NULL
                    ORDER BY p.draw_number DESC LIMIT %s
                """, (n,))
        else:
            cur.execute("""
                SELECT p.vote_breakdown, 'NHO' as actual_size
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE p.vote_breakdown IS NOT NULL AND pr.actual_numbers IS NOT NULL
                ORDER BY p.draw_number DESC LIMIT ?
            """, (n,))
        rows = cur.fetchall()
        conn.close()

        from collections import defaultdict
        voter_acc = defaultdict(lambda: {'correct': 0, 'total': 0,
                                          'votes_nho': 0, 'votes_hoa': 0, 'votes_lon': 0})
        markov_abstain_count = 0
        for vb_raw, actual_size in rows:
            try:
                vb = json.loads(vb_raw) if isinstance(vb_raw, str) else vb_raw
                if not vb:
                    continue
                if vb.get('markov_abstained'):
                    markov_abstain_count += 1
                # Use per-voter SIZE from all_votes if available, else fall back to majority voters list
                all_votes = vb.get('all_votes')
                if all_votes:
                    for voter_name, voted_size in all_votes.items():
                        voter_acc[voter_name]['total'] += 1
                        voter_acc[voter_name][f'votes_{voted_size.lower()}'] += 1
                        if voted_size == actual_size:
                            voter_acc[voter_name]['correct'] += 1
                else:
                    # Legacy: only majority voters recorded; use majority_size as their vote
                    m_size = vb.get('majority_size', 'NHO')
                    for voter_name in vb.get('voters', []):
                        voter_acc[voter_name]['total'] += 1
                        voter_acc[voter_name][f'votes_{m_size.lower()}'] += 1
                        if m_size == actual_size:
                            voter_acc[voter_name]['correct'] += 1
            except Exception:
                continue

        baseline = 0.375
        voter_list = []
        for name, acc in voter_acc.items():
            t = acc['total']
            if t < 5:
                continue
            wr = round(acc['correct'] / t, 4) if t > 0 else 0
            voter_list.append({
                'voter':    name,
                'total':    t,
                'wins':     acc['correct'],
                'win_rate': wr,
                'edge':     round(wr - baseline, 4),
                'size_votes': {
                    'NHO': acc['votes_nho'],
                    'HOA': acc['votes_hoa'],
                    'LON': acc['votes_lon'],
                }
            })
        voter_list.sort(key=lambda x: x['win_rate'], reverse=True)

        markov_abstain_rate = round(markov_abstain_count / len(rows), 3) if rows else 0
        note = f"Phân tích {len(rows)} kỳ. Mỗi voter: SIZE vote của họ vs actual SIZE (baseline ~37.5%)"
        resp = {'n': n, 'rows_analyzed': len(rows), 'voters': voter_list, 'note': note,
                'markov_abstain_rate': markov_abstain_rate,
                'markov_abstain_count': markov_abstain_count}
        if since:
            resp['since'] = since
        return jsonify(resp)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/size-accuracy')
@limiter.limit("30 per minute")
def size_accuracy():
    """Win rate broken down by predicted SIZE — last N evaluated predictions."""
    try:
        n    = min(int(request.args.get('n', 200)), 1000)
        conn = db.get_connection()
        cur  = conn.cursor()

        if USE_POSTGRES:
            cur.execute("""
                SELECT pred_size, COUNT(*) AS total,
                       SUM(CASE WHEN is_win THEN 1 ELSE 0 END) AS wins
                FROM (
                    SELECT pr.is_win,
                        CASE
                            WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 9  THEN 'NHO'
                            WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 11 THEN 'HOA'
                            ELSE 'LON'
                        END AS pred_size
                    FROM predictions p
                    JOIN prediction_results pr ON pr.prediction_id = p.id
                    WHERE pr.actual_numbers IS NOT NULL
                    ORDER BY p.draw_number DESC LIMIT %s
                ) sub
                GROUP BY pred_size
            """, (n,))
        else:
            cur.execute("""
                SELECT pred_size, COUNT(*) AS total,
                       SUM(CASE WHEN is_win THEN 1 ELSE 0 END) AS wins
                FROM (
                    SELECT pr.is_win,
                        CASE
                            WHEN (SELECT SUM(CAST(je.value AS INTEGER)) FROM json_each(p.predicted_numbers) je) <= 9  THEN 'NHO'
                            WHEN (SELECT SUM(CAST(je.value AS INTEGER)) FROM json_each(p.predicted_numbers) je) <= 11 THEN 'HOA'
                            ELSE 'LON'
                        END AS pred_size
                    FROM predictions p
                    JOIN prediction_results pr ON pr.prediction_id = p.id
                    WHERE pr.actual_numbers IS NOT NULL
                    ORDER BY p.draw_number DESC LIMIT ?
                ) sub
                GROUP BY pred_size
            """, (n,))

        rows = cur.fetchall()
        conn.close()

        result = {}
        total_all = sum(r[1] for r in rows)
        for pred_size, total, wins in rows:
            wr = round(wins / total, 4) if total > 0 else 0
            result[pred_size] = {
                'total':    total,
                'wins':     wins,
                'win_rate': wr,
                'pct_of_predictions': round(total / total_all * 100, 1) if total_all else 0,
            }
        # Baseline for each SIZE: P(correct SIZE guess) = P(actual SIZE)
        # NHO ≈ 37.5%, HOA ≈ 25%, LON ≈ 37.5%
        baselines = {'NHO': 0.375, 'HOA': 0.25, 'LON': 0.375}
        for cat, base in baselines.items():
            if cat in result:
                result[cat]['baseline'] = base
                result[cat]['edge'] = round(result[cat]['win_rate'] - base, 4)

        return jsonify({'n': n, 'by_size': result, 'total': total_all})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/size-confusion')
@limiter.limit("30 per minute")
def size_confusion():
    """3×3 confusion matrix: predicted SIZE vs actual SIZE for last N evaluated predictions."""
    try:
        n = min(int(request.args.get('n', 300)), 1000)
        conn = db.get_connection()
        cur  = conn.cursor()
        if USE_POSTGRES:
            cur.execute("""
                SELECT
                    CASE
                        WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 9  THEN 'NHO'
                        WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 11 THEN 'HOA'
                        ELSE 'LON'
                    END AS pred_size,
                    CASE
                        WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 9  THEN 'NHO'
                        WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 11 THEN 'HOA'
                        ELSE 'LON'
                    END AS actual_size
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE pr.actual_numbers IS NOT NULL
                ORDER BY p.draw_number DESC LIMIT %s
            """, (n,))
        else:
            cur.execute("""
                SELECT 'NHO' AS pred_size, 'NHO' AS actual_size
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE pr.actual_numbers IS NOT NULL
                ORDER BY p.draw_number DESC LIMIT ?
            """, (n,))
        rows = cur.fetchall()
        conn.close()

        from collections import defaultdict
        matrix = defaultdict(lambda: defaultdict(int))
        for pred_size, actual_size in rows:
            matrix[pred_size][actual_size] += 1

        # Total per predicted SIZE (row totals)
        sizes = ['NHO', 'HOA', 'LON']
        result = {}
        for ps in sizes:
            row_total = sum(matrix[ps].values())
            result[ps] = {
                'total': row_total,
                'actual': {asz: matrix[ps][asz] for asz in sizes},
                'win_rate': round(matrix[ps][ps] / row_total, 4) if row_total else 0,
            }
        return jsonify({'n': len(rows), 'matrix': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/daily-trend')
@limiter.limit("30 per minute")
def daily_trend():
    """Win rate per VN day for the last 7 days."""
    try:
        days = min(int(request.args.get('days', 7)), 30)
        conn = db.get_connection()
        cur  = conn.cursor()

        if USE_POSTGRES:
            cur.execute("""
                SELECT
                    (pr.created_at AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date AS vn_date,
                    COUNT(*) AS total,
                    SUM(CASE WHEN pr.is_win THEN 1 ELSE 0 END) AS wins
                FROM prediction_results pr
                WHERE pr.created_at >= NOW() - INTERVAL '1 day' * %s
                GROUP BY vn_date
                ORDER BY vn_date ASC
            """, (days,))
        else:
            cur.execute("""
                SELECT date(created_at, '+7 hours') AS vn_date,
                       COUNT(*) AS total,
                       SUM(CASE WHEN is_win THEN 1 ELSE 0 END) AS wins
                FROM prediction_results
                WHERE created_at >= datetime('now', '-' || ? || ' days')
                GROUP BY vn_date
                ORDER BY vn_date ASC
            """, (days,))

        rows = cur.fetchall()
        conn.close()

        trend = []
        for vn_date, total, wins in rows:
            wr = round(wins / total, 4) if total > 0 else 0
            trend.append({
                'date':     str(vn_date),
                'total':    total,
                'wins':     wins,
                'losses':   total - wins,
                'win_rate': wr,
            })
        return jsonify({'days': days, 'trend': trend})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/rolling-win-rate')
@limiter.limit("30 per minute")
def rolling_win_rate():
    """Win rate theo cửa sổ trượt theo số kỳ (rolling draw windows)."""
    try:
        window      = max(10, min(int(request.args.get('window', 30)), 100))
        num_windows = max(3,  min(int(request.args.get('windows', 7)), 20))
        limit_n     = window * num_windows
        conn = db.get_connection()
        cur  = conn.cursor()
        if USE_POSTGRES:
            cur.execute("""
                SELECT draw_number, is_win
                FROM prediction_results
                ORDER BY draw_number DESC
                LIMIT %s
            """, (limit_n,))
        else:
            cur.execute("""
                SELECT draw_number, is_win
                FROM prediction_results
                ORDER BY draw_number DESC
                LIMIT ?
            """, (limit_n,))
        rows = cur.fetchall()
        conn.close()

        windows_out = []
        for i in range(0, len(rows), window):
            batch = rows[i:i + window]
            if len(batch) < window // 2:
                continue
            wins  = sum(1 for _, w in batch if w)
            total = len(batch)
            windows_out.append({
                'window_idx': i // window + 1,
                'from_draw':  int(batch[-1][0]),
                'to_draw':    int(batch[0][0]),
                'total':      total,
                'wins':       wins,
                'win_rate':   round(wins / total, 4) if total else 0.0,
            })
        windows_out.reverse()  # oldest window first for trend chart
        return jsonify({'window_size': window, 'num_windows': num_windows,
                        'windows': windows_out, 'baseline': 0.375})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/confidence-trend')
@limiter.limit("20 per minute")
def confidence_trend():
    """P69: Rolling avg confidence vs actual SIZE win rate per batch."""
    try:
        n     = min(int(request.args.get('n', 500)), 2000)
        batch = max(10, min(int(request.args.get('batch', 50)), 200))
        conn  = db.get_connection()
        cur   = conn.cursor()

        if USE_POSTGRES:
            cur.execute("""
                SELECT p.draw_number, p.confidence, pr.is_win_size
                FROM predictions p
                LEFT JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE p.confidence IS NOT NULL
                ORDER BY p.draw_number DESC LIMIT %s
            """, (n,))
        else:
            cur.execute("""
                SELECT p.draw_number, p.confidence, pr.is_win_size
                FROM predictions p
                LEFT JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE p.confidence IS NOT NULL
                ORDER BY p.draw_number DESC LIMIT ?
            """, (n,))

        rows = list(reversed(cur.fetchall()))
        conn.close()

        if not rows:
            return jsonify({'batches': [], 'n': 0})

        batches = []
        for i in range(0, len(rows), batch):
            chunk = rows[i:i + batch]
            confs = [float(r[1]) for r in chunk if r[1] is not None]
            evals = [r[2] for r in chunk if r[2] is not None]
            avg_conf = round(sum(confs) / len(confs), 4) if confs else None
            avg_wr   = round(sum(1 for v in evals if v) / len(evals), 4) if evals else None
            batches.append({
                'draw_from':  chunk[0][0],
                'draw_to':    chunk[-1][0],
                'count':      len(chunk),
                'evaluated':  len(evals),
                'avg_conf':   avg_conf,
                'avg_wr':     avg_wr,
            })

        return jsonify({'batches': batches, 'n': len(rows), 'batch_size': batch})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def _check_size_bias():
    """Gửi Telegram alert nếu HOA predicted > threshold. Gọi từ health check."""
    global _last_bias_alert_ts
    try:
        now_t = _time.monotonic()
        if now_t - _last_bias_alert_ts < _BIAS_ALERT_COOLDOWN_SEC:
            return
        conn = db.get_connection()
        cur  = conn.cursor()
        if USE_POSTGRES:
            cur.execute("""
                SELECT
                    COUNT(*) as total,
                    SUM(CASE WHEN s<=9 THEN 1 ELSE 0 END) as nho,
                    SUM(CASE WHEN s BETWEEN 10 AND 11 THEN 1 ELSE 0 END) as hoa,
                    SUM(CASE WHEN s>=12 THEN 1 ELSE 0 END) as lon
                FROM (
                    SELECT (SELECT SUM(v::int) FROM json_array_elements_text(predicted_numbers::json) v) AS s
                    FROM predictions ORDER BY draw_number DESC LIMIT 50
                ) sub
            """)
        else:
            cur.execute("""
                SELECT COUNT(*) as total,
                    SUM(CASE WHEN s<=9 THEN 1 ELSE 0 END),
                    SUM(CASE WHEN s BETWEEN 10 AND 11 THEN 1 ELSE 0 END),
                    SUM(CASE WHEN s>=12 THEN 1 ELSE 0 END)
                FROM (
                    SELECT SUM(CAST(je.value AS INTEGER)) AS s
                    FROM predictions p, json_each(p.predicted_numbers) je
                    GROUP BY p.id ORDER BY p.draw_number DESC LIMIT 50
                ) sub
            """)
        row = cur.fetchone()
        conn.close()
        if not row or not row[0]:
            return
        total = row[0]
        hoa_pct = round((row[2] or 0) / total * 100, 1)
        if hoa_pct <= _BIAS_HOA_THRESHOLD:
            return
        _last_bias_alert_ts = now_t
        nho_pct = round((row[1] or 0) / total * 100, 1)
        lon_pct = round((row[3] or 0) / total * 100, 1)
        from telegram_bot import TelegramBot
        TelegramBot().send_message(
            f"⚠️ <b>Size Bias Alert</b>\n"
            f"HOA dự đoán: <b>{hoa_pct}%</b> (thực tế ~20%)\n"
            f"50 kỳ gần nhất: NHO={nho_pct}% HOA={hoa_pct}% LON={lon_pct}%\n"
            f"SizePredictor có thể cần điều chỉnh."
        )
    except Exception:
        pass


@app.route('/api/model-performance')
@limiter.limit("30 per minute")
@cache_resp(ttl=120)
def model_performance():
    """Win rate theo từng model."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT p.model_name,
                   COUNT(*)                                                     AS total,
                   COALESCE(SUM(CASE WHEN pr.is_win THEN 1 ELSE 0 END), 0)    AS wins
            FROM predictions p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            GROUP BY p.model_name
            ORDER BY wins DESC, total DESC
        """)
        rows = cur.fetchall()
        conn.close()

        models = []
        for name, total, wins in rows:
            wins = int(wins or 0)
            models.append({
                "model_name": name,
                "total":      total,
                "wins":       wins,
                "win_rate":   round(wins / total * 100, 1) if total else 0.0,
            })
        return jsonify({"models": models})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/sync-predictions', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def sync_predictions_endpoint():
    """
    Backfill predictions cho tất cả kỳ bị thiếu.
    Chạy ngầm (background thread) để không block request.
    """
    secret = request.headers.get("X-Trigger-Secret", "")
    if secret and secret != config.TRIGGER_SECRET:
        return jsonify({"error": "Unauthorized"}), 401

    # Kiểm tra gap trước
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM draw_history dh
            WHERE NOT EXISTS (
                SELECT 1 FROM predictions p WHERE p.draw_number = dh.draw_number
            )
        """)
        gap_count = cur.fetchone()[0]
        cur.execute("SELECT MAX(draw_number) FROM draw_history")
        max_draw = cur.fetchone()[0] or 0
        cur.execute("SELECT MAX(draw_number) FROM predictions")
        max_pred = cur.fetchone()[0] or 0
        conn.close()
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    if gap_count == 0:
        return jsonify({"status": "ok", "message": "Không có kỳ nào bị thiếu prediction.", "gap": 0})

    limit = int(request.args.get("limit", min(gap_count, 1000)))

    def _run_sync():
        try:
            from sync_predictions import run as sync_run
            sync_run(limit=limit)
        except Exception as e:
            import logging
            logging.getLogger(__name__).error("sync_predictions error: %s", e)

    import threading
    threading.Thread(target=_run_sync, daemon=True).start()

    return jsonify({
        "status":   "started",
        "message":  f"Đang backfill {limit} kỳ bị thiếu trong background...",
        "gap":       gap_count,
        "max_draw":  max_draw,
        "max_pred":  max_pred,
        "limit":     limit,
    })


@app.route('/api/ai-predict', methods=['GET'])
@limiter.limit("10 per minute")
def ai_predict_endpoint():
    """
    Dự đoán kỳ tiếp theo bằng LLM (OpenRouter / Groq / Gemini).
    Trả về: {"prediction":"NHO"|"HOA"|"LON", "confidence":75, "reason":"...", "pattern":"...", "next_draw":N}
    """
    try:
        from ai_predictor import ask_ai, load_recent_draws
        draws = load_recent_draws(n=30)
        if len(draws) < 15:
            return jsonify({"error": "Không đủ dữ liệu lịch sử (cần ít nhất 15 kỳ)"}), 400

        result = ask_ai(draws)
        if not result:
            return jsonify({
                "error": "Không có AI API khả dụng. Kiểm tra OPENROUTER_API_KEY / GROQ_API_KEY / GEMINI_API_KEY trong .env"
            }), 503

        result["next_draw"]    = draws[-1]["draw_number"] + 1
        result["last_n_draws"] = len(draws)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/prediction-gap')
def prediction_gap():
    """Kiểm tra số kỳ chưa có prediction."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("SELECT MAX(draw_number) FROM draw_history")
        max_draw = cur.fetchone()[0] or 0
        cur.execute("SELECT MAX(draw_number) FROM predictions")
        max_pred = cur.fetchone()[0] or 0
        cur.execute("""
            SELECT COUNT(*) FROM draw_history dh
            WHERE NOT EXISTS (SELECT 1 FROM predictions p WHERE p.draw_number = dh.draw_number)
        """)
        gap = cur.fetchone()[0]
        conn.close()
        return jsonify({"max_draw": max_draw, "max_pred": max_pred, "gap": gap})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/admin/test-triple-alert', methods=['POST'])
@limiter.limit("5 per hour")
def admin_test_triple_alert():
    """Test triple drought alert với gap giả. Body: {"gap": 25}"""
    if request.headers.get('X-Admin-Key') != config.ADMIN_SECRET_KEY:
        return jsonify({"error": "unauthorized"}), 401
    body = request.get_json(silent=True) or {}
    gap  = int(body.get('gap', 25))
    cum_prob = round((1 - 0.9722 ** gap) * 100, 1)
    next10   = round((1 - 0.9722 ** 10) * 100, 1)
    if gap >= _TRIPLE_DROUGHT_P95:
        level, note = "🚨", "cực hiếm — top 5% hạn hán lịch sử"
    elif gap >= _TRIPLE_DROUGHT_P90:
        level, note = "🔥", "hiếm — top 10% hạn hán lịch sử"
    elif gap >= 47:
        level, note = "⚠️", "trên trung bình (avg=47kỳ)"
    else:
        level, note = "🎲", "đang tích lũy (median=25kỳ)"
    from telegram_bot import TelegramBot
    from zoneinfo import ZoneInfo
    msg = (
        f"[TEST] {level} <b>Triple Alert · {gap} kỳ chưa có triple</b>\n"
        f"━━━━━━━━━━━━━━━━\n"
        f"Kỳ triple cuối: <b>#TEST</b> · {note}\n"
        f"📊 Xác suất tích lũy đã thoát: <b>{cum_prob}%</b>\n"
        f"🎲 P(triple trong 10 kỳ tới): <b>{next10}%</b>\n"
        f"💡 Mỗi kỳ luôn là 2.78% — cân nhắc <b>Bộ 3 bất kỳ ×20</b>\n"
        f"⏱ {datetime.now(ZoneInfo('Asia/Ho_Chi_Minh')).strftime('%H:%M %d/%m/%Y')} VN"
    )
    sent = TelegramBot().send_message(msg)
    return jsonify({"ok": sent, "gap": gap, "message": msg})


@app.route('/api/admin/reset-checkpoint', methods=['POST'])
@limiter.limit("10 per hour")
def admin_reset_checkpoint():
    """Reset checkpoint_ts và checkpoint_n trong system_config."""
    if request.headers.get('X-Admin-Key') != config.ADMIN_SECRET_KEY:
        return jsonify({"error": "unauthorized"}), 401
    body = request.get_json(silent=True) or {}
    n    = int(body.get('n', 200))
    ts   = body.get('ts') or datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        _save_checkpoint_config(cur, ts, n)
        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'checkpoint_ts': ts, 'checkpoint_n': n})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/backfill-draw-times', methods=['POST'])
@limiter.limit("2 per hour")
def backfill_draw_times():
    """
    One-time fix: recalculate draw_time for ALL draws.
    Groups by UTC date, sorts by draw_number, assigns 6-min offsets.
    Protected by ADMIN_SECRET_KEY header.
    """
    if request.headers.get('X-Admin-Key') != config.ADMIN_SECRET_KEY:
        return jsonify({"error": "unauthorized"}), 401

    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            UPDATE draw_history SET draw_time = sub.new_time
            FROM (
                SELECT
                    draw_number,
                    date_trunc('day', draw_time) +
                        (ROW_NUMBER() OVER (
                            PARTITION BY date_trunc('day', draw_time)
                            ORDER BY draw_number
                        ) - 1) * INTERVAL '6 minutes'
                    AS new_time
                FROM draw_history
            ) sub
            WHERE draw_history.draw_number = sub.draw_number
        """)
        updated = cur.rowcount
        conn.commit()
        conn.close()
        return jsonify({"status": "ok", "updated": updated,
                        "message": f"Recalculated draw_time for {updated} draws (6-min intervals per day)"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/retrain', methods=['GET', 'POST'])
@limiter.limit("3 per minute")
def retrain_models():
    """
    Force retrain toàn bộ Hybrid model (Markov + Cold + ML Ensemble)
    với 500 kỳ gần nhất. Chạy background để không block request.
    """
    secret = request.headers.get("X-Trigger-Secret", "")
    if secret and secret != config.TRIGGER_SECRET:
        return jsonify({"error": "Unauthorized"}), 401

    def _run_retrain():
        try:
            from prediction_service import _background_retrain
            _background_retrain()
        except Exception as e:
            import logging
            logging.getLogger(__name__).error("retrain error: %s", e)

    import threading
    threading.Thread(target=_run_retrain, daemon=True).start()

    try:
        from prediction_service import _last_retrain_time
        last_rt = _last_retrain_time.isoformat() if _last_retrain_time else None
    except Exception:
        last_rt = None

    return jsonify({
        "status":          "started",
        "message":         "Đang retrain model trong background (HybridModel + VotingEnsemble + SizePredictor)...",
        "last_retrain_at": last_rt,
    })


@app.route('/api/weekly-summary', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def weekly_summary():
    """Tổng kết tuần (7 ngày qua theo giờ VN). Cloud Scheduler có thể gọi mỗi Chủ nhật."""
    from datetime import datetime, timedelta, timezone
    from telegram_bot import TelegramBot

    try:
        from database import USE_POSTGRES as _USE_PG
        vn_tz   = timezone(timedelta(hours=7))
        now_vn  = datetime.now(vn_tz)

        conn = db.get_connection()
        cur  = conn.cursor()

        # Tổng 7 ngày
        if _USE_PG:
            cur.execute("""
                SELECT COUNT(*) AS total,
                       SUM(CASE WHEN pr.is_win THEN 1 ELSE 0 END) AS wins,
                       SUM(CASE WHEN COALESCE(pr.is_win_sum, FALSE) THEN 1 ELSE 0 END) AS sum_wins
                FROM prediction_results pr
                WHERE pr.created_at >= NOW() - INTERVAL '7 days'
            """)
        else:
            cur.execute("""
                SELECT COUNT(*), SUM(CASE WHEN is_win THEN 1 ELSE 0 END),
                       SUM(CASE WHEN COALESCE(is_win_sum, 0) THEN 1 ELSE 0 END)
                FROM prediction_results
                WHERE created_at >= datetime('now', '-7 days')
            """)
        row = cur.fetchone()
        total = row[0] or 0
        wins  = row[1] or 0
        sum_wins = row[2] or 0
        losses = total - wins
        wr = wins / total if total > 0 else 0
        sum_wr = sum_wins / total if total > 0 else 0

        # Ngày tốt nhất trong tuần
        if _USE_PG:
            cur.execute("""
                SELECT (created_at AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date AS d,
                       COUNT(*) AS t,
                       SUM(CASE WHEN is_win THEN 1 ELSE 0 END) AS w
                FROM prediction_results
                WHERE created_at >= NOW() - INTERVAL '7 days'
                GROUP BY d ORDER BY w::float/NULLIF(t,0) DESC LIMIT 1
            """)
        else:
            cur.execute("""
                SELECT date(created_at,'+7 hours') AS d, COUNT(*) AS t,
                       SUM(CASE WHEN is_win THEN 1 ELSE 0 END) AS w
                FROM prediction_results
                WHERE created_at >= datetime('now','-7 days')
                GROUP BY d ORDER BY CAST(w AS FLOAT)/MAX(t) DESC LIMIT 1
            """)
        best_row = cur.fetchone()
        best_day = f"{best_row[0]} ({best_row[2]}/{best_row[1]} = {best_row[2]/best_row[1]*100:.0f}%)" if best_row and best_row[1] else "N/A"

        conn.close()

        if total == 0:
            return jsonify({"status": "empty", "message": "Không có dữ liệu 7 ngày qua."})

        week_icon = "🏆" if wr >= 0.42 else ("✅" if wr >= 0.37 else "⚠️")
        msg = (
            f"{week_icon} <b>TỔNG KẾT TUẦN {now_vn.strftime('%d/%m/%Y')}</b>\n"
            f"━━━━━━━━━━━━━━━━━━━\n"
            f"📊 <b>Tổng kỳ (7 ngày):</b> {total}\n"
            f"✅ Thắng SIZE: <b>{wins}</b>  ({wr*100:.1f}%)\n"
            f"🎯 Đúng tổng: <b>{sum_wins}</b>  ({sum_wr*100:.1f}%)\n"
            f"❌ Thua: <b>{losses}</b>\n"
            f"━━━━━━━━━━━━━━━━━━━\n"
            f"🌟 Ngày tốt nhất: {best_day}\n"
            f"━━━━━━━━━━━━━━━━━━━\n"
            f"🕐 Báo cáo lúc {now_vn.strftime('%H:%M')} giờ VN"
        )

        tg   = TelegramBot()
        sent = tg.send_message(msg)

        return jsonify({
            "status":       "ok" if sent else "partial",
            "total":        total,
            "wins":         wins,
            "losses":       losses,
            "win_rate":     round(wr, 4),
            "sum_wins":     sum_wins,
            "sum_win_rate": round(sum_wr, 4),
            "best_day":     best_day,
            "telegram_sent": sent,
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/combo-stats', methods=['GET', 'POST'])
@limiter.limit("10 per minute")
def combo_stats():
    """
    Thống kê combo/sum/size gần đây và gửi qua Telegram.
    Có thể gọi thủ công hoặc schedule qua Cloud Scheduler.
    Query param: window=100 (số kỳ, mặc định 100)
    """
    from telegram_bot import TelegramBot
    try:
        window = int(request.args.get('window', 100))
        top_n  = int(request.args.get('top', 6))
        window = max(20, min(window, 500))
        top_n  = max(3, min(top_n, 20))
        draws = db.get_recent_draws(window)
        if draws.empty:
            return jsonify({"status": "error", "message": "Không có data"}), 500
        bot = TelegramBot()
        ok = bot.send_combo_stats(draws, top_n=top_n)
        return jsonify({"status": "ok" if ok else "telegram_error", "draws": len(draws), "top_n": top_n})
    except Exception as e:
        import logging as _logging
        _logging.getLogger(__name__).error("combo_stats error: %s", e)
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/daily-summary', methods=['GET', 'POST'])
@limiter.limit("10 per minute")
def daily_summary():
    """
    Tính tổng kết ngày (hôm nay theo giờ VN) và gửi qua Telegram.
    Cloud Scheduler có thể gọi lúc 23:55 hàng ngày.
    """
    import logging as _logging
    logger = _logging.getLogger('daily_summary')
    from datetime import datetime, timedelta, timezone
    from telegram_bot import TelegramBot

    try:
        from database import USE_POSTGRES as _USE_PG
        # Query cho hôm nay theo giờ Việt Nam (UTC+7)
        vn_tz = timezone(timedelta(hours=7))
        now_vn = datetime.now(vn_tz)
        date_str = now_vn.strftime('%Y-%m-%d')

        conn = db.get_connection()
        cur  = conn.cursor()

        _TODAY_FILTER = (
            "(pr.created_at AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date "
            "= (NOW() AT TIME ZONE 'Asia/Ho_Chi_Minh')::date"
        ) if _USE_PG else "date(created_at, '+7 hours') = date('now', '+7 hours')"
        _YEST_FILTER = (
            "(pr.created_at AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date "
            "= (NOW() AT TIME ZONE 'Asia/Ho_Chi_Minh')::date - INTERVAL '1 day'"
        ) if _USE_PG else "date(created_at, '+7 hours') = date('now', '+7 hours', '-1 day')"

        # ── Tổng kết hôm nay (dùng is_win_size cho SIZE win) ──────
        cur.execute(f"""
            SELECT COUNT(*),
                   SUM(CASE WHEN COALESCE(pr.is_win_size, pr.is_win, FALSE) THEN 1 ELSE 0 END),
                   SUM(CASE WHEN COALESCE(pr.is_win_sum, FALSE) THEN 1 ELSE 0 END)
            FROM prediction_results pr
            JOIN predictions p ON p.id = pr.prediction_id
            WHERE {_TODAY_FILTER}
        """)
        row = cur.fetchone()
        total    = row[0] or 0
        wins     = row[1] or 0
        sum_wins = row[2] or 0
        win_rate     = wins / total if total > 0 else 0
        sum_win_rate = sum_wins / total if total > 0 else 0
        losses   = total - wins

        if total == 0:
            conn.close()
            return jsonify({"status": "empty",
                            "message": f"Hôm nay ({date_str}) chưa có prediction nào.",
                            "total_predictions": 0})

        # ── Hôm qua (so sánh) ─────────────────────────────────────
        cur.execute(f"""
            SELECT COUNT(*),
                   SUM(CASE WHEN COALESCE(pr.is_win_size, pr.is_win, FALSE) THEN 1 ELSE 0 END)
            FROM prediction_results pr
            JOIN predictions p ON p.id = pr.prediction_id
            WHERE {_YEST_FILTER}
        """)
        yrow = cur.fetchone()
        ytotal = yrow[0] or 0
        ywins  = yrow[1] or 0
        ywr    = ywins / ytotal if ytotal > 0 else None

        # ── SIZE breakdown hôm nay ────────────────────────────────
        _size_expr = (
            "CASE WHEN SUM(v::int) OVER () <= 9 THEN 'NHO' "
            "WHEN SUM(v::int) OVER () <= 11 THEN 'HOA' ELSE 'LON' END"
        )
        if _USE_PG:
            cur.execute("""
                SELECT
                    CASE WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 9  THEN 'NHO'
                         WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 11 THEN 'HOA'
                         ELSE 'LON' END AS pred_size,
                    COUNT(*),
                    SUM(CASE WHEN COALESCE(pr.is_win_size, pr.is_win, FALSE) THEN 1 ELSE 0 END)
                FROM prediction_results pr
                JOIN predictions p ON p.id = pr.prediction_id
                WHERE (pr.created_at AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                      = (NOW() AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                GROUP BY pred_size ORDER BY COUNT(*) DESC
            """)
            size_rows = {r[0]: (r[1], r[2] or 0) for r in cur.fetchall()}
        else:
            size_rows = {}

        # ── Rolling win rate (last 100) ───────────────────────────
        cur.execute("""
            SELECT COALESCE(pr.is_win_size, pr.is_win, FALSE)
            FROM prediction_results pr
            WHERE pr.actual_numbers IS NOT NULL
            ORDER BY pr.draw_number DESC LIMIT 100
        """)
        last100 = [r[0] for r in cur.fetchall()]
        wr100 = sum(1 for w in last100 if w) / len(last100) if last100 else None

        # ── W/L trail + streak hôm nay ───────────────────────────
        cur.execute(f"""
            SELECT COALESCE(pr.is_win_size, pr.is_win, FALSE)
            FROM prediction_results pr
            WHERE {_TODAY_FILTER}
            ORDER BY pr.draw_number ASC
        """)
        wl_seq = [r[0] for r in cur.fetchall()]

        # ── P76: HOA stats hôm nay ────────────────────────────────
        hoa_pred = hoa_wins = 0
        if _USE_PG:
            cur.execute(f"""
                SELECT
                    SUM(CASE WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) BETWEEN 10 AND 11 THEN 1 ELSE 0 END),
                    SUM(CASE WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) BETWEEN 10 AND 11
                             AND  (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) BETWEEN 10 AND 11
                             THEN 1 ELSE 0 END)
                FROM prediction_results pr
                JOIN predictions p ON p.id = pr.prediction_id
                WHERE {_TODAY_FILTER}
            """)
            hr = cur.fetchone()
            hoa_pred = hr[0] or 0
            hoa_wins = hr[1] or 0

        # ── P76: Confidence gap hôm nay ───────────────────────────
        avg_conf_today = None
        if _USE_PG:
            cur.execute(f"""
                SELECT AVG(p.confidence)
                FROM prediction_results pr
                JOIN predictions p ON p.id = pr.prediction_id
                WHERE {_TODAY_FILTER} AND p.confidence IS NOT NULL
            """)
            cr = cur.fetchone()
            if cr and cr[0]: avg_conf_today = float(cr[0])

        # ── P76: Best/worst voter hôm nay ─────────────────────────
        best_voter = worst_voter = None
        if _USE_PG:
            cur.execute(f"""
                SELECT p.vote_breakdown,
                    CASE WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 9  THEN 'NHO'
                         WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 11 THEN 'HOA'
                         ELSE 'LON' END AS actual_size
                FROM prediction_results pr
                JOIN predictions p ON p.id = pr.prediction_id
                WHERE {_TODAY_FILTER} AND p.vote_breakdown IS NOT NULL AND pr.actual_numbers IS NOT NULL
            """)
            vrows = cur.fetchall()
            from collections import defaultdict as _dd
            vacc = _dd(lambda: {'w': 0, 't': 0})
            for vb_raw, act_sz in vrows:
                try:
                    vb  = json.loads(vb_raw) if isinstance(vb_raw, str) else vb_raw
                    avs = (vb or {}).get('all_votes') or {}
                    for vname, vsz in avs.items():
                        vacc[vname]['t'] += 1
                        if vsz == act_sz: vacc[vname]['w'] += 1
                except Exception:
                    continue
            rated_voters = [(n, d['w']/d['t'], d['t']) for n, d in vacc.items() if d['t'] >= 5]
            if rated_voters:
                best_voter  = max(rated_voters, key=lambda x: x[1])
                worst_voter = min(rated_voters, key=lambda x: x[1])

        # ── P76: LON excess từ adaptive cache ─────────────────────
        try:
            from prediction_service import _adaptive_thresh_cache as _atc
            _at = _atc or {}
        except Exception:
            _at = {}

        # ── P106: SIZE bias delta (pred vs actual today) ───────────
        size_actual_rows = {}
        if _USE_PG:
            cur.execute(f"""
                SELECT
                    CASE WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 9  THEN 'NHO'
                         WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 11 THEN 'HOA'
                         ELSE 'LON' END AS actual_size,
                    COUNT(*) AS cnt
                FROM prediction_results pr
                JOIN predictions p ON p.id = pr.prediction_id
                WHERE {_TODAY_FILTER} AND pr.actual_numbers IS NOT NULL
                GROUP BY actual_size
            """)
            size_actual_rows = {r[0]: r[1] for r in cur.fetchall()}

        # ── P106: Voter conf drift today (last 50 vote_breakdowns) ─
        voter_drift_lines = []
        if _USE_PG:
            cur.execute("""
                SELECT vote_breakdown FROM predictions
                WHERE vote_breakdown IS NOT NULL ORDER BY draw_number DESC LIMIT 50
            """)
            vb_drift_rows = cur.fetchall()
            from collections import defaultdict as _ddd
            _vc_d = _ddd(lambda: [[], []])
            for _idx, (vb_raw,) in enumerate(vb_drift_rows):
                try:
                    vb = json.loads(vb_raw) if isinstance(vb_raw, str) else (vb_raw or {})
                    _det = vb.get('all_votes_detail') or {}
                    _bkt = 0 if _idx < 25 else 1
                    for _vn, _vi in _det.items():
                        _c = _vi.get('conf', 0)
                        if _c: _vc_d[_vn][_bkt].append(float(_c))
                except Exception:
                    pass
            for _voter, (_rec, _pri) in _vc_d.items():
                if len(_rec) < 10 or len(_pri) < 10:
                    continue
                _rec_avg = sum(_rec) / len(_rec)
                _pri_avg = sum(_pri) / len(_pri)
                _drop = _pri_avg - _rec_avg
                if abs(_drop) >= 0.05:
                    arrow = '↓' if _drop > 0 else '↑'
                    badge = '🔴' if _drop > 0.10 else '🟡'
                    voter_drift_lines.append(f"  {badge} {_voter}: {_pri_avg*100:.0f}%→{_rec_avg*100:.0f}% {arrow}{abs(_drop)*100:.0f}pp")

        # ── P106: Alert digest today ────────────────────────────────
        alert_today_count = 0
        alert_recent_msg  = None
        if _USE_PG:
            cur.execute("""
                SELECT COUNT(*), MAX(message), MAX(fired_at)
                FROM alert_log
                WHERE fired_at AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh' >= NOW() AT TIME ZONE 'Asia/Ho_Chi_Minh' - INTERVAL '24 hours'
            """)
            ar = cur.fetchone()
            if ar:
                alert_today_count = ar[0] or 0
                alert_recent_msg  = ar[1]

        # ── WR theo giờ VN hôm nay ───────────────────────────────────
        hourly_rows = []
        if _USE_PG:
            cur.execute(f"""
                SELECT
                    EXTRACT(HOUR FROM dh.draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::int AS vn_hour,
                    COUNT(*) AS total,
                    SUM(CASE WHEN COALESCE(pr.is_win_size, pr.is_win, FALSE) THEN 1 ELSE 0 END) AS wins
                FROM prediction_results pr
                JOIN predictions p ON p.id = pr.prediction_id
                JOIN draw_history dh ON dh.draw_number = p.draw_number
                WHERE {_TODAY_FILTER}
                GROUP BY vn_hour
                ORDER BY vn_hour
            """)
            hourly_rows = [(r[0], r[1], r[2] or 0) for r in cur.fetchall()]

        # ── Streak cuối ngày (hiện tại) ───────────────────────────
        cur_streak_val = 0
        cur_streak_type = None
        if wl_seq:
            last_val = wl_seq[-1]
            cur_streak_type = 'win' if last_val else 'loss'
            for w in reversed(wl_seq):
                if w == last_val:
                    cur_streak_val += 1
                else:
                    break

        # ── Combo bộ 3 số ra ≥2 lần hôm nay ─────────────────────────
        combo_freq_today: list = []
        if _USE_PG:
            cur.execute("""
                SELECT
                    array_to_string(
                        ARRAY(SELECT v::int FROM jsonb_array_elements_text(numbers::jsonb) v ORDER BY v::int),
                        '-'
                    ) AS combo,
                    COUNT(*) AS freq
                FROM draw_history
                WHERE (draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                      = (NOW() AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                GROUP BY combo
                HAVING COUNT(*) >= 2
                ORDER BY freq DESC, combo
            """)
            combo_freq_today = [(r[0], r[1]) for r in cur.fetchall()]
        else:
            cur.execute("SELECT numbers FROM draw_history WHERE date(draw_time, '+7 hours') = date('now', '+7 hours')")
            import json as _json2
            _cf = {}
            for (nums_raw,) in cur.fetchall():
                try:
                    _nums = sorted(json.loads(nums_raw) if isinstance(nums_raw, str) else list(nums_raw))
                    _key = '-'.join(str(n) for n in _nums)
                    _cf[_key] = _cf.get(_key, 0) + 1
                except Exception:
                    pass
            combo_freq_today = sorted(
                [(k, v) for k, v in _cf.items() if v >= 2],
                key=lambda x: (-x[1], x[0])
            )

        # ── Query draws hôm nay để xuất Excel ────────────────────
        _draws_today = []
        if _USE_PG:
            cur.execute("""
                SELECT draw_number,
                       draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh',
                       numbers
                FROM draw_history
                WHERE (draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                      = (NOW() AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                ORDER BY draw_number ASC
            """)
            _draws_today = cur.fetchall()
            # Fallback: lấy từ prediction_results nếu draw_history trống
            if not _draws_today:
                cur.execute("""
                    SELECT pr.draw_number,
                           pr.created_at AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh',
                           pr.actual_numbers
                    FROM prediction_results pr
                    WHERE (pr.created_at AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                          = (NOW() AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                    ORDER BY pr.draw_number ASC
                """)
                _draws_today = cur.fetchall()
        else:
            cur.execute("""
                SELECT draw_number, datetime(draw_time, '+7 hours'), numbers
                FROM draw_history
                WHERE date(draw_time, '+7 hours') = date('now', '+7 hours')
                ORDER BY draw_number ASC
            """)
            _draws_today = cur.fetchall()

        conn.close()

        max_win = cur_win = max_loss = cur_loss = 0
        for w in wl_seq:
            if w:
                cur_win += 1; max_win = max(max_win, cur_win); cur_loss = 0
            else:
                cur_loss += 1; max_loss = max(max_loss, cur_loss); cur_win = 0

        # W/L trail string (last 20, newest right)
        trail = "".join("✅" if w else "❌" for w in wl_seq[-20:])

        # ── Build message ─────────────────────────────────────────
        BASELINE = 0.375
        status_icon = "🔥" if win_rate >= 0.44 else ("✅" if win_rate >= BASELINE else "⚠️")
        diff_today = (win_rate - BASELINE) * 100
        diff_str   = f"({'+'if diff_today>=0 else ''}{diff_today:.1f}% vs 37.5%)"

        # Yesterday comparison
        yest_line = ""
        if ywr is not None:
            yd = (win_rate - ywr) * 100
            yest_line = f"📅 Hôm qua: {ywr*100:.1f}%  →  {('↑' if yd>=0 else '↓')}{abs(yd):.1f}%\n"

        # SIZE breakdown + bias delta
        _SL = {'NHO': '🔵NHỎ', 'HOA': '🟡HÒA', 'LON': '🔴LỚN'}
        size_line = ""
        if size_rows:
            parts = []
            for sz in ['NHO', 'HOA', 'LON']:
                if sz in size_rows:
                    t, w = size_rows[sz]
                    parts.append(f"{_SL[sz]} {w}/{t}({w/t*100:.0f}%)" if t else f"{_SL[sz]} 0")
            size_line = "📐 SIZE: " + "  ".join(parts) + "\n"
            # Bias delta
            bias_parts = []
            act_total = sum(size_actual_rows.values()) or 1
            pred_total = sum(t for t, _ in size_rows.values()) or 1
            for sz in ['NHO', 'HOA', 'LON']:
                pred_pct = (size_rows.get(sz, (0, 0))[0] / pred_total * 100) if sz in size_rows else 0
                act_pct  = (size_actual_rows.get(sz, 0) / act_total * 100)
                delta    = pred_pct - act_pct
                arrow    = '↑' if delta > 2 else ('↓' if delta < -2 else '→')
                bias_parts.append(f"{_SL[sz]} Δ{delta:+.0f}%{arrow}")
            size_line += "  Δ bias: " + "  ".join(bias_parts) + "\n"

        # Rolling context
        wr100_line = f"📈 Rolling 100 kỳ: <b>{wr100*100:.1f}%</b>  {('✅' if wr100 >= BASELINE else '⚠️')}\n" if wr100 else ""

        # ── P76: Build extra lines ─────────────────────────────────
        hoa_line = ""
        if hoa_pred > 0:
            hoa_wr_str = f"{hoa_wins/hoa_pred*100:.0f}%" if hoa_pred >= 5 else "n<5"
            hoa_icon   = '🟢' if hoa_pred >= 5 and hoa_wins/hoa_pred >= 0.35 else '🔴'
            hoa_line   = f"{hoa_icon} HOA hôm nay: dự {hoa_pred} lần · thắng {hoa_wins} ({hoa_wr_str})\n"

        conf_line = ""
        if avg_conf_today is not None:
            gap      = avg_conf_today - win_rate
            gap_icon = '🟡' if abs(gap) < 0.05 else ('🔴' if gap > 0.05 else '🟢')
            conf_line = (f"{gap_icon} Confidence gap: avg conf {avg_conf_today*100:.1f}% "
                         f"vs WR {win_rate*100:.1f}% → gap <b>{gap*100:+.1f}%</b>\n")

        voter_line = ""
        if best_voter and worst_voter:
            bn, bwr, bt = best_voter
            wn, wwr, wt = worst_voter
            import html as _html2
            voter_line = (f"🏆 Best voter: <b>{_html2.escape(str(bn))}</b> {bwr*100:.0f}% ({bt} kỳ)\n"
                          f"📉 Worst voter: <b>{_html2.escape(str(wn))}</b> {wwr*100:.0f}% ({wt} kỳ)\n")

        excess_line = ""
        consec = _at.get('consecutive_excess', 0)
        tune_k = _at.get('tune_k', 0.0)
        lon_ex = _at.get('pred_lon_excess', 0.0)
        if consec > 0 or tune_k > 0.3:
            ex_icon     = '🔴' if consec >= 3 else ('🟡' if consec >= 1 else '🟢')
            excess_line = (f"{ex_icon} AutoTune: tune_k={tune_k:.2f} · "
                           f"LON Δ={lon_ex*100:+.1f}% · consec={consec}\n")

        # Voter drift digest
        drift_section = ""
        if voter_drift_lines:
            drift_section = "📉 <b>Voter conf drift (25→25 kỳ):</b>\n" + "\n".join(voter_drift_lines) + "\n"

        # Combo frequency today (≥2 lần)
        combo_section = ""
        if combo_freq_today:
            lines = [f"  {combo} ×{freq}" for combo, freq in combo_freq_today]
            combo_section = "🎲 <b>Bộ 3 số ra ≥2 lần hôm nay:</b>\n" + "\n".join(lines) + "\n"

        # Alert digest
        alert_section = ""
        if alert_today_count > 0:
            al_icon = '🔴' if alert_today_count >= 5 else ('🟡' if alert_today_count >= 2 else '🔵')
            alert_section = f"{al_icon} Alerts hôm nay: <b>{alert_today_count}</b>"
            if alert_recent_msg:
                alert_section += f" · Gần nhất: {alert_recent_msg[:60]}"
            alert_section += "\n"

        # WR theo giờ — top 3 tốt + top 3 tệ
        hourly_section = ""
        if hourly_rows and len(hourly_rows) >= 3:
            rated = [(h, t, w, w/t) for h, t, w in hourly_rows if t >= 5]
            if rated:
                best3  = sorted(rated, key=lambda x: -x[3])[:3]
                worst3 = sorted(rated, key=lambda x: x[3])[:3]
                best_str  = "  ".join(f"{h}h {wr*100:.0f}%({'✅' if wr>=0.375 else '⚠️'})" for h, t, w, wr in best3)
                worst_str = "  ".join(f"{h}h {wr*100:.0f}%({'❌'})" for h, t, w, wr in worst3)
                hourly_section = (
                    f"⏰ <b>WR theo giờ:</b>\n"
                    f"  🏆 Tốt:  {best_str}\n"
                    f"  💀 Tệ:   {worst_str}\n"
                )

        # Streak cuối ngày
        streak_eod_line = ""
        if cur_streak_type == 'loss' and cur_streak_val >= 2:
            streak_eod_line = f"❄️ Đang thua <b>{cur_streak_val} kỳ cuối</b> — chú ý đầu ngày mai\n"
        elif cur_streak_type == 'win' and cur_streak_val >= 3:
            streak_eod_line = f"🔥 Đang thắng <b>{cur_streak_val} kỳ cuối</b> — momentum tốt\n"

        # Checkpoint progress (show until reached) — include fresh WR + ml_controls
        checkpoint_line = ""
        if _USE_PG:
            import math as _math2
            try:
                conn2 = db.get_connection()
                cur2  = conn2.cursor()
                cur2.execute("SELECT COUNT(*) FROM predictions WHERE created_at > %s", (_CHECKPOINT_TS,))
                n_fresh = int(cur2.fetchone()[0])
                # Fresh WR + ml_controls + ML-LON
                cur2.execute("""
                    SELECT COUNT(*),
                           SUM(CASE WHEN pr.is_win_size THEN 1 ELSE 0 END),
                           SUM(CASE WHEN (p.vote_breakdown->'all_votes'->>'ml') =
                                        (p.vote_breakdown->>'final_size') THEN 1 ELSE 0 END),
                           SUM(CASE WHEN (p.vote_breakdown->'all_votes'->>'ml') =
                                        (p.vote_breakdown->>'final_size')
                                        AND pr.is_win_size THEN 1 ELSE 0 END),
                           SUM(CASE WHEN p.vote_breakdown->'all_votes'->>'ml' = 'LON'
                                        AND p.vote_breakdown->>'final_size' = 'LON' THEN 1 ELSE 0 END),
                           SUM(CASE WHEN p.vote_breakdown->'all_votes'->>'ml' = 'LON'
                                        AND p.vote_breakdown->>'final_size' = 'LON'
                                        AND pr.is_win_size THEN 1 ELSE 0 END)
                    FROM predictions p
                    JOIN prediction_results pr ON pr.prediction_id = p.id
                    WHERE p.created_at > %s AND pr.is_win_size IS NOT NULL
                      AND p.vote_breakdown IS NOT NULL
                """, (_CHECKPOINT_TS,))
                r2 = cur2.fetchone()
                conn2.close()
                n_eval2  = int(r2[0] or 0)
                n_wins2  = int(r2[1] or 0)
                ctrl_t2  = int(r2[2] or 0)
                ctrl_w2  = int(r2[3] or 0)
                lon_t2   = int(r2[4] or 0)
                lon_w2   = int(r2[5] or 0)
                wr_f     = n_wins2 / n_eval2 if n_eval2 else None
                wr_ctrl  = ctrl_w2 / ctrl_t2 if ctrl_t2 else None
                wr_f_str = f" · WR {wr_f*100:.1f}%" if wr_f is not None else ""
                ctrl_str = ""
                if wr_ctrl is not None and ctrl_t2 >= 5:
                    z_ctrl = (wr_ctrl - 0.375) / _math2.sqrt(0.375 * 0.625 / ctrl_t2)
                    ctrl_icon = '🔴' if wr_ctrl < 0.30 else ('🟡' if wr_ctrl < 0.375 else '🟢')
                    ctrl_str = f" · {ctrl_icon}ctrl {wr_ctrl*100:.0f}%(z{z_ctrl:+.1f})"
                if lon_t2 >= 10:
                    wr_lon2 = lon_w2 / lon_t2
                    z_lon2  = (wr_lon2 - 0.375) / _math2.sqrt(0.375 * 0.625 / lon_t2)
                    ctrl_str += f" LON z{z_lon2:+.1f}"
                if n_fresh < _CHECKPOINT_N:
                    pct = int(n_fresh / _CHECKPOINT_N * 100)
                    checkpoint_line = f"🔬 Checkpoint: <b>{n_fresh}/{_CHECKPOINT_N}</b> ({pct}%){wr_f_str}{ctrl_str}\n"
                else:
                    checkpoint_line = f"✅ Checkpoint {_CHECKPOINT_N}/{_CHECKPOINT_N}{wr_f_str}{ctrl_str} — sẵn sàng!\n"
            except Exception:
                pass

        # Escape các field động tránh lỗi HTML parse
        import html as _html
        def _esc(s): return _html.escape(str(s)) if s else ''

        msg = (
            f"{status_icon} <b>TỔNG KẾT {now_vn.strftime('%d/%m/%Y')}</b>\n"
            f"━━━━━━━━━━━━━━━━━━━\n"
            f"📊 Tổng: <b>{total} kỳ</b>  |  Thắng SIZE: <b>{wins}</b>  Thua: <b>{losses}</b>\n"
            f"🎯 Win rate: <b>{win_rate*100:.1f}%</b>  {diff_str}\n"
            f"{yest_line}"
            f"🎲 Đúng tổng: {sum_wins}/{total} ({sum_win_rate*100:.1f}%)\n"
            f"━━━━━━━━━━━━━━━━━━━\n"
            f"{size_line}"
            f"{hoa_line}"
            f"{wr100_line}"
            f"🔥 Win streak max: <b>{max_win}</b>  ❄️ Loss streak max: <b>{max_loss}</b>\n"
            f"{streak_eod_line}"
            f"━━━━━━━━━━━━━━━━━━━\n"
            f"{hourly_section}"
            f"{conf_line}"
            f"{voter_line}"
            f"{drift_section}"
            f"{excess_line}"
            f"{alert_section}"
            f"{combo_section}"
            f"{checkpoint_line}"
            f"━━━━━━━━━━━━━━━━━━━\n"
            f"20 kỳ cuối: {trail}\n"
            f"🕐 {now_vn.strftime('%H:%M')} VN"
        )

        # Gửi Telegram
        tg = TelegramBot()
        sent = tg.send_message(msg)

        # ── Build & gửi Excel file ────────────────────────────────
        excel_sent = False
        if _draws_today:
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                from io import BytesIO

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = f"Bingo18_{date_str}"

                # Header
                headers = ["Kỳ", "Giờ VN", "Số 1", "Số 2", "Số 3", "Tổng", "Size"]
                header_fill = PatternFill("solid", fgColor="1F4E79")
                header_font = Font(bold=True, color="FFFFFF")
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=h)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

                from openpyxl.styles import Border, Side

                # Size fill colors (nền hàng)
                size_fills = {
                    "NHO": PatternFill("solid", fgColor="DAEEF3"),
                    "HOA": PatternFill("solid", fgColor="FFFACD"),
                    "LON": PatternFill("solid", fgColor="FFE4E1"),
                }
                size_labels = {"NHO": "NHỎ", "HOA": "HÒA", "LON": "LỚN"}

                # Màu ô số trùng — hoàn toàn khác nhau
                fill_2dup = PatternFill("solid", fgColor="FF6F00")   # cam cháy   — 2 trùng
                fill_3dup = PatternFill("solid", fgColor="AD1457")   # hồng đậm   — 3 trùng
                font_dup  = Font(bold=True, color="FFFFFF")

                # Màu cột Tổng — mỗi giá trị 1 màu riêng biệt hoàn toàn
                sum_colors = {
                    3:  "4A148C",   # tím đậm
                    4:  "1565C0",   # xanh dương đậm
                    5:  "00838F",   # xanh ngọc
                    15: "1B5E20",   # xanh lá đậm
                    16: "E65100",   # cam đỏ đậm
                    17: "B71C1C",   # đỏ đậm
                    18: "212121",   # đen — cực hiếm
                }

                for row_idx, (draw_num, draw_time_vn, nums_raw) in enumerate(_draws_today, 2):
                    try:
                        nums = json.loads(nums_raw) if isinstance(nums_raw, str) else list(nums_raw)
                        nums = [int(n) for n in nums]
                    except Exception:
                        nums = []
                    n1 = nums[0] if len(nums) > 0 else ""
                    n2 = nums[1] if len(nums) > 1 else ""
                    n3 = nums[2] if len(nums) > 2 else ""
                    total_sum = sum(nums) if nums else ""
                    if isinstance(total_sum, int):
                        sz = "NHO" if total_sum <= 9 else ("HOA" if total_sum <= 11 else "LON")
                    else:
                        sz = ""

                    # Phát hiện số trùng
                    from collections import Counter as _Counter
                    cnt = _Counter(nums) if nums else {}
                    max_dup = max(cnt.values()) if cnt else 0
                    dup_nums = {n for n, c in cnt.items() if c >= 2}

                    time_str = str(draw_time_vn)[:16] if draw_time_vn else ""
                    row_data = [draw_num, time_str, n1, n2, n3, total_sum, size_labels.get(sz, sz)]
                    base_fill = size_fills.get(sz)

                    for col, val in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col, value=val)
                        cell.alignment = Alignment(horizontal="center")

                        # Cột số (3,4,5) — tô nếu số đó nằm trong tập trùng
                        if col in (3, 4, 5) and val != "" and max_dup >= 2:
                            if max_dup == 3:
                                cell.fill = fill_3dup
                                cell.font = font_dup
                            elif int(val) in dup_nums:
                                cell.fill = fill_2dup
                                cell.font = font_dup
                            elif base_fill:
                                cell.fill = base_fill
                        # Cột Tổng (6) — màu đặc biệt nếu hiếm
                        elif col == 6 and isinstance(val, int) and val in sum_colors:
                            cell.fill = PatternFill("solid", fgColor=sum_colors[val])
                            cell.font = Font(bold=True, color="FFFFFF", size=12)
                        elif base_fill:
                            cell.fill = base_fill

                # Column widths
                for col, width in zip("ABCDEFG", [8, 16, 7, 7, 7, 8, 8]):
                    ws.column_dimensions[col].width = width

                buf = BytesIO()
                wb.save(buf)
                excel_bytes = buf.getvalue()

                filename = f"Bingo18_{date_str}.xlsx"
                caption = f"📊 Dữ liệu kỳ quay ngày <b>{now_vn.strftime('%d/%m/%Y')}</b> · {len(_draws_today)} kỳ"
                excel_sent = tg.send_document(excel_bytes, filename, caption)
            except Exception as _ex:
                import traceback as _tb; _tb.print_exc()

        return jsonify({
            "status":            "ok" if sent else "partial",
            "date":              date_str,
            "total_predictions": total,
            "wins":              wins,
            "losses":            losses,
            "win_rate":          round(win_rate, 4),
            "sum_wins":          sum_wins,
            "sum_win_rate":      round(sum_win_rate, 4),
            "max_win_streak":    max_win,
            "max_loss_streak":   max_loss,
            "wr_last_100":       round(wr100, 4) if wr100 else None,
            "yesterday_wr":      round(ywr, 4) if ywr else None,
            "telegram_sent":     sent,
            "excel_sent":        excel_sent,
            "message":           msg if not sent else "Đã gửi Telegram",
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/learning-status')
def learning_status():
    """Trả về trạng thái learning của hệ thống."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()

        # Số kỳ học được trong 24h qua
        if db.__class__.__name__ == 'DatabaseManager':
            from database import USE_POSTGRES
            if USE_POSTGRES:
                cur.execute("""
                    SELECT COUNT(*) FROM prediction_results
                    WHERE created_at >= NOW() - INTERVAL '24 hours'
                """)
            else:
                cur.execute("""
                    SELECT COUNT(*) FROM prediction_results
                    WHERE created_at >= datetime('now', '-24 hours')
                """)
        learned_24h = cur.fetchone()[0]

        # Win rate 24h qua
        if USE_POSTGRES:
            cur.execute("""
                SELECT COUNT(*), SUM(CASE WHEN is_win THEN 1 ELSE 0 END)
                FROM prediction_results
                WHERE created_at >= NOW() - INTERVAL '24 hours'
            """)
        else:
            cur.execute("""
                SELECT COUNT(*), SUM(CASE WHEN is_win THEN 1 ELSE 0 END)
                FROM prediction_results
                WHERE created_at >= datetime('now', '-24 hours')
            """)
        row = cur.fetchone()
        total_24h = row[0] or 0
        wins_24h  = row[1] or 0
        win_rate_24h = (wins_24h / total_24h) if total_24h > 0 else 0

        # Win rate last 50 draws (draw-count window, not time — reflects post-fix accuracy)
        if USE_POSTGRES:
            cur.execute("""
                SELECT COUNT(*), SUM(CASE WHEN is_win THEN 1 ELSE 0 END),
                       SUM(CASE WHEN COALESCE(is_win_sum, FALSE) THEN 1 ELSE 0 END)
                FROM (
                    SELECT is_win, is_win_sum FROM prediction_results
                    ORDER BY draw_number DESC LIMIT 50
                ) sub
            """)
        else:
            cur.execute("""
                SELECT COUNT(*), SUM(CASE WHEN is_win THEN 1 ELSE 0 END),
                       SUM(CASE WHEN COALESCE(is_win_sum, 0) THEN 1 ELSE 0 END)
                FROM (
                    SELECT is_win, is_win_sum FROM prediction_results
                    ORDER BY draw_number DESC LIMIT 50
                ) sub
            """)
        row50 = cur.fetchone()
        total_50  = row50[0] or 0
        wins_50   = row50[1] or 0
        sum_wins_50 = row50[2] or 0
        win_rate_50 = (wins_50 / total_50) if total_50 > 0 else 0

        # Current win/loss streak (last draw decides streak type)
        cur.execute("""
            SELECT is_win FROM prediction_results
            ORDER BY draw_number DESC LIMIT 30
        """)
        streak_rows = [r[0] for r in cur.fetchall()]
        current_streak = 0
        streak_type = None
        if streak_rows:
            streak_type = 'win' if streak_rows[0] else 'loss'
            for w in streak_rows:
                if bool(w) == (streak_type == 'win'):
                    current_streak += 1
                else:
                    break

        # Markov transitions đã học
        cur.execute("SELECT COUNT(*) FROM markov_transitions")
        markov_states = cur.fetchone()[0]

        # Thời điểm prediction mới nhất
        cur.execute("SELECT MAX(prediction_time) FROM predictions")
        last_pred_time = cur.fetchone()[0]

        # Kỳ tiếp theo đang predict
        cur.execute("SELECT MAX(draw_number) FROM predictions")
        next_predicting = (cur.fetchone()[0] or 0)

        conn.close()

        return jsonify({
            "learned_last_24h":      learned_24h,
            "win_rate_last_24h":     round(win_rate_24h, 4),
            "total_wins_24h":        wins_24h,
            "win_rate_last_50":      round(win_rate_50, 4),
            "wins_last_50":          wins_50,
            "total_last_50":         total_50,
            "sum_wins_last_50":      sum_wins_50,
            "current_streak":        current_streak,
            "current_streak_type":   streak_type,
            "markov_states_learned": markov_states,
            "last_prediction_time":  str(last_pred_time) if last_pred_time else None,
            "currently_predicting":  next_predicting,
            "auto_retrain_interval": getattr(config, 'AUTO_RETRAIN_INTERVAL', 100),
            "last_retrain_at":       (lambda rt: rt.isoformat() if rt else None)(
                                         __import__('prediction_service')._last_retrain_time
                                     ),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500



@app.route('/api/backtest', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def backtest_models():
    """Backtest all models on a draw_number range. Params: from_draw, to_draw, step, window_size."""
    try:
        from models import MarkovModel, ColdNumberModel, MLEnsembleModel, HybridModel

        body = request.get_json(silent=True) or {}
        def _p(key, default):
            return int(request.args.get(key, body.get(key, default)))

        conn0 = db.get_connection()
        try:
            cur0 = conn0.cursor()
            cur0.execute("SELECT MAX(draw_number) FROM draw_history")
            max_draw = cur0.fetchone()[0] or 0
        finally:
            conn0.close()

        n_range   = _p('n_range',   0)          # convenience: last N draws
        to_draw   = _p('to_draw',   max_draw)
        from_draw = _p('from_draw', max(1, to_draw - (n_range if n_range > 0 else 999)))
        step      = max(1, _p('step', 1))
        win_sz    = max(100, _p('window_size', 1000))

        if to_draw <= from_draw:
            return jsonify({"error": "to_draw phải lớn hơn from_draw"}), 400
        if (to_draw - from_draw) > 10000:
            return jsonify({"error": "Range tối đa 10 000 kỳ"}), 400

        conn1 = db.get_connection()
        try:
            cur1 = conn1.cursor()
            ph   = db._ph()
            cur1.execute(
                f"SELECT draw_number, numbers FROM draw_history "
                f"WHERE draw_number BETWEEN {ph} AND {ph} ORDER BY draw_number ASC",
                (from_draw - 300, to_draw)
            )
            all_rows = cur1.fetchall()
        finally:
            conn1.close()

        draws_map   = {int(dn): (json.loads(raw) if isinstance(raw, str) else raw)
                       for dn, raw in all_rows}
        sorted_keys = sorted(draws_map)

        pre_keys = [k for k in sorted_keys if k < from_draw][-300:]
        pre_df   = pd.DataFrame(
            [{"draw_number": k, "numbers": draws_map[k]} for k in pre_keys]
        ) if pre_keys else pd.DataFrame(columns=["draw_number", "numbers"])

        ml_model = MLEnsembleModel()
        hybrid   = HybridModel()
        if len(pre_df) >= 50:
            ml_model.train(pre_df)
            hybrid.train(pre_df)

        model_configs = {
            "markov_order_2":        MarkovModel(order=2),
            "cold_number_window_30": ColdNumberModel(window_size=30),
            "ml_ensemble":           ml_model,
            "hybrid_model":          hybrid,
        }

        test_keys     = [k for k in sorted_keys if from_draw <= k <= to_draw][::step]
        WIN_THRESHOLD = getattr(config, 'WIN_THRESHOLD', 1)
        BASELINE      = 0.875

        overall    = {n: {"wins": 0, "total": 0} for n in model_configs}
        rolling    = {n: [] for n in model_configs}
        win_bucket = {n: {"wins": 0, "total": 0} for n in model_configs}
        win_start  = test_keys[0] if test_keys else from_draw

        for idx, draw_num in enumerate(test_keys):
            actual       = draws_map[draw_num]
            history_list = [draws_map[k] for k in sorted_keys if k < draw_num][-300:]
            if len(history_list) < 20:
                continue

            for name, model in model_configs.items():
                try:
                    preds = model.predict(history_list, draw_num)
                    if not preds:
                        continue
                    predicted, _ = preds[0]
                    is_win = Counter(predicted) == Counter(actual) if len(predicted) == 3 and len(actual) == 3 else False
                    overall[name]["total"]      += 1
                    win_bucket[name]["total"]   += 1
                    if is_win:
                        overall[name]["wins"]    += 1
                        win_bucket[name]["wins"] += 1
                except Exception:
                    pass

            if (idx + 1) % win_sz == 0 or idx == len(test_keys) - 1:
                label = f"{win_start}-{draw_num}"
                for n in model_configs:
                    t = win_bucket[n]["total"]
                    w = win_bucket[n]["wins"]
                    rolling[n].append({
                        "window":   label,
                        "wins":     w,
                        "total":    t,
                        "win_rate": round(w / t, 4) if t else None,
                    })
                    win_bucket[n] = {"wins": 0, "total": 0}
                if idx < len(test_keys) - 1:
                    win_start = test_keys[idx + 1]

        summary = {}
        for n, d in overall.items():
            t  = d["total"]
            w  = d["wins"]
            wr = w / t if t else 0
            rw  = [r["win_rate"] for r in rolling[n] if r["win_rate"] is not None]
            std = float(pd.Series(rw).std()) if len(rw) > 1 else 0.0
            summary[n] = {
                "win_rate":        round(wr, 4),
                "wins":            w,
                "total":           t,
                "vs_baseline":     round(wr - BASELINE, 4),
                "stability_std":   round(std, 4),
                "rolling_windows": rolling[n],
            }

        return jsonify({
            "from_draw":    from_draw,
            "to_draw":      to_draw,
            "step":         step,
            "draws_tested": len(test_keys),
            "window_size":  win_sz,
            "baseline":     BASELINE,
            "models":       summary,
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route('/api/combo-predict', methods=['GET'])
def combo_predict():
    """
    Dự đoán tổ hợp 2 số & 3 số (có thể trùng nhau) cho kỳ tiếp theo.

    Query params:
        window  — số kỳ lịch sử phân tích (mặc định 200)
        top_n   — số kết quả trả về mỗi loại (mặc định 10)

    Response:
        {
            "next_draw": 163168,
            "top_pairs":   [{combo, label, score, freq_long, confidence}, ...],
            "top_triples": [{combo, label, score, freq_long, confidence}, ...],
            "stats": { analyzed_draws, unique_pairs_seen, ... },
            "mode": "combo_filter"
        }
    """
    try:
        from combo_filter import run_combo_prediction
        window = int(request.args.get('window', 200))
        top_n  = int(request.args.get('top_n', 10))
        window = max(10, min(window, 2000))   # clamp 10..2000
        top_n  = max(1,  min(top_n,  56))     # clamp 1..56 (max bộ 3)
        result = run_combo_prediction(db, window=window, top_n=top_n)
        if "error" in result:
            return jsonify(result), 400
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/time_analysis')
def get_time_analysis():
    """
    FIX: PostgreSQL không có strftime() — dùng EXTRACT(HOUR FROM ...) thay thế.
    """
    try:
        conn   = db.get_connection()
        cursor = conn.cursor()
        if USE_POSTGRES:
            cursor.execute("""
                SELECT EXTRACT(HOUR FROM draw_time)::INTEGER AS hour,
                       COUNT(*), size_category
                FROM draw_history
                GROUP BY hour, size_category
                ORDER BY hour
            """)
        else:
            cursor.execute("""
                SELECT CAST(strftime('%H', draw_time) AS INTEGER) AS hour,
                       COUNT(*), size_category
                FROM draw_history
                GROUP BY hour, size_category
                ORDER BY hour
            """)
        results = cursor.fetchall()
        conn.close()

        time_data: dict = {}
        for row in results:
            hour = row[0]
            if hour not in time_data:
                time_data[hour] = {'hour': hour, 'total': 0, 'NHO': 0, 'HOA': 0, 'LON': 0}
            time_data[hour]['total']    += row[1]
            time_data[hour][row[2]]      = row[1]
        return jsonify(list(time_data.values()))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Analytics ────────────────────────────────────────────────
@app.route('/analytics')
def analytics_page():
    return render_template('analytics.html')


@app.route('/api/analytics/frequency-heatmap')
@limiter.limit("30 per minute")
def analytics_frequency_heatmap():
    """Frequency of each number (1-6) overall and per draw position."""
    try:
        window = min(int(request.args.get('window', 500)), 5000)
        conn = db.get_connection()
        cur  = conn.cursor()
        ph   = db._ph()
        cur.execute(
            f"SELECT numbers FROM draw_history ORDER BY draw_number DESC LIMIT {ph}",
            (window,)
        )
        rows = cur.fetchall()
        conn.close()

        overall  = {n: 0 for n in range(1, 7)}
        by_pos   = [{n: 0 for n in range(1, 7)} for _ in range(3)]
        total    = 0
        for (raw,) in rows:
            nums = json.loads(raw) if isinstance(raw, str) else raw
            total += 1
            for i, n in enumerate(nums[:3]):
                n = int(n)
                overall[n] = overall.get(n, 0) + 1
                if i < 3:
                    by_pos[i][n] = by_pos[i].get(n, 0) + 1

        return jsonify({
            "window": window,
            "total_draws": total,
            "overall": [{"number": n, "count": overall[n],
                         "pct": round(overall[n] / (total * 3) * 100, 2) if total else 0}
                        for n in range(1, 7)],
            "by_position": [
                {"position": i + 1,
                 "numbers": [{"number": n, "count": by_pos[i][n],
                              "pct": round(by_pos[i][n] / total * 100, 2) if total else 0}
                             for n in range(1, 7)]}
                for i in range(3)
            ],
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/analytics/patterns')
@limiter.limit("30 per minute")
def analytics_patterns():
    """Pair co-occurrence, size category distribution, gap (consecutive-draw streak) analysis."""
    try:
        window = min(int(request.args.get('window', 500)), 5000)
        conn = db.get_connection()
        cur  = conn.cursor()
        ph   = db._ph()
        cur.execute(
            f"SELECT numbers, size_category FROM draw_history ORDER BY draw_number DESC LIMIT {ph}",
            (window,)
        )
        rows = cur.fetchall()
        conn.close()

        from itertools import combinations
        pair_counts = {}
        size_dist   = {"NHO": 0, "HOA": 0, "LON": 0}
        sum_dist    = {}
        total = 0

        for raw, size in rows:
            nums = json.loads(raw) if isinstance(raw, str) else raw
            total += 1
            size_dist[size] = size_dist.get(size, 0) + 1
            s = sum(int(n) for n in nums)
            sum_dist[s] = sum_dist.get(s, 0) + 1
            for a, b in combinations(sorted(set(int(n) for n in nums)), 2):
                key = (a, b)
                pair_counts[key] = pair_counts.get(key, 0) + 1

        top_pairs = sorted(pair_counts.items(), key=lambda x: -x[1])[:15]
        return jsonify({
            "window": window,
            "total_draws": total,
            "top_pairs": [{"pair": list(k), "count": v,
                           "pct": round(v / total * 100, 2)} for k, v in top_pairs],
            "size_distribution": [
                {"category": k, "count": size_dist.get(k, 0),
                 "pct": round(size_dist.get(k, 0) / total * 100, 2) if total else 0}
                for k in ["NHO", "HOA", "LON"]
            ],
            "sum_distribution": sorted(
                [{"sum": s, "count": c, "pct": round(c / total * 100, 2)}
                 for s, c in sum_dist.items()],
                key=lambda x: x["sum"]
            ),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/analytics/accuracy-by-time')
@limiter.limit("30 per minute")
def analytics_accuracy_by_time():
    """Win rate over time (daily) and per model."""
    try:
        days = min(int(request.args.get('days', 30)), 180)
        conn = db.get_connection()
        cur  = conn.cursor()

        cur.execute("""
            SELECT DATE(p.prediction_time) AS day,
                   p.model_name,
                   COUNT(*) AS total,
                   COALESCE(SUM(CASE WHEN pr.is_win      THEN 1 ELSE 0 END), 0) AS wins,
                   COALESCE(SUM(CASE WHEN pr.is_win_size THEN 1 ELSE 0 END), 0) AS wins_size
            FROM predictions p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            WHERE p.prediction_time >= NOW() - INTERVAL '1 day' * %s
              AND pr.actual_numbers IS NOT NULL
            GROUP BY DATE(p.prediction_time), p.model_name
            ORDER BY day ASC
        """, (days,))
        rows = cur.fetchall()

        cur.execute("""
            SELECT p.model_name,
                   COUNT(*) AS total,
                   COALESCE(SUM(CASE WHEN pr.is_win      THEN 1 ELSE 0 END), 0) AS wins,
                   COALESCE(SUM(CASE WHEN pr.is_win_size THEN 1 ELSE 0 END), 0) AS wins_size
            FROM predictions p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            WHERE pr.actual_numbers IS NOT NULL
            GROUP BY p.model_name
            ORDER BY wins DESC
        """)
        model_rows = cur.fetchall()
        conn.close()

        day_agg: dict = {}
        for day, model, total, wins, wins_size in rows:
            d = str(day)
            if d not in day_agg:
                day_agg[d] = {"date": d, "total": 0, "wins": 0, "wins_size": 0}
            day_agg[d]["total"]      += total
            day_agg[d]["wins"]       += int(wins or 0)
            day_agg[d]["wins_size"]  += int(wins_size or 0)

        timeline = []
        for d, v in sorted(day_agg.items()):
            timeline.append({
                "date":          d,
                "total":         v["total"],
                "wins":          v["wins"],
                "win_rate":      round(v["wins"] / v["total"] * 100, 1) if v["total"] else 0,
                "win_rate_size": round(v["wins_size"] / v["total"] * 100, 1) if v["total"] else 0,
            })

        by_model = []
        for name, total, wins, wins_size in model_rows:
            total = total or 0
            by_model.append({
                "model_name":    name,
                "total":         total,
                "wins":          int(wins or 0),
                "wins_size":     int(wins_size or 0),
                "win_rate":      round(int(wins or 0) / total * 100, 1) if total else 0,
                "win_rate_size": round(int(wins_size or 0) / total * 100, 1) if total else 0,
            })

        return jsonify({"days": days, "timeline": timeline, "by_model": by_model,
                        "baseline": 87.5})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/analytics/export')
@limiter.limit("10 per minute")
def analytics_export():
    """Export draws or predictions as JSON (use browser download for CSV)."""
    import csv, io
    fmt   = request.args.get('format', 'json')   # json | csv
    kind  = request.args.get('type',   'draws')  # draws | predictions | results
    limit = min(int(request.args.get('limit', 1000)), 10000)

    try:
        conn = db.get_connection()
        cur  = conn.cursor()

        if kind == 'draws':
            cur.execute(
                "SELECT draw_number, numbers, draw_time, size_category "
                "FROM draw_history ORDER BY draw_number DESC LIMIT %s", (limit,)
            )
            cols = ["draw_number", "numbers", "draw_time", "size_category"]
            raw  = cur.fetchall()
            data = [dict(zip(cols, r)) for r in raw]
            for d in data:
                d["numbers"] = json.loads(d["numbers"]) if isinstance(d["numbers"], str) else d["numbers"]
                d["draw_time"] = str(d["draw_time"])
        elif kind == 'predictions':
            cur.execute(
                "SELECT p.draw_number, p.predicted_numbers, p.model_name, p.confidence, "
                "       p.prediction_time, pr.is_win, pr.is_win_size, pr.match_count "
                "FROM predictions p "
                "LEFT JOIN prediction_results pr ON pr.prediction_id = p.id "
                "ORDER BY p.draw_number DESC LIMIT %s", (limit,)
            )
            cols = ["draw_number", "predicted_numbers", "model_name", "confidence",
                    "prediction_time", "is_win", "is_win_size", "match_count"]
            raw  = cur.fetchall()
            data = [dict(zip(cols, r)) for r in raw]
            for d in data:
                d["predicted_numbers"] = json.loads(d["predicted_numbers"]) if isinstance(d["predicted_numbers"], str) else d["predicted_numbers"]
                d["prediction_time"] = str(d["prediction_time"])
        else:
            return jsonify({"error": "type must be draws or predictions"}), 400
        conn.close()

        if fmt == 'csv':
            if not data:
                return jsonify({"error": "no data"}), 404
            si  = io.StringIO()
            writer = csv.DictWriter(si, fieldnames=list(data[0].keys()))
            writer.writeheader()
            for row in data:
                flat = {k: (json.dumps(v) if isinstance(v, list) else v) for k, v in row.items()}
                writer.writerow(flat)
            output = si.getvalue()
            from flask import Response
            return Response(
                output,
                mimetype='text/csv',
                headers={"Content-Disposition": f"attachment; filename=bingo18_{kind}.csv"}
            )
        return jsonify({"type": kind, "count": len(data), "data": data})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Weekly leaderboard ────────────────────────────────────────
@app.route('/api/weekly-leaderboard')
@limiter.limit("30 per minute")
def weekly_leaderboard():
    """#25 Model WR for last 7 days, sorted by wr desc."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT p.model_name,
                   COUNT(*) AS total,
                   SUM(CASE WHEN pr.is_win_size THEN 1 ELSE 0 END) AS wins,
                   ROUND(AVG(p.confidence)::numeric, 3) AS avg_conf
            FROM predictions p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            JOIN draw_history dh ON dh.draw_number = p.draw_number
            WHERE pr.is_win_size IS NOT NULL
              AND dh.draw_time >= NOW() - INTERVAL '7 days'
            GROUP BY p.model_name
            ORDER BY (SUM(CASE WHEN pr.is_win_size THEN 1 ELSE 0 END)::float
                      / NULLIF(COUNT(*), 0)) DESC NULLS LAST
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        models = []
        for name, total, wins, avg_conf in rows:
            total = int(total or 0); wins = int(wins or 0)
            models.append({
                "model_name": name, "total": total, "wins": wins,
                "wr": round(wins / total * 100, 1) if total else 0,
                "avg_conf": float(avg_conf or 0),
            })
        return jsonify({"models": models})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Draw speed monitor ─────────────────────────────────────────
@app.route('/api/draw-speed')
@limiter.limit("30 per minute")
def draw_speed():
    """#26 Intervals between consecutive draws (minutes). Detect anomalies."""
    n = min(int(request.args.get('n', 100)), 500)
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT draw_number,
                   draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh' AS vn_time,
                   EXTRACT(EPOCH FROM (draw_time - LAG(draw_time) OVER (ORDER BY draw_number))) / 60
                       AS gap_min
            FROM draw_history
            ORDER BY draw_number DESC LIMIT %s
        """, (n,))
        rows = cur.fetchall()
        cur.close(); conn.close()
        gaps = [float(r[2]) for r in rows if r[2] is not None]
        if not gaps:
            return jsonify({"error": "no data"}), 404
        avg_g  = sum(gaps) / len(gaps)
        recent = gaps[:10]  # most recent 10 intervals
        anomalies = [g for g in gaps if g > avg_g * 2.5 or g < 1]
        points = [{"draw": int(r[0]), "gap_min": round(float(r[2]), 2)}
                  for r in rows if r[2] is not None]
        return jsonify({
            "avg_gap_min":    round(avg_g, 2),
            "min_gap_min":    round(min(gaps), 2),
            "max_gap_min":    round(max(gaps), 2),
            "recent_avg_min": round(sum(recent) / len(recent), 2) if recent else None,
            "anomaly_count":  len(anomalies),
            "points":         list(reversed(points[:50])),
            "n":              len(gaps),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Combo frequency rank ───────────────────────────────────────
@app.route('/api/combo-rank')
@limiter.limit("20 per minute")
@cache_resp(ttl=120)
def combo_rank():
    """#28 Top 10 + Bottom 10 combos by frequency in last N draws."""
    n = min(int(request.args.get('n', 5000)), 20000)
    limit = min(int(request.args.get('limit', 10)), 20)
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT numbers, COUNT(*) AS cnt
            FROM draw_history
            WHERE draw_number > (SELECT MAX(draw_number) - %s FROM draw_history)
            GROUP BY numbers
            ORDER BY cnt DESC
        """, (n,))
        rows = cur.fetchall()
        cur.close(); conn.close()
        def parse(r):
            nums = json.loads(r[0]) if isinstance(r[0], str) else r[0]
            return {"numbers": nums, "count": int(r[1])}
        top    = [parse(r) for r in rows[:limit]]
        bottom = [parse(r) for r in rows[-limit:] if int(r[1]) <= rows[-1][1] + 1]
        bottom = list(reversed(bottom))
        return jsonify({"top": top, "bottom": bottom, "n": n,
                        "total_combos": len(rows)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Pair frequency heatmap ────────────────────────────────────
@app.route('/api/pair-heatmap')
@limiter.limit("30 per minute")
@cache_resp(ttl=120)
def pair_heatmap():
    """Return 6×6 matrix: how often numbers i and j appear in same draw (last N draws)."""
    n = min(int(request.args.get('n', 500)), 5000)
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute(
            "SELECT numbers FROM draw_history ORDER BY draw_number DESC LIMIT %s", (n,)
        )
        rows = cur.fetchall()
        cur.close()
        conn.close()

        matrix = [[0]*6 for _ in range(6)]  # 0-indexed: num-1
        total  = 0
        for (numbers_raw,) in rows:
            nums = json.loads(numbers_raw) if isinstance(numbers_raw, str) else numbers_raw
            if not nums or len(nums) < 2:
                continue
            total += 1
            seen = sorted(set(int(x) for x in nums if 1 <= int(x) <= 6))
            for i in range(len(seen)):
                for j in range(i, len(seen)):
                    a, b = seen[i] - 1, seen[j] - 1
                    matrix[a][b] += 1
                    if a != b:
                        matrix[b][a] += 1

        # Also count single-number occurrences on diagonal (appears at all)
        # Recompute diagonal as: how many draws contain this number
        for i in range(6):
            matrix[i][i] = 0
        for (numbers_raw,) in rows:
            nums = json.loads(numbers_raw) if isinstance(numbers_raw, str) else numbers_raw
            for x in set(int(v) for v in (nums or []) if 1 <= int(v) <= 6):
                matrix[x-1][x-1] += 1

        return jsonify({"matrix": matrix, "n": total, "labels": list(range(1, 7))})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Today WR live ──────────────────────────────────────────────
@app.route('/api/today-wr')
@limiter.limit("60 per minute")
def today_wr():
    """Live WR for current VN day: predictions made for today's draws."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT
                COUNT(*)                                                AS total,
                SUM(CASE WHEN pr.is_win_size THEN 1 ELSE 0 END)       AS wins,
                SUM(CASE WHEN pr.is_win_size IS NULL THEN 1 ELSE 0 END) AS pending
            FROM predictions p
            JOIN draw_history dh ON dh.draw_number = p.draw_number
            LEFT JOIN prediction_results pr ON pr.prediction_id = p.id
            WHERE (dh.draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                  = (NOW() AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
        """)
        row = cur.fetchone()
        cur.close()
        conn.close()
        total, wins, pending = (int(row[0] or 0), int(row[1] or 0), int(row[2] or 0))
        evaluated = total - pending
        wr = round(wins / evaluated * 100, 1) if evaluated > 0 else None
        return jsonify({
            "total": total,
            "wins": wins,
            "evaluated": evaluated,
            "pending": pending,
            "wr": wr,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── P/L simulation ────────────────────────────────────────────
@app.route('/api/pl-simulation')
@limiter.limit("30 per minute")
def pl_simulation():
    """#19 Cumulative P/L if betting `bet` VND per kỳ on predicted SIZE."""
    n   = min(int(request.args.get('n', 500)), 5000)
    bet = max(1000, min(int(request.args.get('bet', 10000)), 1000000))
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT p.draw_number, COALESCE(pr.is_win_size, FALSE)
            FROM predictions p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            WHERE pr.is_win_size IS NOT NULL
            ORDER BY p.draw_number DESC LIMIT %s
        """, (n,))
        rows = list(reversed(cur.fetchall()))
        cur.close(); conn.close()
        cumulative, points = 0, []
        wins, total = 0, 0
        for draw_number, is_win in rows:
            total += 1
            if is_win:
                cumulative += bet; wins += 1
            else:
                cumulative -= bet
            points.append({"draw": int(draw_number), "pl": cumulative})
        wr = round(wins / total * 100, 1) if total else 0
        return jsonify({"points": points, "bet": bet, "final_pl": cumulative,
                        "wins": wins, "total": total, "wr": wr})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Number heat trend ─────────────────────────────────────────
@app.route('/api/number-trend')
@limiter.limit("30 per minute")
def number_trend():
    """#20 Compare number frequency: last 20 kỳ vs last 100 kỳ baseline."""
    recent_n   = int(request.args.get('recent', 20))
    baseline_n = int(request.args.get('baseline', 100))
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("SELECT numbers FROM draw_history ORDER BY draw_number DESC LIMIT %s",
                    (max(recent_n, baseline_n),))
        rows = cur.fetchall()
        cur.close(); conn.close()

        def freq(subset):
            counts = {i: 0 for i in range(1, 7)}
            total  = 0
            for (raw,) in subset:
                nums = json.loads(raw) if isinstance(raw, str) else raw
                for v in (nums or []):
                    n = int(v)
                    if 1 <= n <= 6:
                        counts[n] += 1; total += 1
            return counts, total

        r_c, r_t = freq(rows[:recent_n])
        b_c, b_t = freq(rows[:baseline_n])
        result = []
        for num in range(1, 7):
            rf = r_c[num] / r_t if r_t else 0
            bf = b_c[num] / b_t if b_t else 0
            delta = rf - bf
            result.append({
                "number": num,
                "recent_freq": round(rf, 3),
                "baseline_freq": round(bf, 3),
                "delta": round(delta, 3),
                "recent_count": r_c[num],
                "baseline_count": b_c[num],
                "trend": "hot" if delta > 0.015 else "cold" if delta < -0.015 else "neutral",
            })
        return jsonify({"numbers": result, "recent_n": recent_n, "baseline_n": baseline_n})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── WR by weekday ─────────────────────────────────────────────
@app.route('/api/wr-by-weekday')
@limiter.limit("20 per minute")
def wr_by_weekday():
    """#21 Win rate (is_win_size) grouped by VN weekday. DOW: 0=Sun…6=Sat."""
    n = min(int(request.args.get('n', 2000)), 10000)
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT EXTRACT(DOW FROM dh.draw_time AT TIME ZONE 'UTC'
                           AT TIME ZONE 'Asia/Ho_Chi_Minh')::int AS dow,
                   COALESCE(pr.is_win_size, FALSE) AS is_win
            FROM prediction_results pr
            JOIN draw_history dh ON dh.draw_number = pr.draw_number
            WHERE pr.is_win_size IS NOT NULL
            ORDER BY pr.draw_number DESC LIMIT %s
        """, (n,))
        rows = cur.fetchall()
        cur.close(); conn.close()
        from collections import defaultdict
        data = defaultdict(lambda: {"wins": 0, "total": 0})
        for dow, is_win in rows:
            d = int(dow)
            data[d]["total"] += 1
            if is_win: data[d]["wins"] += 1
        DAYS = ["CN", "T2", "T3", "T4", "T5", "T6", "T7"]
        days_out = []
        for d in range(7):
            t, w = data[d]["total"], data[d]["wins"]
            days_out.append({
                "dow": d, "label": DAYS[d], "total": t, "wins": w,
                "wr": round(w / t * 100, 1) if t >= 5 else None,
            })
        return jsonify({"days": days_out, "n": len(rows)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Current size streak (#33) ────────────────────────────────
@app.route('/api/current-streak')
@limiter.limit("60 per minute")
def current_streak():
    """#33 Trả về streak size hiện tại (chuỗi liên tiếp mới nhất)."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT size_category FROM draw_history
            ORDER BY draw_number DESC LIMIT 10
        """)
        rows = [r[0] for r in cur.fetchall()]
        conn.close()
        if not rows:
            return jsonify({'size': None, 'count': 0})
        current = rows[0]
        count = 1
        for sz in rows[1:]:
            if sz == current:
                count += 1
            else:
                break
        return jsonify({'size': current, 'count': count})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Alert config (read-only from env) ─────────────────────────
@app.route('/api/alert-config')
@limiter.limit("30 per minute")
def alert_config():
    """#22 Expose current alert thresholds (set via env vars on sync_to_supabase.py)."""
    import os as _os
    return jsonify({
        "ANNOUNCE_MIN_CONF":  float(_os.environ.get('ANNOUNCE_MIN_CONF', '0.0')),
        "streak_threshold":   4,    # hard-coded in check_size_streak
        "cold_threshold":     30,   # hard-coded in check_cold_numbers
        "hot_combo_min":      3,    # hard-coded in check_and_alert_hot_combo
        "gap_alert_minutes":  15,   # hard-coded in draw gap detector
        "_note": "Thay đổi bằng env var ANNOUNCE_MIN_CONF trên sync_to_supabase.py",
    })


# ── Combo comeback ────────────────────────────────────────────
@app.route('/api/combo-comeback')
@limiter.limit("20 per minute")
def combo_comeback():
    """#23 Combos that appeared ≥2x historically but absent for longest time."""
    history_n = min(int(request.args.get('history', 5000)), 20000)
    absence_n = int(request.args.get('absence', 100))
    limit_n   = min(int(request.args.get('limit', 10)), 20)
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT numbers,
                   COUNT(*)                                                   AS total_count,
                   MAX(draw_number)                                           AS last_seen,
                   (SELECT MAX(draw_number) FROM draw_history) - MAX(draw_number) AS absent_kỳ
            FROM draw_history
            WHERE draw_number > (SELECT MAX(draw_number) - %s FROM draw_history)
            GROUP BY numbers
            HAVING COUNT(*) >= 2
               AND MAX(draw_number) < (SELECT MAX(draw_number) - %s FROM draw_history)
            ORDER BY absent_kỳ DESC, total_count DESC
            LIMIT %s
        """, (history_n, absence_n, limit_n))
        rows = cur.fetchall()
        cur.close(); conn.close()
        combos = []
        for numbers_raw, total, last_seen, absent in rows:
            nums = json.loads(numbers_raw) if isinstance(numbers_raw, str) else numbers_raw
            combos.append({
                "numbers": nums,
                "total_count": int(total),
                "last_seen": int(last_seen),
                "absent_ky": int(absent),
            })
        return jsonify({"combos": combos, "history_n": history_n, "absence_min": absence_n})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── SSE: live draw stream ─────────────────────────────────────
@app.route('/api/sse/draws')
def sse_draws():
    """Server-Sent Events: push new draws to dashboard in real-time (poll DB every 6s)."""
    def generate():
        conn = None
        try:
            conn = db.get_connection()
            cur  = conn.cursor()
            cur.execute("SELECT MAX(draw_number) FROM draw_history")
            row = cur.fetchone()
            last_id = int(row[0]) if row and row[0] else 0
            cur.close()

            hb = 0  # heartbeat counter
            while True:
                try:
                    cur = conn.cursor()
                    cur.execute(
                        "SELECT draw_number, numbers, size_category, sum_value, draw_time "
                        "FROM draw_history WHERE draw_number > %s ORDER BY draw_number",
                        (last_id,)
                    )
                    rows = cur.fetchall()
                    cur.close()
                    for dn, numbers, size_cat, sum_val, draw_time in rows:
                        nums = json.loads(numbers) if isinstance(numbers, str) else numbers
                        payload = json.dumps({
                            "type": "new_draw",
                            "draw_number": dn,
                            "numbers": nums,
                            "size_category": size_cat,
                            "sum_value": sum_val,
                            "draw_time": str(draw_time),
                        })
                        yield f"data: {payload}\n\n"
                        last_id = dn
                except Exception:
                    try:
                        conn = db.get_connection()
                    except Exception:
                        pass

                hb += 1
                if hb >= 5:  # ~30s heartbeat
                    yield ": ping\n\n"
                    hb = 0
                _time.sleep(6)
        except GeneratorExit:
            pass
        finally:
            try:
                if conn:
                    conn.close()
            except Exception:
                pass

    return Response(
        stream_with_context(generate()),
        mimetype='text/event-stream',
        headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'},
    )


# ── Model win rate breakdown ───────────────────────────────────
@app.route('/api/model-wr')
@limiter.limit("30 per minute")
def model_wr_breakdown():
    """Win rate by model (is_win_size), last N predictions, with confidence avg."""
    n = min(int(request.args.get('n', 500)), 5000)
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT model_name,
                   COUNT(*)  AS total,
                   SUM(CASE WHEN is_win_size THEN 1 ELSE 0 END) AS wins,
                   ROUND(AVG(conf)::numeric, 3) AS avg_conf
            FROM (
                SELECT p.model_name,
                       COALESCE(pr.is_win_size, FALSE) AS is_win_size,
                       COALESCE(p.confidence, 0)        AS conf
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE pr.is_win_size IS NOT NULL
                ORDER BY p.id DESC
                LIMIT %s
            ) sub
            GROUP BY model_name
            ORDER BY (SUM(CASE WHEN is_win_size THEN 1 ELSE 0 END)::float / NULLIF(COUNT(*),0)) DESC NULLS LAST
        """, (n,))
        rows = cur.fetchall()
        cur.close()
        conn.close()
        models = []
        for name, total, wins, avg_conf in rows:
            wins  = int(wins or 0)
            total = int(total or 0)
            wr    = round(wins / total * 100, 1) if total else 0.0
            models.append({
                "model_name": name,
                "total":      total,
                "wins":       wins,
                "wr":         wr,
                "avg_conf":   float(avg_conf or 0),
            })
        return jsonify({"models": models, "n": n})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Hot combos today ──────────────────────────────────────────
@app.route('/api/hot-combos-today')
@limiter.limit("60 per minute")
def hot_combos_today():
    """Return combos that appeared ≥2 times in current VN date, sorted by count desc."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT
              (SELECT array_agg(x ORDER BY x::int) FROM jsonb_array_elements_text(numbers::jsonb) AS x) AS sorted_nums,
              COUNT(*) AS cnt
            FROM draw_history
            WHERE (draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                  = (NOW() AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
            GROUP BY sorted_nums
            HAVING COUNT(*) >= 2
            ORDER BY cnt DESC, sorted_nums
        """)
        rows = cur.fetchall()
        cur.close()
        combos = []
        for numbers, cnt in rows:
            nums = json.loads(numbers) if isinstance(numbers, str) else numbers
            combos.append({"numbers": nums, "count": int(cnt)})
        return jsonify({"date": None, "combos": combos, "total": len(combos)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Tiện ích dashboard ────────────────────────────────────────
def _load_dashboard() -> str:
    # Chỉ dùng templates/dashboard.html (bingo18_page.html đã xóa - file cũ)
    candidates = [
        os.path.join(os.path.dirname(__file__), "templates", "dashboard.html"),
        "/app/templates/dashboard.html",
    ]
    for path in candidates:
        if os.path.exists(path):
            with open(path, encoding="utf-8") as f:
                return f.read()
    return _fallback_dashboard()


def _fallback_dashboard() -> str:
    return '''<!DOCTYPE html>
<html lang="vi">
<head><meta charset="UTF-8"><title>Bingo 18 Predictor</title></head>
<body><h1>Dashboard đang tải... (Chưa tìm thấy giao diện HTML)</h1></body>
</html>'''


# ── Telegram helpers ──────────────────────────────────────────
def _tg_reply(bot_token: str, chat_id: str, text: str,
              reply_markup: dict = None, parse_mode: str = "HTML"):
    payload = {"chat_id": chat_id, "text": text, "parse_mode": parse_mode}
    if reply_markup:
        payload["reply_markup"] = reply_markup
    try:
        requests.post(
            f"https://api.telegram.org/bot{bot_token}/sendMessage",
            json=payload, timeout=10
        )
    except Exception:
        pass


def _tg_main_keyboard():
    return {
        "inline_keyboard": [
            [
                {"text": "🎯 Dự đoán", "callback_data": "/predict"},
                {"text": "📈 Win rate", "callback_data": "/winrate"},
            ],
            [
                {"text": "📋 10 kỳ gần nhất", "callback_data": "/history 10"},
                {"text": "🔍 Trạng thái", "callback_data": "/status"},
            ],
            [
                {"text": "📊 Thống kê", "callback_data": "/stats"},
                {"text": "🗳️ Voters", "callback_data": "/voters"},
                {"text": "🚫 Abstain", "callback_data": "/abstain"},
            ],
            [
                {"text": "📊 Compare 100", "callback_data": "/compare 100"},
                {"text": "📊 Compare 200", "callback_data": "/compare 200"},
            ],
            [
                {"text": "📈 Trend", "callback_data": "/trend"},
                {"text": "📅 Day/Hour", "callback_data": "/dow"},
                {"text": "🕐 Hourly WR", "callback_data": "/hourly"},
                {"text": "📐 Calibration", "callback_data": "/calibration"},
                {"text": "🔍 Explain", "callback_data": "/explain"},
            ],
            [
                {"text": "🏥 Health", "callback_data": "/health"},
                {"text": "🏆 Top", "callback_data": "/top"},
                {"text": "📋 Recap 100", "callback_data": "/recap 100"},
                {"text": "⚙️ AutoTune", "callback_data": "/autotune"},
                {"text": "🚨 Alerts", "callback_data": "/alerts"},
            ],
            [
                {"text": "🔬 Checkpoint", "callback_data": "/checkpoint"},
            ],
        ]
    }


def _tg_cmd_predict(conn, reply):
    import ast as _ast, json as _json
    cur = conn.cursor()
    cur.execute("""
        SELECT p.draw_number, p.predicted_numbers, p.model_name, p.confidence,
               p.created_at, pr.is_win, pr.actual_numbers, p.vote_breakdown
        FROM predictions p
        LEFT JOIN prediction_results pr ON p.draw_number = pr.draw_number
        ORDER BY p.draw_number DESC LIMIT 1
    """)
    row = cur.fetchone()
    if not row:
        reply("⚠️ Chưa có dự đoán nào trong DB.")
        return
    draw_num, pred_nums, model, conf, created_at, is_win, actual, vb_raw = row
    if isinstance(pred_nums, str):
        pred_nums = _ast.literal_eval(pred_nums)
    nums_str = " · ".join(str(n) for n in pred_nums)
    pred_sum = sum(pred_nums)
    size = "NHỎ" if pred_sum <= 9 else ("HÒA" if pred_sum <= 11 else "LỚN")
    size_emoji = {"NHỎ": "🔵", "HÒA": "🟡", "LỚN": "🔴"}.get(size, "⚪")
    conf_str = f"{conf:.1%}" if conf else "N/A"
    # Vote breakdown
    vote_line = ""
    try:
        vb = _json.loads(vb_raw) if isinstance(vb_raw, str) else (vb_raw or {})
        sw = vb.get('size_weights', {})
        sw_total = sum(sw.values()) or 1
        parts = [f"{'NHỎ' if k=='NHO' else ('HÒA' if k=='HOA' else 'LỚN')} {round(v/sw_total*100)}%"
                 for k, v in sw.items() if v > 0]
        if parts:
            vote_line = f"🗳 {' · '.join(parts)}\n"
    except Exception:
        pass
    # Result
    if is_win is None:
        verdict = "⏳ Chưa có kết quả"
    else:
        actual_list = actual if isinstance(actual, list) else _ast.literal_eval(actual or "[]")
        actual_str = " · ".join(str(n) for n in actual_list)
        actual_size = "NHỎ" if sum(actual_list) <= 9 else ("HÒA" if sum(actual_list) <= 11 else "LỚN")
        verdict = (f"✅ THẮNG  Thực tế: <b>{actual_str}</b> ({actual_size})"
                   if is_win else
                   f"❌ THUA  Thực tế: <b>{actual_str}</b> ({actual_size})")
    reply(
        f"🎯 <b>DỰ ĐOÁN KỲ #{draw_num}</b>\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"🔢 Số: <b>{nums_str}</b>  Tổng: {pred_sum}\n"
        f"{size_emoji} SIZE: <b>{size}</b>\n"
        f"📊 Tin cậy: <b>{conf_str}</b>\n"
        f"{vote_line}"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"{verdict}"
    )


def _tg_cmd_status(conn, reply):
    from zoneinfo import ZoneInfo
    vn_now = datetime.now(ZoneInfo("Asia/Ho_Chi_Minh"))
    cur = conn.cursor()
    # Kỳ mới nhất
    cur.execute("SELECT draw_number, draw_time FROM draw_history ORDER BY draw_number DESC LIMIT 1")
    row = cur.fetchone()
    last_draw, last_time = (row[0], row[1]) if row else (0, None)
    if last_time:
        if hasattr(last_time, 'tzinfo') and last_time.tzinfo is None:
            last_time = last_time.replace(tzinfo=timezone.utc)
        lag = round((datetime.now(timezone.utc) - last_time.astimezone(timezone.utc)).total_seconds() / 60, 1)
        last_time_vn = last_time.astimezone(ZoneInfo("Asia/Ho_Chi_Minh")).strftime("%H:%M %d/%m")
    else:
        lag, last_time_vn = None, "N/A"
    # Win rate 24h
    cur.execute("""
        SELECT COUNT(*),
               COALESCE(SUM(CASE WHEN is_win THEN 1 ELSE 0 END), 0)
        FROM prediction_results
        WHERE actual_numbers IS NOT NULL
          AND created_at > NOW() - INTERVAL '24 hours'
    """)
    r24 = cur.fetchone()
    n24, w24 = (r24[0], r24[1]) if r24 else (0, 0)
    wr24 = f"{w24/n24:.1%}" if n24 else "N/A"
    # Tổng dự đoán + checkpoint progress
    cur.execute("SELECT COUNT(*) FROM predictions")
    total_preds = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM predictions WHERE created_at > %s", (_CHECKPOINT_TS,))
    n_fresh = cur.fetchone()[0]
    vn_hour = vn_now.hour
    game_status = "🟢 Đang quay" if 6 <= vn_hour < 22 else "🔴 Đã đóng"
    lag_str = f"{lag} phút" if lag else "N/A"
    if n_fresh >= _CHECKPOINT_N:
        cp_line = f"🔬 Checkpoint: <b>✅ {n_fresh}/{_CHECKPOINT_N}</b> — sẵn sàng phân tích\n"
    else:
        pct_cp = round(n_fresh / _CHECKPOINT_N * 100)
        cp_line = f"🔬 Checkpoint: <b>{n_fresh}/{_CHECKPOINT_N}</b> ({pct_cp}%) — đợi thêm\n"
    reply(
        f"🔍 <b>TRẠNG THÁI HỆ THỐNG</b>\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"🕐 VN: {vn_now.strftime('%H:%M %d/%m/%Y')}\n"
        f"🎰 Game: {game_status}\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"📌 Kỳ cuối: <b>#{last_draw}</b> ({last_time_vn})\n"
        f"⏱ Lag sync: {lag_str}\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"📈 Win rate 24h: <b>{wr24}</b> ({w24}/{n24})\n"
        f"📦 Tổng dự đoán: {total_preds:,}\n"
        f"{cp_line}"
        f"🗄 DB: ✅ Online"
    )


def _tg_cmd_stats(conn, reply):
    cur = conn.cursor()
    # 24h
    cur.execute("""
        SELECT COUNT(*),
               COALESCE(SUM(CASE WHEN is_win THEN 1 ELSE 0 END),0),
               COALESCE(SUM(CASE WHEN is_win_size THEN 1 ELSE 0 END),0)
        FROM prediction_results
        WHERE actual_numbers IS NOT NULL
          AND created_at > NOW() - INTERVAL '24 hours'
    """)
    r = cur.fetchone(); n24,w24,ws24 = r if r else (0,0,0)
    # 7 ngày
    cur.execute("""
        SELECT COUNT(*),
               COALESCE(SUM(CASE WHEN is_win THEN 1 ELSE 0 END),0),
               COALESCE(SUM(CASE WHEN is_win_size THEN 1 ELSE 0 END),0)
        FROM prediction_results
        WHERE actual_numbers IS NOT NULL
          AND created_at > NOW() - INTERVAL '7 days'
    """)
    r = cur.fetchone(); n7,w7,ws7 = r if r else (0,0,0)
    # Model tốt nhất
    cur.execute("""
        SELECT model_name, win_rate, total_predictions
        FROM model_stats
        WHERE total_predictions > 0
        ORDER BY win_rate DESC LIMIT 3
    """)
    models = cur.fetchall()
    model_lines = "\n".join(
        f"  {'🥇' if i==0 else '🥈' if i==1 else '🥉'} {m[0]}: <b>{m[1]:.1%}</b> ({m[2]} kỳ)"
        for i, m in enumerate(models)
    ) if models else "  Chưa có dữ liệu"
    def pct(w, n): return f"{w/n:.1%}" if n else "N/A"
    reply(
        f"📊 <b>THỐNG KÊ HIỆU SUẤT</b>\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"<b>24 giờ qua:</b>\n"
        f"  Số/Size: <b>{pct(w24,n24)}</b> / {pct(ws24,n24)}  ({n24} kỳ)\n"
        f"<b>7 ngày qua:</b>\n"
        f"  Số/Size: <b>{pct(w7,n7)}</b> / {pct(ws7,n7)}  ({n7} kỳ)\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"<b>Top model:</b>\n{model_lines}"
    )


def _tg_cmd_history(conn, reply, n: int = 10):
    import ast as _ast
    from zoneinfo import ZoneInfo
    _SIZE_EMO = {"NHO": "🔵", "HOA": "🟡", "LON": "🔴"}
    def _size(nums): s = sum(nums); return "NHO" if s <= 9 else ("HOA" if s <= 11 else "LON")
    n = max(1, min(n, 20))
    cur = conn.cursor()
    cur.execute("""
        SELECT dh.draw_number, dh.draw_time, dh.numbers,
               p.predicted_numbers, pr.is_win_size
        FROM draw_history dh
        LEFT JOIN predictions p ON dh.draw_number = p.draw_number
        LEFT JOIN prediction_results pr ON dh.draw_number = pr.draw_number
        ORDER BY dh.draw_number DESC LIMIT %s
    """, (n,))
    rows = cur.fetchall()
    if not rows:
        reply("⚠️ Không có dữ liệu.")
        return
    lines = [f"📋 <b>{n} KỲ GẦN NHẤT</b>  (✅❌ = SIZE win/loss)"]
    for draw_num, draw_time, numbers, pred, is_win_size in rows:
        if isinstance(numbers, str):
            numbers = _ast.literal_eval(numbers)
        actual_emo = _SIZE_EMO.get(_size(numbers), "⚪")
        nums_str = "-".join(str(x) for x in numbers)
        if draw_time:
            if hasattr(draw_time, 'tzinfo') and draw_time.tzinfo is None:
                draw_time = draw_time.replace(tzinfo=timezone.utc)
            t = draw_time.astimezone(ZoneInfo("Asia/Ho_Chi_Minh")).strftime("%H:%M")
        else:
            t = "--:--"
        if pred is not None:
            try:
                pred_list = pred if isinstance(pred, list) else _ast.literal_eval(pred)
                pred_emo = _SIZE_EMO.get(_size(pred_list), "⚪")
            except Exception:
                pred_emo = "⚪"
            win_icon = "✅" if is_win_size else ("❌" if is_win_size is False else "⏳")
            lines.append(f"{win_icon} <b>#{draw_num}</b> {t}  {pred_emo}→{actual_emo}  <code>{nums_str}</code>")
        else:
            lines.append(f"➖ <b>#{draw_num}</b> {t}  {actual_emo}  <code>{nums_str}</code>")
    reply("\n".join(lines))


def _tg_cmd_winrate(conn, reply):
    """Rolling win rate theo 3 window 50-draw gần nhất + streak."""
    cur = conn.cursor()
    cur.execute("""
        SELECT p.draw_number, pr.is_win_size
        FROM predictions p
        JOIN prediction_results pr ON p.draw_number = pr.draw_number
        WHERE pr.actual_numbers IS NOT NULL
          AND pr.is_win_size IS NOT NULL
        ORDER BY p.draw_number DESC LIMIT 200
    """)
    rows = cur.fetchall()
    if not rows:
        reply("⚠️ Chưa đủ dữ liệu.")
        return

    results = [(dn, w) for dn, w in rows]  # newest first

    # Streak
    streak_type = "thắng" if results[0][1] else "thua"
    streak = 0
    for _, w in results:
        if bool(w) == (streak_type == "thắng"):
            streak += 1
        else:
            break
    streak_icon = "🔥" if streak_type == "thắng" else "❄️"

    # Windows of 50
    window = 50
    baseline = 0.375
    lines = [f"📈 <b>ROLLING WIN RATE (SIZE)</b>", f"Baseline: {baseline:.1%}", ""]
    for i in range(min(3, len(results) // window)):
        chunk = results[i * window:(i + 1) * window]
        wins = sum(1 for _, w in chunk if w)
        wr = wins / len(chunk)
        lo, hi = chunk[-1][0], chunk[0][0]
        bar = "▓" * round(wr * 20) + "░" * (20 - round(wr * 20))
        flag = "✅" if wr >= baseline else "❌"
        lines.append(f"{flag} #{lo}–{hi}\n   {bar} {wr:.1%} ({wins}/{len(chunk)})")

    # Last 20 trail
    trail = "".join("✅" if w else "❌" for _, w in results[:20])
    lines += ["", f"<b>20 kỳ gần nhất:</b>", trail,
              "", f"{streak_icon} Streak hiện tại: <b>{streak} {streak_type}</b>"]
    reply("\n".join(lines))


# ── Telegram /abstain ─────────────────────────────────────────
def _tg_cmd_abstain(conn, reply):
    """Markov abstain rate + effect on predicted SIZE distribution."""
    import json as _json
    cur = conn.cursor()

    if not config.DATABASE_URL:
        reply("⚠️ Chỉ hỗ trợ PostgreSQL.")
        return

    cur.execute("""
        SELECT p.vote_breakdown,
            CASE
                WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 9  THEN 'NHO'
                WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 11 THEN 'HOA'
                ELSE 'LON'
            END AS actual_size
        FROM predictions p
        JOIN prediction_results pr ON pr.prediction_id = p.id
        WHERE p.vote_breakdown IS NOT NULL AND pr.actual_numbers IS NOT NULL
        ORDER BY p.draw_number DESC LIMIT 200
    """)
    rows = cur.fetchall()
    if not rows:
        reply("⚠️ Chưa đủ dữ liệu.")
        return

    total = len(rows)
    abs_count   = {'all': 0, 'r50': 0}
    pred_cnt    = {'all': {'NHO': 0, 'HOA': 0, 'LON': 0},
                   'abs': {'NHO': 0, 'HOA': 0, 'LON': 0},
                   'voted': {'NHO': 0, 'HOA': 0, 'LON': 0}}
    act_cnt     = {'NHO': 0, 'HOA': 0, 'LON': 0}

    for i, (vb_raw, actual_size) in enumerate(rows):
        try:
            vb = _json.loads(vb_raw) if isinstance(vb_raw, str) else vb_raw
            if not vb:
                continue
            majority = vb.get('majority_size')
            abstained = vb.get('markov_abstained', False)
            if abstained:
                abs_count['all'] += 1
                if i < 50:
                    abs_count['r50'] += 1
            if majority:
                pred_cnt['all'][majority] = pred_cnt['all'].get(majority, 0) + 1
                if abstained:
                    pred_cnt['abs'][majority] = pred_cnt['abs'].get(majority, 0) + 1
                else:
                    pred_cnt['voted'][majority] = pred_cnt['voted'].get(majority, 0) + 1
            if actual_size:
                act_cnt[actual_size] = act_cnt.get(actual_size, 0) + 1
        except Exception:
            continue

    r50 = min(50, total)
    abs_rate_all = abs_count['all'] / total if total else 0
    abs_rate_50  = abs_count['r50'] / r50 if r50 else 0

    def dist_str(d):
        t = sum(d.values()) or 1
        return f"NHO {d.get('NHO',0)/t:.0%} · HOA {d.get('HOA',0)/t:.0%} · LON {d.get('LON',0)/t:.0%}"

    abs_icon = "🔴" if abs_rate_all >= 0.80 else "🟡" if abs_rate_all >= 0.50 else "🟢"
    trend = "↑" if abs_rate_50 > abs_rate_all + 0.05 else ("↓" if abs_rate_50 < abs_rate_all - 0.05 else "→")

    lines = [
        f"🚫 <b>MARKOV ABSTAIN TRACKER</b> (n={total})",
        "━━━━━━━━━━━━━━━━━━",
        f"{abs_icon} Abstain rate: <b>{abs_rate_all:.0%}</b> overall · <b>{abs_rate_50:.0%}</b> last 50 {trend}",
        f"  ({abs_count['all']}/{total} kỳ bị loại — fallback mode conf≤0.25)",
        "",
        "📊 <b>Predicted SIZE distribution:</b>",
        f"  Tất cả ({total}):  {dist_str(pred_cnt['all'])}",
        f"  Khi abstain ({abs_count['all']}): {dist_str(pred_cnt['abs'])}",
        f"  Khi voted  ({total-abs_count['all']}): {dist_str(pred_cnt['voted'])}",
        f"  Actual ({total}):    {dist_str(act_cnt)}",
        "",
        "💡 Mục tiêu: predicted LON ≈ actual LON (~40%)",
    ]
    reply("\n".join(lines))


# ── Telegram /voters ──────────────────────────────────────────
def _tg_cmd_voters(conn, reply, n: int = 200, since: str = None):
    """P99: Compact voter table — overall WR, per-SIZE WR, avg conf, decay, streak.
    Pass since=timestamp to filter to fresh-only data (post-checkpoint mode).
    """
    import json as _json
    from collections import defaultdict
    cur = conn.cursor()
    n   = max(50, min(n, 500))
    BASELINE = 0.375

    if not USE_POSTGRES:
        reply("⚠️ Chỉ hỗ trợ PostgreSQL.")
        return

    if since:
        cur.execute("""
            SELECT p.vote_breakdown,
                CASE
                    WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 9  THEN 'NHO'
                    WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 11 THEN 'HOA'
                    ELSE 'LON'
                END AS actual_size
            FROM predictions p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            WHERE p.vote_breakdown IS NOT NULL AND pr.actual_numbers IS NOT NULL
              AND p.created_at > %s
            ORDER BY p.draw_number DESC LIMIT %s
        """, (since, n))
    else:
        cur.execute("""
            SELECT p.vote_breakdown,
                CASE
                    WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 9  THEN 'NHO'
                    WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 11 THEN 'HOA'
                    ELSE 'LON'
                END AS actual_size
            FROM predictions p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            WHERE p.vote_breakdown IS NOT NULL AND pr.actual_numbers IS NOT NULL
            ORDER BY p.draw_number DESC LIMIT %s
        """, (n,))
    rows = cur.fetchall()
    if not rows:
        reply("⚠️ Chưa đủ dữ liệu vote_breakdown.")
        return

    # acc[voter] = {total, wins, conf_sum, conf_n,
    #               size_votes: {NHO:{t,w}, HOA:{t,w}, LON:{t,w}}}
    acc = defaultdict(lambda: {
        'total': 0, 'wins': 0,
        'conf_sum': 0.0, 'conf_n': 0,
        'size_votes': {'NHO': {'t': 0, 'w': 0},
                       'HOA': {'t': 0, 'w': 0},
                       'LON': {'t': 0, 'w': 0}},
    })
    streak_cnt  = {}
    streak_done = set()

    for vb_raw, actual_size in rows:
        try:
            vb      = _json.loads(vb_raw) if isinstance(vb_raw, str) else (vb_raw or {})
            votes   = vb.get('all_votes') or {}
            detail  = vb.get('all_votes_detail') or {}
            for vname, vsize in votes.items():
                a = acc[vname]
                a['total'] += 1
                sv = a['size_votes'].get(vsize)
                if sv is not None:
                    sv['t'] += 1
                    if vsize == actual_size:
                        sv['w'] += 1
                        a['wins'] += 1
                        streak_done.add(vname)
                    elif vname not in streak_done:
                        streak_cnt[vname] = streak_cnt.get(vname, 0) + 1
                # conf from all_votes_detail
                d = detail.get(vname) or {}
                c = d.get('conf', 0)
                if c:
                    a['conf_sum'] += float(c)
                    a['conf_n']   += 1
        except Exception:
            continue

    # Live decay from prediction_service cache
    try:
        from prediction_service import _voter_decay_cache as _vdc
        decay_cache = _vdc or {}
    except Exception:
        decay_cache = {}

    VOTER_ORDER = ['ml', 'ml_ensemble', 'markov', 'prior_nho', 'prior_lon', 'prior_hoa']
    # Include any voters not in fixed order
    all_voters = VOTER_ORDER + [v for v in acc if v not in VOTER_ORDER]

    fresh_tag = " · 🔬 FRESH post-checkpoint" if since else ""
    lines = [f"🗳️ <b>VOTERS v2</b> · {len(rows)} kỳ{fresh_tag}", "━━━━━━━━━━━━━━━━━━"]
    SIZE_ICON = {'NHO': '🔵', 'HOA': '🟡', 'LON': '🔴'}

    for vname in all_voters:
        a = acc.get(vname)
        if not a or a['total'] == 0:
            continue
        wr      = a['wins'] / a['total']
        avg_c   = a['conf_sum'] / a['conf_n'] if a['conf_n'] else 0
        streak  = streak_cnt.get(vname, 0)
        dc_info = decay_cache.get(vname, {})
        decay   = dc_info.get('decay', 1.0) if isinstance(dc_info, dict) else 1.0

        wr_icon  = '✅' if wr >= BASELINE else '❌'
        dc_icon  = '' if decay >= 0.95 else (' ⚠️' if decay >= 0.70 else ' 🔴')
        sk_str   = f' · ❄️−{streak}' if streak >= 3 else ''

        # per-SIZE accuracy: when voter voted S, how often was it right?
        sv = a['size_votes']
        size_parts = []
        for sz in ('NHO', 'HOA', 'LON'):
            t = sv[sz]['t']
            if t == 0:
                continue
            sz_wr = sv[sz]['w'] / t
            arrow = '↑' if sz_wr > BASELINE else ('↓' if sz_wr < BASELINE - 0.03 else '')
            size_parts.append(f"{SIZE_ICON[sz]}{sz} <b>{sz_wr*100:.0f}%</b>{arrow}")

        lines.append(
            f"{wr_icon} <b>{vname}</b>{sk_str}\n"
            f"   WR <b>{wr*100:.1f}%</b> · conf <b>{avg_c*100:.0f}%</b> · "
            f"decay <b>{decay:.2f}</b>{dc_icon}\n"
            f"   {' · '.join(size_parts)}"
        )

    # Adaptive state footer
    try:
        from prediction_service import _adaptive_thresh_cache as _atc
        at = _atc or {}
        if at:
            h = at.get('tod_hour')
            hs = f" h{h:02d}" if h is not None else ""
            # ML mult from latest vote_breakdown
            _ml_mult = 1.0
            try:
                cur.execute(
                    "SELECT vote_breakdown FROM predictions "
                    "WHERE vote_breakdown IS NOT NULL ORDER BY draw_number DESC LIMIT 1"
                )
                _vbr = cur.fetchone()
                if _vbr:
                    import json as _jv
                    _vb2 = _jv.loads(_vbr[0]) if isinstance(_vbr[0], str) else (_vbr[0] or {})
                    _ml_mult = float((_vb2.get('all_votes_detail') or {}).get('ml', {}).get('mult', 1.0))
            except Exception:
                pass
            lines += [
                "━━━━━━━━━━━━━━━━━━",
                f"⚙️ <b>Adaptive{hs}:</b> "
                f"nho_min {at.get('nho_share_min',0):.0%} · "
                f"LON Δ {at.get('pred_lon_excess',0)*100:+.0f}% · "
                f"ML×{_ml_mult:.3f}"
            ]
    except Exception:
        pass

    reply("\n".join(lines))


# ── Telegram /trend ───────────────────────────────────────────
def _tg_cmd_trend(conn, reply):
    """P97: WR trend (200/100/50), SIZE bias delta, voter conf shift, verdict."""
    import json as _json
    from zoneinfo import ZoneInfo
    from collections import defaultdict
    cur = conn.cursor()
    BASELINE = 0.375

    # ── 1. WR trend: windows 200 / 100 / 50 ──
    cur.execute(
        "SELECT draw_number, COALESCE(is_win_size, is_win, FALSE) "
        "FROM prediction_results ORDER BY draw_number DESC LIMIT 200"
    )
    pr_rows = cur.fetchall()  # newest first

    def _wr_window(rows, n):
        chunk = rows[:n]
        if not chunk:
            return None
        wins = sum(1 for _, w in chunk if w)
        return wins, len(chunk), wins / len(chunk)

    wr200 = _wr_window(pr_rows, 200)
    wr100 = _wr_window(pr_rows, 100)
    wr50  = _wr_window(pr_rows, 50)

    def _wr_icon(wr):
        return '🟢' if wr >= 0.40 else ('🟡' if wr >= BASELINE else '🔴')

    def _trend_arrow(a, b):
        if a is None or b is None:
            return ''
        diff = b[2] - a[2]
        return ' ↑↑' if diff > 0.04 else (' ↑' if diff > 0.01 else (' ↓↓' if diff < -0.04 else (' ↓' if diff < -0.01 else ' →')))

    wr_lines = []
    for label, w in [('200', wr200), ('100', wr100), ('50', wr50)]:
        if w:
            icon = _wr_icon(w[2])
            diff = (w[2] - BASELINE) * 100
            wr_lines.append(f"{icon} {label:>3} kỳ: <b>{w[2]*100:.1f}%</b> ({diff:+.0f}%){_trend_arrow({'200':None,'100':wr200,'50':wr100}[label], w)}")

    # ── 2. SIZE bias: predicted vs actual (last 100 evaluated) ──
    if USE_POSTGRES:
        cur.execute("""
            SELECT p.predicted_size, pr.actual_size
            FROM predictions p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            WHERE p.predicted_size IS NOT NULL AND pr.actual_size IS NOT NULL
            ORDER BY p.draw_number DESC LIMIT 100
        """)
    else:
        cur.execute("""
            SELECT p.predicted_size, pr.actual_size
            FROM predictions p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            WHERE p.predicted_size IS NOT NULL AND pr.actual_size IS NOT NULL
            ORDER BY p.draw_number DESC LIMIT 100
        """)
    size_rows = cur.fetchall()
    pred_counts: dict = defaultdict(int)
    act_counts:  dict = defaultdict(int)
    for ps, as_ in size_rows:
        pred_counts[ps] += 1
        act_counts[as_]  += 1
    n_size = len(size_rows) or 1
    bias_lines = []
    for sz, color in [('NHO', '🔵'), ('HOA', '🟡'), ('LON', '🔴')]:
        p_pct = pred_counts[sz] / n_size * 100
        a_pct = act_counts[sz]  / n_size * 100
        delta = p_pct - a_pct
        arrow = '↑' if delta > 2 else ('↓' if delta < -2 else '→')
        bias_lines.append(f"{color} {sz}: pred <b>{p_pct:.0f}%</b> vs act <b>{a_pct:.0f}%</b>  Δ<b>{delta:+.0f}%</b>{arrow}")

    # ── 3. Voter conf shift: last 50 preds split into 2 batches of 25 ──
    if USE_POSTGRES:
        cur.execute("""
            SELECT p.vote_breakdown
            FROM predictions p
            WHERE p.vote_breakdown IS NOT NULL
            ORDER BY p.draw_number DESC LIMIT 50
        """)
    else:
        cur.execute("""
            SELECT p.vote_breakdown
            FROM predictions p
            WHERE p.vote_breakdown IS NOT NULL
            ORDER BY p.draw_number DESC LIMIT 50
        """)
    vb_rows = cur.fetchall()  # newest first
    voter_confs: dict = defaultdict(lambda: [[], []])  # {voter: [[recent_25], [prior_25]]}
    for idx, (vb_raw,) in enumerate(vb_rows):
        try:
            vb = _json.loads(vb_raw) if isinstance(vb_raw, str) else vb_raw
            detail = vb.get('all_votes_detail', {})
            bucket = 0 if idx < 25 else 1
            for voter, info in detail.items():
                c = info.get('conf', 0)
                if c:
                    voter_confs[voter][bucket].append(float(c))
        except Exception:
            pass

    VOTER_ORDER = ['ml_ensemble', 'markov', 'prior_nho', 'prior_lon', 'prior_hoa']
    conf_lines = []
    for voter in VOTER_ORDER:
        buckets = voter_confs.get(voter)
        if not buckets or not buckets[0]:
            continue
        rec  = sum(buckets[0]) / len(buckets[0])
        pri  = sum(buckets[1]) / len(buckets[1]) if buckets[1] else None
        if pri is not None:
            delta = rec - pri
            arrow = '↑' if delta > 0.01 else ('↓' if delta < -0.01 else '→')
            badge = '🔴' if rec < 0.45 else ('🟡' if rec < 0.55 else '🟢')
            conf_lines.append(f"{badge} {voter:<13} {pri*100:.0f}%→<b>{rec*100:.0f}%</b> {arrow}")
        else:
            badge = '🔴' if rec < 0.45 else ('🟡' if rec < 0.55 else '🟢')
            conf_lines.append(f"{badge} {voter:<13} <b>{rec*100:.0f}%</b>")

    # ── 4. Verdict ──
    wr_now  = wr50[2] if wr50 else BASELINE
    wr_prev = wr100[2] if wr100 else BASELINE
    wr_trend_up = wr_now > wr_prev + 0.01
    wr_trend_dn = wr_now < wr_prev - 0.01
    nho_delta = pred_counts['NHO'] / n_size - act_counts['NHO'] / n_size
    lon_delta = pred_counts['LON'] / n_size - act_counts['LON'] / n_size
    bias_ok   = abs(nho_delta) < 0.05 and abs(lon_delta) < 0.05

    if wr_now >= 0.42 and wr_trend_up:
        verdict = "✅ Hệ thống đang TĂNG — tiếp tục theo dõi"
    elif wr_now < BASELINE and wr_trend_dn:
        verdict = "⚠️ Hệ thống đang SUY — xem xét /calibration"
    elif not bias_ok:
        worst = 'LON' if abs(lon_delta) > abs(nho_delta) else 'NHO'
        verdict = f"🔶 WR ổn nhưng SIZE BIAS {worst} — adaptive đang xử lý"
    else:
        verdict = "🔵 Hệ thống BÌNH THƯỜNG — không có cảnh báo"

    now_str = datetime.now(ZoneInfo('Asia/Ho_Chi_Minh')).strftime('%H:%M %d/%m')
    msg = (
        f"📈 <b>TREND</b> · {now_str}\n"
        f"━━━━━━━━━━━━━━━━\n"
        f"<b>Win Rate (cửa sổ):</b>\n"
        + '\n'.join(wr_lines) + '\n'
        + f"━━━━━━━━━━━━━━━━\n"
        f"<b>SIZE Bias (100 kỳ):</b>\n"
        + '\n'.join(bias_lines) + '\n'
        + f"━━━━━━━━━━━━━━━━\n"
        f"<b>Voter Conf (25→25 kỳ):</b>\n"
        + '\n'.join(conf_lines) + '\n'
        + f"━━━━━━━━━━━━━━━━\n"
        + verdict
    )
    reply(msg)


# ── Telegram /alerts ──────────────────────────────────────────
def _tg_cmd_alerts(conn, reply, n: int = 5):
    """P108: Last N alerts from alert_log + 24h count summary."""
    from zoneinfo import ZoneInfo
    import json as _json
    cur = conn.cursor()
    n   = max(3, min(n, 20))

    if not USE_POSTGRES:
        reply("⚠️ Chỉ hỗ trợ PostgreSQL.")
        return

    # Recent alerts
    cur.execute(
        "SELECT alert_key, fired_at, message, metadata "
        "FROM alert_log ORDER BY fired_at DESC LIMIT %s",
        (n,)
    )
    rows = cur.fetchall()

    # 24h count
    cur.execute("""
        SELECT COUNT(*) FROM alert_log
        WHERE fired_at >= NOW() - INTERVAL '24 hours'
    """)
    count_24h = (cur.fetchone() or [0])[0]

    # 7d count
    cur.execute("""
        SELECT COUNT(*) FROM alert_log
        WHERE fired_at >= NOW() - INTERVAL '7 days'
    """)
    count_7d = (cur.fetchone() or [0])[0]

    KEY_ICON = {
        'wr_drop':        ('📉', 'WR Drop'),
        'gap':            ('⏰', 'Sync Gap'),
        'momentum_NHO':   ('🔄', 'Bias NHO'),
        'momentum_HOA':   ('🔄', 'Bias HOA'),
        'momentum_LON':   ('🔄', 'Bias LON'),
    }
    def _icon_label(key):
        if key in KEY_ICON:
            return KEY_ICON[key]
        if key.startswith('voter_drift_'):
            return ('📉', f'Drift {key[12:]}')
        if key.startswith('momentum_'):
            return ('🔄', f'Bias {key[9:]}')
        return ('⚠️', key)

    now_str = datetime.now(ZoneInfo('Asia/Ho_Chi_Minh')).strftime('%H:%M %d/%m')
    lines   = [f"🚨 <b>ALERT LOG</b> · {now_str}",
               f"24h: <b>{count_24h}</b> · 7 ngày: <b>{count_7d}</b>",
               "━━━━━━━━━━━━━━━━"]

    if not rows:
        lines.append("✅ Không có alert nào được ghi.")
    else:
        for key, fired_at, msg, meta in rows:
            icon, label = _icon_label(key)
            # Convert fired_at to VN time
            try:
                if hasattr(fired_at, 'astimezone'):
                    vn_ts = fired_at.astimezone(ZoneInfo('Asia/Ho_Chi_Minh')).strftime('%d/%m %H:%M')
                else:
                    vn_ts = str(fired_at)[:16]
            except Exception:
                vn_ts = str(fired_at)[:16]

            detail = (msg or '').strip()
            if not detail and meta:
                try:
                    m = _json.loads(meta) if isinstance(meta, str) else meta
                    detail = ', '.join(f"{k}={v}" for k, v in list(m.items())[:3])
                except Exception:
                    pass

            lines.append(f"{icon} <b>{label}</b> · {vn_ts}\n   {detail or '—'}")

    lines.append("━━━━━━━━━━━━━━━━")
    lines.append("Dùng /alerts 10 để xem thêm")
    reply("\n".join(lines))


# ── Telegram /autotune ────────────────────────────────────────
def _tg_cmd_autotune(conn, reply):
    """P101: Current adaptive state + 3-batch trend + commentary."""
    import json as _json
    from zoneinfo import ZoneInfo
    cur = conn.cursor()

    # ── 1. Current state from cache ──
    try:
        from prediction_service import _adaptive_thresh_cache as _atc
        cur_state = dict(_atc or {})
    except Exception:
        cur_state = {}

    # ── 2. Batch trend: last 75 predictions → 3 batches of 25 ──
    cur.execute(
        "SELECT draw_number, vote_breakdown FROM predictions "
        "WHERE vote_breakdown IS NOT NULL ORDER BY draw_number DESC LIMIT 75"
        if USE_POSTGRES else
        "SELECT draw_number, vote_breakdown FROM predictions "
        "WHERE vote_breakdown IS NOT NULL ORDER BY draw_number DESC LIMIT 75"
    )
    vb_rows = list(reversed(cur.fetchall()))  # oldest first

    FIELDS  = ['tune_k', 'nho_share_min', 'pred_lon_excess', 'consecutive_excess']
    batches = []
    for i in range(0, len(vb_rows), 25):
        chunk = vb_rows[i:i + 25]
        if len(chunk) < 10:
            continue
        sums  = {f: 0.0 for f in FIELDS}
        count = 0
        for dn, vb_raw in chunk:
            try:
                vb = _json.loads(vb_raw) if isinstance(vb_raw, str) else (vb_raw or {})
                at = vb.get('adaptive') or {}
                for f in FIELDS:
                    v = at.get(f)
                    if v is not None:
                        sums[f] += float(v)
                count += 1
            except Exception:
                pass
        if count:
            batches.append({f: sums[f] / count for f in FIELDS} | {'draw_start': chunk[0][0], 'draw_end': chunk[-1][0]})

    # ── 3. Format ──
    now_str = datetime.now(ZoneInfo('Asia/Ho_Chi_Minh')).strftime('%H:%M %d/%m')

    # Current state block
    tk   = cur_state.get('tune_k', 0)
    nm   = cur_state.get('nho_share_min', 0)
    lx   = cur_state.get('pred_lon_excess', 0)
    nx   = cur_state.get('pred_nho_excess', 0)
    ce   = cur_state.get('consecutive_excess', 0)
    hour = cur_state.get('tod_hour')
    hr_s = f" (h{hour:02d})" if hour is not None else ""

    # Fetch ML multiplier from latest prediction vote_breakdown
    ml_mult = 1.0
    try:
        cur.execute(
            "SELECT vote_breakdown FROM predictions WHERE vote_breakdown IS NOT NULL "
            "ORDER BY draw_number DESC LIMIT 1"
        )
        _vb_r = cur.fetchone()
        if _vb_r:
            import json as _jj
            _vb2 = _jj.loads(_vb_r[0]) if isinstance(_vb_r[0], str) else (_vb_r[0] or {})
            ml_mult = float((_vb2.get('all_votes_detail') or {}).get('ml', {}).get('mult', 1.0))
    except Exception:
        pass

    tk_icon  = '🔴' if tk >= 0.80 else ('🟡' if tk >= 0.40 else '🟢')
    ce_icon  = '🔴' if ce >= 3 else ('🟡' if ce >= 1 else '🟢')
    lx_icon  = '🔴' if lx > 0.08 else ('🟡' if lx > 0.03 else '🟢')
    ml_icon  = '🔴' if ml_mult < 0.75 else ('🟡' if ml_mult < 0.95 else '🟢')

    cur_lines = [
        f"⚙️ <b>AUTO-TUNE</b>{hr_s} · {now_str}",
        "━━━━━━━━━━━━━━━━",
        "<b>Trạng thái hiện tại:</b>",
        f"  {tk_icon} tune_k:      <b>{tk:.2f}</b>",
        f"  📊 nho_min:    <b>{nm:.1%}</b>",
        f"  {lx_icon} LON Δ:      <b>{lx*100:+.1f}%</b>",
        f"  📊 NHO Δ:     <b>{nx*100:+.1f}%</b>",
        f"  {ce_icon} consec:     <b>{ce}</b> chu kỳ excess",
        f"  {ml_icon} ML mult:    <b>{ml_mult:.3f}×</b>  (1.0=OK · <0.8=bị phạt)",
    ]

    # Trend table (up to 3 batches)
    if len(batches) >= 2:
        cur_lines += ["━━━━━━━━━━━━━━━━", "<b>Xu hướng (lô 25 kỳ):</b>"]
        labels = [f"#{b['draw_start']}" for b in batches[-3:]]
        hdr    = "  " + "  ".join(f"{l:>8}" for l in labels)
        cur_lines.append(hdr)

        ROWS = [
            ('tune_k',         'tune_k',   lambda v: f"{v:.2f}"),
            ('nho_share_min',  'nho_min',  lambda v: f"{v:.0%}"),
            ('pred_lon_excess','LON Δ',    lambda v: f"{v*100:+.1f}%"),
            ('consecutive_excess','consec',lambda v: f"{v:.0f}"),
        ]
        for key, label, fmt in ROWS:
            vals = [b.get(key, 0) for b in batches[-3:]]
            # pad to 3 if fewer batches
            while len(vals) < 3:
                vals.insert(0, None)
            cells = [f"{fmt(v):>8}" if v is not None else "       –" for v in vals]
            # trend arrow between last two
            if vals[-1] is not None and vals[-2] is not None:
                diff = vals[-1] - vals[-2]
                arrow = '↑' if diff > 0.005 else ('↓' if diff < -0.005 else '→')
            else:
                arrow = ''
            cur_lines.append(f"  {label:<9} {'  '.join(cells)} {arrow}")

    # ── 4. Commentary ──
    if not cur_state:
        verdict = "⚪ Chưa có dữ liệu — chạy ít nhất 1 prediction cycle."
    elif ce >= 3 and tk >= 0.70:
        verdict = "🔴 Hệ thống đang tích cực chống LON bias — tune_k cao."
    elif ce >= 1 and lx > 0.05:
        verdict = "🟡 LON excess đang được xử lý — theo dõi vài kỳ nữa."
    elif nx > 0.15:
        verdict = "🟡 NHO đang được predict quá nhiều so với thực tế."
    elif tk < 0.10 and ce == 0:
        verdict = "✅ Hệ thống ổn định — không có bias đáng kể."
    else:
        verdict = "🔵 Adaptive bình thường — không có cảnh báo."

    cur_lines += ["━━━━━━━━━━━━━━━━", verdict]
    reply("\n".join(cur_lines))


# ── Telegram /dow ─────────────────────────────────────────────
def _tg_cmd_dow(conn, reply, n: int = 500):
    """P74: WR by day-of-week + HOA by hour summary."""
    from zoneinfo import ZoneInfo
    from collections import defaultdict
    cur = conn.cursor()
    n   = max(200, min(n, 2000))

    if not USE_POSTGRES:
        reply("⚠️ Chỉ hỗ trợ PostgreSQL.")
        return

    # ── 1. WR by DOW ──
    cur.execute("""
        SELECT
            EXTRACT(DOW FROM dh.draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::int AS dow,
            COALESCE(pr.is_win_size, pr.is_win, FALSE) AS is_win
        FROM prediction_results pr
        JOIN draw_history dh ON dh.draw_number = pr.draw_number
        WHERE pr.actual_numbers IS NOT NULL
        ORDER BY pr.draw_number DESC LIMIT %s
    """, (n,))
    dow_rows = cur.fetchall()

    dow_data: dict = defaultdict(lambda: {'total': 0, 'wins': 0})
    for dow, is_win in dow_rows:
        d = dow_data[int(dow)]
        d['total'] += 1
        if is_win: d['wins'] += 1

    DOW_LABELS = {0: 'CN', 1: 'T2', 2: 'T3', 3: 'T4', 4: 'T5', 5: 'T6', 6: 'T7'}
    DOW_ORDER  = [1, 2, 3, 4, 5, 6, 0]
    BASELINE   = 0.375
    today_dow  = int(datetime.now(ZoneInfo('Asia/Ho_Chi_Minh')).strftime('%w'))

    dow_lines = []
    for d in DOW_ORDER:
        dd   = dow_data.get(d, {'total': 0, 'wins': 0})
        t, w = dd['total'], dd['wins']
        if t < 5:
            continue
        wr   = w / t
        diff = (wr - BASELINE) * 100
        icon = '🟢' if wr >= 0.40 else ('🟡' if wr >= BASELINE else '🔴')
        today_mark = ' ◀ hôm nay' if d == today_dow else ''
        dow_lines.append(
            f"{icon} <b>{DOW_LABELS[d]}</b>: {wr*100:.1f}% ({diff:+.1f}%) · {w}/{t}{today_mark}"
        )

    # ── 2. HOA WR by hour — top 3 best + top 3 worst ──
    cur.execute("""
        SELECT
            EXTRACT(HOUR FROM dh.draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::int AS vn_hour,
            CASE WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 11
                 AND  (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) >= 10
                 THEN TRUE ELSE FALSE END AS pred_hoa,
            CASE WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 11
                 AND  (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) >= 10
                 THEN TRUE ELSE FALSE END AS actual_hoa
        FROM predictions p
        JOIN prediction_results pr ON pr.prediction_id = p.id
        JOIN draw_history dh ON dh.draw_number = p.draw_number
        WHERE p.predicted_numbers IS NOT NULL AND pr.actual_numbers IS NOT NULL
        ORDER BY p.draw_number DESC LIMIT %s
    """, (n,))
    hour_rows = cur.fetchall()

    hour_data: dict = defaultdict(lambda: {'pred_hoa': 0, 'hoa_wins': 0})
    for vn_hour, pred_hoa, actual_hoa in hour_rows:
        h = hour_data[int(vn_hour)]
        if pred_hoa:
            h['pred_hoa'] += 1
            if actual_hoa: h['hoa_wins'] += 1

    # Compute HOA WR per hour, filter ≥ 5 predictions
    hour_rated = []
    for h in range(6, 23):
        hd = hour_data.get(h, {'pred_hoa': 0, 'hoa_wins': 0})
        p  = hd['pred_hoa']
        if p < 5: continue
        wr = hd['hoa_wins'] / p
        hour_rated.append((h, p, wr))

    hour_rated.sort(key=lambda x: x[2], reverse=True)
    top3_best  = hour_rated[:3]
    top3_worst = hour_rated[-3:][::-1]  # lowest WR first when reversed

    def fmt_hour(hr_list, icon):
        lines = []
        for h, p, wr in hr_list:
            lines.append(f"  {icon} {h}h: <b>{wr*100:.0f}%</b> ({p} lần dự HOA)")
        return lines

    now_str = datetime.now(ZoneInfo('Asia/Ho_Chi_Minh')).strftime('%H:%M %d/%m')
    msg = (
        f"📅 <b>WR BY DAY · HOA BY HOUR</b> · {now_str}\n"
        f"<i>n={n} kỳ · baseline {BASELINE*100:.0f}%</i>\n"
        f"━━━━━━━━━━━━━━━━\n"
        f"<b>Win rate theo thứ:</b>\n"
        + '\n'.join(dow_lines) + '\n'
        + f"━━━━━━━━━━━━━━━━\n"
        + f"<b>HOA WR theo giờ:</b>\n"
        + (f"Top tốt:\n" + '\n'.join(fmt_hour(top3_best, '🟢')) + '\n' if top3_best else '')
        + (f"Top yếu:\n" + '\n'.join(fmt_hour(top3_worst, '🔴')) if top3_worst else '')
    )
    reply(msg)


# ── Telegram /hourly ──────────────────────────────────────────
def _tg_cmd_hourly(conn, reply, n: int = 1000):
    """P115: WR by hour of day (6-22) with block summary."""
    from zoneinfo import ZoneInfo
    from collections import defaultdict

    if not USE_POSTGRES:
        reply("⚠️ Chỉ hỗ trợ PostgreSQL.")
        return

    n = max(200, min(n, 3000))
    cur = conn.cursor()
    cur.execute("""
        SELECT
            EXTRACT(HOUR FROM dh.draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::int AS vn_hour,
            COALESCE(pr.is_win_size, pr.is_win, FALSE) AS is_win
        FROM prediction_results pr
        JOIN draw_history dh ON dh.draw_number = pr.draw_number
        WHERE pr.actual_numbers IS NOT NULL
        ORDER BY pr.draw_number DESC LIMIT %s
    """, (n,))

    rows = cur.fetchall()
    hour_data: dict = defaultdict(lambda: {'total': 0, 'wins': 0})
    for vn_hour, is_win in rows:
        h = hour_data[int(vn_hour)]
        h['total'] += 1
        if is_win: h['wins'] += 1

    BASELINE = 0.375
    now_hour = datetime.now(ZoneInfo('Asia/Ho_Chi_Minh')).hour

    valid_hours = []
    for h in range(6, 23):
        d = hour_data.get(h, {'total': 0, 'wins': 0})
        t, w = d['total'], d['wins']
        if t >= 5:
            valid_hours.append((h, t, w, w / t))

    if not valid_hours:
        reply("⚠️ Không đủ dữ liệu theo giờ.")
        return

    best_h  = max(valid_hours, key=lambda x: x[3])
    worst_h = min(valid_hours, key=lambda x: x[3])

    lines = []
    for h, t, w, wr in valid_hours:
        diff    = (wr - BASELINE) * 100
        icon    = '🟢' if wr >= 0.42 else ('🟡' if wr >= BASELINE else '🔴')
        bar_len = round(wr * 16)
        bar     = '█' * bar_len + '░' * (16 - bar_len)
        marks   = ''
        if h == best_h[0]:  marks += ' ★'
        if h == worst_h[0]: marks += ' ▼'
        if h == now_hour:   marks += ' ◀'
        lines.append(f"{icon} <b>{h:02d}h</b> {bar} <b>{wr*100:.1f}%</b> ({diff:+.0f}%) {w}/{t}{marks}")

    BLOCKS = [('☀️ Sáng', range(6, 10)), ('🌤 Trưa', range(10, 14)),
              ('🌇 Chiều', range(14, 18)), ('🌙 Tối', range(18, 23))]
    block_lines = []
    for name, hrs in BLOCKS:
        bt = bw = 0
        for h in hrs:
            d = hour_data.get(h, {})
            bt += d.get('total', 0); bw += d.get('wins', 0)
        if bt >= 10:
            bwr  = bw / bt
            icon = '🟢' if bwr >= 0.42 else ('🟡' if bwr >= BASELINE else '🔴')
            block_lines.append(f"  {icon} {name}: <b>{bwr*100:.1f}%</b> ({bw}/{bt})")

    now_str = datetime.now(ZoneInfo('Asia/Ho_Chi_Minh')).strftime('%H:%M %d/%m')
    msg = (
        f"🕐 <b>WIN RATE THEO GIỜ</b> · {now_str}\n"
        f"<i>n={len(rows)} kỳ · baseline {BASELINE*100:.0f}% · ★tốt nhất · ▼yếu nhất · ◀hiện tại</i>\n"
        f"━━━━━━━━━━━━━━━━\n"
        + '\n'.join(lines) + '\n'
        + f"━━━━━━━━━━━━━━━━\n"
        + f"<b>Theo buổi:</b>\n"
        + '\n'.join(block_lines)
    )
    reply(msg)


# ── Telegram /explain ─────────────────────────────────────────
def _tg_cmd_explain(conn, reply):
    """P84: Detailed vote breakdown for the latest prediction."""
    import json as _json, ast as _ast
    cur = conn.cursor()
    cur.execute("""
        SELECT p.draw_number, p.predicted_numbers, p.confidence,
               p.vote_breakdown, pr.is_win, pr.actual_numbers
        FROM predictions p
        LEFT JOIN prediction_results pr ON pr.prediction_id = p.id
        ORDER BY p.draw_number DESC LIMIT 1
    """)
    row = cur.fetchone()
    if not row:
        reply("⚠️ Chưa có dự đoán nào.")
        return

    draw_num, pred_nums, conf, vb_raw, is_win, actual = row

    try:
        vb = _json.loads(vb_raw) if isinstance(vb_raw, str) else (vb_raw or {})
    except Exception:
        vb = {}

    if isinstance(pred_nums, str):
        pred_nums = _ast.literal_eval(pred_nums)

    pred_sum = sum(pred_nums)
    pred_size = 'NHO' if pred_sum <= 9 else ('HOA' if pred_sum <= 11 else 'LON')
    SIZE_VI   = {'NHO': 'NHỎ 🔵', 'HOA': 'HÒA 🟡', 'LON': 'LỚN 🔴'}
    SIZE_EMJ  = {'NHO': '🔵', 'HOA': '🟡', 'LON': '🔴'}

    conf_str   = f"{conf:.1%}" if conf else "N/A"
    nums_str   = " · ".join(str(n) for n in pred_nums)
    vote_share = vb.get('vote_share', 0)
    majority   = vb.get('majority_size', pred_size)
    size_w     = vb.get('size_weights', {})
    sw_total   = sum(size_w.values()) or 1

    # ── Result line ──
    if is_win is None:
        result_line = "⏳ Chưa có kết quả"
    else:
        if isinstance(actual, str):
            actual = _ast.literal_eval(actual or "[]")
        act_str  = " · ".join(str(n) for n in actual)
        act_size = 'NHO' if sum(actual) <= 9 else ('HOA' if sum(actual) <= 11 else 'LON')
        result_line = (f"✅ THẮNG — thực tế: <b>{act_str}</b> ({SIZE_VI.get(act_size, act_size)})"
                       if is_win else
                       f"❌ THUA — thực tế: <b>{act_str}</b> ({SIZE_VI.get(act_size, act_size)})")

    # ── Per-voter detail ──
    detail = vb.get('all_votes_detail', {})
    VOTER_ORDER = ['ml', 'markov', 'prior_nho', 'prior_hoa', 'prior_lon']
    voter_lines = []
    for vname in VOTER_ORDER:
        d = detail.get(vname)
        if d is None:
            continue
        sz       = d.get('size', '?')
        c        = d.get('conf', 0)
        mult     = d.get('mult', 1.0)
        eff_pct  = d.get('eff_w_pct', 0)
        streak   = d.get('streak', 0)
        decay    = d.get('decay', 1.0)
        won      = d.get('winner', False)
        icon     = '✅' if won else '❌'
        s_emj    = SIZE_EMJ.get(sz, '⚪')
        decay_str = f" decay {decay:.2f}×" if decay < 1.0 else ""
        streak_str = f" {streak}L" if streak > 0 else ""
        voter_lines.append(
            f"{icon} <b>{vname}</b>: {s_emj}{sz}  conf {c:.0%}  ×{mult:.2f}{decay_str}{streak_str}  → <b>{eff_pct:.1f}%</b> tổng weight"
        )

    if vb.get('markov_abstained'):
        voter_lines.append("⏸ <i>markov abstained (low confidence)</i>")

    # ── SIZE weight bar ──
    sw_parts = []
    for sz in ['NHO', 'HOA', 'LON']:
        w = size_w.get(sz, 0)
        pct = round(w / sw_total * 100)
        bar = '█' * max(1, pct // 10)
        emj = SIZE_EMJ.get(sz, '⚪')
        bold_open  = '<b>' if sz == majority else ''
        bold_close = '</b>' if sz == majority else ''
        sw_parts.append(f"{emj}{bold_open}{sz} {pct}%{bold_close} {bar}")

    # ── Adaptive thresholds ──
    adaptive = vb.get('adaptive', {})
    adapt_parts = []
    if adaptive:
        if 'tune_k' in adaptive:
            adapt_parts.append(f"tune_k={adaptive['tune_k']:.2f}")
        if 'nho_share_min' in adaptive:
            adapt_parts.append(f"nho_min={adaptive['nho_share_min']:.2f}")
        if 'hoa_suppress' in adaptive:
            adapt_parts.append(f"hoa_sup={adaptive['hoa_suppress']:.2f}")
        if 'consecutive_excess' in adaptive:
            adapt_parts.append(f"excess={int(adaptive['consecutive_excess'])}")

    lines = [
        f"🔍 <b>EXPLAIN KỲ #{draw_num}</b>",
        f"━━━━━━━━━━━━━━━━━━",
        f"🔢 Số: <b>{nums_str}</b>  Tổng: {pred_sum}",
        f"{SIZE_EMJ.get(majority,'⚪')} SIZE: <b>{SIZE_VI.get(majority, majority)}</b>  conf <b>{conf_str}</b>  consensus <b>{vote_share:.0%}</b>",
        "",
        "📊 <b>SIZE weights:</b>",
        "  " + "  ".join(sw_parts),
        "",
        "🗳 <b>Voters:</b>",
    ] + [f"  {l}" for l in voter_lines]

    if adapt_parts:
        lines += ["", f"⚙️ <b>Adaptive:</b> {' · '.join(adapt_parts)}"]

    lines += ["━━━━━━━━━━━━━━━━━━", result_line]

    reply("\n".join(lines))


# ── Telegram /calibration ─────────────────────────────────────
def _tg_cmd_calibration(conn, reply, n: int = 500):
    """P83: Brier score, ECE, confidence vs WR gap per SIZE for last N draws."""
    cur = conn.cursor()
    n   = max(100, min(n, 2000))

    if not USE_POSTGRES:
        reply("⚠️ Chỉ hỗ trợ PostgreSQL.")
        return

    cur.execute("""
        SELECT p.confidence, p.predicted_size,
               COALESCE(pr.is_win_size, pr.is_win, FALSE) AS is_win
        FROM predictions p
        JOIN prediction_results pr ON pr.prediction_id = p.id
        WHERE p.confidence IS NOT NULL AND p.predicted_size IS NOT NULL
          AND pr.actual_numbers IS NOT NULL
        ORDER BY p.draw_number DESC LIMIT %s
    """, (n,))
    rows = cur.fetchall()

    if not rows:
        reply("⚠️ Không đủ dữ liệu để tính calibration.")
        return

    from collections import defaultdict
    size_stats: dict = defaultdict(lambda: {'conf_sum': 0.0, 'wins': 0, 'total': 0, 'brier': 0.0})
    brier_total = 0.0
    conf_bins: dict = defaultdict(lambda: {'conf_sum': 0.0, 'wins': 0, 'total': 0})

    for conf, sz, is_win in rows:
        c   = float(conf)
        win = 1 if is_win else 0
        brier_total += (c - win) ** 2

        d = size_stats[sz]
        d['conf_sum'] += c
        d['wins']     += win
        d['total']    += 1
        d['brier']    += (c - win) ** 2

        # 5 bins: [0,0.2) [0.2,0.4) [0.4,0.6) [0.6,0.8) [0.8,1.0]
        bin_idx = min(int(c * 5), 4)
        b = conf_bins[bin_idx]
        b['conf_sum'] += c
        b['wins']     += win
        b['total']    += 1

    total = len(rows)
    overall_brier  = brier_total / total
    baseline_brier = 0.375 * (1 - 0.375) ** 2 + 0.625 * 0.375 ** 2  # baseline SIZE brier
    brier_skill    = round((1 - overall_brier / baseline_brier) * 100, 1)

    # ECE
    ece = 0.0
    for b in conf_bins.values():
        if b['total'] == 0:
            continue
        mean_conf = b['conf_sum'] / b['total']
        actual_wr = b['wins']    / b['total']
        ece += (b['total'] / total) * abs(mean_conf - actual_wr)
    ece_pct = round(ece * 100, 2)

    BASELINE_WR = 0.375
    SIZE_ORDER  = ['NHO', 'HOA', 'LON']
    SIZE_ICONS  = {'NHO': '🔵', 'HOA': '🟡', 'LON': '🔴'}

    size_lines = []
    for sz in SIZE_ORDER:
        d = size_stats.get(sz)
        if not d or d['total'] < 5:
            continue
        avg_conf = d['conf_sum'] / d['total']
        wr       = d['wins']     / d['total']
        gap      = avg_conf - wr  # positive = overconfident
        sz_brier = d['brier']    / d['total']
        gap_icon = '⬆️' if gap > 0.05 else ('⬇️' if gap < -0.05 else '✅')
        size_lines.append(
            f"{SIZE_ICONS.get(sz,'·')} <b>{sz}</b>: conf {avg_conf*100:.1f}% → WR {wr*100:.1f}%"
            f" | gap <b>{gap*100:+.1f}%</b> {gap_icon} | Brier {sz_brier:.3f} ({d['wins']}/{d['total']})"
        )

    skill_icon = '🟢' if brier_skill > 5 else ('🟡' if brier_skill > 0 else '🔴')
    ece_icon   = '🟢' if ece_pct < 5 else ('🟡' if ece_pct < 10 else '🔴')

    lines = [
        f"📐 <b>CALIBRATION REPORT</b> · {total:,} kỳ gần nhất",
        "",
        f"{skill_icon} Brier score: <b>{overall_brier:.4f}</b>  |  Skill: <b>{brier_skill:+.1f}%</b> vs baseline",
        f"{ece_icon} ECE: <b>{ece_pct:.2f}%</b>  (càng nhỏ càng tốt)",
        "",
        "📊 <b>Per SIZE:</b>",
    ] + size_lines + [
        "",
        "<i>gap > 0 = model quá tự tin | gap < 0 = model chưa đủ tự tin</i>",
    ]

    reply("\n".join(lines))


# ── Telegram /health ──────────────────────────────────────────
def _tg_cmd_health(conn, reply):
    """P87: System health digest — voter state, adaptive tune, rolling WR, streak."""
    from zoneinfo import ZoneInfo
    from collections import defaultdict
    from prediction_service import _adaptive_thresh_cache as _atc, _get_voter_decay

    cur    = conn.cursor()
    vn_now = datetime.now(ZoneInfo('Asia/Ho_Chi_Minh'))

    # ── 1. Last draw + lag ──
    cur.execute("SELECT draw_number, draw_time FROM draw_history ORDER BY draw_number DESC LIMIT 1")
    row = cur.fetchone()
    last_draw, last_time = (row[0], row[1]) if row else (0, None)
    if last_time:
        if hasattr(last_time, 'tzinfo') and last_time.tzinfo is None:
            last_time = last_time.replace(tzinfo=timezone.utc)
        lag_min = round((datetime.now(timezone.utc) - last_time.astimezone(timezone.utc)).total_seconds() / 60, 1)
        last_vn = last_time.astimezone(ZoneInfo('Asia/Ho_Chi_Minh')).strftime('%H:%M %d/%m')
    else:
        lag_min, last_vn = None, 'N/A'

    # ── 2. Rolling WR last 50 + current streak ──
    cur.execute("""
        SELECT COALESCE(is_win_size, is_win, FALSE)
        FROM prediction_results
        ORDER BY draw_number DESC LIMIT 50
    """)
    seq = [r[0] for r in cur.fetchall()]
    wr50 = sum(seq) / len(seq) if seq else None
    streak_val = loss_streak = win_streak = 0
    for w in seq:
        if w:
            if loss_streak > 0: break
            win_streak += 1
        else:
            if win_streak > 0: break
            loss_streak += 1
    streak_val = win_streak if win_streak else -loss_streak  # pos=win, neg=loss

    # ── 3. Voter health (last 200 vote_breakdowns) ──
    if USE_POSTGRES:
        cur.execute("""
            SELECT p.vote_breakdown,
                CASE
                    WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 9  THEN 'NHO'
                    WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 11 THEN 'HOA'
                    ELSE 'LON'
                END AS actual_size
            FROM predictions p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            WHERE p.vote_breakdown IS NOT NULL AND pr.actual_numbers IS NOT NULL
            ORDER BY p.draw_number DESC LIMIT 200
        """)
    else:
        cur.execute("SELECT NULL, NULL WHERE 0=1")
    vb_rows = cur.fetchall()

    acc = defaultdict(lambda: {'correct': 0, 'total': 0})
    for vb_raw, actual_size in vb_rows:
        try:
            vb = json.loads(vb_raw) if isinstance(vb_raw, str) else vb_raw
            for vname, vsize in ((vb or {}).get('all_votes') or {}).items():
                acc[vname]['total'] += 1
                if vsize == actual_size:
                    acc[vname]['correct'] += 1
        except Exception:
            continue

    BASELINE    = 0.375
    decay_cache = _get_voter_decay()
    voter_health = []
    for vname, a in acc.items():
        t, w = a['total'], a['correct']
        if t < 10:
            continue
        wr      = w / t
        wr_mult = round(max(0.4, min(wr / BASELINE, 2.5)), 3) if t >= 20 else 1.0
        dk      = decay_cache.get(vname, {'streak': 0, 'decay': 1.0})
        eff     = round(wr_mult * dk['decay'], 3)
        voter_health.append({'name': vname, 'wr': wr, 'streak': dk['streak'], 'eff': eff})

    voter_health.sort(key=lambda v: v['eff'])
    worst = voter_health[0] if voter_health else None
    alerted = [v for v in voter_health if v['streak'] >= 5 or v['eff'] < 0.6]

    # ── 4. Adaptive thresholds ──
    at = _atc or {}

    # ── Build message ──
    lag_str = f"{lag_min}p" if lag_min is not None else 'N/A'
    sync_icon = '🟢' if lag_min and lag_min < 10 else ('🟡' if lag_min and lag_min < 20 else '🔴')

    wr50_str = f"{wr50*100:.1f}%" if wr50 is not None else 'N/A'
    wr50_icon = '🟢' if wr50 and wr50 >= 0.40 else ('🟡' if wr50 and wr50 >= BASELINE else '🔴')

    if streak_val > 0:
        streak_str = f"🔥 Win streak <b>{streak_val}</b> kỳ"
    elif streak_val < 0:
        streak_str = f"❄️ Loss streak <b>{abs(streak_val)}</b> kỳ"
    else:
        streak_str = "➖ Streak: 0"

    voter_lines = []
    for v in voter_health:
        eff_icon = '🟢' if v['eff'] >= 1.2 else ('🔵' if v['eff'] >= 0.8 else ('🟡' if v['eff'] >= 0.6 else '🔴'))
        sk_str   = f" {v['streak']}L" if v['streak'] > 0 else ""
        voter_lines.append(
            f"  {eff_icon} <b>{v['name']}</b>: WR {v['wr']*100:.1f}%  Eff {v['eff']:.2f}×{sk_str}"
        )

    adapt_parts = []
    for k, label in [('consecutive_excess', 'excess'), ('tune_k', 'tune_k'),
                     ('nho_share_min', 'nho_min'), ('pred_lon_excess', 'lon_ex')]:
        if k in at:
            v = at[k]
            adapt_parts.append(f"{label}={'%d' % v if k == 'consecutive_excess' else '%.2f' % v}")

    lines = [
        f"🏥 <b>SYSTEM HEALTH</b> · {vn_now.strftime('%H:%M %d/%m')}",
        "━━━━━━━━━━━━━━━━━━",
        f"{sync_icon} Kỳ cuối: <b>#{last_draw}</b> ({last_vn}) · lag {lag_str}",
        f"{wr50_icon} WR 50 kỳ: <b>{wr50_str}</b>  |  {streak_str}",
        "",
        "🗳 <b>Voter health:</b>",
    ] + voter_lines

    if alerted:
        lines += ["", f"⚠️ Cần chú ý: {', '.join(v['name'] for v in alerted)}"]

    if adapt_parts:
        lines += ["", f"⚙️ Adaptive: {' · '.join(adapt_parts)}"]

    if at.get('consecutive_excess', 0) >= 3:
        lines.append(f"🔴 LON excess kéo dài {int(at['consecutive_excess'])} chu kỳ!")

    reply("\n".join(lines))


# ── Telegram /top ─────────────────────────────────────────────
def _tg_cmd_top(conn, reply, n: int = 500):
    """P90: Top 5 high-confidence wins + top 5 high-confidence losses."""
    cur = conn.cursor()
    n   = max(100, min(n, 2000))
    if not USE_POSTGRES:
        reply("⚠️ Chỉ hỗ trợ PostgreSQL.")
        return

    cur.execute("""
        SELECT p.draw_number, p.confidence,
            CASE
                WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 9  THEN 'NHO'
                WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 11 THEN 'HOA'
                ELSE 'LON'
            END AS pred_size,
            CASE
                WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 9  THEN 'NHO'
                WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 11 THEN 'HOA'
                ELSE 'LON'
            END AS actual_size,
            COALESCE(pr.is_win_size, pr.is_win, FALSE) AS is_win
        FROM predictions p
        JOIN prediction_results pr ON pr.prediction_id = p.id
        WHERE p.confidence IS NOT NULL AND pr.actual_numbers IS NOT NULL
        ORDER BY p.draw_number DESC LIMIT %s
    """, (n,))
    rows = cur.fetchall()

    SIZE_VI  = {'NHO': 'NHỎ 🔵', 'HOA': 'HÒA 🟡', 'LON': 'LỚN 🔴'}
    wins     = sorted([r for r in rows if r[4]],     key=lambda r: -r[1])[:5]
    losses   = sorted([r for r in rows if not r[4]], key=lambda r: -r[1])[:5]

    def fmt_row(i, r):
        dn, conf, pred, actual, _ = r
        match = '✅' if pred == actual else '❌'
        return (f"  {i+1}. #{dn}  conf <b>{conf:.0%}</b>  "
                f"{SIZE_VI.get(pred, pred)} → {SIZE_VI.get(actual, actual)} {match}")

    lines = [
        f"🏆 <b>TOP · {n} kỳ gần nhất</b>",
        "━━━━━━━━━━━━━━━━━━",
        "✅ <b>Top 5 confident WINS</b>",
    ] + [fmt_row(i, r) for i, r in enumerate(wins)] + [
        "",
        "❌ <b>Top 5 confident LOSSES</b>",
    ] + [fmt_row(i, r) for i, r in enumerate(losses)] + [
        "",
        "<i>Kỳ thua confidence cao = model tự tin nhưng sai → nên review /explain</i>",
    ]
    reply("\n".join(lines))



# ── Telegram /recap ───────────────────────────────────────────
def _tg_cmd_recap(conn, reply, n: int = 100):
    """P92: Compact recap of last N draws — WR, SIZE dist, best hour, top voter, streak."""
    from collections import defaultdict
    cur = conn.cursor()
    n   = max(20, min(n, 500))

    if not USE_POSTGRES:
        reply("⚠️ Chỉ hỗ trợ PostgreSQL.")
        return

    # ── 1. WR + SIZE distribution + best hour ──
    cur.execute("""
        SELECT
            COALESCE(pr.is_win_size, pr.is_win, FALSE)   AS is_win,
            CASE
                WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 9  THEN 'NHO'
                WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 11 THEN 'HOA'
                ELSE 'LON'
            END AS pred_size,
            CASE
                WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 9  THEN 'NHO'
                WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 11 THEN 'HOA'
                ELSE 'LON'
            END AS act_size,
            EXTRACT(HOUR FROM dh.draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::int AS vn_hour,
            p.confidence
        FROM predictions p
        JOIN prediction_results pr ON pr.prediction_id = p.id
        JOIN draw_history dh        ON dh.draw_number  = pr.draw_number
        WHERE pr.actual_numbers IS NOT NULL
        ORDER BY p.draw_number DESC LIMIT %s
    """, (n,))
    rows = cur.fetchall()

    if not rows:
        reply("⚠️ Không đủ dữ liệu.")
        return

    total  = len(rows)
    wins   = sum(1 for r in rows if r[0])
    wr     = wins / total

    pred_sz: dict = defaultdict(int)
    act_sz:  dict = defaultdict(int)
    hour_w:  dict = defaultdict(int)
    hour_t:  dict = defaultdict(int)
    conf_sum = 0.0; conf_n = 0

    for is_win, ps, as_, hour, conf in rows:
        pred_sz[ps] += 1
        act_sz[as_]  += 1
        hour_t[hour] += 1
        if is_win: hour_w[hour] += 1
        if conf:
            conf_sum += float(conf); conf_n += 1

    avg_conf = conf_sum / conf_n if conf_n else None

    # Best + worst hour (min 5 draws)
    hour_wr = {h: hour_w[h] / hour_t[h] for h in hour_t if hour_t[h] >= 5}
    best_h  = max(hour_wr, key=hour_wr.get) if hour_wr else None
    worst_h = min(hour_wr, key=hour_wr.get) if hour_wr else None

    # ── 2. Current streak ──
    streak_val = 0; streak_type = None
    for is_win, *_ in rows:
        if streak_type is None:
            streak_type = is_win; streak_val = 1
        elif is_win == streak_type:
            streak_val += 1
        else:
            break

    # ── 3. Top voter from vote_breakdown ──
    cur.execute("""
        SELECT p.vote_breakdown,
            CASE
                WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 9  THEN 'NHO'
                WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 11 THEN 'HOA'
                ELSE 'LON'
            END AS actual_size
        FROM predictions p
        JOIN prediction_results pr ON pr.prediction_id = p.id
        WHERE p.vote_breakdown IS NOT NULL AND pr.actual_numbers IS NOT NULL
        ORDER BY p.draw_number DESC LIMIT %s
    """, (n,))
    vb_rows = cur.fetchall()

    voter_acc: dict = defaultdict(lambda: {'correct': 0, 'total': 0})
    for vb_raw, actual_size in vb_rows:
        try:
            vb = json.loads(vb_raw) if isinstance(vb_raw, str) else vb_raw
            for vname, vsize in ((vb or {}).get('all_votes') or {}).items():
                voter_acc[vname]['total'] += 1
                if vsize == actual_size:
                    voter_acc[vname]['correct'] += 1
        except Exception:
            continue

    top_voter = None
    if voter_acc:
        top_voter = max(
            ((name, a) for name, a in voter_acc.items() if a['total'] >= 10),
            key=lambda x: x[1]['correct'] / x[1]['total'],
            default=None
        )

    # ── Build message ──
    BASELINE  = 0.375
    SIZE_VI   = {'NHO': 'NHỎ 🔵', 'HOA': 'HÒA 🟡', 'LON': 'LỚN 🔴'}
    wr_icon   = '🟢' if wr >= 0.40 else ('🟡' if wr >= BASELINE else '🔴')
    wr_delta  = (wr - BASELINE) * 100

    streak_str = ''
    if streak_type is True:
        streak_str = f"🔥 Win streak {streak_val} kỳ"
    elif streak_type is False:
        streak_str = f"❄️ Loss streak {streak_val} kỳ"

    conf_str = f"{avg_conf*100:.1f}%" if avg_conf else "—"

    # SIZE distribution lines
    size_lines = []
    for sz in ['NHO', 'HOA', 'LON']:
        p_pct = pred_sz.get(sz, 0) / total * 100
        a_pct = act_sz.get(sz,  0) / total * 100
        diff  = p_pct - a_pct
        d_str = f"{diff:+.1f}%"
        d_col = '' if abs(diff) < 3 else ('📈' if diff > 0 else '📉')
        size_lines.append(
            f"  {SIZE_VI[sz]}: pred {p_pct:.1f}% / actual {a_pct:.1f}% {d_col}{d_str}"
        )

    lines = [
        f"📋 <b>RECAP · {total} kỳ gần nhất</b>",
        "━━━━━━━━━━━━━━━━━━",
        f"{wr_icon} WR: <b>{wr*100:.1f}%</b> ({wins}/{total})"
        + (f"  {wr_delta:+.1f}% vs baseline" if True else ""),
        f"📊 Conf trung bình: <b>{conf_str}</b>",
        f"{('  ' + streak_str) if streak_str else ''}",
        "",
        "📐 <b>SIZE distribution:</b>",
    ] + size_lines

    if best_h is not None:
        lines += [
            "",
            f"⏰ <b>Giờ tốt nhất:</b> {best_h}h → WR {hour_wr[best_h]*100:.1f}% ({hour_w[best_h]}/{hour_t[best_h]})",
            f"⏰ <b>Giờ tệ nhất:</b> {worst_h}h → WR {hour_wr[worst_h]*100:.1f}% ({hour_w[worst_h]}/{hour_t[worst_h]})",
        ]

    if top_voter:
        vname, va = top_voter
        vwr = va['correct'] / va['total']
        lines += [
            "",
            f"🥇 <b>Voter dẫn đầu:</b> {vname} → WR {vwr*100:.1f}% ({va['correct']}/{va['total']})",
        ]

    reply("\n".join(l for l in lines if l is not None))


# ── Telegram /wincal ──────────────────────────────────────────
def _tg_cmd_wincal(conn, reply, n_days: int = 28):
    """P128: Win-rate calendar — last N days as text grid (Mon–Sun columns)."""
    import datetime as _dt
    n_days = max(7, min(n_days, 84))

    cur = conn.cursor()
    if USE_POSTGRES:
        cur.execute("""
            SELECT
                (dh.draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date AS vn_date,
                COUNT(*) AS total,
                SUM(CASE WHEN COALESCE(pr.is_win_size, pr.is_win, FALSE) THEN 1 ELSE 0 END) AS wins
            FROM prediction_results pr
            JOIN draw_history dh ON dh.draw_number = pr.draw_number
            WHERE pr.actual_numbers IS NOT NULL
              AND dh.draw_time >= NOW() - (%s || ' days')::INTERVAL
            GROUP BY 1
            ORDER BY 1
        """, (n_days + 7,))
    else:
        cur.execute("""
            SELECT DATE(datetime(dh.draw_time, '+7 hours')) AS vn_date,
                   COUNT(*) AS total,
                   SUM(COALESCE(pr.is_win_size, pr.is_win, 0)) AS wins
            FROM prediction_results pr
            JOIN draw_history dh ON dh.draw_number = pr.draw_number
            WHERE pr.actual_numbers IS NOT NULL
              AND dh.draw_time >= DATE(datetime('now', ?))
            GROUP BY 1 ORDER BY 1
        """, (f'-{n_days + 7} days',))

    rows = cur.fetchall()
    day_map = {}
    for vn_date, total, wins in rows:
        ds = str(vn_date)[:10]
        day_map[ds] = {'total': int(total), 'wins': int(wins)}

    today_vn = _dt.datetime.utcnow() + _dt.timedelta(hours=7)
    today = today_vn.date()

    # Build grid: fill Mon–Sun weeks
    # Find the Monday of the week containing (today - n_days + 1)
    start_day = today - _dt.timedelta(days=n_days - 1)
    # Align back to Monday
    start_mon = start_day - _dt.timedelta(days=start_day.weekday())

    BASELINE = 0.375
    above = below = no_data = 0
    week_rows = []
    d = start_mon
    while d <= today:
        week = []
        for _ in range(7):  # Mon..Sun
            ds = d.strftime('%Y-%m-%d')
            info = day_map.get(ds)
            if d > today or d < start_day:
                week.append(('  ', None))
            elif info and info['total'] >= 5:
                wr = info['wins'] / info['total']
                if wr >= BASELINE + 0.04:
                    cell = '🟢'
                    above += 1
                elif wr >= BASELINE - 0.04:
                    cell = '🟡'
                    above += 1
                else:
                    cell = '🔴'
                    below += 1
                week.append((cell, wr))
            else:
                week.append(('⬜', None))
                no_data += 1
            d += _dt.timedelta(days=1)
        week_rows.append(week)

    hdr = "T2  T3  T4  T5  T6  T7  CN"
    grid_lines = [hdr]
    for wk in week_rows:
        parts = []
        for cell, wr in wk:
            if wr is not None:
                parts.append(f"{cell}")
            else:
                parts.append(f"{cell}")
        grid_lines.append("  ".join(parts))

    total_known = above + below
    avg_line = ""
    if total_known:
        avg_wr_list = [day_map[d.strftime('%Y-%m-%d')]['wins'] / day_map[d.strftime('%Y-%m-%d')]['total']
                       for d in (today - _dt.timedelta(days=i) for i in range(n_days))
                       if d.strftime('%Y-%m-%d') in day_map and day_map[d.strftime('%Y-%m-%d')]['total'] >= 5]
        if avg_wr_list:
            overall_avg = sum(avg_wr_list) / len(avg_wr_list)
            avg_line = f"Avg WR ({len(avg_wr_list)} ngày): <b>{overall_avg*100:.1f}%</b>"

    lines = [
        f"📅 <b>WIN-RATE CALENDAR ({n_days} ngày)</b>",
        f"🟢≥{(BASELINE+0.04)*100:.0f}%  🟡≈{BASELINE*100:.0f}%  🔴<{(BASELINE-0.04)*100:.0f}%  ⬜n<5",
        "━━━━━━━━━━━━━━━━",
        *grid_lines,
        "━━━━━━━━━━━━━━━━",
        f"✅ Trên baseline: <b>{above}</b>  ❌ Dưới: <b>{below}</b>  ⬜ Ít data: <b>{no_data}</b>",
    ]
    if avg_line:
        lines.append(avg_line)

    reply("\n".join(lines))


# ── Telegram /next ────────────────────────────────────────────
def _tg_cmd_next(conn, reply):
    """P132: Multi-draw preview — N+1 (current prediction) + N+2/N+3 Markov projections."""
    import json as _json
    cur = conn.cursor()
    cur.execute(
        "SELECT predicted_numbers, confidence, model_name, draw_number, vote_breakdown "
        "FROM predictions ORDER BY draw_number DESC LIMIT 1"
    )
    row = cur.fetchone()
    if not row:
        reply("⚠️ Chưa có dự đoán.")
        return
    pred_raw, conf, model_name, draw_n1, vb_raw = row
    try:
        pred_nums = _json.loads(pred_raw) if isinstance(pred_raw, str) else pred_raw
    except Exception:
        pred_nums = []
    pred_sum  = sum(pred_nums) if pred_nums else 0
    pred_size = 'NHO' if pred_sum <= 9 else ('HOA' if pred_sum <= 11 else 'LON')

    n1_probs = {'NHO': 0.333, 'HOA': 0.333, 'LON': 0.334}
    try:
        vb = _json.loads(vb_raw) if isinstance(vb_raw, str) else (vb_raw or {})
        sw = vb.get('size_weights', {})
        total_w = sum(sw.values()) if sw else 0
        if total_w > 0:
            n1_probs = {sz: sw.get(sz, 0) / total_w for sz in ('NHO', 'HOA', 'LON')}
    except Exception:
        n1_probs = {sz: (0.7 if sz == pred_size else 0.15) for sz in ('NHO', 'HOA', 'LON')}

    from collections import defaultdict
    raw = defaultdict(lambda: {'NHO': 0, 'HOA': 0, 'LON': 0})
    if USE_POSTGRES:
        cur.execute("""
            WITH sized AS (
                SELECT draw_number, size_category AS sz
                FROM draw_history ORDER BY draw_number DESC LIMIT 2000
            )
            SELECT a.sz, b.sz, COUNT(*)::int
            FROM sized a JOIN sized b ON b.draw_number = a.draw_number + 1
            GROUP BY a.sz, b.sz
        """)
    else:
        cur.execute("""
            WITH sized AS (
                SELECT draw_number, size_category AS sz
                FROM draw_history ORDER BY draw_number DESC LIMIT 2000
            )
            SELECT a.sz, b.sz, COUNT(*)
            FROM sized a JOIN sized b ON b.draw_number = a.draw_number + 1
            GROUP BY a.sz, b.sz
        """)
    for prev_sz, next_sz, cnt in cur.fetchall():
        if prev_sz in raw and next_sz in ('NHO', 'HOA', 'LON'):
            raw[prev_sz][next_sz] += cnt
    trans = {}
    for sz in ('NHO', 'HOA', 'LON'):
        total = sum(raw[sz].values())
        trans[sz] = {k: v / total if total else 1/3 for k, v in raw[sz].items()}

    def apply_trans(probs):
        out = {'NHO': 0.0, 'HOA': 0.0, 'LON': 0.0}
        for fs, pf in probs.items():
            for ts, pt in trans.get(fs, {}).items():
                out[ts] += pf * pt
        t = sum(out.values())
        return {k: v / t if t else 1/3 for k, v in out.items()}

    n2_probs = apply_trans(n1_probs)
    n3_probs = apply_trans(n2_probs)

    def top(p): return max(p, key=p.get)
    def bar(p):
        return (f"NHO {p['NHO']*100:.0f}% │ HOA {p['HOA']*100:.0f}% │ LON {p['LON']*100:.0f}%")

    conf_pct = f"{float(conf)*100:.0f}%" if conf else "—"
    lines = [
        "🔮 <b>MULTI-DRAW PREVIEW</b>",
        "━━━━━━━━━━━━━━━━━━",
        f"<b>N+1 #{draw_n1}</b>  [{', '.join(str(x) for x in pred_nums)}]  <b>{pred_size}</b>",
        f"  conf {conf_pct}  model: {model_name or '—'}",
        f"  {bar(n1_probs)}",
        "",
        f"<b>N+2 #{draw_n1+1}</b>  <b>{top(n2_probs)}</b>  <i>(Markov¹)</i>",
        f"  {bar(n2_probs)}",
        "",
        f"<b>N+3 #{draw_n1+2}</b>  <b>{top(n3_probs)}</b>  <i>(Markov²)</i>",
        f"  {bar(n3_probs)}",
        "━━━━━━━━━━━━━━━━━━",
        "⚠️ N+2/N+3 chỉ là xác suất Markov, độ tin cậy thấp.",
    ]
    reply("\n".join(lines))


# ── Telegram /votertrend ───────────────────────────────────────
def _tg_cmd_votertrend(conn, reply, n: int = 500, batch: int = 25):
    """P132: Per-voter WR trend — last 5 batches table with trend arrows."""
    import json as _json
    n     = max(50, min(n, 2000))
    batch = max(10, min(batch, 100))
    cur   = conn.cursor()
    if USE_POSTGRES:
        cur.execute("""
            SELECT p.draw_number, p.vote_breakdown, pr.actual_numbers
            FROM predictions p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            WHERE p.vote_breakdown IS NOT NULL AND pr.actual_numbers IS NOT NULL
            ORDER BY p.draw_number DESC LIMIT %s
        """, (n,))
    else:
        cur.execute("""
            SELECT p.draw_number, p.vote_breakdown, pr.actual_numbers
            FROM predictions p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            WHERE p.vote_breakdown IS NOT NULL AND pr.actual_numbers IS NOT NULL
            ORDER BY p.draw_number DESC LIMIT ?
        """, (n,))
    rows = list(reversed(cur.fetchall()))
    if not rows:
        reply("⚠️ Chưa có đủ dữ liệu.")
        return

    def _sz(raw):
        try:
            ns = raw if not isinstance(raw, str) else __import__('ast').literal_eval(raw)
            s = sum(int(x) for x in ns)
            return 'NHO' if s <= 9 else ('HOA' if s <= 11 else 'LON')
        except Exception:
            return None

    batches = []
    for i in range(0, len(rows), batch):
        chunk = rows[i:i + batch]
        if len(chunk) < max(3, batch // 3):
            continue
        vstats: dict = {}
        for dn, vb_raw, act_raw in chunk:
            act_sz = _sz(act_raw)
            if not act_sz:
                continue
            try:
                vb = _json.loads(vb_raw) if isinstance(vb_raw, str) else (vb_raw or {})
                for vname, vsize in (vb.get('all_votes') or {}).items():
                    if vname not in vstats:
                        vstats[vname] = [0, 0]
                    vstats[vname][1] += 1
                    if vsize == act_sz:
                        vstats[vname][0] += 1
            except Exception:
                pass
        if vstats:
            batches.append({v: w/t if t >= 3 else None for v, (w, t) in vstats.items()})

    if not batches:
        reply("⚠️ Không đủ batch.")
        return

    # Collect voters present in all batches (or ≥ half)
    voter_counts: dict = {}
    for b in batches:
        for v, wr in b.items():
            if wr is not None:
                voter_counts[v] = voter_counts.get(v, 0) + 1
    min_b = max(1, len(batches) // 2)
    voters = sorted([v for v, c in voter_counts.items() if c >= min_b],
                    key=lambda v: -voter_counts[v])

    # Show last 5 batches
    show = batches[-5:]
    total_batches = len(batches)

    lines = [
        f"📈 <b>VOTER WR TREND</b>  (n={n}, batch={batch})",
        f"Tổng {total_batches} batch | hiển thị 5 batch gần nhất",
        "━━━━━━━━━━━━━━━━━━━━━",
    ]
    hdr_cols = "  ".join(f"B{total_batches - len(show) + i + 1:>2}" for i in range(len(show)))
    lines.append(f"{'Voter':<20}  {hdr_cols}  Xu hướng")
    lines.append("─" * 44)

    BASELINE = 0.375
    for v in voters[:12]:
        cells = []
        wrs = []
        for b in show:
            wr = b.get(v)
            if wr is None:
                cells.append("  — ")
            else:
                wrs.append(wr)
                diff = wr - BASELINE
                marker = "▲" if diff >= 0.04 else ("▼" if diff <= -0.04 else "─")
                cells.append(f"{wr*100:>3.0f}%{marker}")
        # Trend: compare last 2 wrs
        if len(wrs) >= 2:
            delta = wrs[-1] - wrs[-2]
            trend = "↑↑" if delta >= 0.06 else ("↑" if delta >= 0.02 else ("↓↓" if delta <= -0.06 else ("↓" if delta <= -0.02 else "→")))
        else:
            trend = "—"
        short = v[:18]
        lines.append(f"{short:<20}  {'  '.join(cells)}  {trend}")

    lines += [
        "━━━━━━━━━━━━━━━━━━━━━",
        f"Baseline: 37.5%  ▲≥41.5%  ▼≤33.5%",
    ]
    reply("\n".join(lines))


# ── Telegram /compare ─────────────────────────────────────────
def _tg_cmd_compare(conn, reply, n: int = 100, since: str = None):
    """P66+P95: Ranked voter WR table + per-SIZE WR breakdown for last N draws.
    Pass since= to filter to fresh-only (post-checkpoint mode).
    """
    import json as _json
    n = max(20, min(n, 500))
    cur = conn.cursor()

    if not config.DATABASE_URL:
        reply("⚠️ Chỉ hỗ trợ PostgreSQL.")
        return

    if since:
        cur.execute("""
            SELECT p.vote_breakdown,
                CASE
                    WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 9  THEN 'NHO'
                    WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 11 THEN 'HOA'
                    ELSE 'LON'
                END AS actual_size
            FROM predictions p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            WHERE p.vote_breakdown IS NOT NULL AND pr.actual_numbers IS NOT NULL
              AND p.created_at > %s
            ORDER BY p.draw_number DESC LIMIT %s
        """, (since, n))
    else:
        cur.execute("""
            SELECT p.vote_breakdown,
                CASE
                    WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 9  THEN 'NHO'
                    WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 11 THEN 'HOA'
                    ELSE 'LON'
                END AS actual_size
            FROM predictions p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            WHERE p.vote_breakdown IS NOT NULL AND pr.actual_numbers IS NOT NULL
            ORDER BY p.draw_number DESC LIMIT %s
        """, (n,))
    rows = cur.fetchall()

    if not rows:
        reply("⚠️ Chưa đủ dữ liệu vote_breakdown.")
        return

    from collections import defaultdict
    # acc[voter][size] = {'w': wins, 't': total} — P95: per-SIZE WR
    acc = defaultdict(lambda: {
        'total': 0, 'wins': 0,
        'NHO': {'w': 0, 't': 0},
        'HOA': {'w': 0, 't': 0},
        'LON': {'w': 0, 't': 0},
    })

    for vb_raw, actual_size in rows:
        try:
            vb = _json.loads(vb_raw) if isinstance(vb_raw, str) else vb_raw
            for vname, vsize in ((vb or {}).get('all_votes') or {}).items():
                a = acc[vname]
                a['total'] += 1
                is_win = vsize == actual_size
                if is_win: a['wins'] += 1
                sz = a.get(vsize, {'w': 0, 't': 0})
                sz['t'] += 1
                if is_win: sz['w'] += 1
                a[vsize] = sz
        except Exception:
            continue

    if not acc:
        reply("⚠️ Không tìm thấy vote nào.")
        return

    BASELINE = 0.375
    SIZE_EMJ = {'NHO': '🔵', 'HOA': '🟡', 'LON': '🔴'}
    voters_sorted = sorted(acc.items(),
                           key=lambda x: x[1]['wins'] / max(x[1]['total'], 1), reverse=True)

    fresh_tag = " · 🔬 FRESH" if since else ""
    lines = [
        f"📊 <b>COMPARE VOTER WR</b> ({len(rows)} kỳ){fresh_tag}",
        "━━━━━━━━━━━━━━━━━━",
    ]

    for rank, (name, a) in enumerate(voters_sorted, 1):
        vn  = a['total']
        wr  = a['wins'] / vn if vn else 0
        edge = wr - BASELINE
        flag = '🟢' if edge > 0.02 else ('🟡' if edge > -0.02 else '🔴')

        # Per-SIZE WR strings
        sz_parts = []
        for sz in ['NHO', 'HOA', 'LON']:
            sd = a.get(sz, {'w': 0, 't': 0})
            if sd['t'] >= 5:
                sz_wr   = sd['w'] / sd['t']
                sz_icon = '↑' if sz_wr >= BASELINE else '↓'
                sz_parts.append(f"{SIZE_EMJ[sz]}{sz_wr*100:.0f}%{sz_icon}")
            else:
                sz_parts.append(f"{SIZE_EMJ[sz]}—")

        lines.append(
            f"{flag} <b>#{rank} {name}</b>  {wr*100:.1f}% ({a['wins']}/{vn}) edge {edge:+.1%}\n"
            f"   {'  '.join(sz_parts)}"
        )

    lines += [
        "━━━━━━━━━━━━━━━━━━",
        "↑ ≥ baseline 37.5% · ↓ dưới baseline",
    ]
    reply("\n".join(lines))


# ── Telegram /checkpoint ──────────────────────────────────────
def _tg_cmd_checkpoint(conn, reply):
    """P150: Validation checkpoint — fresh predictions post-p128, ML LON signal."""
    import math as _math
    from collections import defaultdict
    CHECKPOINT_TS = _CHECKPOINT_TS
    CHECKPOINT_N  = _CHECKPOINT_N
    BASELINE      = 0.375

    if not config.DATABASE_URL:
        reply("⚠️ Chỉ hỗ trợ PostgreSQL.")
        return

    cur = conn.cursor()
    CHECKPOINT_TS, CHECKPOINT_N = _get_checkpoint_config(cur)

    cur.execute("SELECT COUNT(*) FROM predictions WHERE created_at > %s", (CHECKPOINT_TS,))
    n_fresh = cur.fetchone()[0]

    cur.execute("""
        SELECT p.vote_breakdown->'all_votes'->>'ml' AS ml_vote,
               p.vote_breakdown->>'final_size'       AS final_size,
               p.vote_breakdown->>'majority_size'    AS majority_size,
               pr.is_win_size
        FROM predictions p
        JOIN prediction_results pr ON pr.prediction_id = p.id
        WHERE p.created_at > %s AND pr.is_win_size IS NOT NULL
          AND p.vote_breakdown IS NOT NULL
        ORDER BY p.draw_number DESC
    """, (CHECKPOINT_TS,))
    eval_rows = cur.fetchall()

    n_eval = len(eval_rows)
    n_wins = sum(1 for _, _, _, w in eval_rows if w)
    wr_fresh = n_wins / n_eval if n_eval else None

    ml_acc = defaultdict(lambda: {'w': 0, 't': 0})
    ctrl_w, ctrl_t, over_w, over_t = 0, 0, 0, 0
    size_acc = defaultdict(lambda: {'w': 0, 't': 0})  # pre-reg #3: WR by majority_size
    for ml_vote, final_size, majority_size, is_win in eval_rows:
        if ml_vote:
            ml_acc[ml_vote]['t'] += 1
            if is_win:
                ml_acc[ml_vote]['w'] += 1
        if ml_vote and final_size:
            if ml_vote == final_size:
                ctrl_t += 1
                ctrl_w += (1 if is_win else 0)
            else:
                over_t += 1
                over_w += (1 if is_win else 0)
        maj = majority_size or final_size
        if maj:
            size_acc[maj]['t'] += 1
            if is_win:
                size_acc[maj]['w'] += 1

    progress = min(n_fresh / CHECKPOINT_N, 1.0)
    filled   = round(progress * 20)
    bar      = '▓' * filled + '░' * (20 - filled)
    ready    = n_fresh >= CHECKPOINT_N

    lines = [
        "🔬 <b>CHECKPOINT P142–P149</b>",
        "━━━━━━━━━━━━━━━━━━",
        f"📊 Fresh predictions: <b>{n_fresh}/{CHECKPOINT_N}</b>",
        f"{bar} {progress:.0%}",
        "",
    ]

    if ready:
        lines.append("✅ <b>SẴN SÀNG PHÂN TÍCH</b> — đủ 200 kỳ, chạy /compare /voters")
    else:
        remain  = CHECKPOINT_N - n_fresh
        eta_h   = remain / 10  # ~10 draws/hour during game hours
        eta_str = (f"~{int(eta_h*60)}p" if eta_h < 1
                   else f"~{eta_h:.0f}h" if eta_h < 24
                   else f"~{eta_h/24:.1f} ngày")
        lines.append(f"⏳ ETA: {eta_str}  (ước ~10 kỳ/h trong giờ mở)")

    lines.append("")

    if n_eval > 0:
        wr_icon = '🟢' if wr_fresh and wr_fresh >= BASELINE else '🔴'
        lines += [
            f"📈 WR fresh ({n_eval} kỳ có kết quả):",
            f"  {wr_icon} Tổng: <b>{wr_fresh:.1%}</b>  baseline {BASELINE:.1%}",
        ]

        # Batch trend (25/batch, oldest→newest)
        if n_eval >= 25:
            chrono   = list(reversed(eval_rows))
            batches  = [chrono[i:i + 25] for i in range(0, len(chrono), 25)]
            trend_lines = []
            for bi, batch in enumerate(batches):
                bw  = sum(1 for _, _, _, w in batch if w)
                bwr = bw / len(batch)
                bi_icon = '🟢' if bwr >= BASELINE else '🔴'
                trend_lines.append(f"  {bi_icon} B{bi+1}: {bwr:.0%} ({bw}/{len(batch)})")
            lines += ["", "📉 Trend (25/batch cũ→mới):"] + trend_lines

        lines += ["", "🤖 ML vote split (fresh):"]
        for sz in ['NHO', 'LON', 'HOA']:
            a = ml_acc.get(sz)
            if sz == 'HOA' and (not a or a['t'] == 0):
                lines.append("  ✅ HOA: 0 votes (blocked)")
                continue
            if not a or a['t'] == 0:
                continue
            wr_sz   = a['w'] / a['t']
            icon    = '✅' if wr_sz >= BASELINE else '❌'
            signal  = ' ← pre-reg #1' if sz == 'LON' else ''
            lines.append(f"  {icon} {sz}: {wr_sz:.1%} (n={a['t']}){signal}")

        ship_p171 = ship_p172 = False
        ml_lon = ml_acc.get('LON', {'w': 0, 't': 0})
        if ml_lon['t'] >= 10:
            wr_lon = ml_lon['w'] / ml_lon['t']
            z = (wr_lon - BASELINE) / _math.sqrt(BASELINE * (1 - BASELINE) / ml_lon['t'])
            if z <= -2.0 and n_fresh >= CHECKPOINT_N:
                act = " → <b>SHIP P171: redirect ml LON→NHO</b>"
                ship_p171 = True
            else:
                act = f" (cần z≤−2.0, hiện {z:+.2f})" if n_fresh < CHECKPOINT_N else f" (z={z:+.2f} chưa đủ)"
            lines += ["", f"   z ML LON: <b>{z:+.2f}</b>  (pre-reg #1){act}"]

        # Pre-reg hypothesis #2: ml_controls vs ml_overridden
        if ctrl_t >= 5 or over_t >= 5:
            lines.append("")
            lines.append("📐 ML controls vs overridden (pre-reg #2):")
            if ctrl_t >= 5:
                wr_c = ctrl_w / ctrl_t
                z_c  = (wr_c - BASELINE) / _math.sqrt(BASELINE * (1 - BASELINE) / ctrl_t)
                ic   = '🔴' if wr_c < BASELINE else '🟢'
                if z_c <= -2.0 and n_fresh >= CHECKPOINT_N:
                    act2 = " → <b>SHIP P171: xóa ml voter</b>"
                    ship_p171 = True
                else:
                    act2 = f" (cần z≤−2.0+n=200, z={z_c:+.2f})"
                lines.append(f"  {ic} controls: {wr_c:.1%} (n={ctrl_t}, z={z_c:+.2f}){act2}")
            if over_t >= 5:
                wr_o = over_w / over_t
                io   = '🟢' if wr_o >= BASELINE else '🔴'
                lines.append(f"  {io} overridden: {wr_o:.1%} (n={over_t})")

        # Pre-reg hypothesis #3: WR by predicted SIZE (majority_size)
        size_data = [(sz, size_acc[sz]) for sz in ['NHO', 'LON', 'HOA'] if size_acc[sz]['t'] >= 5]
        if size_data:
            lines.append("")
            lines.append("📏 WR by pred SIZE (pre-reg #3):")
            sz_wrs = {}
            for sz, sd in size_data:
                wr_s = sd['w'] / sd['t']
                ic_s = '🟢' if wr_s >= BASELINE else '🔴'
                lines.append(f"  {ic_s} {sz}: {wr_s:.1%} (n={sd['t']})")
                sz_wrs[sz] = wr_s
            # Action hint when LON WR is significantly below NHO on sufficient data
            nho_t = size_acc['NHO']['t']
            lon_t = size_acc['LON']['t']
            if nho_t >= 20 and lon_t >= 20 and n_fresh >= CHECKPOINT_N:
                nho_wr = sz_wrs.get('NHO', 0)
                lon_wr = sz_wrs.get('LON', 0)
                if lon_wr < nho_wr - 0.08:
                    lines.append(f"  → <b>SHIP P172: block LON majority predictions</b>")
                    ship_p172 = True

        # Conclusion block when checkpoint is complete
        if ready and n_eval > 0:
            lines.append("")
            lines.append("━━━━━━━━━━━━━━━━━━")
            if ship_p171 or ship_p172:
                lines.append("🚨 <b>KẾT LUẬN: CÓ THAY ĐỔI CẦN SHIP (xem hint trên)</b>")
            else:
                wr_delta = (wr_fresh - BASELINE) * 100 if wr_fresh else 0
                delta_str = f"+{wr_delta:.1f}pp" if wr_delta >= 0 else f"{wr_delta:.1f}pp"
                lines.append(f"✅ <b>KẾT LUẬN: KHÔNG CẦN THAY ĐỔI</b>")
                lines.append(f"   WR fresh {wr_fresh:.1%} ({delta_str} vs baseline) — hệ thống ổn định")
                lines.append(f"   3/3 pre-reg hypothesis không đạt ngưỡng → giữ nguyên thuật toán")
    else:
        lines.append("📈 Chưa có kết quả  (game chưa mở hoặc chưa sync)")

    lines += ["", f"⏱ Mốc: {_CHECKPOINT_TS} UTC (post-p128)"]
    reply("\n".join(lines))


# ── Telegram AI Chat ──────────────────────────────────────────
def _tg_ai_chat(conn, user_message: str, reply):
    """Gọi LLM với context DB, trả lời tiếng Việt qua Telegram."""
    import ast as _ast

    # 1. Thu thập context từ DB
    cur = conn.cursor()
    # 10 kỳ gần nhất
    cur.execute("""
        SELECT draw_number, numbers, size_category, sum_value
        FROM draw_history ORDER BY draw_number DESC LIMIT 10
    """)
    rows = cur.fetchall()
    draws_ctx = ""
    for dn, nums, cat, sv in reversed(rows):
        if isinstance(nums, str):
            try: nums = _ast.literal_eval(nums)
            except Exception: pass
        nums_str = "-".join(str(x) for x in nums) if isinstance(nums, list) else str(nums)
        draws_ctx += f"  #{dn}: {nums_str} tổng={sv} ({cat})\n"

    # Win rate 24h
    cur.execute("""
        SELECT COUNT(*),
               COALESCE(SUM(CASE WHEN is_win THEN 1 ELSE 0 END),0)
        FROM prediction_results
        WHERE actual_numbers IS NOT NULL
          AND created_at > NOW() - INTERVAL '24 hours'
    """)
    r = cur.fetchone(); n24, w24 = (r[0], r[1]) if r else (0, 0)
    wr24 = f"{w24/n24:.1%}" if n24 else "N/A"

    # Dự đoán mới nhất
    cur.execute("""
        SELECT draw_number, predicted_numbers, model_name, confidence
        FROM predictions ORDER BY draw_number DESC LIMIT 1
    """)
    pr = cur.fetchone()
    if pr:
        pnums = pr[1]
        if isinstance(pnums, str):
            try: pnums = _ast.literal_eval(pnums)
            except Exception: pass
        pred_ctx = (f"Dự đoán kỳ #{pr[0]}: "
                    f"{'-'.join(str(x) for x in pnums) if isinstance(pnums, list) else pnums} "
                    f"(model={pr[2]}, conf={pr[3]:.1%})" if pr[3] else f"(model={pr[2]})")
    else:
        pred_ctx = "Chưa có dự đoán."

    from zoneinfo import ZoneInfo
    vn_now = datetime.now(ZoneInfo("Asia/Ho_Chi_Minh")).strftime("%H:%M %d/%m/%Y")

    system_prompt = (
        "Bạn là trợ lý AI của hệ thống dự đoán Bingo18 (xổ số Việt Nam).\n"
        "Luật game: mỗi kỳ rút 3 số (1–6), tổng 3–9=NHO, 10–11=HOA, 12–18=LON.\n"
        "Trả lời ngắn gọn, thân thiện, bằng tiếng Việt. Tối đa 200 từ.\n"
        "Không bịa số liệu ngoài context được cung cấp."
    )
    user_prompt = (
        f"[Context hệ thống - {vn_now}]\n"
        f"10 kỳ gần nhất:\n{draws_ctx}"
        f"Win rate 24h: {wr24} ({w24}/{n24} kỳ)\n"
        f"Dự đoán mới nhất: {pred_ctx}\n\n"
        f"[Câu hỏi của người dùng]\n{user_message}"
    )

    # 2. Gọi LLM: Groq (nhanh) → Gemini (fallback)
    groq_key   = config.GROQ_API_KEY   if hasattr(config, 'GROQ_API_KEY')   else os.environ.get('GROQ_API_KEY', '')
    gemini_key = config.GEMINI_API_KEY if hasattr(config, 'GEMINI_API_KEY') else os.environ.get('GEMINI_API_KEY', '')
    openrouter_key = config.OPENROUTER_API_KEY if hasattr(config, 'OPENROUTER_API_KEY') else os.environ.get('OPENROUTER_API_KEY', '')

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user",   "content": user_prompt},
    ]
    answer = None

    # Thử Groq trước (latency ~1s)
    if groq_key and not answer:
        try:
            r = requests.post(
                "https://api.groq.com/openai/v1/chat/completions",
                json={"model": "llama-3.3-70b-versatile", "messages": messages,
                      "max_tokens": 400, "temperature": 0.5},
                headers={"Authorization": f"Bearer {groq_key}",
                         "Content-Type": "application/json"},
                timeout=15,
            )
            if r.ok:
                answer = r.json()["choices"][0]["message"]["content"].strip()
        except Exception:
            pass

    # Fallback: OpenRouter
    if openrouter_key and not answer:
        try:
            r = requests.post(
                "https://openrouter.ai/api/v1/chat/completions",
                json={"model": "openrouter/auto", "messages": messages,
                      "max_tokens": 400, "temperature": 0.5},
                headers={"Authorization": f"Bearer {openrouter_key}",
                         "Content-Type": "application/json",
                         "HTTP-Referer": "https://bingo18-633959711537.asia-southeast1.run.app"},
                timeout=20,
            )
            if r.ok:
                answer = r.json()["choices"][0]["message"]["content"].strip()
        except Exception:
            pass

    # Fallback: Gemini
    if gemini_key and not answer:
        try:
            r = requests.post(
                f"https://generativelanguage.googleapis.com/v1beta/models/"
                f"gemini-2.0-flash:generateContent?key={gemini_key}",
                json={"contents": [{"parts": [{"text": f"{system_prompt}\n\n{user_prompt}"}]}],
                      "generationConfig": {"temperature": 0.5, "maxOutputTokens": 400}},
                headers={"Content-Type": "application/json"},
                timeout=20,
            )
            if r.ok:
                answer = r.json()["candidates"][0]["content"]["parts"][0]["text"].strip()
        except Exception:
            pass

    if answer:
        reply(f"🤖 {answer}")
    else:
        reply("⚠️ AI tạm thời không khả dụng. Thử lại sau hoặc dùng /status, /stats.")


# ── Telegram Webhook: commands + ảnh → vision AI ─────────────
@app.route('/telegram/webhook', methods=['POST'])
def telegram_webhook():
    import logging as _log
    _wlog = _log.getLogger("telegram_webhook")
    try:
        update = request.get_json(force=True, silent=True) or {}

        # Dedup: bỏ qua nếu update_id đã xử lý (Telegram có thể retry)
        uid = update.get("update_id")
        if uid:
            if uid in _PROCESSED_UPDATES:
                return jsonify({"ok": True})
            _PROCESSED_UPDATES.add(uid)
            if len(_PROCESSED_UPDATES) > _PROCESSED_MAX:
                _PROCESSED_UPDATES.pop()

        # Hỗ trợ cả message và callback_query (inline keyboard)
        cb = update.get("callback_query")
        if cb:
            message  = cb.get("message") or {}
            chat_id  = str(message.get("chat", {}).get("id", ""))
            text_msg = cb.get("data", "")
            # Ack callback
            try:
                requests.post(
                    f"https://api.telegram.org/bot{config.TELEGRAM_BOT_TOKEN}/answerCallbackQuery",
                    json={"callback_query_id": cb["id"]}, timeout=5
                )
            except Exception:
                pass
        else:
            message  = update.get("message") or update.get("edited_message") or {}
            chat_id  = str(message.get("chat", {}).get("id", ""))
            text_msg = message.get("text", "")

        bot_token = config.TELEGRAM_BOT_TOKEN

        def reply(text: str, markup=None):
            _tg_reply(bot_token, chat_id, text, reply_markup=markup)

        # ── Auth: chỉ phản hồi chat được cấu hình ────────────
        if chat_id and str(config.TELEGRAM_CHAT_ID) and chat_id != str(config.TELEGRAM_CHAT_ID):
            return jsonify({"ok": True})

        # ── Per-user command throttle ─────────────────────────
        if cmd_raw := (text_msg.split()[0].lower().split("@")[0] if text_msg.strip() else ""):
            if cmd_raw.startswith("/") and cmd_raw not in ("/start", "/help"):
                import time as _ttime
                _now_t = _ttime.monotonic()
                _cool  = _TG_SLOW_COOLDOWN_SEC if cmd_raw in _TG_SLOW_CMDS else _TG_CMD_COOLDOWN_SEC
                _last  = _TG_CMD_COOLDOWN.get(chat_id, 0.0)
                if _now_t - _last < _cool:
                    _wait = int(_cool - (_now_t - _last)) + 1
                    reply(f"⏱ Vui lòng chờ {_wait}s trước khi dùng lệnh tiếp theo.")
                    return jsonify({"ok": True})
                _TG_CMD_COOLDOWN[chat_id] = _now_t
                if len(_TG_CMD_COOLDOWN) > 500:   # cap memory
                    oldest = min(_TG_CMD_COOLDOWN, key=_TG_CMD_COOLDOWN.get)
                    del _TG_CMD_COOLDOWN[oldest]

        # ── Xử lý lệnh text ──────────────────────────────────
        cmd = cmd_raw  # already parsed by throttle block above

        if cmd in ("/start", "/help"):
            reply(
                "🤖 <b>BINGO18 BOT</b>\n"
                "━━━━━━━━━━━━━━━━━━\n"
                "Chọn lệnh hoặc gõ:\n"
                "/predict  — Dự đoán kỳ tiếp theo\n"
                "/winrate  — Rolling win rate + streak\n"
                "/history  — 10 kỳ gần nhất\n"
                "/voters [N|fresh] — WR + per-SIZE WR + conf + decay per voter; 'fresh' → post-checkpoint only\n"
                "/abstain  — Markov abstain rate + SIZE distribution\n"
                "/compare [N|fresh] — Ranked voter WR bảng; 'fresh' → post-checkpoint only\n"
                "/stats    — Thống kê tổng hợp\n"
                "/status   — Trạng thái hệ thống\n"
                "/calibration [N] — Brier score, ECE, confidence gap theo SIZE\n"
                "/explain  — Giải thích vote breakdown kỳ gần nhất\n"
                "/health   — Tóm tắt sức khỏe hệ thống + voter + adaptive\n"
                "/top [N]  — Top 5 high-conf wins + losses (mặc định 500 kỳ)\n"
                "/recap [N] — Tóm tắt N kỳ: WR, SIZE dist, giờ tốt, voter dẫn đầu\n"
                "/autotune  — Adaptive state hiện tại + xu hướng 3 lô × 25 kỳ\n"
                "/alerts [N] — N alert gần nhất từ DB + tổng 24h/7d (mặc định 5)\n"
                "/hourly [N] — Win rate theo từng giờ 6h-22h + tổng theo buổi (n mặc định 1000)\n"
                "/wincal [N] — Lịch WR N ngày qua dạng lưới 7 cột (mặc định 28 ngày)\n"
                "/next       — Preview N+1 (hiện tại) + N+2/N+3 theo Markov\n"
                "/votertrend [n] [batch] — WR từng voter qua 5 batch gần nhất (mặc định n=500 batch=25)\n"
                "/checkpoint  — Trạng thái validation checkpoint (X/200 kỳ, WR, z-score ML)\n"
                "/new_checkpoint [N] — Reset checkpoint mới từ bây giờ (mặc định N=200)",
                markup=_tg_main_keyboard()
            )
            return jsonify({"ok": True})

        if cmd in ("/predict", "/winrate", "/stats", "/status", "/history", "/voters", "/abstain", "/compare", "/trend", "/dow", "/hourly", "/calibration", "/explain", "/health", "/top", "/recap", "/autotune", "/alerts", "/wincal", "/next", "/votertrend", "/checkpoint", "/new_checkpoint"):
            conn2 = None
            try:
                conn2 = db.get_connection()
                if cmd == "/predict":
                    _tg_cmd_predict(conn2, reply)
                elif cmd == "/winrate":
                    _tg_cmd_winrate(conn2, reply)
                elif cmd == "/voters":
                    parts = text_msg.split()
                    fresh_mode = any(p.lower() == 'fresh' for p in parts[1:])
                    n_v = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 200
                    _tg_cmd_voters(conn2, reply, n_v,
                                   since=_CHECKPOINT_TS if fresh_mode else None)
                elif cmd == "/abstain":
                    _tg_cmd_abstain(conn2, reply)
                elif cmd == "/stats":
                    _tg_cmd_stats(conn2, reply)
                elif cmd == "/status":
                    _tg_cmd_status(conn2, reply)
                elif cmd == "/history":
                    parts = text_msg.split()
                    n = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 10
                    _tg_cmd_history(conn2, reply, n)
                elif cmd == "/compare":
                    parts = text_msg.split()
                    fresh_mode = any(p.lower() == 'fresh' for p in parts[1:])
                    n = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 100
                    _tg_cmd_compare(conn2, reply, n,
                                    since=_CHECKPOINT_TS if fresh_mode else None)
                elif cmd == "/trend":
                    _tg_cmd_trend(conn2, reply)
                elif cmd == "/dow":
                    parts = text_msg.split()
                    n = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 500
                    _tg_cmd_dow(conn2, reply, n)
                elif cmd == "/hourly":
                    parts = text_msg.split()
                    n = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 1000
                    _tg_cmd_hourly(conn2, reply, n)
                elif cmd == "/calibration":
                    parts = text_msg.split()
                    n = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 500
                    _tg_cmd_calibration(conn2, reply, n)
                elif cmd == "/explain":
                    _tg_cmd_explain(conn2, reply)
                elif cmd == "/health":
                    _tg_cmd_health(conn2, reply)
                elif cmd == "/top":
                    parts = text_msg.split()
                    n = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 500
                    _tg_cmd_top(conn2, reply, n)
                elif cmd == "/recap":
                    parts = text_msg.split()
                    n = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 100
                    _tg_cmd_recap(conn2, reply, n)
                elif cmd == "/autotune":
                    _tg_cmd_autotune(conn2, reply)
                elif cmd == "/alerts":
                    parts = text_msg.split()
                    n_al = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 5
                    _tg_cmd_alerts(conn2, reply, n_al)
                elif cmd == "/wincal":
                    parts = text_msg.split()
                    n_wc = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 28
                    _tg_cmd_wincal(conn2, reply, n_wc)
                elif cmd == "/next":
                    _tg_cmd_next(conn2, reply)
                elif cmd == "/votertrend":
                    parts = text_msg.split()
                    n_vt    = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 500
                    batch_vt = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 25
                    _tg_cmd_votertrend(conn2, reply, n_vt, batch_vt)
                elif cmd == "/checkpoint":
                    _tg_cmd_checkpoint(conn2, reply)
                elif cmd == "/new_checkpoint":
                    parts = text_msg.split()
                    n_cp = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 200
                    ts_cp = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
                    cur2 = conn2.cursor()
                    _save_checkpoint_config(cur2, ts_cp, n_cp)
                    conn2.commit()
                    from zoneinfo import ZoneInfo
                    ts_vn = datetime.utcnow().astimezone(ZoneInfo('Asia/Ho_Chi_Minh')).strftime('%H:%M %d/%m/%Y')
                    reply(
                        f"✅ <b>Checkpoint mới đã được đặt!</b>\n"
                        f"📅 Từ: <code>{ts_cp} UTC</code> ({ts_vn} VN)\n"
                        f"🎯 Target: <b>{n_cp}</b> kỳ\n"
                        f"Dùng /checkpoint để theo dõi tiến độ."
                    )
            except Exception as ce:
                _wlog.error("Command %s error: %s", cmd, ce)
                reply(f"❌ Lỗi: {str(ce)[:200]}")
            finally:
                if conn2:
                    try: conn2.close()
                    except Exception: pass
            return jsonify({"ok": True})

        if not text_msg.startswith("/") and text_msg.strip():
            # Tin nhắn tự do → AI chat
            conn2 = None
            try:
                conn2 = db.get_connection()
                _tg_ai_chat(conn2, text_msg.strip(), reply)
            except Exception as ce:
                _wlog.error("AI chat error: %s", ce)
                reply("❌ Lỗi xử lý tin nhắn.")
            finally:
                if conn2:
                    try: conn2.close()
                    except Exception: pass
            return jsonify({"ok": True})

    except Exception as e:
        _log.getLogger("telegram_webhook").error("Webhook error: %s", e)
        try:
            reply(f"❌ Lỗi xử lý: {str(e)[:200]}")
        except Exception:
            pass

    return jsonify({"ok": True})


@app.route('/api/max3d/recent')
@limiter.limit("30 per minute")
def api_max3d_recent():
    """P138: Last 10 Max3D draws with jackpot numbers and prediction result."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT d.draw_number, d.draw_date::text, d.jackpot_1, d.jackpot_2,
                   d.prize1_numbers, d.prize2_numbers, d.prize3_numbers,
                   p.predicted_numbers, p.confidence, p.model_name,
                   r.is_win_jackpot, r.is_win_any, r.match_count
            FROM max3d_draw_history d
            LEFT JOIN max3d_predictions p ON p.draw_number = d.draw_number
            LEFT JOIN max3d_prediction_results r ON r.draw_number = d.draw_number
            ORDER BY d.draw_number DESC
            LIMIT 10
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        draws = []
        for row in rows:
            draws.append({
                "draw_number":      row[0],
                "draw_date":        row[1],
                "jackpot_1":        row[2],
                "jackpot_2":        row[3],
                "prize1":           row[4] if row[4] else [],
                "prize2":           row[5] if row[5] else [],
                "prize3":           row[6] if row[6] else [],
                "predicted":        row[7] if row[7] else [],
                "confidence":       round(row[8] * 100, 1) if row[8] is not None else None,
                "model_name":       row[9],
                "is_win_jackpot":   row[10],
                "is_win_any":       row[11],
                "match_count":      row[12] if row[12] is not None else 0,
            })
        return jsonify({"draws": draws})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/max3d/stats')
@limiter.limit("30 per minute")
def api_max3d_stats():
    """P138: Max3D prediction win-rate summary + latest prediction."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT
                COUNT(*)                                         AS total_evaluated,
                SUM(CASE WHEN is_win_any     THEN 1 ELSE 0 END) AS any_wins,
                SUM(CASE WHEN is_win_jackpot THEN 1 ELSE 0 END) AS jackpot_wins
            FROM max3d_prediction_results
        """)
        row = cur.fetchone()
        cur.close()

        cur2 = conn.cursor()
        cur2.execute("""
            SELECT draw_number, predicted_numbers, confidence, model_name
            FROM max3d_predictions
            ORDER BY draw_number DESC LIMIT 1
        """)
        pred = cur2.fetchone()
        cur2.close()

        cur3 = conn.cursor()
        cur3.execute("SELECT COUNT(*) FROM max3d_predictions")
        total_predicted = (cur3.fetchone() or [0])[0]
        cur3.close()

        cur4 = conn.cursor()
        cur4.execute("SELECT draw_number FROM max3d_draw_history ORDER BY draw_number DESC LIMIT 1")
        latest_draw = (cur4.fetchone() or [None])[0]
        cur4.close()
        conn.close()

        total_eval   = row[0] or 0
        any_wins     = row[1] or 0
        jackpot_wins = row[2] or 0

        return jsonify({
            "latest_draw":           latest_draw,
            "total_draws":           latest_draw,
            "total_predicted":       total_predicted,
            "total_evaluated":       total_eval,
            "any_win_count":         any_wins,
            "jackpot_win_count":     jackpot_wins,
            "any_win_rate_pct":      round(any_wins / total_eval * 100, 1) if total_eval > 0 else 0,
            "jackpot_win_rate_pct":  round(jackpot_wins / total_eval * 100, 1) if total_eval > 0 else 0,
            "next_prediction": {
                "draw_number": pred[0],
                "numbers":     pred[1] if pred[1] else [],
                "confidence":  round(pred[2] * 100, 1) if pred[2] is not None else None,
                "model":       pred[3],
            } if pred else None,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/search-history')
@limiter.limit("30 per minute")
def search_history():
    """Tìm kiếm lịch sử: sau khi ra bộ số / tổng X thì kỳ tiếp theo thường ra gì."""
    import re as _re
    q    = request.args.get('q', '').strip()
    mode = request.args.get('mode', 'numbers')  # 'numbers' | 'sum'
    if not q:
        return jsonify({"error": "Thiếu tham số q"}), 400

    query_nums = None  # sorted list[int] len=3, each 1-6
    query_sum  = None  # int 3-18

    if mode == 'sum':
        try:
            s = int(q)
            if 3 <= s <= 18:
                query_sum = s
        except ValueError:
            pass
    else:
        raw_digits = [int(c) for c in _re.sub(r'[^\d]', '', q) if c.isdigit()]
        if len(raw_digits) == 3 and all(1 <= n <= 6 for n in raw_digits):
            query_nums = raw_digits  # P148: ordered — 1-1-4 ≠ 4-1-1

    if query_nums is None and query_sum is None:
        return jsonify({"error": "Query không hợp lệ. Ví dụ: '356' hoặc tổng '14'"}), 400

    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT d.draw_number, d.numbers, d.sum_value, d.size_category,
                   n.numbers, n.sum_value, n.size_category
            FROM draw_history d
            LEFT JOIN draw_history n ON n.draw_number = d.draw_number + 1
            ORDER BY d.draw_number
        """)
        rows = cur.fetchall()
        conn.close()

        from collections import Counter
        matches, has_next = [], []
        for draw_num, nums_raw, sum_val, size, nn_raw, nn_sum, nn_size in rows:
            nums = json.loads(nums_raw) if isinstance(nums_raw, str) else (nums_raw or [])
            nums_int = [int(n) for n in nums]
            if query_nums is not None:
                if nums_int != query_nums:  # P148: exact order match
                    continue
            else:
                if sum_val != query_sum:
                    continue
            next_nums = []
            if nn_raw:
                next_nums = json.loads(nn_raw) if isinstance(nn_raw, str) else (nn_raw or [])
                next_nums = [int(n) for n in next_nums]
            m = dict(draw_number=draw_num, numbers=nums_int, sum=sum_val, size=size,
                     next_numbers=next_nums, next_sum=nn_sum, next_size=nn_size)
            matches.append(m)
            if next_nums:
                has_next.append(m)

        n_found, n_next = len(matches), len(has_next)
        if n_found == 0:
            return jsonify({"query": q, "mode": mode, "occurrences": 0,
                            "message": "Không tìm thấy kỳ nào khớp"})

        num_freq  = Counter()
        sum_freq  = Counter()
        size_freq = Counter()
        combo_freq = Counter()
        for m in has_next:
            nn = m['next_numbers']
            for n in nn:
                num_freq[n] += 1
            if m['next_sum']:
                sum_freq[m['next_sum']] += 1
            if m['next_size']:
                size_freq[m['next_size']] += 1
            combo_freq[tuple(sorted(nn))] += 1

        return jsonify({
            "query": q, "mode": mode,
            "query_value": query_nums or query_sum,
            "occurrences": n_found,
            "has_next": n_next,
            "next_number_freq": [
                {"number": n, "count": num_freq.get(n, 0),
                 "pct": round(num_freq.get(n, 0) / (n_next * 3) * 100, 1) if n_next else 0}
                for n in range(1, 7)
            ],
            "next_sum_top": [
                {"sum": s, "count": c,
                 "size": "NHO" if s <= 9 else ("HOA" if s <= 11 else "LON"),
                 "pct": round(c / n_next * 100, 1)}
                for s, c in sum_freq.most_common(8)
            ],
            "next_size_dist": {s: round(size_freq.get(s, 0) / n_next * 100, 1)
                               for s in ["NHO", "HOA", "LON"]} if n_next else {},
            "next_combo_top": [
                {"numbers": list(k), "count": v,
                 "pct": round(v / n_next * 100, 1)}
                for k, v in combo_freq.most_common(8)
            ],
            "recent_examples": [
                {"draw_number": m['draw_number'], "numbers": m['numbers'],
                 "sum": m['sum'], "size": m['size'],
                 "next_numbers": m['next_numbers'],
                 "next_sum": m['next_sum'], "next_size": m['next_size']}
                for m in reversed(matches[-5:])
            ],
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/telegram/set-webhook', methods=['POST'])
def set_telegram_webhook():
    """One-time setup: register webhook URL with Telegram."""
    secret = request.headers.get("X-Admin-Secret", "")
    if secret != config.ADMIN_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 401

    webhook_url = f"https://bingo18-633959711537.asia-southeast1.run.app/telegram/webhook"
    bot_token   = config.TELEGRAM_BOT_TOKEN
    r = requests.post(
        f"https://api.telegram.org/bot{bot_token}/setWebhook",
        json={"url": webhook_url, "drop_pending_updates": True},
        timeout=10
    )
    return jsonify(r.json())


@app.route('/api/recent-outcomes')
@limiter.limit("60 per minute")
def recent_outcomes():
    """10 kỳ gần nhất với kết quả thực tế và dự đoán (majority_vote)."""
    def _size(nums):
        s = sum(nums)
        return 'HOA' if 10 <= s <= 11 else ('NHO' if s <= 9 else 'LON')
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        if USE_POSTGRES:
            cur.execute("""
                SELECT d.draw_number,
                       d.numbers,
                       d.size_category,
                       d.sum_value,
                       d.draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh' AS draw_vn,
                       p.predicted_numbers,
                       p.confidence,
                       pr.is_win_size,
                       pr.match_count
                FROM draw_history d
                LEFT JOIN predictions p
                  ON p.draw_number = d.draw_number AND p.model_name = 'majority_vote'
                LEFT JOIN prediction_results pr ON pr.prediction_id = p.id
                ORDER BY d.draw_number DESC LIMIT 10
            """)
        else:
            cur.execute("""
                SELECT d.draw_number, d.numbers, d.size_category, d.sum_value,
                       d.draw_time, p.predicted_numbers, p.confidence,
                       pr.is_win_size, pr.match_count
                FROM draw_history d
                LEFT JOIN predictions p
                  ON p.draw_number = d.draw_number AND p.model_name = 'majority_vote'
                LEFT JOIN prediction_results pr ON pr.prediction_id = p.id
                ORDER BY d.draw_number DESC LIMIT 10
            """)
        rows = cur.fetchall()
        conn.close()
        results = []
        for r in rows:
            draw_n, nums_raw, sz, total, draw_vn, pred_raw, conf, is_win, match = r
            try: nums = json.loads(nums_raw) if isinstance(nums_raw, str) else (nums_raw or [])
            except Exception: nums = []
            try: pred = json.loads(pred_raw) if isinstance(pred_raw, str) else (pred_raw or [])
            except Exception: pred = []
            pred_sz = _size(pred) if pred else None
            results.append({
                'draw_number': draw_n,
                'numbers': nums,
                'size': sz,
                'total': total,
                'draw_time': str(draw_vn)[:16] if draw_vn else None,
                'pred_numbers': pred,
                'pred_size': pred_sz,
                'is_win': is_win,
                'match_count': match,
                'confidence': round(float(conf), 4) if conf else None,
            })
        return jsonify({'draws': results})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/size-streak')
@limiter.limit("60 per minute")
def size_streak():
    """SIZE nào đang streak (liên tiếp) và streak dài bao nhiêu kỳ."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute(
            "SELECT draw_number, size_category FROM draw_history "
            "ORDER BY draw_number DESC LIMIT 50"
        )
        rows = cur.fetchall()
        conn.close()
        if not rows:
            return jsonify({'streak_size': None, 'streak_len': 0, 'recent': []})
        last_sz = rows[0][1]
        streak = 0
        for _, sz in rows:
            if sz == last_sz:
                streak += 1
            else:
                break
        recent = [{'draw_number': r[0], 'size': r[1]} for r in rows[:20]]
        return jsonify({'streak_size': last_sz, 'streak_len': streak, 'recent': recent})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/wl-streak')
@limiter.limit("60 per minute")
def wl_streak():
    """Win/Loss streak hiện tại từ majority_vote predictions."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT p.draw_number, pr.is_win_size
            FROM predictions p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            WHERE p.model_name = 'majority_vote'
            ORDER BY p.draw_number DESC LIMIT 50
        """)
        rows = cur.fetchall()
        conn.close()
        if not rows:
            return jsonify({'result': None, 'streak_len': 0, 'recent': []})
        last_win = rows[0][1]
        streak = 0
        for _, w in rows:
            if w == last_win:
                streak += 1
            else:
                break
        recent = [{'draw_number': r[0], 'win': r[1]} for r in rows[:20]]
        return jsonify({'result': 'WIN' if last_win else 'LOSS', 'streak_len': streak, 'recent': recent})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/today-draws')
@limiter.limit("60 per minute")
def today_draws():
    """Tất cả kỳ hôm nay (VN time) với kết quả và dự đoán majority_vote."""
    def _size(nums):
        s = sum(nums)
        return 'HOA' if 10 <= s <= 11 else ('NHO' if s <= 9 else 'LON')
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        if USE_POSTGRES:
            cur.execute("""
                SELECT d.draw_number,
                       d.numbers,
                       d.size_category,
                       d.draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh' AS draw_vn,
                       p.predicted_numbers,
                       p.confidence,
                       pr.is_win_size,
                       pr.match_count
                FROM draw_history d
                LEFT JOIN predictions p
                  ON p.draw_number = d.draw_number AND p.model_name = 'majority_vote'
                LEFT JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE (d.draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                      = (NOW() AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                ORDER BY d.draw_number DESC
            """)
        else:
            cur.execute("""
                SELECT d.draw_number, d.numbers, d.size_category, d.draw_time,
                       p.predicted_numbers, p.confidence, pr.is_win_size, pr.match_count
                FROM draw_history d
                LEFT JOIN predictions p
                  ON p.draw_number = d.draw_number AND p.model_name = 'majority_vote'
                LEFT JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE date(d.draw_time) = date('now')
                ORDER BY d.draw_number DESC
            """)
        rows = cur.fetchall()
        conn.close()
        results = []
        for r in rows:
            draw_n, nums_raw, sz, draw_vn, pred_raw, conf, is_win, match = r
            try: nums = json.loads(nums_raw) if isinstance(nums_raw, str) else (nums_raw or [])
            except Exception: nums = []
            try: pred = json.loads(pred_raw) if isinstance(pred_raw, str) else (pred_raw or [])
            except Exception: pred = []
            from collections import Counter
            has_repeat = len(nums) > len(set(nums)) if nums else False
            results.append({
                'draw_number': draw_n,
                'numbers': nums,
                'size': sz,
                'draw_time': str(draw_vn)[11:16] if draw_vn else None,
                'pred_numbers': pred,
                'pred_size': _size(pred) if pred else None,
                'is_win': is_win,
                'match_count': match,
                'confidence': round(float(conf), 3) if conf else None,
                'has_repeat': has_repeat,
            })
        wins = sum(1 for r in results if r['is_win'])
        total_with_pred = sum(1 for r in results if r['is_win'] is not None)
        repeats = sum(1 for r in results if r['has_repeat'])
        return jsonify({
            'draws': results,
            'total': len(results),
            'wins': wins,
            'total_with_pred': total_with_pred,
            'wr': round(wins / total_with_pred, 4) if total_with_pred else None,
            'repeat_count': repeats,
            'repeat_pct': round(repeats / len(results) * 100, 1) if results else 0,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/repeat-stats')
@limiter.limit("60 per minute")
def repeat_stats():
    """Thống kê số lặp trong draw (≥2 số giống nhau): tỉ lệ, streak hiện tại, phân bố 100 kỳ."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute(
            "SELECT draw_number, numbers FROM draw_history "
            "ORDER BY draw_number DESC LIMIT 200"
        )
        rows = cur.fetchall()
        conn.close()
        results = []
        for draw_n, nums_raw in rows:
            try: nums = json.loads(nums_raw) if isinstance(nums_raw, str) else (nums_raw or [])
            except Exception: nums = []
            has_repeat = len(nums) > len(set(nums))
            triple = len(nums) == 3 and len(set(nums)) == 1
            results.append({'draw_number': draw_n, 'has_repeat': has_repeat, 'triple': triple})
        # streak hiện tại
        last_rep = results[0]['has_repeat'] if results else None
        streak = 0
        for r in results:
            if r['has_repeat'] == last_rep:
                streak += 1
            else:
                break
        n100 = results[:100]
        rep100 = sum(1 for r in n100 if r['has_repeat'])
        tri100 = sum(1 for r in n100 if r['triple'])
        recent20 = [{'draw_number': r['draw_number'], 'has_repeat': r['has_repeat'], 'triple': r['triple']} for r in results[:20]]
        return jsonify({
            'streak_type': 'repeat' if last_rep else 'no_repeat',
            'streak_len': streak,
            'repeat_pct_100': round(rep100 / len(n100) * 100, 1) if n100 else 0,
            'triple_pct_100': round(tri100 / len(n100) * 100, 1) if n100 else 0,
            'repeat_count_100': rep100,
            'triple_count_100': tri100,
            'recent': recent20,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/number-hot-today')
@limiter.limit("60 per minute")
def number_hot_today():
    """Tần suất mỗi số 1-6 xuất hiện trong các kỳ hôm nay (VN time)."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        if USE_POSTGRES:
            cur.execute("""
                SELECT numbers FROM draw_history
                WHERE (draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                      = (NOW() AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                ORDER BY draw_number DESC
            """)
        else:
            cur.execute("SELECT numbers FROM draw_history WHERE date(draw_time)=date('now') ORDER BY draw_number DESC")
        rows = cur.fetchall()
        conn.close()
        from collections import Counter
        cnt = Counter()
        total_draws = 0
        for (nums_raw,) in rows:
            try: nums = json.loads(nums_raw) if isinstance(nums_raw, str) else (nums_raw or [])
            except Exception: nums = []
            for n in nums:
                cnt[n] += 1
            total_draws += 1
        freq = [{'number': i, 'count': cnt.get(i, 0)} for i in range(1, 7)]
        max_cnt = max(f['count'] for f in freq) if freq else 1
        for f in freq:
            f['pct'] = round(f['count'] / max_cnt * 100) if max_cnt else 0
        return jsonify({'freq': freq, 'total_draws': total_draws, 'total_slots': total_draws * 3})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/bet-signal')
@limiter.limit("30 per minute")
def bet_signal():
    """Tín hiệu đánh: tổng hợp WL streak + WR50 + WR today + confidence thành score 0-100."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        # WR last 50 (majority_vote) — subquery to avoid ORDER BY inside aggregate
        cur.execute("""
            SELECT COUNT(*), SUM(CASE WHEN is_win_size THEN 1 ELSE 0 END)
            FROM (
                SELECT pr.is_win_size
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE p.model_name='majority_vote'
                ORDER BY p.draw_number DESC LIMIT 50
            ) sub50
        """)
        r = cur.fetchone()
        n50, w50 = (r[0] or 0), (r[1] or 0)
        wr50 = w50 / n50 if n50 else 0.375

        # WR today (majority_vote)
        if USE_POSTGRES:
            cur.execute("""
                SELECT COUNT(*), SUM(CASE WHEN pr.is_win_size THEN 1 ELSE 0 END)
                FROM predictions p
                JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE p.model_name='majority_vote'
                  AND (p.prediction_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                      = (NOW() AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
            """)
        else:
            cur.execute("""
                SELECT COUNT(*), SUM(CASE WHEN pr.is_win_size THEN 1 ELSE 0 END)
                FROM predictions p JOIN prediction_results pr ON pr.prediction_id=p.id
                WHERE p.model_name='majority_vote' AND date(p.prediction_time)=date('now')
            """)
        r = cur.fetchone()
        n_today, w_today = (r[0] or 0), (r[1] or 0)
        wr_today = w_today / n_today if n_today else None

        # W/L streak
        cur.execute("""
            SELECT p.draw_number, pr.is_win_size
            FROM predictions p JOIN prediction_results pr ON pr.prediction_id=p.id
            WHERE p.model_name='majority_vote'
            ORDER BY p.draw_number DESC LIMIT 20
        """)
        wl_rows = cur.fetchall()

        # Latest confidence
        cur.execute("""
            SELECT confidence FROM predictions
            WHERE model_name='majority_vote'
            ORDER BY draw_number DESC LIMIT 1
        """)
        r = cur.fetchone()
        latest_conf = float(r[0]) if r else 0.5

        # Avg confidence last 100 — subquery
        cur.execute("""
            SELECT AVG(confidence) FROM (
                SELECT confidence FROM predictions
                WHERE model_name='majority_vote'
                ORDER BY draw_number DESC LIMIT 100
            ) sub_conf
        """)
        r = cur.fetchone()
        avg_conf = float(r[0]) if r and r[0] else 0.5

        conn.close()

        # W/L streak
        wl_streak = 0
        is_win_streak = None
        if wl_rows:
            is_win_streak = wl_rows[0][1]
            for _, w in wl_rows:
                if w == is_win_streak:
                    wl_streak += 1
                else:
                    break

        BASELINE = 0.375
        signals = []
        score = 50  # start neutral

        # Signal 1: WR50 vs baseline
        delta50 = wr50 - BASELINE
        if delta50 >= 0.10:
            score += 20; signals.append({'label': 'WR50', 'value': f'{wr50*100:.0f}%', 'sign': '+'})
        elif delta50 >= 0.04:
            score += 10; signals.append({'label': 'WR50', 'value': f'{wr50*100:.0f}%', 'sign': '+'})
        elif delta50 <= -0.10:
            score -= 20; signals.append({'label': 'WR50', 'value': f'{wr50*100:.0f}%', 'sign': '-'})
        elif delta50 <= -0.04:
            score -= 10; signals.append({'label': 'WR50', 'value': f'{wr50*100:.0f}%', 'sign': '-'})
        else:
            signals.append({'label': 'WR50', 'value': f'{wr50*100:.0f}%', 'sign': '='})

        # Signal 2: WR today
        if wr_today is not None and n_today >= 5:
            delta_today = wr_today - BASELINE
            if delta_today >= 0.08:
                score += 15; signals.append({'label': 'WR hôm nay', 'value': f'{wr_today*100:.0f}%', 'sign': '+'})
            elif delta_today <= -0.08:
                score -= 15; signals.append({'label': 'WR hôm nay', 'value': f'{wr_today*100:.0f}%', 'sign': '-'})
            else:
                signals.append({'label': 'WR hôm nay', 'value': f'{wr_today*100:.0f}%', 'sign': '='})

        # Signal 3: W/L streak
        if is_win_streak and wl_streak >= 4:
            score += 15; signals.append({'label': 'Streak', 'value': f'WIN ×{wl_streak}', 'sign': '+'})
        elif is_win_streak and wl_streak >= 2:
            score += 7; signals.append({'label': 'Streak', 'value': f'WIN ×{wl_streak}', 'sign': '+'})
        elif not is_win_streak and wl_streak >= 4:
            score -= 15; signals.append({'label': 'Streak', 'value': f'LOSS ×{wl_streak}', 'sign': '-'})
        elif not is_win_streak and wl_streak >= 2:
            score -= 7; signals.append({'label': 'Streak', 'value': f'LOSS ×{wl_streak}', 'sign': '-'})
        else:
            signals.append({'label': 'Streak', 'value': ('WIN' if is_win_streak else 'LOSS') + f' ×{wl_streak}', 'sign': '='})

        # Signal 4: confidence bucket (MED 50-65% is sweet spot, HIGH >65% NOT better)
        if 0.50 <= latest_conf < 0.65:
            score += 15; signals.append({'label': 'Conf', 'value': f'{latest_conf*100:.0f}% MED', 'sign': '+'})
        elif latest_conf < 0.50:
            score -= 12; signals.append({'label': 'Conf', 'value': f'{latest_conf*100:.0f}% LOW', 'sign': '-'})
        else:
            # HIGH conf (>65%): WR=38.6% historically — only slight positive vs baseline
            score += 3; signals.append({'label': 'Conf', 'value': f'{latest_conf*100:.0f}% HIGH', 'sign': '='})

        score = max(0, min(100, score))
        if score >= 75:
            label, color = 'MẠNH', 'green'
        elif score >= 55:
            label, color = 'TỐT', 'cyan'
        elif score >= 40:
            label, color = 'TRUNG BÌNH', 'gold'
        else:
            label, color = 'YẾU', 'red'

        return jsonify({
            'score': score,
            'label': label,
            'color': color,
            'signals': signals,
            'wr50': round(wr50, 4),
            'wr_today': round(wr_today, 4) if wr_today is not None else None,
            'wl_streak': wl_streak,
            'is_win_streak': is_win_streak,
            'confidence': round(latest_conf, 4),
            'avg_conf': round(avg_conf, 4),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/voter-current')
@limiter.limit("60 per minute")
def voter_current():
    """Vote breakdown của kỳ dự đoán gần nhất (majority_vote)."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT draw_number, predicted_numbers, confidence, vote_breakdown
            FROM predictions
            WHERE model_name = 'majority_vote'
            ORDER BY draw_number DESC LIMIT 1
        """)
        row = cur.fetchone()
        conn.close()
        if not row:
            return jsonify({'error': 'no data'})
        draw_n, pred_raw, conf, vb_raw = row
        try: pred = json.loads(pred_raw) if isinstance(pred_raw, str) else (pred_raw or [])
        except Exception: pred = []
        try:
            vb = json.loads(vb_raw) if isinstance(vb_raw, str) else (vb_raw or {})
        except Exception: vb = {}
        detail = vb.get('all_votes_detail', {})
        tally  = vb.get('size_tally', {})
        weights_ema = vb.get('size_weights_ema', {})
        voters = []
        for name, v in detail.items():
            voters.append({
                'name': name,
                'size': v.get('size'),
                'conf': round(float(v.get('conf', 0)), 3),
                'eff_w_pct': round(float(v.get('eff_w_pct', 0)), 1),
                'winner': v.get('winner', False),
                'streak': v.get('streak', 0),
                'decay': round(float(v.get('decay', 1)), 2),
            })
        voters.sort(key=lambda x: -x['eff_w_pct'])
        s = sum(pred)
        pred_size = 'HOA' if 10 <= s <= 11 else ('NHO' if s <= 9 else 'LON')
        return jsonify({
            'draw_number': draw_n,
            'predicted_numbers': pred,
            'predicted_size': pred_size,
            'confidence': round(float(conf), 4) if conf else None,
            'voters': voters,
            'size_tally': tally,
            'size_weights_ema': {k: round(float(v), 3) for k, v in weights_ema.items()},
            'majority_count': vb.get('majority_count'),
            'total_models': vb.get('total_models'),
            'markov_abstained': vb.get('markov_abstained', False),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/pl-running')
@limiter.limit("30 per minute")
def pl_running():
    """Cumulative P/L (±1 per kỳ) nếu theo đúng dự đoán majority_vote — 200 kỳ gần nhất."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT p.draw_number, pr.is_win_size, p.confidence
            FROM (
                SELECT id, draw_number, confidence FROM predictions
                WHERE model_name = 'majority_vote'
                ORDER BY draw_number DESC LIMIT 200
            ) p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            ORDER BY p.draw_number ASC
        """)
        rows = cur.fetchall()
        conn.close()
        cumulative = 0
        points = []
        for draw_n, is_win, conf in rows:
            cumulative += 1 if is_win else -1
            points.append({
                'draw': draw_n,
                'pl': cumulative,
                'win': is_win,
                'conf': round(float(conf), 3) if conf else None,
            })
        wins   = sum(1 for r in rows if r[1])
        losses = sum(1 for r in rows if not r[1])
        return jsonify({
            'points': points,
            'final_pl': cumulative,
            'wins': wins,
            'losses': losses,
            'n': len(rows),
            'wr': round(wins / len(rows), 4) if rows else None,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/combo-gap')
@limiter.limit("30 per minute")
def combo_gap():
    """56 bộ số (sorted, repeats allowed), xếp theo số kỳ từ lần ra cuối."""
    from itertools import combinations_with_replacement
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute(
            "SELECT draw_number, numbers FROM draw_history "
            "ORDER BY draw_number DESC LIMIT 2000"
        )
        rows = cur.fetchall()
        conn.close()
        # Build last_seen map: sorted_tuple -> (draw_number, gap_index)
        last_seen = {}  # sorted_tuple -> draw_number
        last_gap  = {}  # sorted_tuple -> position index (0=most recent draw)
        latest_draw = rows[0][0] if rows else 0
        for idx, (draw_n, nums_raw) in enumerate(rows):
            try: nums = json.loads(nums_raw) if isinstance(nums_raw, str) else (nums_raw or [])
            except Exception: nums = []
            key = tuple(sorted(nums))
            if key not in last_seen:
                last_seen[key] = draw_n
                last_gap[key]  = idx
        # All 56 combos
        all_combos = list(combinations_with_replacement(range(1, 7), 3))
        result = []
        for c in all_combos:
            gap = last_gap.get(c)
            ls  = last_seen.get(c)
            s   = sum(c)
            size = 'HOA' if 10 <= s <= 11 else ('NHO' if s <= 9 else 'LON')
            is_triple = len(set(c)) == 1
            is_pair   = len(set(c)) == 2
            result.append({
                'combo': list(c),
                'size': size,
                'sum': s,
                'is_triple': is_triple,
                'is_pair': is_pair,
                'last_draw': ls,
                'gap': gap,  # None if not seen in last 2000
            })
        result.sort(key=lambda x: (-(x['gap'] if x['gap'] is not None else 9999)))
        return jsonify({'combos': result, 'latest_draw': latest_draw, 'window': 2000})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/combo-heatmap')
@limiter.limit("30 per minute")
@cache_resp(ttl=120)
def combo_heatmap():
    """56 bộ số với tần suất + khoảng cách, cửa sổ N kỳ gần nhất."""
    from itertools import combinations_with_replacement
    try:
        n = min(int(request.args.get('n', 500)), 2000)
        conn = db.get_connection()
        cur  = conn.cursor()
        ph = '%s' if USE_POSTGRES else '?'
        cur.execute(
            f"SELECT draw_number, numbers FROM draw_history "
            f"ORDER BY draw_number DESC LIMIT {ph}", (n,)
        )
        rows = cur.fetchall()
        conn.close()
        freq_map = {}
        last_gap = {}
        latest_draw = rows[0][0] if rows else 0
        for idx, (draw_n, nums_raw) in enumerate(rows):
            try: nums = json.loads(nums_raw) if isinstance(nums_raw, str) else (nums_raw or [])
            except Exception: nums = []
            key = tuple(sorted(nums))
            freq_map[key] = freq_map.get(key, 0) + 1
            if key not in last_gap:
                last_gap[key] = idx
        all_combos = list(combinations_with_replacement(range(1, 7), 3))
        result = []
        for c in all_combos:
            freq = freq_map.get(c, 0)
            gap  = last_gap.get(c)
            s    = sum(c)
            size = 'HOA' if 10 <= s <= 11 else ('NHO' if s <= 9 else 'LON')
            result.append({
                'combo': list(c),
                'freq': freq,
                'gap': gap,
                'size': size,
                'sum': s,
            })
        max_freq = max((r['freq'] for r in result), default=1) or 1
        for r in result:
            r['heat'] = round(r['freq'] / max_freq, 3)
        return jsonify({'combos': result, 'latest_draw': latest_draw, 'n': n, 'max_freq': max_freq})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/number-gap')
@limiter.limit("60 per minute")
def number_gap():
    """Bao nhiêu kỳ từ lần cuối mỗi số 1-6 xuất hiện (kỳ gần nhất lên trước)."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute(
            "SELECT draw_number, numbers FROM draw_history "
            "ORDER BY draw_number DESC LIMIT 100"
        )
        rows = cur.fetchall()
        conn.close()
        last_seen = {}  # number -> draw_number
        last_gap  = {}  # number -> gap (draws since last seen)
        for i, (draw_n, nums_raw) in enumerate(rows):
            try: nums = json.loads(nums_raw) if isinstance(nums_raw, str) else (nums_raw or [])
            except Exception: nums = []
            for n in nums:
                if n not in last_seen:
                    last_seen[n] = draw_n
                    last_gap[n]  = i  # gap = position index (0 = appeared in latest draw)
        result = []
        latest_draw = rows[0][0] if rows else 0
        for n in range(1, 7):
            gap = last_gap.get(n)
            ls  = last_seen.get(n)
            result.append({
                'number': n,
                'last_draw': ls,
                'gap': gap,  # None if not seen in last 100
                'gap_label': f'{gap}k' if gap is not None else '>100k',
            })
        return jsonify({'numbers': result, 'latest_draw': latest_draw})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/sum-today')
@limiter.limit("60 per minute")
def sum_today():
    """Phân bố tổng (3-18) trong các kỳ hôm nay, so sánh với 100 kỳ gần nhất."""
    from collections import Counter
    def _parse(raw):
        try: return json.loads(raw) if isinstance(raw, str) else (raw or [])
        except Exception: return []
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        # Today
        if USE_POSTGRES:
            cur.execute("""
                SELECT numbers FROM draw_history
                WHERE (draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                      = (NOW() AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
            """)
        else:
            cur.execute("SELECT numbers FROM draw_history WHERE date(draw_time)=date('now')")
        today_rows = cur.fetchall()
        # Last 100
        cur.execute("SELECT numbers FROM draw_history ORDER BY draw_number DESC LIMIT 100")
        hist_rows = cur.fetchall()
        conn.close()
        def _count(rows):
            c = Counter()
            for (raw,) in rows:
                nums = _parse(raw)
                if nums: c[sum(nums)] += 1
            return c
        today_c = _count(today_rows)
        hist_c  = _count(hist_rows)
        n_today = sum(today_c.values()) or 1
        n_hist  = sum(hist_c.values())  or 1
        dist = []
        for s in range(3, 19):
            size = 'HOA' if 10 <= s <= 11 else ('NHO' if s <= 9 else 'LON')
            dist.append({
                'sum': s,
                'size': size,
                'today': today_c.get(s, 0),
                'today_pct': round(today_c.get(s, 0) / n_today * 100, 1),
                'hist': hist_c.get(s, 0),
                'hist_pct': round(hist_c.get(s, 0) / n_hist * 100, 1),
            })
        sums_today = [s for (raw,) in today_rows for nums in [_parse(raw)] if nums for s in [sum(nums)]]
        avg = round(sum(sums_today)/len(sums_today), 2) if sums_today else None
        mode_sum = today_c.most_common(1)[0][0] if today_c else None
        return jsonify({'dist': dist, 'avg': avg, 'mode': mode_sum, 'n': len(today_rows)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/combo-hot-today')
@limiter.limit("60 per minute")
def combo_hot_today():
    """Bộ số nào ra nhiều lần nhất hôm nay + bộ chưa ra hôm nay."""
    from collections import Counter
    from itertools import combinations_with_replacement
    def _size(s):
        return 'HOA' if 10 <= s <= 11 else ('NHO' if s <= 9 else 'LON')
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        if USE_POSTGRES:
            cur.execute("""
                SELECT numbers FROM draw_history
                WHERE (draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                      = (NOW() AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                ORDER BY draw_number DESC
            """)
        else:
            cur.execute("SELECT numbers FROM draw_history WHERE date(draw_time)=date('now') ORDER BY draw_number DESC")
        rows = cur.fetchall()
        conn.close()
        cnt = Counter()
        for (raw,) in rows:
            try: nums = json.loads(raw) if isinstance(raw, str) else (raw or [])
            except Exception: nums = []
            key = tuple(sorted(nums))
            if len(key) == 3:
                cnt[key] += 1
        all_combos = set(combinations_with_replacement(range(1, 7), 3))
        seen_today = set(cnt.keys())
        not_seen   = sorted(all_combos - seen_today)
        hot = [{'combo': list(k), 'count': v, 'size': _size(sum(k)),
                'is_triple': len(set(k)) == 1, 'is_pair': len(set(k)) == 2}
               for k, v in cnt.most_common(15)]
        cold_today = [{'combo': list(k), 'size': _size(sum(k)),
                       'is_triple': len(set(k)) == 1, 'is_pair': len(set(k)) == 2}
                      for k in not_seen[:20]]
        return jsonify({
            'hot': hot,
            'not_seen_today': cold_today,
            'total_draws': len(rows),
            'unique_combos_today': len(seen_today),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/wr-by-size')
@limiter.limit("30 per minute")
def wr_by_size():
    """WR lịch sử khi model dự đoán mỗi SIZE (NHO/HOA/LON) — 500 kỳ gần nhất."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT
                CASE WHEN arr_sum BETWEEN 10 AND 11 THEN 'HOA'
                     WHEN arr_sum <= 9 THEN 'NHO'
                     ELSE 'LON' END AS pred_size,
                COUNT(*) AS n,
                SUM(CASE WHEN pr.is_win_size THEN 1 ELSE 0 END) AS wins
            FROM (
                SELECT p.id,
                       (SELECT SUM(v::int)
                        FROM json_array_elements_text(p.predicted_numbers::json) v) AS arr_sum
                FROM predictions p
                WHERE p.model_name = 'majority_vote'
                ORDER BY p.draw_number DESC LIMIT 500
            ) sub
            JOIN prediction_results pr ON pr.prediction_id = sub.id
            GROUP BY pred_size
        """ if USE_POSTGRES else """
            SELECT pred_size, COUNT(*) n, SUM(is_win) wins
            FROM (
                SELECT p.predicted_numbers, pr.is_win_size AS is_win,
                       CASE WHEN json_extract(p.predicted_numbers,'$[0]')+
                                 json_extract(p.predicted_numbers,'$[1]')+
                                 json_extract(p.predicted_numbers,'$[2]') BETWEEN 10 AND 11 THEN 'HOA'
                            WHEN json_extract(p.predicted_numbers,'$[0]')+
                                 json_extract(p.predicted_numbers,'$[1]')+
                                 json_extract(p.predicted_numbers,'$[2]') <= 9 THEN 'NHO'
                            ELSE 'LON' END AS pred_size
                FROM predictions p JOIN prediction_results pr ON pr.prediction_id=p.id
                WHERE p.model_name='majority_vote'
                ORDER BY p.draw_number DESC LIMIT 500
            ) sub GROUP BY pred_size
        """)
        rows = cur.fetchall()
        conn.close()
        result = {}
        for sz, n, wins in rows:
            result[sz] = {'n': n, 'wins': wins, 'wr': round(wins / n, 4) if n else 0}
        return jsonify({'sizes': result, 'window': 500})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/hourly-today')
@limiter.limit("30 per minute")
def hourly_today():
    """WR theo giờ hôm nay (VN time) — số kỳ và win rate mỗi giờ."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        if USE_POSTGRES:
            cur.execute("""
                SELECT EXTRACT(HOUR FROM d.draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh') AS hr,
                       COUNT(*) AS total,
                       SUM(CASE WHEN pr.is_win_size THEN 1 ELSE 0 END) AS wins
                FROM draw_history d
                LEFT JOIN predictions p
                  ON p.draw_number = d.draw_number AND p.model_name = 'majority_vote'
                LEFT JOIN prediction_results pr ON pr.prediction_id = p.id
                WHERE (d.draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                      = (NOW() AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                GROUP BY hr ORDER BY hr
            """)
        else:
            cur.execute("""
                SELECT strftime('%H', d.draw_time) AS hr, COUNT(*) AS total,
                       SUM(pr.is_win_size) AS wins
                FROM draw_history d
                LEFT JOIN predictions p ON p.draw_number=d.draw_number AND p.model_name='majority_vote'
                LEFT JOIN prediction_results pr ON pr.prediction_id=p.id
                WHERE date(d.draw_time)=date('now')
                GROUP BY hr ORDER BY hr
            """)
        rows = cur.fetchall()
        conn.close()
        hours = []
        for hr, total, wins in rows:
            wins = wins or 0
            wr = round(wins / total, 3) if total else None
            hours.append({'hour': int(hr), 'total': total, 'wins': wins, 'wr': wr})
        return jsonify({'hours': hours})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/confidence-buckets')
@limiter.limit("30 per minute")
def confidence_buckets():
    """WR theo mức confidence: LOW(<50%) / MED(50-65%) / HIGH(>65%) — 1000 kỳ gần nhất."""
    BUCKETS = [
        ('LOW',  0.00, 0.50),
        ('MED',  0.50, 0.65),
        ('HIGH', 0.65, 1.01),
    ]
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT p.confidence, pr.is_win_size
            FROM (
                SELECT id, confidence FROM predictions
                WHERE model_name = 'majority_vote'
                ORDER BY draw_number DESC LIMIT 1000
            ) p
            JOIN prediction_results pr ON pr.prediction_id = p.id
        """)
        rows = cur.fetchall()
        conn.close()
        result = []
        for label, lo, hi in BUCKETS:
            subset = [(conf, win) for conf, win in rows if lo <= float(conf or 0) < hi]
            n = len(subset)
            wins = sum(1 for _, w in subset if w)
            wr = round(wins / n, 4) if n else None
            avg_conf = round(sum(float(c) for c, _ in subset) / n, 4) if n else None
            result.append({
                'label': label, 'lo': lo, 'hi': hi,
                'n': n, 'wins': wins, 'wr': wr, 'avg_conf': avg_conf,
            })
        return jsonify({'buckets': result, 'window': 1000})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/pair-today')
@limiter.limit("60 per minute")
def pair_today():
    """Cặp số xuất hiện cùng nhau nhiều nhất hôm nay (VN time)."""
    from collections import Counter
    from itertools import combinations
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        if USE_POSTGRES:
            cur.execute("""
                SELECT numbers FROM draw_history
                WHERE (draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                      = (NOW() AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                ORDER BY draw_number DESC
            """)
        else:
            cur.execute("SELECT numbers FROM draw_history WHERE date(draw_time)=date('now') ORDER BY draw_number DESC")
        rows = cur.fetchall()
        conn.close()
        pair_cnt = Counter()
        single_cnt = Counter()
        total = 0
        for (raw,) in rows:
            try: nums = json.loads(raw) if isinstance(raw, str) else (raw or [])
            except Exception: nums = []
            if len(nums) != 3: continue
            total += 1
            for n in nums:
                single_cnt[n] += 1
            for a, b in combinations(sorted(set(nums)), 2):
                pair_cnt[(a, b)] += 1
        pairs = [{'pair': list(k), 'count': v, 'pct': round(v / total * 100, 1) if total else 0}
                 for k, v in pair_cnt.most_common(15)]
        singles = [{'number': k, 'count': v} for k, v in sorted(single_cnt.items())]
        return jsonify({'pairs': pairs, 'singles': singles, 'total_draws': total})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/streak-history')
@limiter.limit("30 per minute")
@cache_resp(ttl=60)
def streak_history():
    """Phân bố độ dài streak W/L từ lịch sử majority_vote predictions."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT pr.is_win_size
            FROM (
                SELECT id FROM predictions
                WHERE model_name = 'majority_vote'
                ORDER BY draw_number DESC LIMIT 2000
            ) p
            JOIN prediction_results pr ON pr.prediction_id = p.id
        """)
        rows = [r[0] for r in cur.fetchall()]
        conn.close()
        if not rows:
            return jsonify({'win_dist': {}, 'loss_dist': {}, 'avg_win_streak': 0, 'avg_loss_streak': 0})
        # Build streak list
        win_streaks, loss_streaks = [], []
        cur_val, cur_len = rows[0], 1
        for val in rows[1:]:
            if val == cur_val:
                cur_len += 1
            else:
                (win_streaks if cur_val else loss_streaks).append(cur_len)
                cur_val, cur_len = val, 1
        (win_streaks if cur_val else loss_streaks).append(cur_len)
        from collections import Counter
        win_dist  = dict(sorted(Counter(win_streaks).items()))
        loss_dist = dict(sorted(Counter(loss_streaks).items()))
        avg_w = round(sum(win_streaks)  / len(win_streaks),  2) if win_streaks  else 0
        avg_l = round(sum(loss_streaks) / len(loss_streaks), 2) if loss_streaks else 0
        max_w = max(win_streaks)  if win_streaks  else 0
        max_l = max(loss_streaks) if loss_streaks else 0
        return jsonify({
            'win_dist': win_dist, 'loss_dist': loss_dist,
            'avg_win_streak': avg_w, 'avg_loss_streak': avg_l,
            'max_win_streak': max_w, 'max_loss_streak': max_l,
            'total_win_streaks': len(win_streaks), 'total_loss_streaks': len(loss_streaks),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/draw-heatmap')
@limiter.limit("30 per minute")
def draw_heatmap():
    """40 kỳ gần nhất × 6 số: mỗi ô = 1 nếu số đó xuất hiện trong kỳ đó."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute(
            "SELECT draw_number, numbers FROM draw_history "
            "ORDER BY draw_number DESC LIMIT 40"
        )
        rows = cur.fetchall()
        conn.close()
        draws = []
        for draw_n, nums_raw in rows:
            try: nums = json.loads(nums_raw) if isinstance(nums_raw, str) else list(nums_raw or [])
            except Exception: nums = []
            nums_set = set(nums)
            draws.append({
                'draw': draw_n,
                'numbers': sorted(nums),
                'grid': [1 if i in nums_set else 0 for i in range(1, 7)],
            })
        return jsonify({'draws': draws, 'n': len(draws)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/size-sequence')
@limiter.limit("30 per minute")
def size_sequence():
    """60 kỳ gần nhất: SIZE thực tế + W/L dự đoán majority_vote."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT dh.draw_number,
                   dh.sum_value,
                   pr.is_win_size,
                   p.predicted_numbers
            FROM (
                SELECT draw_number, sum_value, numbers
                FROM draw_history
                ORDER BY draw_number DESC LIMIT 60
            ) dh
            LEFT JOIN predictions p
                   ON p.draw_number = dh.draw_number
                  AND p.model_name = 'majority_vote'
            LEFT JOIN prediction_results pr
                   ON pr.prediction_id = p.id
            ORDER BY dh.draw_number ASC
        """)
        rows = cur.fetchall()
        conn.close()
        result = []
        for draw_n, s, is_win, pred_raw in rows:
            s = s or 0
            size = 'HOA' if 10 <= s <= 11 else ('NHO' if s <= 9 else 'LON')
            try: pred = json.loads(pred_raw) if isinstance(pred_raw, str) else (pred_raw or [])
            except Exception: pred = []
            ps = sum(pred) if pred else None
            pred_size = ('HOA' if 10 <= ps <= 11 else ('NHO' if ps <= 9 else 'LON')) if ps else None
            result.append({
                'draw': draw_n,
                'sum': s,
                'size': size,
                'pred_size': pred_size,
                'win': is_win,
            })
        return jsonify({'sequence': result, 'n': len(result)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/sum-rolling')
@limiter.limit("30 per minute")
def sum_rolling():
    """100 kỳ gần nhất: sum_value + 10-draw moving average."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute(
            "SELECT draw_number, sum_value FROM draw_history "
            "ORDER BY draw_number DESC LIMIT 100"
        )
        rows = list(reversed(cur.fetchall()))
        conn.close()
        points = []
        win_size = 10
        for i, (draw_n, s) in enumerate(rows):
            window = [r[1] for r in rows[max(0, i - win_size + 1): i + 1] if r[1] is not None]
            ma = round(sum(window) / len(window), 2) if window else None
            points.append({'draw': draw_n, 'sum': s, 'ma': ma})
        sums = [p['sum'] for p in points if p['sum'] is not None]
        return jsonify({
            'points': points,
            'n': len(points),
            'avg': round(sum(sums) / len(sums), 2) if sums else None,
            'min': min(sums) if sums else None,
            'max': max(sums) if sums else None,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/number-deviation')
@limiter.limit("30 per minute")
@cache_resp(ttl=120)
def number_deviation():
    """Tần suất thực tế vs lý thuyết của mỗi số 1-6 trong N kỳ gần nhất."""
    n = min(int(request.args.get('n', 300)), 1000)
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute(
            "SELECT numbers FROM draw_history "
            "ORDER BY draw_number DESC LIMIT %s", (n,)
        )
        rows = cur.fetchall()
        conn.close()
        counts = {i: 0 for i in range(1, 7)}
        total_appearances = 0
        for (nums_raw,) in rows:
            try: nums = json.loads(nums_raw) if isinstance(nums_raw, str) else list(nums_raw or [])
            except Exception: nums = []
            for x in nums:
                if 1 <= x <= 6:
                    counts[x] += 1
                    total_appearances += 1
        # Expected per number = total_appearances / 6 (uniform distribution)
        expected = total_appearances / 6 if total_appearances else 0
        result = []
        for num in range(1, 7):
            actual = counts[num]
            delta  = actual - expected
            pct    = round(actual / total_appearances * 100, 1) if total_appearances else 0
            result.append({
                'number': num,
                'count': actual,
                'expected': round(expected, 1),
                'delta': round(delta, 1),
                'pct': pct,
                'status': 'due' if delta < -5 else ('hot' if delta > 5 else 'normal'),
            })
        return jsonify({
            'numbers': result,
            'draws': len(rows),
            'total_appearances': total_appearances,
            'expected_per_num': round(expected, 1),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/wl-calendar')
@limiter.limit("20 per minute")
@cache_resp(ttl=300)
def wl_calendar():
    """WR per ngày (VN time) trong 30 ngày gần nhất."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT
                (dh.draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date AS vn_date,
                COUNT(*) AS total,
                SUM(CASE WHEN pr.is_win_size THEN 1 ELSE 0 END) AS wins
            FROM (
                SELECT id, draw_number, draw_time FROM draw_history
                WHERE draw_time >= NOW() - INTERVAL '30 days'
            ) dh
            JOIN predictions p
                ON p.draw_number = dh.draw_number
               AND p.model_name = 'majority_vote'
            JOIN prediction_results pr ON pr.prediction_id = p.id
            GROUP BY vn_date
            ORDER BY vn_date ASC
        """)
        rows = cur.fetchall()
        conn.close()
        days = []
        for vn_date, total, wins in rows:
            wr = round(wins / total, 4) if total else None
            days.append({
                'date': str(vn_date),
                'total': int(total),
                'wins': int(wins),
                'wr': wr,
            })
        return jsonify({'days': days, 'n': len(days)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/momentum')
@limiter.limit("60 per minute")
def momentum():
    """Weighted momentum score dựa trên 20 kỳ gần nhất (kỳ mới hơn = trọng số cao hơn)."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT pr.is_win_size, p.confidence
            FROM (
                SELECT id, confidence FROM predictions
                WHERE model_name = 'majority_vote'
                ORDER BY draw_number DESC LIMIT 20
            ) p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            ORDER BY p.id DESC
        """)
        rows = cur.fetchall()
        conn.close()
        if not rows:
            return jsonify({'error': 'no data'})
        # Weighted score: most recent = highest weight
        n = len(rows)
        score = 0.0
        total_w = 0.0
        sequence = []
        for i, (is_win, conf) in enumerate(rows):
            w = (n - i)  # weight: n for most recent, 1 for oldest
            score += w * (1 if is_win else -1)
            total_w += w
            sequence.append({'win': is_win, 'conf': round(float(conf), 3) if conf else None})
        norm_score = round(score / total_w * 100, 1)  # -100 to +100
        # Current streak
        streak_val = rows[0][0]
        streak_len = 0
        for is_win, _ in rows:
            if is_win == streak_val:
                streak_len += 1
            else:
                break
        wins = sum(1 for r in rows if r[0])
        return jsonify({
            'score': norm_score,
            'direction': 'up' if norm_score > 10 else ('down' if norm_score < -10 else 'neutral'),
            'streak_win': streak_val,
            'streak_len': streak_len,
            'sequence': sequence,
            'wins': wins,
            'losses': n - wins,
            'n': n,
            'wr': round(wins / n, 4) if n else None,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/pred-window-wr')
@limiter.limit("20 per minute")
def pred_window_wr():
    """WR của majority_vote theo nhiều cửa sổ: 20/50/100/200/500/1000 kỳ gần nhất."""
    windows = [20, 50, 100, 200, 500, 1000]
    max_n   = max(windows)
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT pr.is_win_size
            FROM (
                SELECT id FROM predictions
                WHERE model_name = 'majority_vote'
                ORDER BY draw_number DESC LIMIT %s
            ) p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            ORDER BY p.id DESC
        """, (max_n,))
        rows = [r[0] for r in cur.fetchall()]
        conn.close()
        result = []
        for w in windows:
            subset = rows[:w]
            if not subset:
                continue
            wins = sum(1 for x in subset if x)
            result.append({
                'window': w,
                'wins': wins,
                'total': len(subset),
                'wr': round(wins / len(subset), 4),
            })
        return jsonify({'windows': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/number-pair-matrix')
@limiter.limit("20 per minute")
@cache_resp(ttl=120)
def number_pair_matrix():
    """6×6 co-occurrence matrix: số lần số i và số j cùng xuất hiện trong 1 kỳ (N kỳ gần nhất)."""
    n = min(int(request.args.get('n', 500)), 2000)
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute(
            "SELECT numbers FROM draw_history ORDER BY draw_number DESC LIMIT %s", (n,)
        )
        rows = cur.fetchall()
        conn.close()
        matrix = [[0] * 6 for _ in range(6)]
        total_draws = 0
        for (nums_raw,) in rows:
            try: nums = json.loads(nums_raw) if isinstance(nums_raw, str) else list(nums_raw or [])
            except Exception: nums = []
            nums_valid = [x for x in nums if 1 <= x <= 6]
            total_draws += 1
            seen = set()
            for x in nums_valid:
                for y in nums_valid:
                    if x != y:
                        matrix[x - 1][y - 1] += 1
                seen.add(x)
            # Self (diagonal): count appearances of each number
            for x in nums_valid:
                matrix[x - 1][x - 1] += 1
        # Normalize to percentage of draws
        result = []
        for i in range(6):
            row = []
            for j in range(6):
                count = matrix[i][j]
                pct = round(count / total_draws * 100, 1) if total_draws else 0
                row.append({'count': count, 'pct': pct})
            result.append(row)
        return jsonify({'matrix': result, 'draws': total_draws, 'n': n})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/size-by-hour')
@limiter.limit("20 per minute")
@cache_resp(ttl=300)
def size_by_hour():
    """Phân bố SIZE (NHO/HOA/LON) theo từng giờ trong ngày — lịch sử toàn bộ."""
    try:
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT
                EXTRACT(HOUR FROM draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::int AS hour,
                SUM(CASE WHEN sum_value BETWEEN 3 AND 9  THEN 1 ELSE 0 END) AS nho,
                SUM(CASE WHEN sum_value BETWEEN 10 AND 11 THEN 1 ELSE 0 END) AS hoa,
                SUM(CASE WHEN sum_value BETWEEN 12 AND 18 THEN 1 ELSE 0 END) AS lon,
                COUNT(*) AS total
            FROM draw_history
            WHERE draw_time IS NOT NULL
            GROUP BY hour
            ORDER BY hour
        """)
        rows = cur.fetchall()
        conn.close()
        hours = []
        for hour, nho, hoa, lon, total in rows:
            total = total or 1
            hours.append({
                'hour': int(hour),
                'nho': int(nho), 'nho_pct': round(nho / total * 100, 1),
                'hoa': int(hoa), 'hoa_pct': round(hoa / total * 100, 1),
                'lon': int(lon), 'lon_pct': round(lon / total * 100, 1),
                'total': int(total),
            })
        return jsonify({'hours': hours})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/size-transition')
@limiter.limit("20 per minute")
@cache_resp(ttl=120)
def size_transition():
    """3×3 transition matrix: P(size_next | size_curr) trên N kỳ gần nhất."""
    try:
        n = min(int(request.args.get('n', 1000)), 5000)
    except ValueError:
        n = 1000
    try:
        conn = db.get_connection()
        cur = conn.cursor()
        cur.execute(
            "SELECT size_category FROM draw_history ORDER BY draw_number DESC LIMIT %s" % (n + 1,)
        )
        rows = [r[0] for r in cur.fetchall() if r[0]]
        conn.close()

        SIZES = ['NHO', 'HOA', 'LON']
        counts = {s: {t: 0 for t in SIZES} for s in SIZES}
        totals = {s: 0 for s in SIZES}

        # rows is DESC, so rows[i+1] is older, rows[i] is newer
        for i in range(len(rows) - 1):
            curr = rows[i + 1]  # current draw (older)
            nxt  = rows[i]      # next draw (newer)
            if curr in SIZES and nxt in SIZES:
                counts[curr][nxt] += 1
                totals[curr] += 1

        matrix = {}
        for s in SIZES:
            total = totals[s] or 1
            matrix[s] = {
                t: {'count': counts[s][t], 'pct': round(counts[s][t] / total * 100, 1)}
                for t in SIZES
            }

        return jsonify({'matrix': matrix, 'totals': totals, 'n': len(rows)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/smart-summary')
@limiter.limit("30 per minute")
def smart_summary():
    """Banner tóm tắt: WR20, streak, hot SIZE 10 kỳ, coldest combo."""
    from itertools import combinations_with_replacement
    from collections import Counter
    try:
        conn = db.get_connection()
        cur = conn.cursor()

        # WR last 20 + W/L streak (majority_vote)
        cur.execute("""
            SELECT pr.is_win_size
            FROM predictions p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            WHERE p.model_name = 'majority_vote'
            ORDER BY p.draw_number DESC LIMIT 20
        """)
        wl_rows = [r[0] for r in cur.fetchall()]
        wr20 = round(sum(1 for w in wl_rows if w) / len(wl_rows), 4) if wl_rows else None
        streak_len = 0
        if wl_rows:
            last_w = wl_rows[0]
            for w in wl_rows:
                if w == last_w:
                    streak_len += 1
                else:
                    break

        # Hot SIZE last 10 draws
        cur.execute("SELECT size_category FROM draw_history ORDER BY draw_number DESC LIMIT 10")
        sizes = [r[0] for r in cur.fetchall() if r[0]]
        size_ctr = Counter(sizes)
        hot_size = size_ctr.most_common(1)[0][0] if size_ctr else None

        # Coldest combo from last 2000 draws
        cur.execute("SELECT draw_number, numbers FROM draw_history ORDER BY draw_number DESC LIMIT 2000")
        rows = cur.fetchall()
        conn.close()

        last_seen = {}
        latest_draw = rows[0][0] if rows else 0
        for idx, (draw_n, nums_raw) in enumerate(rows):
            try:
                nums = json.loads(nums_raw) if isinstance(nums_raw, str) else (nums_raw or [])
            except Exception:
                nums = []
            key = tuple(sorted(int(n) for n in nums))
            if key not in last_seen:
                last_seen[key] = (draw_n, idx)

        all_combos = list(combinations_with_replacement(range(1, 7), 3))
        coldest = max(all_combos, key=lambda c: last_seen.get(c, (0, 9999))[1])
        coldest_gap = last_seen.get(coldest, (None, None))[1]
        coldest_draw = last_seen.get(coldest, (None, None))[0]

        status = 'hot' if (wr20 or 0) >= 0.50 else ('cold' if (wr20 or 0.375) < 0.30 else 'warm')

        return jsonify({
            'wr20': wr20,
            'streak': {'type': 'WIN' if wl_rows[0] else 'LOSS', 'length': streak_len} if wl_rows else None,
            'hot_size': hot_size,
            'hot_size_count': size_ctr.get(hot_size, 0) if hot_size else 0,
            'coldest_combo': {
                'combo': list(coldest),
                'gap': coldest_gap,
                'last_draw': coldest_draw,
            } if coldest_gap is not None else None,
            'status': status,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/combo-detail')
@limiter.limit("30 per minute")
def combo_detail():
    """Chi tiết thống kê combo cụ thể: tần suất, gap, hot/cold — dựa toàn bộ lịch sử."""
    import re as _re
    nums_str = request.args.get('nums', '').strip()
    raw = [int(c) for c in _re.sub(r'[^\d]', '', nums_str) if c.isdigit()]
    if len(raw) != 3 or any(n < 1 or n > 6 for n in raw):
        return jsonify({'error': 'Cần đúng 3 số từ 1 đến 6'}), 400

    combo_key = tuple(sorted(raw))
    s = sum(combo_key)
    size = 'HOA' if 10 <= s <= 11 else ('NHO' if s <= 9 else 'LON')

    try:
        conn = db.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT draw_number, numbers FROM draw_history ORDER BY draw_number")
        rows = cur.fetchall()
        conn.close()

        total_draws = len(rows)
        occurrences = []
        for draw_n, nums_raw in rows:
            try:
                nums = json.loads(nums_raw) if isinstance(nums_raw, str) else (nums_raw or [])
            except Exception:
                nums = []
            if tuple(sorted(int(n) for n in nums)) == combo_key:
                occurrences.append(draw_n)

        n_occ = len(occurrences)
        latest_draw = rows[-1][0] if rows else 0

        gaps = []
        for i in range(1, len(occurrences)):
            gaps.append(occurrences[i] - occurrences[i - 1] - 1)

        last_seen_draw = occurrences[-1] if occurrences else None
        gap_since_last = (latest_draw - last_seen_draw) if last_seen_draw is not None else None
        avg_gap = round(sum(gaps) / len(gaps), 2) if gaps else None
        expected_gap = round(total_draws / n_occ, 1) if n_occ > 0 else None

        is_hot = (gap_since_last is not None and expected_gap is not None
                  and gap_since_last < expected_gap * 0.5)
        is_cold = (last_seen_draw is None
                   or (expected_gap is not None and gap_since_last is not None
                       and gap_since_last > expected_gap * 1.5))

        return jsonify({
            'sorted_combo': list(combo_key),
            'size': size,
            'sum': s,
            'total_draws': total_draws,
            'occurrences': n_occ,
            'frequency_pct': round(n_occ / total_draws * 100, 3) if total_draws else 0,
            'last_seen': last_seen_draw,
            'gap_since_last': gap_since_last,
            'avg_gap': avg_gap,
            'min_gap': min(gaps) if gaps else None,
            'max_gap': max(gaps) if gaps else None,
            'expected_gap': expected_gap,
            'is_hot': is_hot,
            'is_cold': is_cold,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/weight-optimizer')
@limiter.limit("5 per minute")
def weight_optimizer():
    """Grid search over voter effective multipliers to find SIZE-optimal weights.
    Uses stored vote_breakdown.all_votes_detail (conf + size per voter) vs actual SIZE.
    Returns: per-voter standalone accuracy, optimal multiplier combo, WR improvement.
    """
    if not USE_POSTGRES:
        return jsonify({'error': 'PostgreSQL required'}), 400
    try:
        import itertools
        import numpy as np

        n = min(int(request.args.get('n', 500)), 2000)
        conn = db.get_connection()
        cur  = db.get_connection().cursor()
        conn.close()
        conn = db.get_connection()
        cur  = conn.cursor()
        cur.execute("""
            SELECT p.vote_breakdown,
                CASE
                    WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 9  THEN 'NHO'
                    WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 11 THEN 'HOA'
                    ELSE 'LON'
                END AS actual_size
            FROM predictions p
            JOIN prediction_results pr ON pr.prediction_id = p.id
            WHERE p.model_name = 'majority_vote'
              AND p.vote_breakdown IS NOT NULL
              AND pr.actual_numbers IS NOT NULL
            ORDER BY p.draw_number DESC LIMIT %s
        """, (n,))
        rows = cur.fetchall()
        conn.close()

        if len(rows) < 30:
            return jsonify({'error': 'insufficient data', 'n': len(rows)})

        # Parse rows
        data = []
        current_wins = 0
        for vb_raw, actual_size in rows:
            try:
                vb = json.loads(vb_raw) if isinstance(vb_raw, str) else (vb_raw or {})
                detail = vb.get('all_votes_detail') or {}
                final_size = vb.get('final_size') or vb.get('majority_size')
                voter_preds = {}
                for vname, d in detail.items():
                    if not isinstance(d, dict):
                        continue
                    sz = d.get('size', '')
                    if sz not in ('NHO', 'HOA', 'LON'):
                        continue
                    voter_preds[vname] = {'conf': float(d.get('conf', 0.4)), 'size': sz}
                if not voter_preds:
                    continue
                data.append({'v': voter_preds, 'actual': actual_size, 'final': final_size})
                if final_size == actual_size:
                    current_wins += 1
            except Exception:
                continue

        if len(data) < 30:
            return jsonify({'error': 'insufficient parsed data', 'raw': len(rows)})

        current_wr = round(current_wins / len(data) * 100, 2)

        # Collect all voter names seen
        voter_names_set: set = set()
        for row in data:
            voter_names_set.update(row['v'].keys())
        voter_names = sorted(voter_names_set)
        n_voters = len(voter_names)
        n_rows   = len(data)
        v_idx    = {v: i for i, v in enumerate(voter_names)}

        # Per-voter standalone SIZE accuracy
        voter_acc: dict = {v: {'correct': 0, 'total': 0} for v in voter_names}
        for row in data:
            for vname, vd in row['v'].items():
                voter_acc[vname]['total'] += 1
                if vd['size'] == row['actual']:
                    voter_acc[vname]['correct'] += 1

        baseline = 0.375
        voter_stats_out = {}
        for vname in voter_names:
            a = voter_acc[vname]
            t = a['total']
            wr = a['correct'] / t if t > 0 else baseline
            voter_stats_out[vname] = {
                'accuracy': round(wr * 100, 1),
                'edge_pp': round((wr - baseline) * 100, 1),
                'total': t,
            }

        # Build numpy arrays for vectorized scoring
        # nho_contrib[v, j] = conf of voter v on draw j if they voted NHO, else 0
        nho_c = np.zeros((n_voters, n_rows), dtype=np.float32)
        hoa_c = np.zeros((n_voters, n_rows), dtype=np.float32)
        lon_c = np.zeros((n_voters, n_rows), dtype=np.float32)
        actual_idx_arr = np.zeros(n_rows, dtype=np.int8)  # 0=NHO 1=HOA 2=LON

        for j, row in enumerate(data):
            actual_idx_arr[j] = 0 if row['actual'] == 'NHO' else (1 if row['actual'] == 'HOA' else 2)
            for vname, vd in row['v'].items():
                i = v_idx[vname]
                c = vd['conf']
                if vd['size'] == 'NHO': nho_c[i, j] = c
                elif vd['size'] == 'HOA': hoa_c[i, j] = c
                else: lon_c[i, j] = c

        # Grid search over multiplier levels per voter
        weight_levels = np.array([0.0, 0.5, 1.0, 1.5, 2.0, 2.5], dtype=np.float32)
        n_levels = len(weight_levels)

        best_wins = -1
        best_combo = None

        for combo in itertools.product(range(n_levels), repeat=n_voters):
            w = weight_levels[list(combo)]
            if w.sum() == 0:
                continue
            # Weighted scores per draw
            nho_s = w @ nho_c  # (n_rows,)
            hoa_s = w @ hoa_c
            lon_s = w @ lon_c
            # Predict: argmax with HOA block
            stack   = np.stack([nho_s, hoa_s, lon_s])  # (3, n_rows)
            pred_i  = np.argmax(stack, axis=0).astype(np.int8)
            hoa_mask = (pred_i == 1)
            if hoa_mask.any():
                pred_i = np.where(hoa_mask,
                                  np.where(lon_s >= nho_s, np.int8(2), np.int8(0)),
                                  pred_i)
            wins = int(np.sum(pred_i == actual_idx_arr))
            if wins > best_wins:
                best_wins = wins
                best_combo = combo

        best_wr = round(best_wins / n_rows * 100, 2)
        optimal_weights = {voter_names[i]: float(weight_levels[best_combo[i]])
                           for i in range(n_voters)}
        improvement = round(best_wr - current_wr, 2)

        return jsonify({
            'n': n_rows,
            'current_wr': current_wr,
            'best_wr': best_wr,
            'improvement_pp': improvement,
            'optimal_weights': optimal_weights,
            'voter_accuracy': voter_stats_out,
            'baseline': round(baseline * 100, 1),
            'note': f'Grid {n_levels}^{n_voters}={n_levels**n_voters} combos, {n_rows} draws',
        })
    except Exception as e:
        logger.exception("weight_optimizer error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/today-combos')
@limiter.limit("30 per minute")
def today_combos():
    """P55: Combos (3-number sets) appeared today vs not-appeared yet (VN time).

    Returns all 56 possible unordered multisets of {1..6} choose 3.
    'appeared'     = list of combos seen today, sorted by count desc.
    'not_appeared' = list of combos not yet seen today, sorted by sum asc.
    """
    try:
        import itertools
        from collections import defaultdict
        from datetime import datetime, timezone, timedelta

        conn = db.get_connection()
        cur  = conn.cursor()

        if USE_POSTGRES:
            cur.execute("""
                SELECT draw_number, numbers, draw_time
                FROM draw_history
                WHERE (draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                      = (NOW() AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                ORDER BY draw_number ASC
            """)
        else:
            cur.execute("""
                SELECT draw_number, numbers, draw_time
                FROM draw_history
                WHERE date(datetime(draw_time, '+7 hours')) = date('now', '+7 hours')
                ORDER BY draw_number ASC
            """)

        rows = cur.fetchall()
        conn.close()

        combo_draws: dict = defaultdict(list)  # (a,b,c) -> [draw_number, ...]
        for draw_number, numbers_raw, draw_time in rows:
            try:
                nums = json.loads(numbers_raw) if isinstance(numbers_raw, str) else numbers_raw
                key  = tuple(sorted(int(x) for x in nums))
                combo_draws[key].append(int(draw_number))
            except Exception:
                continue

        all_combos = list(itertools.combinations_with_replacement(range(1, 7), 3))

        appeared     = []
        not_appeared = []
        for combo in all_combos:
            s    = sum(combo)
            size = 'NHO' if s <= 9 else ('HOA' if s <= 11 else 'LON')
            entry = {'combo': list(combo), 'label': f"{combo[0]}-{combo[1]}-{combo[2]}",
                     'sum': s, 'size': size}
            if combo in combo_draws:
                entry['count'] = len(combo_draws[combo])
                entry['draws'] = combo_draws[combo]
                appeared.append(entry)
            else:
                not_appeared.append(entry)

        appeared.sort(key=lambda x: -x['count'])
        not_appeared.sort(key=lambda x: x['sum'])

        vn_now = datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=7)))
        total_draws = sum(len(v) for v in combo_draws.values())

        return jsonify({
            'date':               vn_now.strftime('%Y-%m-%d'),
            'vn_time':            vn_now.strftime('%H:%M'),
            'total_draws_today':  total_draws,
            'total_unique_combos': 56,
            'appeared_count':     len(appeared),
            'not_appeared_count': len(not_appeared),
            'coverage_pct':       round(len(appeared) / 56 * 100, 1),
            'appeared':           appeared,
            'not_appeared':       not_appeared,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ─── P56: daily-summary-evening ──────────────────────────────────────────────
@app.route('/api/daily-summary-evening', methods=['GET', 'POST'])
@limiter.limit("10 per minute")
def daily_summary_evening():
    """
    P56: Báo cáo sơ kết 22h — gửi Telegram.
    Cloud Scheduler gọi lúc 22:05 giờ VN (15:05 UTC).
    """
    import itertools
    from collections import defaultdict, Counter as _Counter2
    from datetime import datetime, timedelta, timezone
    from telegram_bot import TelegramBot

    secret = request.args.get('secret') or request.headers.get('X-Trigger-Secret', '')
    if secret and secret != config.TRIGGER_SECRET:
        return jsonify({'error': 'unauthorized'}), 403

    try:
        from database import USE_POSTGRES as _USE_PG
        vn_tz  = timezone(timedelta(hours=7))
        now_vn = datetime.now(vn_tz)
        date_str = now_vn.strftime('%d/%m/%Y')

        conn = db.get_connection()
        cur  = conn.cursor()

        _TODAY_PRED = (
            "(pr.created_at AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date "
            "= (NOW() AT TIME ZONE 'Asia/Ho_Chi_Minh')::date"
        ) if _USE_PG else "date(created_at, '+7 hours') = date('now', '+7 hours')"

        # ── 1. Tổng predictions hôm nay ──────────────────────────
        cur.execute(f"""
            SELECT COUNT(*),
                   SUM(CASE WHEN COALESCE(pr.is_win_size, pr.is_win, FALSE) THEN 1 ELSE 0 END),
                   SUM(CASE WHEN COALESCE(pr.is_win_sum,  FALSE) THEN 1 ELSE 0 END)
            FROM prediction_results pr
            JOIN predictions p ON p.id = pr.prediction_id
            WHERE {_TODAY_PRED}
        """)
        row = cur.fetchone()
        total    = row[0] or 0
        wins     = row[1] or 0
        sum_wins = row[2] or 0
        losses   = total - wins
        win_rate     = wins / total if total > 0 else 0
        sum_win_rate = sum_wins / total if total > 0 else 0

        # ── 2. SIZE breakdown ─────────────────────────────────────
        size_rows        = {}
        size_actual_rows = {}
        if _USE_PG:
            cur.execute(f"""
                SELECT
                    CASE WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 9  THEN 'NHO'
                         WHEN (SELECT SUM(v::int) FROM json_array_elements_text(p.predicted_numbers::json) v) <= 11 THEN 'HOA'
                         ELSE 'LON' END AS pred_size,
                    COUNT(*),
                    SUM(CASE WHEN COALESCE(pr.is_win_size, pr.is_win, FALSE) THEN 1 ELSE 0 END)
                FROM prediction_results pr
                JOIN predictions p ON p.id = pr.prediction_id
                WHERE {_TODAY_PRED}
                GROUP BY pred_size ORDER BY COUNT(*) DESC
            """)
            size_rows = {r[0]: (r[1], r[2] or 0) for r in cur.fetchall()}

            cur.execute(f"""
                SELECT
                    CASE WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 9  THEN 'NHO'
                         WHEN (SELECT SUM(v::int) FROM json_array_elements_text(pr.actual_numbers::json) v) <= 11 THEN 'HOA'
                         ELSE 'LON' END AS actual_size,
                    COUNT(*) AS cnt
                FROM prediction_results pr
                JOIN predictions p ON p.id = pr.prediction_id
                WHERE {_TODAY_PRED} AND pr.actual_numbers IS NOT NULL
                GROUP BY actual_size
            """)
            size_actual_rows = {r[0]: r[1] for r in cur.fetchall()}

        # ── 3. W/L trail + streak ─────────────────────────────────
        cur.execute(f"""
            SELECT COALESCE(pr.is_win_size, pr.is_win, FALSE)
            FROM prediction_results pr
            WHERE {_TODAY_PRED}
            ORDER BY pr.draw_number ASC
        """)
        wl_seq = [r[0] for r in cur.fetchall()]

        # ── 4. Draws hôm nay ─────────────────────────────────────
        if _USE_PG:
            cur.execute("""
                SELECT draw_number, numbers FROM draw_history
                WHERE (draw_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                      = (NOW() AT TIME ZONE 'Asia/Ho_Chi_Minh')::date
                ORDER BY draw_number ASC
            """)
        else:
            cur.execute("""
                SELECT draw_number, numbers FROM draw_history
                WHERE date(datetime(draw_time, '+7 hours')) = date('now', '+7 hours')
                ORDER BY draw_number ASC
            """)
        draw_rows = cur.fetchall()

        # ── 5. Alerts hôm nay ────────────────────────────────────
        alert_count = 0
        if _USE_PG:
            cur.execute("""
                SELECT COUNT(*) FROM alert_log
                WHERE fired_at AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh'
                      >= NOW() AT TIME ZONE 'Asia/Ho_Chi_Minh' - INTERVAL '24 hours'
            """)
            alert_count = (cur.fetchone() or [0])[0] or 0

        # ── 6. Rolling 100 ────────────────────────────────────────
        cur.execute("""
            SELECT COALESCE(pr.is_win_size, pr.is_win, FALSE)
            FROM prediction_results pr
            WHERE pr.actual_numbers IS NOT NULL
            ORDER BY pr.draw_number DESC LIMIT 100
        """)
        last100 = [r[0] for r in cur.fetchall()]
        wr100 = sum(1 for w in last100 if w) / len(last100) if last100 else None

        conn.close()

        # ── Process draws ─────────────────────────────────────────
        number_freq = _Counter2()
        pair_count = triple_count = alldiff_count = 0
        combo_draws_map: dict = defaultdict(list)

        for draw_num, nums_raw in draw_rows:
            try:
                nums = json.loads(nums_raw) if isinstance(nums_raw, str) else list(nums_raw)
                nums = [int(n) for n in nums]
                number_freq.update(nums)
                key = tuple(sorted(nums))
                combo_draws_map[key].append(int(draw_num))
                cnt_nums = _Counter2(nums)
                mx = max(cnt_nums.values()) if cnt_nums else 0
                if mx == 3:
                    triple_count += 1
                elif mx == 2:
                    pair_count += 1
                else:
                    alldiff_count += 1
            except Exception:
                continue

        total_draws_today = len(draw_rows)

        # Combo coverage
        all_combos      = list(itertools.combinations_with_replacement(range(1, 7), 3))
        appeared_combos = [c for c in all_combos if c in combo_draws_map]
        not_appr_combos = [c for c in all_combos if c not in combo_draws_map]
        coverage_pct    = round(len(appeared_combos) / 56 * 100, 1)
        top_appeared    = sorted(appeared_combos, key=lambda c: -len(combo_draws_map[c]))[:3]
        top_not_appeared = not_appr_combos[:5]

        # Hot / cold
        hot_nums  = number_freq.most_common(3)
        cold_nums = sorted(((n, number_freq.get(n, 0)) for n in range(1, 7)), key=lambda x: x[1])[:3]

        # Streaks
        max_win = cur_win = max_loss = cur_loss = 0
        for w in wl_seq:
            if w:
                cur_win += 1; max_win = max(max_win, cur_win); cur_loss = 0
            else:
                cur_loss += 1; max_loss = max(max_loss, cur_loss); cur_win = 0

        cur_streak_val = 0
        cur_streak_type = None
        if wl_seq:
            last_val = wl_seq[-1]
            cur_streak_type = 'win' if last_val else 'loss'
            for w in reversed(wl_seq):
                if w == last_val:
                    cur_streak_val += 1
                else:
                    break

        trail = ''.join('✅' if w else '❌' for w in wl_seq[-20:])

        # ── Build message ─────────────────────────────────────────
        BASELINE    = 0.375
        diff_today  = (win_rate - BASELINE) * 100
        status_icon = '🔥' if win_rate >= 0.44 else ('✅' if win_rate >= BASELINE else '⚠️')

        _SL = {'NHO': '🔵NHỎ', 'HOA': '🟡HÒA', 'LON': '🔴LỚN'}
        size_line = ''
        if size_rows:
            parts = []
            for sz in ['NHO', 'HOA', 'LON']:
                if sz in size_rows:
                    t, w = size_rows[sz]
                    parts.append(f'{_SL[sz]} {w}/{t}({w/t*100:.0f}%)' if t else f'{_SL[sz]} 0')
            size_line = '📐 SIZE dự: ' + '  '.join(parts) + '\n'
        if size_actual_rows:
            act_total = sum(size_actual_rows.values()) or 1
            act_parts = [
                f'{_SL[sz]} {size_actual_rows.get(sz,0)}({size_actual_rows.get(sz,0)/act_total*100:.0f}%)'
                for sz in ['NHO', 'HOA', 'LON']
            ]
            size_line += '📐 SIZE thực: ' + '  '.join(act_parts) + '\n'

        combo_top_str  = '  '.join(
            f"{'-'.join(str(x) for x in c)}(×{len(combo_draws_map[c])})" for c in top_appeared
        )
        combo_miss_str = '  '.join('-'.join(str(x) for x in c) for c in top_not_appeared)
        combo_line = (
            f'🎰 Combo: <b>{len(appeared_combos)}/56</b> = {coverage_pct}%\n'
            f'  🔥 Ra nhiều: {combo_top_str if combo_top_str else "N/A"}\n'
            f'  ❄️ Chưa ra (top5): {combo_miss_str if combo_miss_str else "Tất cả đã ra!"}\n'
        )

        hot_str  = '  '.join(f'[{n}]×{c}' for n, c in hot_nums)
        cold_str = '  '.join(f'[{n}]×{c}' for n, c in cold_nums)
        freq_str = '  '.join(f'{n}:{number_freq.get(n,0)}' for n in range(1, 7))
        hotcold_line = (
            f'🔥 Số nóng: {hot_str}\n'
            f'❄️ Số lạnh: {cold_str}\n'
            f'📊 Tần suất 1-6: {freq_str}\n'
        )

        p_pct = pair_count / total_draws_today * 100 if total_draws_today else 0
        t_pct = triple_count / total_draws_today * 100 if total_draws_today else 0
        pair_line = f'🔗 Pair: {pair_count}({p_pct:.0f}%)  Triple: {triple_count}({t_pct:.0f}%)  AllDiff: {alldiff_count}\n'

        wr100_line = (
            f'📈 Rolling 100: <b>{wr100*100:.1f}%</b> {"✅" if wr100 >= BASELINE else "⚠️"}\n'
            if wr100 else ''
        )

        streak_line = ''
        if cur_streak_type == 'loss' and cur_streak_val >= 2:
            streak_line = f'❄️ Đang thua <b>{cur_streak_val} kỳ cuối</b>\n'
        elif cur_streak_type == 'win' and cur_streak_val >= 3:
            streak_line = f'🔥 Đang thắng <b>{cur_streak_val} kỳ cuối</b>\n'

        alert_line = ''
        if alert_count > 0:
            al_icon    = '🔴' if alert_count >= 5 else ('🟡' if alert_count >= 2 else '🔵')
            alert_line = f'{al_icon} Alerts hôm nay: <b>{alert_count}</b>\n'

        sign = '+' if diff_today >= 0 else ''
        msg = (
            f'📋 <b>SƠ KẾT 22H — {date_str}</b>\n'
            f'━━━━━━━━━━━━━━━━━━━\n'
            f'🎯 Dự đoán: <b>{total} kỳ</b>  |  Lượt quay: <b>{total_draws_today}</b>\n'
            f'{status_icon} Thắng SIZE: <b>{wins}/{total}</b> = <b>{win_rate*100:.1f}%</b>'
            f' ({sign}{diff_today:.1f}% vs 37.5%)\n'
            f'🎲 Đúng tổng: {sum_wins}/{total} ({sum_win_rate*100:.1f}%)\n'
            f'━━━━━━━━━━━━━━━━━━━\n'
            f'{size_line}'
            f'{wr100_line}'
            f'🔥 WIN streak max: <b>{max_win}</b>  ❄️ LOSS streak max: <b>{max_loss}</b>\n'
            f'{streak_line}'
            f'━━━━━━━━━━━━━━━━━━━\n'
            f'{combo_line}'
            f'{hotcold_line}'
            f'{pair_line}'
            f'━━━━━━━━━━━━━━━━━━━\n'
            f'{alert_line}'
            f'20 kỳ cuối: {trail}\n'
            f'🕙 {now_vn.strftime("%H:%M")} VN · Còn ~2h đến tổng kết'
        )

        tg   = TelegramBot()
        sent = tg.send_message(msg)

        chat_id = config.TELEGRAM_CHAT_ID or ''
        chat_hint = f"...{chat_id[-6:]}" if len(chat_id) > 6 else (chat_id or 'NOT_SET')
        token_ok  = bool(config.TELEGRAM_BOT_TOKEN)

        return jsonify({
            'status':            'ok',
            'sent':              sent,
            'date':              date_str,
            'total_predictions': total,
            'wins':              wins,
            'win_rate':          round(win_rate, 4),
            'total_draws_today': total_draws_today,
            'combo_coverage':    f'{len(appeared_combos)}/56',
            'coverage_pct':      coverage_pct,
            'msg_len':           len(msg),
            'msg_preview':       msg[:300],
            'debug_chat_id':     chat_hint,
            'debug_token_set':   token_ok,
        })

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/test-telegram')
@limiter.limit("5 per minute")
def test_telegram():
    """Debug: kiểm tra bot token + chat_id bằng cách gọi thẳng Telegram API."""
    import requests as _req
    token   = config.TELEGRAM_BOT_TOKEN or ''
    chat_id = config.TELEGRAM_CHAT_ID   or ''

    if not token:
        return jsonify({'error': 'TELEGRAM_BOT_TOKEN not set'}), 500

    base = f"https://api.telegram.org/bot{token}"

    # 1. getMe — verify token
    try:
        r_me = _req.get(f"{base}/getMe", timeout=8)
        me   = r_me.json()
    except Exception as e:
        return jsonify({'error': f'getMe failed: {e}'}), 500

    if not me.get('ok'):
        return jsonify({'step': 'getMe', 'ok': False, 'telegram_error': me}), 200

    bot_info = me.get('result', {})

    # 2. sendMessage — gửi tin test
    if not chat_id:
        return jsonify({'step': 'sendMessage', 'ok': False,
                        'error': 'TELEGRAM_CHAT_ID not set',
                        'bot': bot_info.get('username')}), 200

    try:
        r_send = _req.post(f"{base}/sendMessage",
                           json={'chat_id': chat_id, 'text': '🔧 Bingo18 test — bot hoạt động!'},
                           timeout=8)
        send_result = r_send.json()
    except Exception as e:
        return jsonify({'step': 'sendMessage', 'error': str(e)}), 500

    return jsonify({
        'bot_username':  bot_info.get('username'),
        'bot_id':        bot_info.get('id'),
        'chat_id_hint':  f"...{chat_id[-6:]}" if len(chat_id) > 6 else chat_id,
        'send_ok':       send_result.get('ok'),
        'send_error':    send_result.get('description') if not send_result.get('ok') else None,
        'send_error_code': send_result.get('error_code') if not send_result.get('ok') else None,
    })


def create_app():
    import logging
    import threading
    logging.basicConfig(
        level=getattr(logging, config.LOG_LEVEL.upper(), logging.INFO),
        format="%(asctime)s [%(levelname)s] %(name)s - %(message)s")

    try:
        import admin_interface  # noqa — registers /admin and /api/admin/* routes
    except Exception as e:
        logging.getLogger(__name__).warning("admin_interface load failed: %s", e)

    def _warm_transition_cache():
        try:
            from prediction_service import _ensure_transition_cache
            conn = db.get_connection()
            cur  = conn.cursor()
            cur.execute("SELECT MAX(draw_number) FROM draw_history")
            latest = cur.fetchone()[0] or 0
            conn.close()
            _ensure_transition_cache(db, latest)
        except Exception as e:
            logging.getLogger(__name__).warning("Transition cache warm-up failed: %s", e)

    threading.Thread(target=_warm_transition_cache, daemon=True).start()
    return app


def run_dashboard():
    try:
        import admin_interface  # noqa
    except Exception:
        pass
    print(f"Dashboard: http://localhost:{config.DASHBOARD_PORT}")
    app.run(host=config.DASHBOARD_HOST, port=config.DASHBOARD_PORT, debug=False, threaded=True)


if __name__ == '__main__':
    run_dashboard()
